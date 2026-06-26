"""
Modulo Importazione Libro Matricola / Stato di Rischio.

Permette di caricare file Excel/CSV con i veicoli di una flotta (libro matricola)
e di mappare manualmente le colonne sorgente ai campi della nostra entità Veicolo.

Flusso:
  1. POST /api/import/libro-matricola/preview  → upload file, ritorna header + 5 righe preview
  2. POST /api/import/libro-matricola/commit   → riceve mapping (column_name → field_name) e importa

Campi target (Veicolo):
  - targa            (OBBLIGATORIO)
  - data_entrata     (OBBLIGATORIO, data inizio rischio)
  - proprietario     (OBBLIGATORIO)
  - data_uscita      (opzionale)
  - marca, modello, telaio, alimentazione
  - cilindrata, kw, cv_fiscali, quintali, posti
  - classe_cu (classe di merito)
  - valore_assicurato, valore_veicolo, valore_accessori
  - massimale, franchigia
  - settore, uso, provincia
  - codice_oggetto, codice_uso, codice_iur
  - rimorchio, sgombraneve (flag)
  - premio_rca, premio_furto, premio_kasko, premio_tutela_legale, premio_infortuni
  - polizza_id (link opzionale a polizza esistente)
  - note
"""
from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import load_workbook


# Definizione dei campi target e meta-informazioni
CAMPI_VEICOLO: dict[str, dict] = {
    "targa":             {"label": "Targa",                 "required": True,  "type": "str"},
    "data_entrata":      {"label": "Data Inizio Rischio",   "required": True,  "type": "date"},
    "proprietario":      {"label": "Proprietario",          "required": True,  "type": "str"},
    "data_uscita":       {"label": "Data Fine Rischio",     "required": False, "type": "date"},
    "marca":             {"label": "Marca",                 "required": False, "type": "str"},
    "modello":           {"label": "Modello",               "required": False, "type": "str"},
    "telaio":            {"label": "Telaio",                "required": False, "type": "str"},
    "alimentazione":     {"label": "Alimentazione",         "required": False, "type": "str"},
    "data_immatricolazione": {"label": "Data Immatricolazione", "required": False, "type": "date"},
    "cilindrata":        {"label": "Cilindrata",            "required": False, "type": "int"},
    "kw":                {"label": "KW",                    "required": False, "type": "float"},
    "cv_fiscali":        {"label": "CV Fiscali",            "required": False, "type": "int"},
    "quintali":          {"label": "Peso in Quintali",      "required": False, "type": "float"},
    "posti":             {"label": "Posti",                 "required": False, "type": "int"},
    "classe_cu":         {"label": "Classe C.U.",           "required": False, "type": "str"},
    "valore_assicurato": {"label": "Valore Assicurato",     "required": False, "type": "float"},
    "valore_veicolo":    {"label": "Valore Veicolo",        "required": False, "type": "float"},
    "valore_accessori":  {"label": "Valore Accessori",      "required": False, "type": "float"},
    "massimale":         {"label": "Massimale",             "required": False, "type": "float"},
    "franchigia":        {"label": "Franchigia",            "required": False, "type": "float"},
    "settore":           {"label": "Settore RCA",           "required": False, "type": "str"},
    "uso":               {"label": "Uso",                   "required": False, "type": "str"},
    "provincia":         {"label": "Provincia",             "required": False, "type": "str"},
    "codice_oggetto":    {"label": "Codice Oggetto",        "required": False, "type": "str"},
    "codice_uso":        {"label": "Codice Uso",            "required": False, "type": "str"},
    "codice_iur":        {"label": "Codice IUR",            "required": False, "type": "str"},
    "rimorchio":         {"label": "Rimorchio (S/N)",       "required": False, "type": "bool"},
    "sgombraneve":       {"label": "Sgombraneve (S/N)",     "required": False, "type": "bool"},
    "premio_rca":        {"label": "Premio RCA",            "required": False, "type": "float"},
    "premio_furto":      {"label": "Premio Furto",          "required": False, "type": "float"},
    "premio_kasko":      {"label": "Premio Kasko",          "required": False, "type": "float"},
    "premio_tutela_legale": {"label": "Premio Tutela Legale", "required": False, "type": "float"},
    "premio_infortuni":  {"label": "Premio Infortuni",      "required": False, "type": "float"},
    "note":              {"label": "Note",                  "required": False, "type": "str"},
}

# Suggerimenti automatici: header sorgente (lowercase) -> field name
HEADER_HINTS: dict[str, str] = {
    "targa":                                 "targa",
    "data entrata":                          "data_entrata",
    "data inizio":                           "data_entrata",
    "data inizio rischio":                   "data_entrata",
    "data uscita":                           "data_uscita",
    "data fine":                             "data_uscita",
    "data fine rischio":                     "data_uscita",
    "den proprietario":                      "proprietario",
    "proprietario":                          "proprietario",
    "intestatario":                          "proprietario",
    "marca":                                 "marca",
    "modello":                               "modello",
    "telaio":                                "telaio",
    "alimentazione":                         "alimentazione",
    "data immatricolazione":                 "data_immatricolazione",
    "cilindrata":                            "cilindrata",
    "kw":                                    "kw",
    "cavalli fiscali":                       "cv_fiscali",
    "cv fiscali":                            "cv_fiscali",
    "peso in quintali":                      "quintali",
    "quintali":                              "quintali",
    "posti":                                 "posti",
    "classe c.u.":                           "classe_cu",
    "classe cu":                             "classe_cu",
    "classe di merito":                      "classe_cu",
    "valore assicurato":                     "valore_assicurato",
    "valore veicolo":                        "valore_veicolo",
    "valore accessori":                      "valore_accessori",
    "codice massimale":                      "massimale",
    "massimale":                             "massimale",
    "importo di franchigia":                 "franchigia",
    "franchigia":                            "franchigia",
    "settore":                               "settore",
    "uso":                                   "uso",
    "provincia":                             "provincia",
    "codice oggetto":                        "codice_oggetto",
    "codice uso":                            "codice_uso",
    "codice iur":                            "codice_iur",
    "rimorchio si/no":                       "rimorchio",
    "sgombraneve si/no":                     "sgombraneve",
    "rata rca":                              "premio_rca",
    "rata furto":                            "premio_furto",
    "rata kasko":                            "premio_kasko",
    "rata tutela legale":                    "premio_tutela_legale",
    "rata infortuni":                        "premio_infortuni",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _uid() -> str:
    return str(uuid.uuid4())


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_header(h: str) -> str:
    """Normalizza header: lowercase, single space, no newline."""
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def _parse_date(value: Any) -> Optional[str]:
    """Ritorna ISO date YYYY-MM-DD oppure None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value).strip()
    if not s or s == "-":
        return None
    # prova vari formati
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(".", "").replace(",", ".")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(value: Any) -> Optional[int]:
    f = _parse_float(value)
    return int(f) if f is not None else None


def _parse_bool(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    s = str(value).strip().upper()
    if s in ("S", "SI", "Y", "YES", "1", "TRUE"):
        return True
    if s in ("N", "NO", "0", "FALSE"):
        return False
    return None


def _cast_value(raw: Any, field: str) -> Any:
    meta = CAMPI_VEICOLO.get(field, {})
    t = meta.get("type", "str")
    if t == "date":
        return _parse_date(raw)
    if t == "float":
        return _parse_float(raw)
    if t == "int":
        return _parse_int(raw)
    if t == "bool":
        return _parse_bool(raw)
    s = str(raw).strip() if raw is not None else ""
    return s or None


# ---------------------------------------------------------------------------
# File parsing (XLSX + CSV)
# ---------------------------------------------------------------------------
def _read_xlsx(file_bytes: bytes) -> tuple[list[str], list[list[Any]]]:
    """Legge prima sheet. Cerca la prima riga "header" (con almeno 3 celle non vuote
    che contengono parole tipiche da intestazione)."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    # Trova la riga header: prima riga con almeno 3 celle non vuote.
    # Le righe sopra (titolo, sottotitolo, ecc.) vengono ignorate.
    header_idx = 0
    for i, row in enumerate(all_rows):
        non_empty = sum(1 for c in row if c not in (None, ""))
        if non_empty >= 3:
            header_idx = i
            break
    header = [_norm_header(c) for c in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]
    data = [list(r) for r in data_rows if any(c not in (None, "") for c in r)]
    return header, data


def _read_csv_bytes(file_bytes: bytes) -> tuple[list[str], list[list[Any]]]:
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    # auto-detect delimiter
    sample = text[:2048]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return [], []
    header = [_norm_header(c) for c in rows[0]]
    data = [r for r in rows[1:] if any((c or "").strip() for c in r)]
    return header, data


def parse_file(file_bytes: bytes, filename: str) -> tuple[list[str], list[list[Any]]]:
    """Parsa file libro matricola. Auto-detect XLSX / CSV."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _read_xlsx(file_bytes)
    return _read_csv_bytes(file_bytes)


# ---------------------------------------------------------------------------
# Auto-mapping suggestions
# ---------------------------------------------------------------------------
def suggest_mapping(headers: list[str]) -> dict[str, str]:
    """Ritorna {header_originale: field_target} basato su HEADER_HINTS."""
    out: dict[str, str] = {}
    for h in headers:
        key = _norm_header(h)
        if key in HEADER_HINTS:
            out[h] = HEADER_HINTS[key]
        else:
            # fuzzy: cerca hint contenuto
            for hint, field in HEADER_HINTS.items():
                if hint in key or key in hint:
                    out[h] = field
                    break
    return out


# ---------------------------------------------------------------------------
# Preview / Commit
# ---------------------------------------------------------------------------
def build_preview(file_bytes: bytes, filename: str, max_rows: int = 5) -> dict:
    headers, data = parse_file(file_bytes, filename)
    return {
        "headers": headers,
        "preview_rows": data[:max_rows],
        "total_rows": len(data),
        "suggested_mapping": suggest_mapping(headers),
        "campi_target": [
            {"field": k, "label": v["label"], "required": v.get("required", False), "type": v.get("type", "str")}
            for k, v in CAMPI_VEICOLO.items()
        ],
    }


async def commit_import(db, file_bytes: bytes, filename: str,
                        mapping: dict[str, str],
                        utente: dict,
                        polizza_id: Optional[str] = None) -> dict:
    """Importa veicoli applicando il mapping {header_originale -> field_target}.

    Regole:
      - Validazione campi obbligatori (targa, data_entrata, proprietario)
      - Se esiste già un veicolo con stessa targa: aggiornamento (merge non distruttivo).
      - Se polizza_id è fornito, aggiunge il veicolo all'array polizza.veicoli_ids.
    """
    headers, data = parse_file(file_bytes, filename)
    # Normalizza il mapping con header normalizzati
    norm_mapping: dict[str, str] = {}
    for h, field in mapping.items():
        if not field:
            continue
        norm_h = _norm_header(h)
        norm_mapping[norm_h] = field
    norm_headers = [_norm_header(h) for h in headers]
    # Verifica obbligatori
    required = [f for f, meta in CAMPI_VEICOLO.items() if meta.get("required")]
    mapped_fields = set(norm_mapping.values())
    missing_req = [f for f in required if f not in mapped_fields]

    stats = {
        "totale": len(data),
        "creati": 0,
        "aggiornati": 0,
        "scartati": 0,
        "errori": [],
        "missing_required_fields": missing_req,
    }
    if missing_req:
        return stats  # blocco preventivo: l'utente deve mappare i required

    veicoli_ids_da_aggiungere: list[str] = []

    for idx, row in enumerate(data, start=2):  # 2 perché riga 1 è header
        # Costruisci dict {field: value}
        record: dict[str, Any] = {}
        for col_idx, h in enumerate(norm_headers):
            field = norm_mapping.get(h)
            if not field:
                continue
            raw = row[col_idx] if col_idx < len(row) else None
            val = _cast_value(raw, field)
            if val is not None:
                record[field] = val
        # Validazione required
        miss = [f for f in required if not record.get(f)]
        if miss:
            stats["scartati"] += 1
            stats["errori"].append(f"Riga {idx}: mancano {miss}")
            continue
        # Upsert per targa
        targa = (record["targa"] or "").upper().strip()
        record["targa"] = targa
        existing = await db.veicoli.find_one({"targa": targa}, {"_id": 0, "id": 1})
        if existing:
            record["updated_at"] = _now_iso()
            await db.veicoli.update_one({"id": existing["id"]}, {"$set": record})
            stats["aggiornati"] += 1
            veicoli_ids_da_aggiungere.append(existing["id"])
        else:
            record["id"] = _uid()
            record["created_at"] = _now_iso()
            record["updated_at"] = _now_iso()
            record["fonte"] = "import_libro_matricola"
            await db.veicoli.insert_one(record)
            stats["creati"] += 1
            veicoli_ids_da_aggiungere.append(record["id"])

    # Log import
    await db.import_logs.insert_one({
        "id": _uid(),
        "utente_id": utente.get("id"),
        "nome_file": filename,
        "flusso": "libro_matricola",
        "stato": "completato" if not stats["errori"] else "completato_con_warning",
        "polizze_create": 0,
        "polizze_aggiornate": 0,
        "anagrafiche_create": 0,
        "anagrafiche_aggiornate": 0,
        "titoli_creati": 0,
        "sinistri_creati": 0,
        "veicoli_creati": stats["creati"],
        "veicoli_aggiornati": stats["aggiornati"],
        "record_skipped": [{"riga": e.split(":")[0], "motivo": e} for e in stats["errori"][:50]],
        "entita_non_mappate": {},
        "record_types_processati": {"veicoli": stats["totale"]},
        "errori": stats["errori"][:50],
        "durata_ms": 0,
        "created_at": _now_iso(),
    })

    # Eventuale link a polizza esistente
    if polizza_id and veicoli_ids_da_aggiungere:
        await db.polizze.update_one(
            {"id": polizza_id},
            {"$addToSet": {"veicoli_ids": {"$each": veicoli_ids_da_aggiungere}}},
        )

    return stats
