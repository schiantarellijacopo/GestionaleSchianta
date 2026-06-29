"""Marketing — Voucher compagnia + Newsletter + Liste Lead.

Voucher: la compagnia fornisce codici sconto anonimi che l'agenzia assegna ai clienti.
Newsletter: campagne email/sms/whatsapp a liste segmentate di clienti.
Liste Lead: liste di prospect (importate da Excel/CSV forniti dalle compagnie)
            con dispatch WhatsApp/Email diretto.
"""
from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel, Field

from auth import current_user, require_user
from database import db
from db_models import _now_iso


router = APIRouter()


# ============= VOUCHER COMPAGNIA =============
class VoucherBody(BaseModel):
    codice: str
    compagnia_id: Optional[str] = None
    descrizione: Optional[str] = None
    valore: float = 0
    tipo_valore: str = "euro"  # euro | percentuale
    valido_dal: Optional[str] = None
    valido_al: Optional[str] = None
    ramo: Optional[str] = None
    assegnato_a: Optional[str] = None  # anagrafica_id (cliente)
    assegnato_a_collaboratore: Optional[str] = None  # user_id (collaboratore)
    note: Optional[str] = None
    usato: bool = False
    data_uso: Optional[str] = None


@router.get("/voucher")
async def list_voucher(
    stato: Optional[str] = None,  # disponibile | assegnato | usato
    compagnia_id: Optional[str] = None,
    user=Depends(current_user),
) -> list[dict]:
    flt: dict = {}
    if compagnia_id: flt["compagnia_id"] = compagnia_id
    if stato == "disponibile":
        flt["assegnato_a"] = None; flt["assegnato_a_collaboratore"] = None; flt["usato"] = False
    elif stato == "assegnato":
        flt["$or"] = [{"assegnato_a": {"$ne": None}}, {"assegnato_a_collaboratore": {"$ne": None}}]
        flt["usato"] = False
    elif stato == "usato":
        flt["usato"] = True
    items = await db.voucher.find(flt, {"_id": 0}).sort("created_at", -1).to_list(5000)
    ana_ids = list({v["assegnato_a"] for v in items if v.get("assegnato_a")})
    anas = {a["id"]: a async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1, "cognome": 1, "nome": 1})}
    coll_ids = list({v["assegnato_a_collaboratore"] for v in items if v.get("assegnato_a_collaboratore")})
    colls = {u["id"]: u async for u in db.users.find(
        {"id": {"$in": coll_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1})}
    comp_ids = list({v["compagnia_id"] for v in items if v.get("compagnia_id")})
    comps = {c["id"]: c async for c in db.compagnie.find(
        {"id": {"$in": comp_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    for v in items:
        a = anas.get(v.get("assegnato_a"), {})
        v["assegnato_a_nome"] = a.get("ragione_sociale") or f"{a.get('cognome','')} {a.get('nome','')}".strip()
        c = colls.get(v.get("assegnato_a_collaboratore"), {})
        v["assegnato_a_collaboratore_nome"] = c.get("name") or c.get("email")
        v["compagnia_nome"] = comps.get(v.get("compagnia_id"), {}).get("ragione_sociale")
    return items


@router.post("/voucher", status_code=201)
async def create_voucher(body: VoucherBody,
                          user=Depends(require_user("admin", "collaboratore"))) -> dict:
    if not body.codice.strip():
        raise HTTPException(400, "Codice obbligatorio")
    if await db.voucher.find_one({"codice": body.codice}):
        raise HTTPException(400, "Codice voucher già esistente")
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": _now_iso()}
    await db.voucher.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/voucher/{vid}")
async def update_voucher(vid: str, body: VoucherBody,
                          user=Depends(require_user("admin", "collaboratore"))) -> dict:
    data = body.model_dump()
    data["updated_at"] = _now_iso()
    res = await db.voucher.update_one({"id": vid}, {"$set": data})
    if res.matched_count == 0:
        raise HTTPException(404, "Voucher non trovato")
    return await db.voucher.find_one({"id": vid}, {"_id": 0})


@router.delete("/voucher/{vid}")
async def delete_voucher(vid: str, user=Depends(require_user("admin"))) -> dict:
    res = await db.voucher.delete_one({"id": vid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Voucher non trovato")
    return {"ok": True}


@router.post("/voucher/{vid}/assegna")
async def assegna_voucher(vid: str, body: dict,
                          user=Depends(require_user("admin", "collaboratore"))) -> dict:
    """Assegna un voucher a un cliente (anagrafica_id) e/o a un collaboratore (collaboratore_id).
    Almeno uno dei due deve essere presente. Si possono assegnare entrambi contemporaneamente."""
    ana_id = body.get("anagrafica_id") or None
    coll_id = body.get("collaboratore_id") or None
    if not ana_id and not coll_id:
        raise HTTPException(400, "Indicare almeno un cliente o un collaboratore")
    upd = {"data_assegnazione": _now_iso()}
    if ana_id:
        upd["assegnato_a"] = ana_id
    if coll_id:
        upd["assegnato_a_collaboratore"] = coll_id
    await db.voucher.update_one({"id": vid}, {"$set": upd})
    return await db.voucher.find_one({"id": vid}, {"_id": 0})


@router.post("/voucher/bulk-import")
async def bulk_import_voucher(body: dict,
                              user=Depends(require_user("admin", "collaboratore"))) -> dict:
    """Importa una lista di voucher anonimi (codici).

    Body: ``{compagnia_id, ramo?, valore, tipo_valore, valido_dal, valido_al, codici: [str]}``
    """
    codici = body.get("codici") or []
    if not codici:
        raise HTTPException(400, "Lista codici vuota")
    n = 0
    skipped = 0
    for c in codici:
        c = (c or "").strip()
        if not c: continue
        if await db.voucher.find_one({"codice": c}):
            skipped += 1; continue
        await db.voucher.insert_one({
            "id": str(uuid.uuid4()),
            "codice": c,
            "compagnia_id": body.get("compagnia_id"),
            "ramo": body.get("ramo"),
            "valore": float(body.get("valore") or 0),
            "tipo_valore": body.get("tipo_valore") or "euro",
            "valido_dal": body.get("valido_dal"),
            "valido_al": body.get("valido_al"),
            "descrizione": body.get("descrizione"),
            "usato": False,
            "created_at": _now_iso(),
        })
        n += 1
    return {"creati": n, "duplicati": skipped}


# ============= NEWSLETTER =============
class NewsletterBody(BaseModel):
    nome: str
    oggetto: str
    contenuto: str  # HTML / testo
    canale: str = "email"  # email | sms | whatsapp
    target_filtro: dict = Field(default_factory=dict)  # {tags?, ramo?, compagnia_id?}
    stato: str = "bozza"  # bozza | inviata | programmata
    data_programmata: Optional[str] = None


@router.get("/newsletter")
async def list_newsletter(user=Depends(current_user)) -> list[dict]:
    return await db.newsletter.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)


@router.post("/newsletter", status_code=201)
async def create_newsletter(body: NewsletterBody,
                            user=Depends(require_user("admin", "collaboratore"))) -> dict:
    doc = {"id": str(uuid.uuid4()), **body.model_dump(),
           "destinatari_calcolati": 0, "destinatari_inviati": 0,
           "created_at": _now_iso(), "created_by": user.get("id")}
    # Pre-calcolo destinatari
    flt = await _build_target_filter(body.target_filtro)
    doc["destinatari_calcolati"] = await db.anagrafiche.count_documents(flt)
    await db.newsletter.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/newsletter/{nid}")
async def update_newsletter(nid: str, body: NewsletterBody,
                            user=Depends(require_user("admin", "collaboratore"))) -> dict:
    data = body.model_dump()
    data["updated_at"] = _now_iso()
    flt = await _build_target_filter(body.target_filtro)
    data["destinatari_calcolati"] = await db.anagrafiche.count_documents(flt)
    res = await db.newsletter.update_one({"id": nid}, {"$set": data})
    if res.matched_count == 0:
        raise HTTPException(404, "Newsletter non trovata")
    return await db.newsletter.find_one({"id": nid}, {"_id": 0})


@router.delete("/newsletter/{nid}")
async def delete_newsletter(nid: str, user=Depends(require_user("admin"))) -> dict:
    res = await db.newsletter.delete_one({"id": nid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Newsletter non trovata")
    return {"ok": True}


async def _build_target_filter(target: dict) -> dict:
    flt: dict = {"consenso_commerciale": True}
    if target.get("tags"):
        flt["tags"] = {"$in": target["tags"] if isinstance(target["tags"], list) else [target["tags"]]}
    if target.get("tipo"):
        flt["tipo"] = target["tipo"]
    return flt


@router.post("/newsletter/{nid}/invia")
async def invia_newsletter(nid: str,
                            user=Depends(require_user("admin"))) -> dict:
    """Marca la newsletter come inviata e simula invio (in produzione qui
    si dispatcia su Resend/Twilio in base al canale)."""
    nl = await db.newsletter.find_one({"id": nid}, {"_id": 0})
    if not nl:
        raise HTTPException(404, "Newsletter non trovata")
    flt = await _build_target_filter(nl.get("target_filtro") or {})
    destinatari = await db.anagrafiche.find(flt, {"_id": 0, "id": 1, "email": 1,
                                                   "cellulare": 1}).to_list(50000)
    # Log invio nel Diario di ogni cliente
    for d in destinatari:
        await db.diario_cliente.insert_one({
            "id": str(uuid.uuid4()),
            "anagrafica_id": d["id"],
            "tipo": "newsletter_inviata",
            "data": _now_iso(),
            "operatore_id": user.get("id"),
            "contenuto": f"Newsletter '{nl['nome']}' inviata via {nl['canale']}",
            "fonte": "marketing",
        })
    await db.newsletter.update_one({"id": nid}, {"$set": {
        "stato": "inviata",
        "destinatari_inviati": len(destinatari),
        "data_invio": _now_iso(),
    }})
    return {"ok": True, "destinatari": len(destinatari)}


# ============= LISTE LEAD (import Excel/CSV) =============
# Schema generico: nome | cognome | codice_fiscale | email | telefono | cellulare | citta | note
# La colonna è case-insensitive, accetta sinonimi comuni.

COL_ALIASES = {
    "nome": ["nome", "first_name", "firstname", "first name", "name"],
    "cognome": ["cognome", "last_name", "lastname", "last name", "surname"],
    "nome_completo_unico": ["contatto", "nominativo", "cliente", "denominazione"],  # "PEPE ALFONSO" → cognome+nome
    "codice_fiscale": ["codice fiscale", "codice_fiscale", "cf", "fiscal code", "fiscalcode"],
    "email": ["email", "e-mail", "mail", "posta", "indirizzo email"],
    "telefono": ["telefono", "phone", "tel", "fisso"],
    "cellulare": ["cellulare", "cell", "mobile", "telefono cellulare", "cellphone", "whatsapp"],
    "citta": ["città", "citta", "city", "comune", "comune di residenza"],
    "note": ["note", "notes", "descrizione", "annotazioni"],
    "ragione_sociale": ["ragione sociale", "ragione_sociale", "azienda", "company"],
    "data_nascita": ["data nascita", "data_nascita", "birthdate", "data di nascita"],
    "indirizzo": ["indirizzo", "address", "via"],
    "tipo": ["tipo", "categoria", "segmento lista", "partizionamento cliente"],
    "professione": ["professione", "professione/attività", "professione/attivita", "occupazione", "attività"],
    "eta": ["età", "eta", "anni"],
    "codice_agente": ["cod. age.", "cod age", "codice agente", "punto vendita", "rhx", "cod. age. - rhx"],
    "esito": ["esito", "stato attivazione", "stato"],
    "cross_selling": ["cross selling", "cross-selling"],
    # === RHX specific (Cattolica/Generali iniziativa AutoConvenienTe & DNA senza RCA) ===
    "id_contatto": ["idcontatto", "id contatto"],
    "id_gestore": ["idgestore", "id gestore"],
    "gestore_contatto": ["gestore contatto", "gestore"],
    "canale": ["canale"],
    "stato_comunicazione": ["stato comunicazione"],
    "data_scadenza_contatto": ["data scadenza contatto"],
    "codice_sag": ["codice sag", "sag"],
    "privacy_commerciale": ["privacy commerciale"],
    "privacy_posta_telefono": ["privacy posta/telefono", "privacy posta telefono"],
    "privacy_email_sms": ["privacy email/sms", "privacy email sms"],
    "livello_piu_generali": ["livello più generali", "livello piu generali"],
    "stato_ultimo_puc": ["stato ultimo puc"],
    "profilo_cliente": ["profilo cliente"],
    "esito_comunicazione": ["esito comunicazione"],
    "azione_cliente": ["azione cliente"],
    "esito_direzionale": ["esito direzionale"],
    "esito_contatto": ["esito contatto"],
    "esito_trattativa": ["esito trattativa"],
    "attivazione_s": ["attivaz.", "attivaz. - [s]", "attivazione"],
    "aggiorna_attivazione": ["aggiorna attivazione"],
    "aggiorna_data_scadenza": ["aggiorna data scadenza contatto (nel formato gg/mm/aaaa)", "aggiorna data scadenza contatto"],
    "data_comunicazione": ["data comunicazione"],
    "data_azione_cliente": ["data azione cliente"],
    "data_direzionale": ["data direzionale"],
    "aggiorna_contatto": ["aggiorna contatto"],
    "data_contatto": ["data contatto"],
    "aggiorna_trattativa": ["aggiorna trattativa"],
    "data_trattativa": ["data trattativa"],
}


def _normalize_col(c: str) -> Optional[str]:
    if not c: return None
    c = str(c).strip().lower()
    for canonical, alts in COL_ALIASES.items():
        if c in alts or c == canonical:
            return canonical
    return None


def _parse_csv(content: bytes) -> List[dict]:
    txt = content.decode("utf-8", errors="ignore")
    # auto-detect separatore
    sample = txt[:2048]
    sep = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(txt), delimiter=sep)
    rows = []
    for r in reader:
        row = {}
        for k, v in r.items():
            canonical = _normalize_col(k)
            if canonical and v is not None and str(v).strip():
                row[canonical] = str(v).strip()
        if row:
            rows.append(row)
    return rows


def _parse_xlsx(content: bytes) -> List[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl non installato")
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    rows: List[dict] = []
    # Itera su TUTTI i fogli (es. file RHX con "AutoConvenienTe" + "DNA senza RCA")
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        cols = [_normalize_col(c) for c in header]
        if not any(cols):
            continue
        for r in rows_iter:
            if not r or all(v is None or str(v).strip() == "" for v in r):
                continue
            row = {"_sheet": ws.title}
            for i, v in enumerate(r):
                if i >= len(cols): break
                col = cols[i]
                if col and v is not None and str(v).strip():
                    row[col] = str(v).strip()
            if len(row) > 1:
                rows.append(row)
    return rows


def _digits_only(s: str) -> str:
    return re.sub(r"\D", "", s or "")


async def _match_anagrafica(row: dict) -> Optional[dict]:
    """Match euristico: CF → Email → cellulare/telefono."""
    cf = (row.get("codice_fiscale") or "").upper().strip()
    if cf:
        a = await db.anagrafiche.find_one({"codice_fiscale": cf}, {"_id": 0, "id": 1, "ragione_sociale": 1})
        if a: return a
    email = (row.get("email") or "").lower().strip()
    if email:
        a = await db.anagrafiche.find_one({"email": email}, {"_id": 0, "id": 1, "ragione_sociale": 1})
        if a: return a
    tel = _digits_only(row.get("cellulare") or row.get("telefono") or "")
    if tel and len(tel) >= 8:
        a = await db.anagrafiche.find_one({
            "$or": [{"cellulare": {"$regex": tel[-8:]}}, {"telefono": {"$regex": tel[-8:]}}],
        }, {"_id": 0, "id": 1, "ragione_sociale": 1})
        if a: return a
    return None


@router.post("/lead-liste/import")
async def import_lead_lista(
    nome: str = Form(...),
    note: Optional[str] = Form(None),
    fonte: Optional[str] = Form("import"),
    compagnia_id: Optional[str] = Form(None),
    file: UploadFile = File(...),
    user=Depends(require_user("admin", "collaboratore")),
) -> dict:
    """Importa una lista di lead da Excel (.xlsx) o CSV.

    Lo schema atteso accetta colonne con header (case-insensitive, sinonimi):
    Nome, Cognome, Codice Fiscale, Email, Telefono, Cellulare, Città, Note,
    Ragione Sociale, Data Nascita, Indirizzo, Tipo.

    Per ogni riga viene fatto matching automatico con le anagrafiche esistenti
    (CF → Email → telefono). I lead non-clienti vengono salvati in `lead`.
    """
    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith(".xlsx") or fname.endswith(".xlsm"):
        rows = _parse_xlsx(content)
    elif fname.endswith(".csv") or fname.endswith(".txt"):
        rows = _parse_csv(content)
    else:
        # tenta CSV se non riconosciuto
        try:
            rows = _parse_csv(content)
        except Exception:
            raise HTTPException(400, "Formato non supportato. Usa .xlsx o .csv")

    if not rows:
        raise HTTPException(400, "File vuoto o senza dati validi")

    # Post-process: se c'è solo "nome_completo_unico" (es. "PEPE ALFONSO" = cognome nome),
    # spezzalo in cognome + nome.
    for r in rows:
        nc = r.pop("nome_completo_unico", None)
        if nc and not r.get("nome") and not r.get("cognome"):
            parts = str(nc).strip().split()
            if len(parts) >= 2:
                r["cognome"] = parts[0]
                r["nome"] = " ".join(parts[1:])
            elif len(parts) == 1:
                r["cognome"] = parts[0]
        # Pulizia: i file Excel usano "-" come placeholder per "vuoto"
        for k in list(r.keys()):
            if r[k] in ("-", "--", "---"):
                r[k] = None
        # Email lowercase
        if r.get("email"):
            r["email"] = r["email"].lower().strip()
        # Privacy normalization: "S" → True, "N" → False
        for pk in ("privacy_commerciale", "privacy_posta_telefono", "privacy_email_sms"):
            v = r.get(pk)
            if v is not None:
                r[pk] = str(v).strip().upper() in ("S", "SI", "SÌ", "Y", "YES", "1", "TRUE")
        # Indirizzo RHX format: "VIA X 12-23030-CITTA-SO" → splitta in indirizzo + citta + cap + provincia
        ind = r.get("indirizzo")
        if ind and "-" in ind and not r.get("citta"):
            parts = [p.strip() for p in ind.split("-")]
            if len(parts) >= 3:
                r["indirizzo"] = parts[0]
                # cap (5 digits)
                for p in parts[1:]:
                    if p.isdigit() and len(p) == 5:
                        r["cap"] = p
                if len(parts) >= 3:
                    r["citta"] = parts[2] if len(parts) > 2 else parts[1]
                if len(parts) >= 4 and len(parts[-1]) == 2:
                    r["provincia"] = parts[-1].upper()

    lista_id = str(uuid.uuid4())
    matched_clienti, lead_creati = 0, 0
    leads_doc = []
    for r in rows:
        ana = await _match_anagrafica(r)
        if ana:
            matched_clienti += 1
            r["anagrafica_id"] = ana["id"]
            r["matched_nome"] = ana.get("ragione_sociale")
        else:
            lead_creati += 1
        nome_full = r.get("ragione_sociale") or f"{r.get('cognome','')} {r.get('nome','')}".strip() or r.get("email") or r.get("cellulare") or "—"
        leads_doc.append({
            "id": str(uuid.uuid4()),
            "lista_id": lista_id,
            "nome_completo": nome_full,
            **r,
            "created_at": _now_iso(),
        })

    await db.lead_liste.insert_one({
        "id": lista_id,
        "nome": nome,
        "note": note,
        "fonte": fonte,
        "compagnia_id": compagnia_id,
        "totale": len(rows),
        "matched_clienti": matched_clienti,
        "lead_creati": lead_creati,
        "created_at": _now_iso(),
        "created_by": user.get("id"),
    })
    if leads_doc:
        await db.lead.insert_many(leads_doc)
    return {
        "lista_id": lista_id,
        "totale": len(rows),
        "matched_clienti": matched_clienti,
        "lead_creati": lead_creati,
    }


@router.get("/lead-liste")
async def list_lead_liste(user=Depends(current_user)) -> list[dict]:
    return await db.lead_liste.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)


@router.get("/lead-liste/{lid}/lead")
async def list_lead_di_lista(
    lid: str,
    stato: Optional[str] = None,
    user=Depends(current_user),
) -> list[dict]:
    flt: dict = {"lista_id": lid}
    if stato == "matched": flt["anagrafica_id"] = {"$ne": None}
    if stato == "non_matched": flt["anagrafica_id"] = None
    return await db.lead.find(flt, {"_id": 0}).sort("nome_completo", 1).to_list(10000)


@router.delete("/lead-liste/{lid}")
async def delete_lead_lista(
    lid: str, user=Depends(require_user("admin")),
) -> dict:
    await db.lead.delete_many({"lista_id": lid})
    res = await db.lead_liste.delete_one({"id": lid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Lista non trovata")
    return {"ok": True}


class DispatchBody(BaseModel):
    lista_id: str
    canale: str  # "email" | "whatsapp"
    oggetto: Optional[str] = None  # solo email
    messaggio: str
    solo_matched: bool = False  # se True, manda solo a chi è in anagrafica


@router.post("/lead-liste/dispatch")
async def dispatch_lista(
    body: DispatchBody,
    user=Depends(require_user("admin", "collaboratore")),
) -> dict:
    """Spedisce un messaggio (email/whatsapp) a tutta la lista.

    Simulato: logga l'invio + crea un record in `dispatch_log`. Quando
    saranno disponibili le credenziali Twilio/Spoki/SMTP, l'invio reale
    può essere agganciato qui.
    """
    lista = await db.lead_liste.find_one({"id": body.lista_id}, {"_id": 0})
    if not lista:
        raise HTTPException(404, "Lista non trovata")
    flt: dict = {"lista_id": body.lista_id}
    if body.solo_matched: flt["anagrafica_id"] = {"$ne": None}
    leads = await db.lead.find(flt, {"_id": 0}).to_list(10000)
    if not leads:
        raise HTTPException(400, "Nessun lead in lista")
    inviati = 0
    fallliti = 0
    canale = body.canale.lower()
    for l in leads:
        # personalizzazione: sostituisce {nome} {cognome}
        msg = (body.messaggio
               .replace("{nome}", l.get("nome", ""))
               .replace("{cognome}", l.get("cognome", ""))
               .replace("{citta}", l.get("citta", "")))
        target = l.get("email") if canale == "email" else (l.get("cellulare") or l.get("telefono"))
        if not target:
            fallliti += 1; continue
        # log dispatch (mock: in produzione qui si chiama SMTP/Twilio/Spoki)
        await db.dispatch_log.insert_one({
            "id": str(uuid.uuid4()),
            "lista_id": body.lista_id,
            "lead_id": l["id"],
            "anagrafica_id": l.get("anagrafica_id"),
            "canale": canale,
            "target": target,
            "oggetto": body.oggetto,
            "messaggio": msg,
            "stato": "simulato",
            "data": _now_iso(),
            "operatore_id": user.get("id"),
        })
        inviati += 1
        # se collegato a anagrafica → diario
        if l.get("anagrafica_id"):
            await db.diario_cliente.insert_one({
                "id": str(uuid.uuid4()),
                "anagrafica_id": l["anagrafica_id"],
                "tipo": f"marketing_{canale}",
                "data": _now_iso(),
                "operatore_id": user.get("id"),
                "contenuto": f"Lista '{lista['nome']}' · {canale}: {msg[:200]}",
                "fonte": "lead_lista",
            })
    await db.lead_liste.update_one({"id": body.lista_id}, {"$set": {
        "ultimo_dispatch": _now_iso(),
        "ultimo_canale": canale,
        "ultimo_inviati": inviati,
    }})
    return {"ok": True, "inviati": inviati, "falliti": fallliti, "totale_lead": len(leads)}
