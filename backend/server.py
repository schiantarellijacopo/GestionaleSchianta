"""Main FastAPI application for Programma Assicurativo.

Tutti gli endpoint sono sotto /api.
"""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Optional, List, Literal
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Query, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from starlette.middleware.cors import CORSMiddleware

from database import client, db
from shared import (
    _MESI_PER_FRAZIONAMENTO,  # noqa: F401  (re-export)
    _MEZZO_TO_TIPO,           # noqa: F401  (re-export)
    _CORPO_LETTERA_DEFAULT,   # noqa: F401  (re-export)
    log_attivita,
    log_diario_cliente,
    strip_mongo_id,
    visibility_filter,
    calcola_scadenza_titolo as _calcola_scadenza_titolo,
    resolve_conto_cassa as _resolve_conto_cassa,
    intestazione_pdf as _intestazione_pdf,
    assert_giornata_aperta,
)
from auth import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    decode_token, get_token_from_request, require_user, current_user, can_see_all,
)
from db_models import (
    UserCreate, UserPublic, LoginRequest, Compagnia, Anagrafica, Polizza, Titolo,
    Sinistro, MovimentoContabile, Intervista, CalcoloPensione, EmailMessaggio,
    AttivitaLog, ImportLog, Banca, ContoCassa, ProdottoLibreria, RamoLibreria, ApplicazioneLibroMatricola,
    Allegato, DiarioVoce, MessaggioChat, Corso, ProgressoCorso, PagamentoProvvigioni, VoceManualeCollab,
    ChiusuraGiorno,
    AziendaConfig, SchemaProvvigionale, EventoCalendario, ContattoCompagnia,
    PipelineCustom, PipelineColonna, PipelineCard,
    AnalisiCliente,
    Rappel, VoceRicorsivaCollab, MezzoPagamento, TipoPagamento, LetteraAbbuono,
    DiarioNota, EmailInbox, DiarioCliente,
    _now_iso, _uid,
)
import ania_importer
import inps_calculator
import successione_calc
import reddito_calc
import storage as obj_storage
import pdf_report
import pdf_diagnosi
import pdf_brogliaccio
import pdf_lettera_abbuono
import pdf_avviso
import brogliaccio as brog
import cf_calc
import geocoder as geocoder_svc
import avvisi_scadenze
import imap_poller
from fastapi.responses import StreamingResponse
import io as _io

# ---------- DB ----------
# Shared MongoDB connection lives in database.py (imported above).
# Local aliases kept for backward-compat with the rest of this module.
mongo_url = os.environ["MONGO_URL"]

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="Programma Assicurativo")
api = APIRouter(prefix="/api")



# ---------- PDF helpers ----------
# NOTE(iter19): re-aggiunto dopo lo split parziale di server.py: era stato
# inavvertitamente rimosso durante il refactor e tutte le route /pdf/* andavano
# in NameError (es. POST /api/dashboard/stampa-titoli-sospesi). Vedi iter19
# bug report.
# _intestazione_pdf now lives in shared.py — see import at top of file



# ---------- Helpers (extracted to shared.py) ----------

# ============================================================
# AUTH
# ============================================================
@api.post("/auth/login")
async def login(payload: LoginRequest, request: Request, response: Response):
    email = payload.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    tid = user.get("agenzia_tenant_id")
    is_su = bool(user.get("is_super_admin"))
    access = create_access_token(user["id"], user["email"], user["role"],
                                 agenzia_tenant_id=tid, is_super_admin=is_su)
    refresh = create_refresh_token(user["id"])
    response.set_cookie("access_token", access, httponly=True, secure=False,
                        samesite="lax", max_age=60 * 60 * 8, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=False,
                        samesite="lax", max_age=60 * 60 * 24 * 7, path="/")
    user.pop("password_hash", None)
    user.pop("_id", None)
    await log_attivita(user, "login", "auth", user["id"], f"Login utente {email}")
    # Super admin login audit
    if is_su:
        try:
            from audit_super_admin import log_action as _sa_log
            await _sa_log(user=user, action_type="SUPER_ADMIN_LOGIN",
                          details=f"Login super_admin da IP", request=request)
        except Exception:
            pass
    return {"user": user, "access_token": access}


@api.post("/auth/logout")
async def logout(response: Response, user=Depends(current_user)):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    await log_attivita(user, "logout", "auth", user["id"], "Logout")
    return {"ok": True}


@api.get("/auth/me")
async def auth_me(user=Depends(current_user)):
    return user


@api.post("/auth/users", status_code=201)
async def create_user(payload: UserCreate, user=Depends(require_user("admin"))):
    email = payload.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email già registrata")
    doc = UserPublic(email=email, name=payload.name, role=payload.role,
                     anagrafica_id=payload.anagrafica_id).model_dump()
    doc["password_hash"] = hash_password(payload.password)
    await db.users.insert_one(doc)
    doc.pop("password_hash", None)
    doc.pop("_id", None)
    await log_attivita(user, "create", "user", doc["id"], f"Creato utente {email}")
    return doc


@api.get("/auth/users")
async def list_users(user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    # Visibility: solo admin vede tutti. Gli altri vedono solo se stessi.
    flt: dict = {}
    if user["role"] != "admin":
        flt["id"] = user["id"]
    users = await db.users.find(flt, {"password_hash": 0, "_id": 0}).to_list(500)
    return users


@api.put("/auth/users/{uid}")
async def update_user(uid: str, body: dict, user=Depends(require_user("admin"))):
    body.pop("password_hash", None)
    body.pop("id", None)
    if body.get("password"):
        body["password_hash"] = hash_password(body.pop("password"))
    body["updated_at"] = _now_iso()
    res = await db.users.update_one({"id": uid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Utente non trovato")
    await log_attivita(user, "update", "user", uid)
    u = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    return u


@api.delete("/auth/users/{uid}")
async def delete_user(uid: str, user=Depends(require_user("admin"))):
    if uid == user["id"]:
        raise HTTPException(400, "Non puoi eliminare te stesso")
    res = await db.users.delete_one({"id": uid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Utente non trovato")
    await log_attivita(user, "delete", "user", uid)
    return {"ok": True}


# ============================================================
# COLLABORATORI / PROVVIGIONI
# ============================================================
@api.get("/collaboratori")
async def list_collaboratori(user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Tutti gli utenti con ruolo collaboratore/dipendente."""
    items = await db.users.find(
        {"role": {"$in": ["collaboratore", "dipendente"]}},
        {"_id": 0, "password_hash": 0},
    ).sort("name", 1).to_list(500)
    return items


@api.get("/collaboratori/{cid}/estratto-provvigioni")
async def estratto_provvigioni(
    cid: str, dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Estratto conto provvigioni: titoli incassati nel periodo + pagamenti effettuati.

    Restituisce:
      - titoli con provvigione maturata (da pagare o già pagata)
      - pagamenti effettuati (storico)
      - totali del periodo
    """
    if user.get("role") == "collaboratore" and cid != user.get("id"):
        raise HTTPException(403, "Non autorizzato a vedere l'estratto conto di altri collaboratori")
    collab = await db.users.find_one({"id": cid}, {"_id": 0, "password_hash": 0})
    if not collab:
        raise HTTPException(404, "Collaboratore non trovato")

    # auto-materializza voci ricorsive fino ad oggi (idempotente)
    try:
        await _materializza_voci_ricorsive(cid)
    except Exception as e:
        # non bloccare l'estratto conto in caso di errore di materializzazione
        import logging
        logging.getLogger(__name__).exception("Materializzazione voci ricorsive fallita: %s", e)

    pol_filter = {"collaboratore_id": cid}
    pol_ids = [p["id"] async for p in db.polizze.find(pol_filter, {"_id": 0, "id": 1})]

    titolo_filter = {"polizza_id": {"$in": pol_ids}, "stato": "incassato"}
    inc_cond = {}
    if dal: inc_cond["$gte"] = dal
    if al: inc_cond["$lte"] = al
    if inc_cond: titolo_filter["data_incasso"] = inc_cond

    titoli = await db.titoli.find(titolo_filter, {"_id": 0}).to_list(5000)
    # arricchimento polizza/contraente
    pol_map = {p["id"]: p async for p in db.polizze.find(
        {"id": {"$in": pol_ids}}, {"_id": 0, "id": 1, "numero_polizza": 1, "contraente_id": 1, "ramo": 1})}
    ana_ids = list({p.get("contraente_id") for p in pol_map.values() if p.get("contraente_id")})
    ana_map = {a["id"]: a async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}

    # pagamenti già effettuati nel periodo (per evitare conteggi doppi)
    pag_filter = {"collaboratore_id": cid}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        pag_filter["data_pagamento"] = cond
    pagamenti = await db.pagamenti_provvigioni.find(pag_filter, {"_id": 0}).sort("data_pagamento", -1).to_list(500)
    titoli_gia_pagati = set()
    voci_gia_pagate = set()
    for p in pagamenti:
        for tid in p.get("titoli_ids", []):
            titoli_gia_pagati.add(tid)
        for vid in p.get("voci_manuali_ids", []):
            voci_gia_pagate.add(vid)

    rows = []
    tot_lordo = 0.0
    tot_da_pagare = 0.0
    for t in titoli:
        pol = pol_map.get(t["polizza_id"], {})
        ana = ana_map.get(pol.get("contraente_id", ""), {})
        provv = t.get("provvigioni", 0.0) or 0.0
        gia_pagato = t["id"] in titoli_gia_pagati
        rows.append({
            "titolo_id": t["id"],
            "polizza_id": t["polizza_id"],
            "numero_polizza": pol.get("numero_polizza"),
            "contraente": ana.get("ragione_sociale"),
            "ramo": pol.get("ramo"),
            "data_incasso": t.get("data_incasso"),
            "importo_lordo": t.get("importo_lordo", 0.0),
            "provvigione": provv,
            "gia_pagato": gia_pagato,
        })
        tot_lordo += provv
        if not gia_pagato:
            tot_da_pagare += provv

    # calcolo trattenute sul "da pagare"
    rit = collab.get("perc_ritenuta_acconto", 0.0) or 0.0
    inps = collab.get("perc_inps_inarcassa", 0.0) or 0.0
    ritenuta_calc = round(tot_da_pagare * rit / 100.0, 2)
    contributi_calc = round(tot_da_pagare * inps / 100.0, 2)

    # voci manuali nel periodo (bonus / trattenute / acconti)
    voci_filter = {"collaboratore_id": cid}
    voci_cond = {}
    if dal: voci_cond["$gte"] = dal
    if al: voci_cond["$lte"] = al
    if voci_cond: voci_filter["data"] = voci_cond
    voci_manuali = await db.voci_manuali_collab.find(voci_filter, {"_id": 0}).sort("data", -1).to_list(500)
    tot_voci_da_pagare = 0.0
    tot_voci_totale = 0.0
    for v in voci_manuali:
        imp = float(v.get("importo") or 0.0)
        tot_voci_totale += imp
        if not v.get("pagata") and v["id"] not in voci_gia_pagate:
            tot_voci_da_pagare += imp
        # marca pagata se è in un pagamento
        if v["id"] in voci_gia_pagate:
            v["pagata"] = True

    netto_calc = round(tot_da_pagare - ritenuta_calc - contributi_calc + tot_voci_da_pagare, 2)

    return {
        "collaboratore": collab,
        "periodo": {"dal": dal, "al": al},
        "righe": rows,
        "voci_manuali": voci_manuali,
        "totali": {
            "provvigioni_lorde_periodo": round(tot_lordo, 2),
            "provvigioni_da_pagare": round(tot_da_pagare, 2),
            "ritenuta_acconto_calcolata": ritenuta_calc,
            "contributi_calcolati": contributi_calc,
            "voci_manuali_periodo": round(tot_voci_totale, 2),
            "voci_manuali_da_pagare": round(tot_voci_da_pagare, 2),
            "netto_da_pagare": netto_calc,
        },
        "pagamenti_periodo": pagamenti,
    }


# --- Voci manuali (bonus/trattenute/acconti) sull'estratto conto collaboratore ---
@api.get("/collaboratori/{cid}/voci-manuali")
async def list_voci_manuali_collab(
    cid: str, dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    if user.get("role") == "collaboratore" and cid != user.get("id"):
        raise HTTPException(403, "Non autorizzato")
    f = {"collaboratore_id": cid}
    cond = {}
    if dal: cond["$gte"] = dal
    if al: cond["$lte"] = al
    if cond: f["data"] = cond
    items = await db.voci_manuali_collab.find(f, {"_id": 0}).sort("data", -1).to_list(500)
    return items


@api.post("/collaboratori/{cid}/voci-manuali", status_code=201)
async def create_voce_manuale_collab(
    cid: str, body: dict, user=Depends(require_user("admin")),
):
    collab = await db.users.find_one({"id": cid}, {"_id": 0, "id": 1})
    if not collab:
        raise HTTPException(404, "Collaboratore non trovato")
    if not body.get("causale"):
        raise HTTPException(400, "Causale obbligatoria")
    if body.get("importo") in (None, ""):
        raise HTTPException(400, "Importo obbligatorio (positivo = bonus, negativo = trattenuta)")
    voce = VoceManualeCollab(
        collaboratore_id=cid,
        data=body.get("data") or _now_iso()[:10],
        causale=str(body["causale"]).strip(),
        importo=float(body["importo"]),
        note=body.get("note"),
    )
    await db.voci_manuali_collab.insert_one(voce.model_dump())
    await log_attivita(user, "create", "voce_manuale_collab", voce.id,
                       f"Voce manuale {voce.causale} €{voce.importo:.2f} per collaboratore {cid}")
    return voce.model_dump()


@api.delete("/collaboratori/{cid}/voci-manuali/{vid}")
async def delete_voce_manuale_collab(
    cid: str, vid: str, user=Depends(require_user("admin")),
):
    v = await db.voci_manuali_collab.find_one({"id": vid, "collaboratore_id": cid}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Voce non trovata")
    if v.get("pagata"):
        raise HTTPException(400, "Voce già pagata: non eliminabile")
    # blocco se la data della voce ricade in una giornata di prima nota chiusa
    await assert_giornata_aperta(v.get("data"), azione="eliminare la voce manuale")
    await db.voci_manuali_collab.delete_one({"id": vid})
    await log_attivita(user, "delete", "voce_manuale_collab", vid)
    return {"ok": True}


# ====================== VOCI RICORSIVE COLLABORATORI ======================
def _genera_date_ricorsive(regola: dict, fino_a: str) -> list[str]:
    """Restituisce la lista di date (YYYY-MM-DD) in cui la regola genera una voce,
    fra `regola.data_inizio` e `fino_a` (incluso).
    """
    from datetime import date as _date, timedelta as _td
    try:
        di = _date.fromisoformat(regola["data_inizio"])
    except Exception:
        return []
    df_raw = regola.get("data_fine") or fino_a
    try:
        df = _date.fromisoformat(df_raw)
    except Exception:
        df = _date.fromisoformat(fino_a)
    fa = _date.fromisoformat(fino_a)
    end = min(df, fa)
    if di > end:
        return []
    out: list[str] = []
    periodicita = regola.get("periodicita") or "mensile"
    giorno = max(1, min(28, int(regola.get("giorno_mese") or 1)))
    if periodicita == "annuale":
        mese = int(regola.get("mese_anno") or di.month)
        # primo anno utile
        y = di.year
        while True:
            try:
                d = _date(y, mese, giorno)
            except ValueError:
                y += 1
                continue
            if d < di:
                y += 1; continue
            if d > end: break
            out.append(d.isoformat())
            y += 1
    else:  # mensile
        # iterate month by month
        y, m = di.year, di.month
        while True:
            try:
                d = _date(y, m, giorno)
            except ValueError:
                # next month
                m += 1
                if m > 12: m = 1; y += 1
                continue
            if d < di:
                m += 1
                if m > 12: m = 1; y += 1
                continue
            if d > end:
                break
            out.append(d.isoformat())
            m += 1
            if m > 12: m = 1; y += 1
    return out


async def _materializza_voci_ricorsive(collaboratore_id: str, fino_a: Optional[str] = None) -> int:
    """Genera in `voci_manuali_collab` tutte le occorrenze delle regole ricorsive
    attive per il collaboratore, fino alla data `fino_a` (default: oggi).
    Idempotente: una voce per (ricorsiva_id, data).
    """
    today = fino_a or _now_iso()[:10]
    # carica regole attive applicabili al collaboratore (specifiche o __all__)
    rules = await db.voci_ricorsive_collab.find(
        {"attiva": True, "collaboratore_id": {"$in": [collaboratore_id, "__all__"]}},
        {"_id": 0},
    ).to_list(500)
    if not rules:
        return 0
    created = 0
    for r in rules:
        dates = _genera_date_ricorsive(r, today)
        if not dates:
            continue
        # voci già materializzate per questa regola (su questo collaboratore)
        existing = await db.voci_manuali_collab.find(
            {"ricorsiva_id": r["id"], "collaboratore_id": collaboratore_id},
            {"_id": 0, "data": 1},
        ).to_list(2000)
        already = {e["data"] for e in existing}
        for d in dates:
            if d in already:
                continue
            v = VoceManualeCollab(
                collaboratore_id=collaboratore_id,
                data=d,
                causale=r["causale"],
                importo=float(r["importo"] or 0),
                note=r.get("note"),
                ricorsiva_id=r["id"],
            )
            await db.voci_manuali_collab.insert_one(v.model_dump())
            created += 1
    return created


class VoceRicorsivaBody(BaseModel):
    collaboratore_id: str  # specifico o "__all__"
    causale: str
    importo: float
    periodicita: Literal["mensile", "annuale"] = "mensile"
    giorno_mese: int = 1
    mese_anno: Optional[int] = None
    data_inizio: str
    data_fine: Optional[str] = None
    note: Optional[str] = None
    attiva: bool = True


@api.get("/voci-ricorsive-collab")
async def list_voci_ricorsive(
    collaboratore_id: Optional[str] = None,
    solo_attive: bool = False,
    user=Depends(require_user("admin")),
):
    """Lista regole di voci ricorsive (admin only)."""
    flt: dict = {}
    if collaboratore_id:
        flt["collaboratore_id"] = collaboratore_id
    if solo_attive:
        flt["attiva"] = True
    items = await db.voci_ricorsive_collab.find(flt, {"_id": 0}).sort("created_at", -1).to_list(500)
    # enrich con nome collaboratore
    cids = list({r["collaboratore_id"] for r in items if r.get("collaboratore_id") and r["collaboratore_id"] != "__all__"})
    name_map: dict[str, str] = {}
    if cids:
        async for u in db.users.find(
            {"id": {"$in": cids}}, {"_id": 0, "id": 1, "name": 1, "email": 1},
        ):
            name_map[u["id"]] = u.get("name") or u.get("email")
    for r in items:
        if r["collaboratore_id"] == "__all__":
            r["collaboratore_nome"] = "Tutti i collaboratori"
        else:
            r["collaboratore_nome"] = name_map.get(r["collaboratore_id"], "?")
    return items


@api.post("/voci-ricorsive-collab", status_code=201)
async def create_voce_ricorsiva(body: VoceRicorsivaBody, user=Depends(require_user("admin"))):
    if body.collaboratore_id != "__all__":
        u = await db.users.find_one({"id": body.collaboratore_id}, {"_id": 0, "id": 1})
        if not u:
            raise HTTPException(404, "Collaboratore non trovato")
    obj = VoceRicorsivaCollab(**body.model_dump(), created_by=user.get("id"))
    await db.voci_ricorsive_collab.insert_one(obj.model_dump())
    await log_attivita(user, "create", "voce_ricorsiva_collab", obj.id)
    # materializza subito le occorrenze fino a oggi
    targets = []
    if body.collaboratore_id == "__all__":
        async for u in db.users.find(
            {"role": {"$in": ["collaboratore", "dipendente"]}, "attivo": {"$ne": False}},
            {"_id": 0, "id": 1},
        ):
            targets.append(u["id"])
    else:
        targets.append(body.collaboratore_id)
    n = 0
    for cid in targets:
        n += await _materializza_voci_ricorsive(cid)
    return {**obj.model_dump(), "voci_generate": n}


@api.put("/voci-ricorsive-collab/{rid}")
async def update_voce_ricorsiva(
    rid: str, body: VoceRicorsivaBody, user=Depends(require_user("admin")),
):
    r = await db.voci_ricorsive_collab.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Regola non trovata")
    upd = {**body.model_dump(), "updated_at": _now_iso()}
    await db.voci_ricorsive_collab.update_one({"id": rid}, {"$set": upd})
    await log_attivita(user, "update", "voce_ricorsiva_collab", rid)
    return {**r, **upd}


@api.delete("/voci-ricorsive-collab/{rid}")
async def delete_voce_ricorsiva(
    rid: str,
    elimina_voci_non_pagate: bool = False,
    user=Depends(require_user("admin")),
):
    """Elimina la regola. Se `elimina_voci_non_pagate=true` rimuove anche le
    voci già materializzate non ancora pagate."""
    res = await db.voci_ricorsive_collab.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Regola non trovata")
    n_voci = 0
    if elimina_voci_non_pagate:
        del_res = await db.voci_manuali_collab.delete_many(
            {"ricorsiva_id": rid, "pagata": {"$ne": True}, "pagamento_id": None},
        )
        n_voci = del_res.deleted_count
    await log_attivita(user, "delete", "voce_ricorsiva_collab", rid, payload={"n_voci_rimosse": n_voci})
    return {"ok": True, "n_voci_rimosse": n_voci}


@api.post("/voci-ricorsive-collab/{rid}/materializza")
async def materializza_voce_ricorsiva(rid: str, user=Depends(require_user("admin"))):
    """Forza la materializzazione delle occorrenze fino ad oggi."""
    r = await db.voci_ricorsive_collab.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Regola non trovata")
    targets: list[str] = []
    if r["collaboratore_id"] == "__all__":
        async for u in db.users.find(
            {"role": {"$in": ["collaboratore", "dipendente"]}, "attivo": {"$ne": False}},
            {"_id": 0, "id": 1},
        ):
            targets.append(u["id"])
    else:
        targets.append(r["collaboratore_id"])
    n = 0
    for cid in targets:
        n += await _materializza_voci_ricorsive(cid)
    return {"voci_generate": n}



@api.post("/collaboratori/{cid}/paga-provvigioni")
async def paga_provvigioni(cid: str, body: dict, user=Depends(require_user("admin", "collaboratore"))):
    """Esegue pagamento provvigioni: crea PagamentoProvvigioni + movimento contabile uscita.

    body: { titoli_ids: [...], conto_cassa_id, data_pagamento, mezzo_pagamento, note,
            override_provvigioni_lorde?, override_ritenuta?, override_contributi? }
    """
    collab = await db.users.find_one({"id": cid}, {"_id": 0, "password_hash": 0})
    if not collab:
        raise HTTPException(404, "Collaboratore non trovato")
    if user.get("role") == "collaboratore" and cid != user.get("id"):
        raise HTTPException(403, "Non autorizzato a pagare provvigioni di altri collaboratori")
    titoli_ids = body.get("titoli_ids") or []
    voci_ids = body.get("voci_manuali_ids") or []
    if not titoli_ids and not voci_ids:
        raise HTTPException(400, "Seleziona almeno un titolo o una voce manuale")
    conto_id = body.get("conto_cassa_id")
    mezzo_pag = body.get("mezzo_pagamento", "bonifico")
    if not conto_id:
        conto_id = await _resolve_conto_cassa(mezzo_pag)
    data_pag = body.get("data_pagamento") or _now_iso()[:10]
    # Lock: pagare provvigioni crea un movimento in Prima Nota.
    await assert_giornata_aperta(data_pag, azione="pagare provvigioni nel giorno (Prima Nota chiusa)")

    titoli = await db.titoli.find({"id": {"$in": titoli_ids}}, {"_id": 0}).to_list(5000) if titoli_ids else []
    lordo = sum((t.get("provvigioni") or 0.0) for t in titoli)
    if body.get("override_provvigioni_lorde") is not None:
        lordo = float(body["override_provvigioni_lorde"])

    rit_perc = collab.get("perc_ritenuta_acconto", 0.0) or 0.0
    inps_perc = collab.get("perc_inps_inarcassa", 0.0) or 0.0
    ritenuta = float(body.get("override_ritenuta")) if body.get("override_ritenuta") is not None else round(lordo * rit_perc / 100.0, 2)
    contributi = float(body.get("override_contributi")) if body.get("override_contributi") is not None else round(lordo * inps_perc / 100.0, 2)

    # Voci manuali (somma algebrica)
    voci = await db.voci_manuali_collab.find({"id": {"$in": voci_ids}, "collaboratore_id": cid}, {"_id": 0}).to_list(500) if voci_ids else []
    tot_voci = round(sum(float(v.get("importo") or 0.0) for v in voci), 2)

    netto = round(lordo - ritenuta - contributi + tot_voci, 2)

    # determina periodo (min/max data_incasso dei titoli + date voci)
    inc_dates = [t.get("data_incasso") for t in titoli if t.get("data_incasso")]
    inc_dates += [v.get("data") for v in voci if v.get("data")]
    periodo_dal = min(inc_dates) if inc_dates else data_pag
    periodo_al = max(inc_dates) if inc_dates else data_pag

    descr_extra = f" (+ {len(voci)} voci manuali: €{tot_voci:.2f})" if voci else ""
    # Crea movimento contabile (uscita) - va in Brogliaccio
    mov = MovimentoContabile(
        data_movimento=data_pag,
        tipo="uscita",
        categoria="provvigioni",
        importo=netto,
        descrizione=f"Pagamento provvigioni a {collab.get('name')} - periodo {periodo_dal} / {periodo_al}{descr_extra}",
        conto_cassa_id=conto_id,
        mezzo_pagamento=body.get("mezzo_pagamento", "bonifico"),
        note=(f"Lordo {lordo:.2f} - rit. {ritenuta:.2f} - contributi {contributi:.2f} + voci {tot_voci:.2f} = netto {netto:.2f}"),
    )
    await db.movimenti.insert_one(mov.model_dump())

    pag = PagamentoProvvigioni(
        collaboratore_id=cid,
        collaboratore_nome=collab.get("name", ""),
        periodo_dal=periodo_dal,
        periodo_al=periodo_al,
        provvigioni_lorde=round(lordo, 2),
        ritenuta_acconto=ritenuta,
        contributi=contributi,
        netto_pagato=netto,
        conto_cassa_id=conto_id,
        mezzo_pagamento=body.get("mezzo_pagamento", "bonifico"),
        data_pagamento=data_pag,
        movimento_id=mov.id,
        titoli_ids=titoli_ids,
        voci_manuali_ids=voci_ids,
        note=body.get("note"),
    )
    await db.pagamenti_provvigioni.insert_one(pag.model_dump())
    # AUTO-REGISTRA la ritenuta d'acconto in /api/ritenute se > 0
    if ritenuta > 0:
        await db.ritenute.insert_one({
            "id": str(uuid.uuid4()),
            "anno": int(data_pag[:4]),
            "collaboratore_id": cid,
            "imponibile": round(lordo, 2),
            "aliquota": rit_perc,
            "importo_ritenuta": ritenuta,
            "causale": "1040",
            "data": data_pag,
            "descrizione": f"Ritenuta automatica da pagamento provvigioni periodo {periodo_dal}/{periodo_al}",
            "versata": False,
            "pagamento_id": pag.id,
            "movimento_id": mov.id,
            "auto_generata": True,
            "created_at": _now_iso(),
        })
    # Marca voci manuali come pagate
    if voci_ids:
        await db.voci_manuali_collab.update_many(
            {"id": {"$in": voci_ids}, "collaboratore_id": cid},
            {"$set": {"pagata": True, "pagamento_id": pag.id, "updated_at": _now_iso()}},
        )
    await log_attivita(user, "paga_provvigioni", "collaboratore", cid,
                       f"Pagate provvigioni a {collab.get('name')} per €{netto:.2f}")
    return {
        "ok": True,
        "pagamento": pag.model_dump(),
        "movimento": mov.model_dump(),
    }


@api.get("/collaboratori/{cid}/pagamenti")
async def list_pagamenti(cid: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    items = await db.pagamenti_provvigioni.find({"collaboratore_id": cid}, {"_id": 0}) \
        .sort("data_pagamento", -1).to_list(500)
    # enrichment: count allegati on movimento_id, conto cassa name
    mov_ids = [p["movimento_id"] for p in items if p.get("movimento_id")]
    n_all_by_mov: dict[str, int] = {}
    if mov_ids:
        pipeline = [
            {"$match": {"entita_tipo": "movimento", "entita_id": {"$in": mov_ids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]
        async for r in db.allegati.aggregate(pipeline):
            n_all_by_mov[r["_id"]] = r["n"]
    conto_ids = list({p.get("conto_cassa_id") for p in items if p.get("conto_cassa_id")})
    conti_map: dict[str, str] = {}
    if conto_ids:
        async for c in db.conti_cassa.find({"id": {"$in": conto_ids}}, {"_id": 0, "id": 1, "nome": 1}):
            conti_map[c["id"]] = c["nome"]
    for p in items:
        p["n_allegati"] = n_all_by_mov.get(p.get("movimento_id"), 0)
        p["conto_cassa_nome"] = conti_map.get(p.get("conto_cassa_id"))
        p["n_titoli"] = len(p.get("titoli_ids") or [])
        p["n_voci_manuali"] = len(p.get("voci_manuali_ids") or [])
    return items


@api.get("/collaboratori/{cid}/pagamenti/{pid}")
async def get_pagamento_dettaglio(
    cid: str, pid: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Dettaglio di un pagamento provvigioni: include titoli pagati e voci manuali."""
    pag = await db.pagamenti_provvigioni.find_one({"id": pid, "collaboratore_id": cid}, {"_id": 0})
    if not pag:
        raise HTTPException(404, "Pagamento non trovato")
    # carica titoli + polizza/contraente
    tids = pag.get("titoli_ids") or []
    titoli: list[dict] = []
    if tids:
        async for t in db.titoli.find({"id": {"$in": tids}}, {"_id": 0}):
            titoli.append(t)
        pol_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
        pol_map: dict[str, dict] = {}
        if pol_ids:
            async for p in db.polizze.find(
                {"id": {"$in": pol_ids}},
                {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "contraente_id": 1},
            ):
                pol_map[p["id"]] = p
        ana_ids = list({p.get("contraente_id") for p in pol_map.values() if p.get("contraente_id")})
        ana_map: dict[str, str] = {}
        if ana_ids:
            async for a in db.anagrafiche.find(
                {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1},
            ):
                ana_map[a["id"]] = a["ragione_sociale"]
        for t in titoli:
            p = pol_map.get(t.get("polizza_id"), {})
            t["numero_polizza"] = p.get("numero_polizza")
            t["ramo"] = p.get("ramo")
            t["contraente_nome"] = ana_map.get(p.get("contraente_id", ""))
    # voci manuali
    vids = pag.get("voci_manuali_ids") or []
    voci: list[dict] = []
    if vids:
        async for v in db.voci_manuali_collab.find({"id": {"$in": vids}}, {"_id": 0}):
            voci.append(v)
    # conto cassa
    conto_nome = None
    if pag.get("conto_cassa_id"):
        c = await db.conti_cassa.find_one({"id": pag["conto_cassa_id"]}, {"_id": 0, "nome": 1})
        if c:
            conto_nome = c["nome"]
    n_all = await db.allegati.count_documents(
        {"entita_tipo": "movimento", "entita_id": pag.get("movimento_id"), "is_deleted": False},
    ) if pag.get("movimento_id") else 0
    return {
        **pag,
        "titoli": titoli,
        "voci_manuali": voci,
        "conto_cassa_nome": conto_nome,
        "n_allegati": n_all,
    }


@api.get("/stampa/provvigioni/{cid}")
async def stampa_provvigioni(cid: str, dal: Optional[str] = None, al: Optional[str] = None,
                             user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    data = await estratto_provvigioni(cid, dal, al, user)
    collab = data["collaboratore"]
    headers = ["Data", "N. polizza", "Contraente", "Ramo", "Premio €", "Provvigione €", "Pagata"]
    rows = []
    for r in data["righe"]:
        rows.append([r.get("data_incasso", ""), r.get("numero_polizza", ""),
                     r.get("contraente", ""), r.get("ramo", ""),
                     r.get("importo_lordo", 0), r.get("provvigione", 0),
                     "SÌ" if r.get("gia_pagato") else "NO"])
    tot = data["totali"]
    # Voci manuali
    for v in data.get("voci_manuali", []):
        rows.append([v.get("data", ""), "—", "VOCE MANUALE", v.get("causale", ""),
                     "", v.get("importo", 0),
                     "SÌ" if v.get("pagata") else "NO"])
    rows.append(["", "", "", "TOTALE LORDO", "", tot["provvigioni_lorde_periodo"], ""])
    rows.append(["", "", "", "DA PAGARE", "", tot["provvigioni_da_pagare"], ""])
    rows.append(["", "", "", "RITENUTA ACCONTO", "", tot["ritenuta_acconto_calcolata"], ""])
    rows.append(["", "", "", "CONTRIBUTI", "", tot["contributi_calcolati"], ""])
    rows.append(["", "", "", "VOCI MANUALI", "", tot.get("voci_manuali_da_pagare", 0), ""])
    rows.append(["", "", "", "NETTO DA PAGARE", "", tot["netto_da_pagare"], ""])
    pdf = pdf_report.stampa_elenco(
        f"Estratto conto collaboratore - {collab.get('name')}",
        f"Periodo: {dal or '—'} -> {al or '—'}",
        headers, rows,
        col_widths_mm=[25, 35, 60, 25, 30, 30, 22], landscape_mode=False,
        filtri_attivi={"Dal": dal, "Al": al},
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"provvigioni_{collab.get('name')}.pdf")


# ============================================================
# COMPAGNIE
# ============================================================
@api.get("/compagnie")
async def list_compagnie(user=Depends(current_user)):
    items = await db.compagnie.find({}, {"_id": 0}).to_list(1000)
    return items


@api.post("/compagnie", status_code=201)
async def create_compagnia(body: dict, user=Depends(require_user("admin", "collaboratore"))):
    obj = Compagnia(**body)
    await db.compagnie.insert_one(obj.model_dump())
    await log_attivita(user, "create", "compagnia", obj.id, f"Creata compagnia {obj.codice}")
    return obj.model_dump()


@api.put("/compagnie/{cid}")
async def update_compagnia(cid: str, body: dict, user=Depends(require_user("admin", "collaboratore"))):
    body["updated_at"] = _now_iso()
    res = await db.compagnie.update_one({"id": cid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Compagnia non trovata")
    await log_attivita(user, "update", "compagnia", cid)
    doc = await db.compagnie.find_one({"id": cid}, {"_id": 0})
    return doc


@api.delete("/compagnie/{cid}")
async def delete_compagnia(cid: str, user=Depends(require_user("admin"))):
    await db.compagnie.delete_one({"id": cid})
    await log_attivita(user, "delete", "compagnia", cid)
    return {"ok": True}


# ============================================================
# COMPAGNIE — Estratto conto e Saldo cassa
# ============================================================
async def _compagnia_estratto_data(compagnia_id: str, dal: Optional[str], al: Optional[str], collaboratore_id: Optional[str] = None) -> dict:
    """Aggrega dare/avere per una compagnia su un periodo.

    Logica:
      - Titoli incassati nel periodo -> DARE verso compagnia = (lordo - provvigioni se trattiene)
      - Movimenti contabili categoria 'pagamento_compagnia' verso questa compagnia -> AVERE
      - Filtro opzionale collaboratore_id (polizze del collaboratore)
    """
    comp = await db.compagnie.find_one({"id": compagnia_id}, {"_id": 0})
    if not comp:
        raise HTTPException(404, "Compagnia non trovata")
    trattiene = bool(comp.get("trattiene_provvigioni", True))
    tipo_mandato = comp.get("tipo_mandato", "diretto")
    # Mandato collaborazione: saldo = premi puri (non scaliamo le provv,
    # verranno fatturate dalla agenzia partner). Forziamo trattiene=False.
    if tipo_mandato == "collaborazione":
        trattiene = False

    pol_flt: dict = {"compagnia_id": compagnia_id}
    if collaboratore_id:
        pol_flt["collaboratore_id"] = collaboratore_id
    polizze = await db.polizze.find(
        pol_flt,
        {"_id": 0, "id": 1, "numero_polizza": 1, "contraente_id": 1, "ramo": 1, "collaboratore_id": 1},
    ).to_list(20000)
    pol_index = {p["id"]: p for p in polizze}
    if not pol_index:
        return {"compagnia": comp, "righe": [], "totale_dare": 0.0, "totale_avere": 0.0, "saldo": 0.0,
                "periodo": {"dal": dal, "al": al}, "trattiene_provvigioni": trattiene}

    titoli_flt: dict = {
        "polizza_id": {"$in": list(pol_index.keys())},
        "stato": "incassato",
    }
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        titoli_flt["data_incasso"] = cond
    titoli = await db.titoli.find(titoli_flt, {"_id": 0}).sort("data_incasso", 1).to_list(20000)

    # arricchimento contraenti e collaboratori
    contr_ids = list({pol_index[t["polizza_id"]].get("contraente_id") for t in titoli if t["polizza_id"] in pol_index})
    collab_ids = list({pol_index[t["polizza_id"]].get("collaboratore_id") for t in titoli if t["polizza_id"] in pol_index and pol_index[t["polizza_id"]].get("collaboratore_id")})
    contr_map = {}
    if contr_ids:
        async for c in db.anagrafiche.find({"id": {"$in": contr_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1}):
            contr_map[c["id"]] = c["ragione_sociale"]
    collab_map = {}
    if collab_ids:
        async for u in db.users.find({"id": {"$in": collab_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1}):
            collab_map[u["id"]] = u.get("name") or u.get("email")

    righe = []
    totale_dare = 0.0
    for t in titoli:
        pol = pol_index.get(t["polizza_id"], {})
        provv = float(t.get("provvigioni") or 0)
        lordo = float(t.get("importo_lordo") or 0)
        dovuto_alla_compagnia = (lordo - provv) if trattiene else lordo
        righe.append({
            "data": t.get("data_incasso") or t.get("scadenza"),
            "tipo": "incasso",
            "titolo_id": t.get("id"),
            "polizza_id": pol.get("id"),
            "polizza": pol.get("numero_polizza"),
            "contraente_id": pol.get("contraente_id"),
            "contraente": contr_map.get(pol.get("contraente_id", "")),
            "ramo": pol.get("ramo"),
            "collaboratore_id": pol.get("collaboratore_id"),
            "collaboratore": collab_map.get(pol.get("collaboratore_id", "")),
            "dare": round(dovuto_alla_compagnia, 2),
            "avere": 0.0,
            "lordo": lordo, "provvigioni": provv,
            "stato_pagamento": t.get("pagato_alla_compagnia") and "pagato" or "da_versare",
            "data_pagamento_compagnia": t.get("data_pagamento_compagnia"),
            "_movimento_id": t.get("id"),
        })
        totale_dare += dovuto_alla_compagnia

    # pagamenti verso compagnia
    mov_flt: dict = {"compagnia_id": compagnia_id, "categoria": "pagamento_compagnia"}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        mov_flt["data_movimento"] = cond
    movs = await db.movimenti.find(mov_flt, {"_id": 0}).sort("data_movimento", 1).to_list(5000)
    totale_avere = 0.0
    for m in movs:
        imp = float(m.get("importo") or 0)
        righe.append({
            "data": m.get("data_movimento"),
            "tipo": "pagamento",
            "polizza": None,
            "contraente": None,
            "ramo": None,
            "dare": 0.0,
            "avere": imp,
            "descrizione": m.get("descrizione") or m.get("causale"),
            "_movimento_id": m.get("id"),
        })
        totale_avere += imp

    righe.sort(key=lambda r: r["data"] or "")
    saldo = totale_dare - totale_avere
    # ----- Rappel (sovraprovvigioni): accreditate fittiziamente, riducono saldo -----
    rap_flt: dict = {"compagnia_id": compagnia_id}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        rap_flt["data"] = cond
    rappel = await db.rappel.find(rap_flt, {"_id": 0}).sort("data", 1).to_list(5000)
    totale_rappel = 0.0
    for r in rappel:
        imp = float(r.get("importo") or 0)
        righe.append({
            "data": r.get("data"),
            "tipo": "rappel",
            "polizza": None,
            "contraente": None,
            "ramo": None,
            "dare": 0.0,
            "avere": imp,
            "descrizione": f"Rappel {r.get('anno')}: {r.get('descrizione') or ''}".strip(),
            "_movimento_id": r.get("id"),
            "is_rappel": True,
        })
        totale_rappel += imp
    totale_avere += totale_rappel
    saldo -= totale_rappel

    # ----- Ritenute compagnia: AUMENTANO il dare (sono provvigioni che la
    # compagnia ci trattiene -> dobbiamo versarle nel saldo) — solo mandato diretto
    totale_ritenute_comp = 0.0
    if tipo_mandato == "diretto":
        rit_flt: dict = {"compagnia_id": compagnia_id}
        if dal or al:
            cond = {}
            if dal: cond["$gte"] = dal
            if al: cond["$lte"] = al
            rit_flt["data"] = cond
        ritenute_c = await db.ritenute_compagnia.find(rit_flt, {"_id": 0}).sort("data", 1).to_list(5000)
        for r in ritenute_c:
            imp = float(r.get("importo") or 0)
            righe.append({
                "data": r.get("data"),
                "tipo": "ritenuta_compagnia",
                "polizza": None,
                "contraente": None,
                "ramo": None,
                "dare": imp,  # aumenta il dare
                "avere": 0.0,
                "descrizione": f"Ritenuta {r.get('anno')}: {r.get('descrizione') or ''}".strip(),
                "_movimento_id": r.get("id"),
                "is_ritenuta_compagnia": True,
                "stato_ritenuta": r.get("stato"),
            })
            totale_ritenute_comp += imp
        totale_dare += totale_ritenute_comp
        saldo += totale_ritenute_comp

    # ----- Fatture agenzia partner (solo mandato collaborazione) — sono
    # provvigioni che ci pagherà la partner: AVERE che riduce dare verso la compagnia.
    totale_fatture_partner = 0.0
    if tipo_mandato == "collaborazione":
        f_flt: dict = {"compagnia_id": compagnia_id}
        if dal or al:
            cond = {}
            if dal: cond["$gte"] = dal
            if al: cond["$lte"] = al
            f_flt["data"] = cond
        fatture = await db.fatture_agenzia_partner.find(f_flt, {"_id": 0}).sort("data", 1).to_list(2000)
        for f in fatture:
            imp = float(f.get("importo") or 0)
            label = f"Fattura partner n. {f.get('numero_fattura') or '—'} ({f.get('stato')})"
            righe.append({
                "data": f.get("data"),
                "tipo": "fattura_partner",
                "polizza": None,
                "contraente": None,
                "ramo": None,
                "dare": 0.0,
                "avere": imp,
                "descrizione": f"{label}: {f.get('descrizione') or ''}".strip(),
                "_movimento_id": f.get("id"),
                "is_fattura_partner": True,
                "stato_fattura": f.get("stato"),
            })
            totale_fatture_partner += imp
        totale_avere += totale_fatture_partner
        saldo -= totale_fatture_partner

    righe.sort(key=lambda r: r["data"] or "")
    return {
        "compagnia": comp,
        "righe": righe,
        "totale_dare": round(totale_dare, 2),
        "totale_avere": round(totale_avere, 2),
        "totale_rappel": round(totale_rappel, 2),
        "totale_ritenute_compagnia": round(totale_ritenute_comp, 2),
        "totale_fatture_partner": round(totale_fatture_partner, 2),
        "saldo": round(saldo, 2),
        "periodo": {"dal": dal, "al": al},
        "trattiene_provvigioni": trattiene,
        "tipo_mandato": tipo_mandato,
    }


@api.get("/compagnie/{cid}/estratto-conto")
async def compagnia_estratto_conto(
    cid: str, dal: Optional[str] = None, al: Optional[str] = None,
    collaboratore_id: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    return await _compagnia_estratto_data(cid, dal, al, collaboratore_id)


@api.get("/compagnie/{cid}/rimesse-storico")
async def compagnia_rimesse_storico(
    cid: str, dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Storico di tutte le rimesse pagate a una compagnia.

    Ritorna i movimenti `pagamento_compagnia` con i titoli associati e contraenti.
    """
    flt: dict = {"compagnia_id": cid, "categoria": "pagamento_compagnia", "tipo": "uscita"}
    if dal:
        flt["data_movimento"] = flt.get("data_movimento", {})
        flt["data_movimento"]["$gte"] = dal
    if al:
        flt["data_movimento"] = flt.get("data_movimento", {})
        flt["data_movimento"]["$lte"] = al
    rimesse = await db.movimenti.find(flt, {"_id": 0}).sort("data_movimento", -1).to_list(2000)
    # Per ogni rimessa, arricchisce con i titoli legati e nome contraenti
    all_titoli_ids = []
    for r in rimesse:
        all_titoli_ids.extend(r.get("titoli_ids_versati") or [])
    titoli_map = {}
    contr_map = {}
    if all_titoli_ids:
        async for t in db.titoli.find({"id": {"$in": all_titoli_ids}}, {"_id": 0}):
            titoli_map[t["id"]] = t
        pol_ids = list({t.get("polizza_id") for t in titoli_map.values() if t.get("polizza_id")})
        pol_map = {}
        if pol_ids:
            async for p in db.polizze.find(
                {"id": {"$in": pol_ids}},
                {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "contraente_id": 1, "targa": 1},
            ):
                pol_map[p["id"]] = p
        ana_ids = list({p.get("contraente_id") for p in pol_map.values() if p.get("contraente_id")})
        if ana_ids:
            async for a in db.anagrafiche.find(
                {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1},
            ):
                contr_map[a["id"]] = a["ragione_sociale"]
        # Inietta nome polizza/contraente in ogni titolo
        for t in titoli_map.values():
            p = pol_map.get(t.get("polizza_id"), {})
            t["numero_polizza"] = p.get("numero_polizza")
            t["ramo"] = p.get("ramo")
            t["targa"] = p.get("targa")
            t["contraente_nome"] = contr_map.get(p.get("contraente_id", ""))

    # Conta allegati per ogni movimento
    out = []
    for r in rimesse:
        ids = r.get("titoli_ids_versati") or []
        n_all = await db.allegati.count_documents({"entita_tipo": "movimento", "entita_id": r["id"]})
        out.append({
            **r,
            "titoli": [titoli_map.get(i) for i in ids if titoli_map.get(i)],
            "n_titoli": len(ids),
            "n_allegati": n_all,
        })
    totale_pagato = round(sum(r.get("importo", 0) for r in rimesse), 2)
    return {
        "compagnia_id": cid,
        "rimesse": out,
        "totale_pagato": totale_pagato,
        "n_rimesse": len(out),
    }


@api.post("/compagnie/{cid}/paga-titoli")
async def compagnia_paga_titoli(
    cid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore")),
):
    """Registra il pagamento alla compagnia di un insieme di titoli.

    body: {
        titoli_ids: [str],
        conto_cassa_id: str,
        data_movimento?: str (YYYY-MM-DD, default oggi),
        descrizione?: str,
    }
    Crea un MovimentoContabile categoria 'pagamento_compagnia' e marca i titoli
    come 'pagato_alla_compagnia' = True.
    """
    titoli_ids = body.get("titoli_ids") or []
    if not titoli_ids:
        raise HTTPException(400, "titoli_ids obbligatorio")
    conto_cassa_id = body.get("conto_cassa_id")
    if not conto_cassa_id:
        # auto-derive from mezzo or fallback to any active conto
        conto_cassa_id = await _resolve_conto_cassa(body.get("mezzo_pagamento") or "bonifico")
    if not conto_cassa_id:
        raise HTTPException(400, "Nessun conto cassa attivo configurato. Crea un conto in Librerie.")
    comp = await db.compagnie.find_one({"id": cid}, {"_id": 0})
    if not comp:
        raise HTTPException(404, "Compagnia non trovata")
    trattiene = bool(comp.get("trattiene_provvigioni", True))

    titoli = await db.titoli.find({"id": {"$in": titoli_ids}, "stato": "incassato"}, {"_id": 0}).to_list(5000)
    if not titoli:
        raise HTTPException(404, "Nessun titolo incassato trovato")

    totale = 0.0
    for t in titoli:
        lordo = float(t.get("importo_lordo") or 0)
        provv = float(t.get("provvigioni") or 0)
        totale += (lordo - provv) if trattiene else lordo

    data_mov = body.get("data_movimento") or _now_iso()[:10]
    # Lock: pagare la compagnia crea un movimento in Prima Nota.
    await assert_giornata_aperta(data_mov, azione="pagare la compagnia nel giorno (Prima Nota chiusa)")
    descrizione = body.get("descrizione") or f"Versamento {comp.get('ragione_sociale')} — {len(titoli)} titoli"

    mov = {
        "id": str(uuid.uuid4()),
        "data_movimento": data_mov,
        "tipo": "uscita",
        "categoria": "pagamento_compagnia",
        "compagnia_id": cid,
        "conto_cassa_id": conto_cassa_id,
        "importo": round(totale, 2),
        "descrizione": descrizione,
        "titoli_ids_versati": [t["id"] for t in titoli],
        "created_at": _now_iso(),
        "creato_da": user.get("id"),
    }
    await db.movimenti.insert_one(mov)

    # Marca titoli come pagati alla compagnia
    await db.titoli.update_many(
        {"id": {"$in": [t["id"] for t in titoli]}},
        {"$set": {
            "pagato_alla_compagnia": True,
            "data_pagamento_compagnia": data_mov,
            "movimento_pagamento_id": mov["id"],
            "updated_at": _now_iso(),
        }},
    )
    await log_attivita(user, "paga_compagnia", "movimento", mov["id"],
                       f"Compagnia {comp.get('ragione_sociale')} · {len(titoli)} titoli · {round(totale, 2)}€")
    return {"ok": True, "movimento_id": mov["id"], "totale": round(totale, 2),
            "titoli_pagati": len(titoli)}


@api.get("/compagnie/saldi-cassa")
async def saldi_cassa_compagnie(
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Saldo da versare/da incassare per ciascuna compagnia (tutti i periodi)."""
    compagnie = await db.compagnie.find({}, {"_id": 0}).to_list(500)
    risultati = []
    for c in compagnie:
        try:
            data = await _compagnia_estratto_data(c["id"], None, None)
            risultati.append({
                "compagnia_id": c["id"],
                "codice": c.get("codice"),
                "ragione_sociale": c.get("ragione_sociale"),
                "trattiene_provvigioni": c.get("trattiene_provvigioni", True),
                "totale_incassato": data["totale_dare"],
                "totale_versato": data["totale_avere"],
                "saldo_da_versare": data["saldo"],
                "righe_count": len(data["righe"]),
            })
        except Exception:
            continue
    risultati.sort(key=lambda r: -abs(r["saldo_da_versare"] or 0))
    return risultati


# ====================== RAPPEL (sovraprovvigioni compagnia) ======================
class RappelBody(BaseModel):
    compagnia_id: str
    data: str
    importo: float
    descrizione: Optional[str] = None
    anno: Optional[int] = None
    note: Optional[str] = None


@api.get("/rappel")
async def list_rappel(
    compagnia_id: Optional[str] = None,
    anno: Optional[int] = None,
    dal: Optional[str] = None,
    al: Optional[str] = None,
    stato: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Lista rappel filtrabile per compagnia, anno, periodo (data) o stato."""
    flt: dict = {}
    if compagnia_id: flt["compagnia_id"] = compagnia_id
    if anno: flt["anno"] = anno
    if stato: flt["stato"] = stato
    if dal or al:
        cond: dict = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        flt["data"] = cond
    items = await db.rappel.find(flt, {"_id": 0}).sort("data", -1).to_list(5000)
    # enrich con ragione_sociale compagnia + n_allegati
    comp_ids = list({r["compagnia_id"] for r in items if r.get("compagnia_id")})
    if comp_ids:
        comp_map: dict[str, dict] = {}
        async for c in db.compagnie.find(
            {"id": {"$in": comp_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1, "codice": 1},
        ):
            comp_map[c["id"]] = c
        for r in items:
            c = comp_map.get(r["compagnia_id"], {})
            r["compagnia_nome"] = c.get("ragione_sociale")
            r["compagnia_codice"] = c.get("codice")
    # conta allegati per rappel
    ids = [r["id"] for r in items]
    if ids:
        n_all: dict[str, int] = {}
        async for x in db.allegati.aggregate([
            {"$match": {"entita_tipo": "rappel", "entita_id": {"$in": ids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]):
            n_all[x["_id"]] = x["n"]
        for r in items:
            r["n_allegati"] = n_all.get(r["id"], 0)
            r["stato"] = r.get("stato") or "da_incassare"
    return items


@api.get("/rappel/archivio")
async def archivio_rappel(
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Ritorna i totali aggregati per anno + breakdown per compagnia."""
    pipeline = [
        {"$group": {
            "_id": {"anno": "$anno", "compagnia_id": "$compagnia_id"},
            "totale": {"$sum": "$importo"},
            "n_movimenti": {"$sum": 1},
        }},
        {"$sort": {"_id.anno": -1}},
    ]
    out = await db.rappel.aggregate(pipeline).to_list(5000)
    comp_ids = list({o["_id"]["compagnia_id"] for o in out})
    comp_map: dict[str, dict] = {}
    if comp_ids:
        async for c in db.compagnie.find(
            {"id": {"$in": comp_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1, "codice": 1},
        ):
            comp_map[c["id"]] = c
    # aggrega per anno
    per_anno: dict[int, dict] = {}
    for o in out:
        anno = o["_id"]["anno"]
        cid = o["_id"]["compagnia_id"]
        bucket = per_anno.setdefault(anno, {"anno": anno, "totale": 0.0, "n_movimenti": 0, "compagnie": []})
        bucket["totale"] += o["totale"]
        bucket["n_movimenti"] += o["n_movimenti"]
        bucket["compagnie"].append({
            "compagnia_id": cid,
            "compagnia_nome": comp_map.get(cid, {}).get("ragione_sociale"),
            "compagnia_codice": comp_map.get(cid, {}).get("codice"),
            "totale": round(o["totale"], 2),
            "n_movimenti": o["n_movimenti"],
        })
    risultato = sorted(per_anno.values(), key=lambda x: -x["anno"])
    for r in risultato:
        r["totale"] = round(r["totale"], 2)
        r["compagnie"].sort(key=lambda x: -x["totale"])
    return risultato


@api.post("/rappel")
async def create_rappel(
    body: RappelBody,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    comp = await db.compagnie.find_one({"id": body.compagnia_id}, {"_id": 0, "id": 1})
    if not comp:
        raise HTTPException(404, "Compagnia non trovata")
    if body.importo <= 0:
        raise HTTPException(400, "Importo deve essere positivo")
    anno = body.anno or int(body.data[:4])
    r = Rappel(
        compagnia_id=body.compagnia_id,
        data=body.data, anno=anno, importo=round(body.importo, 2),
        descrizione=body.descrizione, note=body.note,
        created_by=user.get("id"),
    )
    await db.rappel.insert_one(r.model_dump())
    await log_attivita(user, "create", "rappel", r.id, payload={"importo": r.importo, "compagnia_id": r.compagnia_id})
    return r.model_dump()


@api.put("/rappel/{rid}")
async def update_rappel(
    rid: str, body: RappelBody,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    r = await db.rappel.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Rappel non trovato")
    # blocco se incassato in giornata chiusa
    if r.get("stato") == "incassato":
        await assert_giornata_aperta(r.get("data_incasso") or r.get("data"), azione="modificare il rappel")
    if body.importo <= 0:
        raise HTTPException(400, "Importo deve essere positivo")
    anno = body.anno or int(body.data[:4])
    update = {
        "compagnia_id": body.compagnia_id,
        "data": body.data, "anno": anno,
        "importo": round(body.importo, 2),
        "descrizione": body.descrizione, "note": body.note,
        "updated_at": _now_iso(),
    }
    await db.rappel.update_one({"id": rid}, {"$set": update})
    await log_attivita(user, "update", "rappel", rid, payload=update)
    return {**r, **update}


@api.delete("/rappel/{rid}")
async def delete_rappel(
    rid: str, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    r = await db.rappel.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Rappel non trovato")
    # blocco se rappel incassato in giornata chiusa
    if r.get("stato") == "incassato":
        await assert_giornata_aperta(r.get("data_incasso") or r.get("data"), azione="eliminare il rappel")
    if r.get("stato") == "incassato" and r.get("movimento_id"):
        # rimuovi anche il movimento collegato
        await db.movimenti.delete_one({"id": r["movimento_id"]})
    res = await db.rappel.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Rappel non trovato")
    await log_attivita(user, "delete", "rappel", rid, payload={})
    return {"ok": True}


@api.post("/rappel/{rid}/incassa")
async def incassa_rappel(
    rid: str, body: dict = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Marca un rappel come incassato e crea un movimento in Prima Nota
    nella categoria 'provvigioni' (NON è un incasso premio normale)."""
    r = await db.rappel.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Rappel non trovato")
    if r.get("stato") == "incassato":
        raise HTTPException(400, "Rappel già incassato")
    body = body or {}
    data_incasso = body.get("data_incasso") or _now_iso()[:10]
    # Lock: incassare un rappel crea un movimento in Prima Nota.
    await assert_giornata_aperta(data_incasso, azione="incassare un rappel nel giorno")
    comp = await db.compagnie.find_one({"id": r["compagnia_id"]}, {"_id": 0, "ragione_sociale": 1})
    importo = float(r.get("importo") or 0)
    # Crea movimento in Prima Nota — categoria "provvigioni" entrata fittizia
    mov_id = _uid()
    mov = {
        "id": mov_id,
        "data_movimento": data_incasso,
        "tipo": "entrata",
        "categoria": "provvigioni",
        "importo": round(importo, 2),
        "descrizione": f"Rappel {r.get('anno')} — {comp.get('ragione_sociale','')}: {r.get('descrizione') or ''}".strip(),
        "compagnia_id": r["compagnia_id"],
        "rappel_id": rid,
        "is_rappel": True,
        "created_at": _now_iso(),
        "created_by": user.get("id"),
    }
    await db.movimenti.insert_one(mov)
    await db.rappel.update_one({"id": rid}, {"$set": {
        "stato": "incassato",
        "data_incasso": data_incasso,
        "movimento_id": mov_id,
        "updated_at": _now_iso(),
    }})
    await log_attivita(user, "incassa", "rappel", rid, payload={"importo": importo})
    return {**r, "stato": "incassato", "data_incasso": data_incasso, "movimento_id": mov_id}


@api.post("/rappel/{rid}/storna")
async def storna_rappel(
    rid: str, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Annulla l'incasso di un rappel: rimuove il movimento in Prima Nota
    e riporta il rappel a stato 'da_incassare'."""
    r = await db.rappel.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Rappel non trovato")
    if r.get("stato") != "incassato":
        raise HTTPException(400, "Rappel non incassato")
    # Lock: lo storno cancella il movimento Prima Nota -> se chiuso, blocca.
    await assert_giornata_aperta(r.get("data_incasso") or r.get("data"), azione="stornare un rappel del giorno")
    if r.get("movimento_id"):
        await db.movimenti.delete_one({"id": r["movimento_id"]})
    await db.rappel.update_one({"id": rid}, {"$set": {
        "stato": "da_incassare", "data_incasso": None, "movimento_id": None,
        "updated_at": _now_iso(),
    }})
    return {"ok": True}


@api.get("/stampa/rappel/{rid}")
async def stampa_rappel(
    rid: str, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """PDF di un singolo rappel."""
    r = await db.rappel.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Rappel non trovato")
    comp = await db.compagnie.find_one({"id": r["compagnia_id"]}, {"_id": 0}) or {}
    stato = r.get("stato") or "da_incassare"
    headers = ["Campo", "Valore"]
    rows = [
        ["Compagnia", comp.get("ragione_sociale") or "—"],
        ["Codice compagnia", comp.get("codice") or "—"],
        ["Data accredito", r.get("data") or "—"],
        ["Anno competenza", str(r.get("anno") or "—")],
        ["Descrizione", r.get("descrizione") or "—"],
        ["Note", r.get("note") or "—"],
        ["Stato", "INCASSATO" if stato == "incassato" else "DA INCASSARE"],
        ["Data incasso", r.get("data_incasso") or "—"],
        ["Importo €", f"{float(r.get('importo') or 0):,.2f}"],
    ]
    pdf = pdf_report.stampa_elenco(
        f"Rappel — {comp.get('ragione_sociale', 'Compagnia')}",
        f"Documento di sovraprovvigione · ID {rid[:8]}",
        headers, rows,
        col_widths_mm=[55, 110], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"rappel_{rid[:8]}.pdf")


@api.get("/stampa/compagnie/{cid}/estratto-conto")
async def stampa_compagnia_estratto(
    cid: str, dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    data = await _compagnia_estratto_data(cid, dal, al)
    comp = data["compagnia"]
    headers = ["Data", "Tipo", "Polizza", "Contraente", "Ramo", "Dare €", "Avere €"]
    rows = []
    for r in data["righe"]:
        rows.append([
            r.get("data") or "", r.get("tipo") or "",
            r.get("polizza") or "", (r.get("contraente") or "")[:35],
            r.get("ramo") or "", r.get("dare") or "", r.get("avere") or "",
        ])
    rows.append(["", "", "", "", "TOTALE", round(data["totale_dare"], 2), round(data["totale_avere"], 2)])
    rows.append(["", "", "", "", "SALDO DA VERSARE", round(data["saldo"], 2), ""])
    pdf = pdf_report.stampa_elenco(
        f"Estratto conto compagnia - {comp.get('ragione_sociale')}",
        f"Periodo: {dal or '—'} -> {al or '—'} · Trattiene provv: {'SI' if data['trattiene_provvigioni'] else 'NO'}",
        headers, rows,
        col_widths_mm=[22, 22, 32, 60, 22, 28, 28], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"estratto_compagnia_{comp.get('codice') or cid}.pdf")


class TitoliSelectionBody(BaseModel):
    titoli_ids: list[str]


@api.post("/stampa/compagnie/{cid}/titoli-selezionati")
async def stampa_compagnia_titoli_selezionati(
    cid: str, body: TitoliSelectionBody,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Stampa PDF dei soli titoli selezionati di una compagnia, con totale somma."""
    if not body.titoli_ids:
        raise HTTPException(400, "Nessun titolo selezionato")
    comp = await db.compagnie.find_one({"id": cid}, {"_id": 0})
    if not comp:
        raise HTTPException(404, "Compagnia non trovata")
    trattiene = bool(comp.get("trattiene_provvigioni", True))
    titoli = await db.titoli.find({"id": {"$in": body.titoli_ids}}, {"_id": 0}).to_list(5000)
    pol_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
    pol_map = {}
    async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0}):
        pol_map[p["id"]] = p
    ana_ids = list({p.get("contraente_id") for p in pol_map.values() if p.get("contraente_id")})
    ana_map = {}
    async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1}):
        ana_map[a["id"]] = a["ragione_sociale"]

    headers = ["Data incasso", "Polizza", "Contraente", "Ramo", "Lordo €", "Provv. €", "Dovuto €"]
    rows = []
    tot_lordo = 0.0
    tot_provv = 0.0
    tot_dovuto = 0.0
    for t in sorted(titoli, key=lambda x: x.get("data_incasso") or ""):
        pol = pol_map.get(t.get("polizza_id"), {})
        lordo = float(t.get("importo_lordo") or 0)
        provv = float(t.get("provvigioni") or 0)
        dovuto = (lordo - provv) if trattiene else lordo
        rows.append([
            t.get("data_incasso") or "",
            pol.get("numero_polizza") or "",
            (ana_map.get(pol.get("contraente_id", "")) or "")[:35],
            pol.get("ramo") or "",
            round(lordo, 2), round(provv, 2), round(dovuto, 2),
        ])
        tot_lordo += lordo
        tot_provv += provv
        tot_dovuto += dovuto
    rows.append(["", "", "", "TOTALE", round(tot_lordo, 2), round(tot_provv, 2), round(tot_dovuto, 2)])
    pdf = pdf_report.stampa_elenco(
        f"Titoli selezionati — {comp.get('ragione_sociale')}",
        f"N. titoli: {len(titoli)} · Totale dovuto: € {tot_dovuto:,.2f} · Trattiene provv: {'SI' if trattiene else 'NO'}",
        headers, rows,
        col_widths_mm=[24, 30, 60, 22, 25, 25, 28], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"titoli_selezionati_{comp.get('codice') or cid}.pdf")


@api.get("/stampa/rimessa/{mov_id}")
async def stampa_rimessa_compagnia(
    mov_id: str, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Stampa PDF di una singola rimessa (pagamento compagnia) con l'elenco dei titoli versati."""
    mov = await db.movimenti.find_one(
        {"id": mov_id, "categoria": "pagamento_compagnia"}, {"_id": 0},
    )
    if not mov:
        raise HTTPException(404, "Rimessa non trovata")
    comp = await db.compagnie.find_one({"id": mov.get("compagnia_id")}, {"_id": 0}) or {}
    titoli_ids = mov.get("titoli_ids_versati") or mov.get("titoli_versati_ids") or mov.get("titoli_ids") or []
    titoli = []
    if titoli_ids:
        async for t in db.titoli.find({"id": {"$in": titoli_ids}}, {"_id": 0}):
            titoli.append(t)
    pol_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
    pol_map = {}
    async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0}):
        pol_map[p["id"]] = p
    ana_ids = list({p.get("contraente_id") for p in pol_map.values() if p.get("contraente_id")})
    ana_map = {}
    async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1}):
        ana_map[a["id"]] = a["ragione_sociale"]

    headers = ["Polizza", "Ramo", "Contraente", "Data incasso", "Lordo €", "Provv. €"]
    rows = []
    tot_lordo = 0.0
    tot_provv = 0.0
    for t in sorted(titoli, key=lambda x: x.get("data_incasso") or ""):
        pol = pol_map.get(t.get("polizza_id"), {})
        rows.append([
            pol.get("numero_polizza") or "",
            pol.get("ramo") or "",
            (ana_map.get(pol.get("contraente_id", "")) or "")[:38],
            t.get("data_incasso") or "",
            round(float(t.get("importo_lordo") or 0), 2),
            round(float(t.get("provvigioni") or 0), 2),
        ])
        tot_lordo += float(t.get("importo_lordo") or 0)
        tot_provv += float(t.get("provvigioni") or 0)
    rows.append(["", "", "", "TOTALE", round(tot_lordo, 2), round(tot_provv, 2)])
    pdf = pdf_report.stampa_elenco(
        f"Rimessa — {comp.get('ragione_sociale', 'Compagnia')}",
        f"Data: {mov.get('data_movimento') or '—'} · {mov.get('descrizione') or ''} · Importo versato: € {float(mov.get('importo') or 0):,.2f}",
        headers, rows,
        col_widths_mm=[28, 22, 70, 26, 22, 22], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"rimessa_{mov_id[:8]}.pdf")


@api.get("/stampa/pagamento-provvigioni/{pid}")
async def stampa_pagamento_provvigioni(
    pid: str, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Stampa PDF di un singolo estratto conto provvigioni pagato a un collaboratore."""
    pag = await db.pagamenti_provvigioni.find_one({"id": pid}, {"_id": 0})
    if not pag:
        raise HTTPException(404, "Pagamento non trovato")
    collab = await db.users.find_one(
        {"id": pag.get("collaboratore_id")}, {"_id": 0, "password_hash": 0},
    ) or {}
    nome_collab = collab.get("name") or collab.get("email") or pag.get("collaboratore_id")

    titoli_ids = pag.get("titoli_ids") or []
    titoli = []
    if titoli_ids:
        async for t in db.titoli.find({"id": {"$in": titoli_ids}}, {"_id": 0}):
            titoli.append(t)
    pol_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
    pol_map = {}
    async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0}):
        pol_map[p["id"]] = p
    ana_ids = list({p.get("contraente_id") for p in pol_map.values() if p.get("contraente_id")})
    ana_map = {}
    async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1}):
        ana_map[a["id"]] = a["ragione_sociale"]

    headers = ["Polizza", "Ramo", "Contraente", "Data incasso", "Lordo €", "Provv. €"]
    rows = []
    tot_lordo = 0.0
    tot_provv = 0.0
    for t in sorted(titoli, key=lambda x: x.get("data_incasso") or ""):
        pol = pol_map.get(t.get("polizza_id"), {})
        rows.append([
            pol.get("numero_polizza") or "",
            pol.get("ramo") or "",
            (ana_map.get(pol.get("contraente_id", "")) or "")[:38],
            t.get("data_incasso") or "",
            round(float(t.get("importo_lordo") or 0), 2),
            round(float(t.get("provvigioni") or 0), 2),
        ])
        tot_lordo += float(t.get("importo_lordo") or 0)
        tot_provv += float(t.get("provvigioni") or 0)
    rows.append(["", "", "", "TOT. TITOLI", round(tot_lordo, 2), round(tot_provv, 2)])

    # Voci manuali
    voci_ids = pag.get("voci_manuali_ids") or []
    voci = []
    if voci_ids:
        async for v in db.voci_manuali_collab.find({"id": {"$in": voci_ids}}, {"_id": 0}):
            voci.append(v)
    if voci:
        rows.append(["", "", "", "", "", ""])
        rows.append(["VOCI MANUALI", "", "", "", "", ""])
        for v in voci:
            rows.append([
                v.get("data") or "",
                v.get("causale") or "",
                (v.get("note") or "")[:38],
                "", "", round(float(v.get("importo") or 0), 2),
            ])

    sottotitolo = (
        f"Collaboratore: {nome_collab} · Data pagamento: {pag.get('data_pagamento')} "
        f"· Periodo: {pag.get('periodo_dal')} -> {pag.get('periodo_al')}"
        f"\nLordo: € {pag.get('provvigioni_lorde', 0):,.2f}"
        f" · Ritenuta: € -{pag.get('ritenuta_acconto', 0):,.2f}"
        f" · Contributi: € -{pag.get('contributi', 0):,.2f}"
        f" · NETTO PAGATO: € {pag.get('netto_pagato', 0):,.2f}"
        f" · Mezzo: {pag.get('mezzo_pagamento', '—')}"
    )
    pdf = pdf_report.stampa_elenco(
        f"Estratto conto provvigioni pagato — {nome_collab}",
        sottotitolo,
        headers, rows,
        col_widths_mm=[28, 22, 70, 26, 22, 22], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"pagamento_provv_{pid[:8]}.pdf")


@api.get("/stampa/compagnie/saldi-cassa")
async def stampa_saldi_compagnie(
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    items = await saldi_cassa_compagnie(user=user)
    headers = ["Codice", "Ragione sociale", "Tratt. provv.", "Incassato €", "Versato €", "Saldo €"]
    rows = []
    tot_inc = tot_vers = tot_sal = 0.0
    for r in items:
        rows.append([
            r.get("codice") or "", (r.get("ragione_sociale") or "")[:50],
            "SI" if r.get("trattiene_provvigioni") else "NO",
            r.get("totale_incassato"), r.get("totale_versato"), r.get("saldo_da_versare"),
        ])
        tot_inc += float(r.get("totale_incassato") or 0)
        tot_vers += float(r.get("totale_versato") or 0)
        tot_sal += float(r.get("saldo_da_versare") or 0)
    rows.append(["", "TOTALI", "", round(tot_inc, 2), round(tot_vers, 2), round(tot_sal, 2)])
    pdf = pdf_report.stampa_elenco(
        "Saldi cassa compagnie", f"{len(items)} compagnie",
        headers, rows,
        col_widths_mm=[22, 70, 22, 30, 30, 30], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "saldi_compagnie.pdf")


# ============================================================
# ANAGRAFICHE
# ============================================================
# UTILITY — Codice Fiscale & Geocoding
# ============================================================
@api.post("/utility/codice-fiscale/calcola")
async def calcola_codice_fiscale(body: dict, user=Depends(current_user)):
    """body: {nome, cognome, sesso (M/F), data_nascita (YYYY-MM-DD), comune_nascita}"""
    try:
        cf = cf_calc.calcola_cf(
            lastname=body.get("cognome", ""),
            firstname=body.get("nome", ""),
            gender=body.get("sesso", ""),
            birthdate=body.get("data_nascita", ""),
            birthplace=body.get("comune_nascita", ""),
        )
        return {"codice_fiscale": cf}
    except Exception as e:
        raise HTTPException(400, f"Impossibile calcolare CF: {e}")


@api.post("/utility/codice-fiscale/decodifica")
async def decodifica_codice_fiscale(body: dict, user=Depends(current_user)):
    """body: {codice_fiscale}"""
    try:
        return cf_calc.decodifica_cf(body.get("codice_fiscale", ""))
    except Exception as e:
        raise HTTPException(400, f"CF non valido: {e}")


@api.post("/utility/geocoding")
async def utility_geocoding(body: dict, user=Depends(current_user)):
    """body: {indirizzo, comune, cap, provincia, nazione?}. Ritorna {lat, lng, display_name}."""
    result = await geocoder_svc.geocoda_indirizzo(
        indirizzo=body.get("indirizzo", ""),
        comune=body.get("comune"),
        cap=body.get("cap"),
        provincia=body.get("provincia"),
        nazione=body.get("nazione") or "Italia",
    )
    if not result:
        return {"trovato": False}
    return {"trovato": True, **result}


@api.post("/utility/ocr-documento-identita")
async def ocr_documento_identita(
    file: UploadFile = File(...),
    file_retro: Optional[UploadFile] = File(None),
    anagrafica_id: Optional[str] = Form(None),
    tipo: str = Form("carta_identita"),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """OCR universale per CI / patente / passaporto.

    Supporta:
    - 1 PDF (multi-pagina automatico)
    - 1 immagine (solo fronte)
    - 2 immagini (fronte + retro) -> combinate verticalmente prima dell'OCR
    """
    import ocr_ci
    valid = {"carta_identita", "patente", "passaporto"}
    if tipo not in valid:
        raise HTTPException(400, f"Tipo non valido: {tipo}")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 10 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    file_for_ocr = contents
    ct_for_ocr = ct
    contents_retro = None
    ct_retro = None
    if file_retro is not None:
        contents_retro = await file_retro.read()
        ct_retro = file_retro.content_type or obj_storage.mime_for(file_retro.filename or "")
        if len(contents_retro) > 10 * 1024 * 1024:
            raise HTTPException(400, "Retro troppo grande (max 10 MB)")
    if ct == "application/pdf":
        try:
            import pdfplumber
            from io import BytesIO
            with pdfplumber.open(BytesIO(contents)) as pdf:
                if not pdf.pages:
                    raise HTTPException(400, "PDF vuoto")
                img = pdf.pages[0].to_image(resolution=200).original
                out = BytesIO()
                img.save(out, format="JPEG", quality=85)
                file_for_ocr = out.getvalue()
                ct_for_ocr = "image/jpeg"
        except Exception as e:
            raise HTTPException(400, f"Errore conversione PDF: {e}")
    elif not ct.startswith("image/"):
        raise HTTPException(400, "Formato non supportato (PDF/JPG/PNG)")

    # Combina fronte + retro in singola immagine verticale
    if contents_retro and ct_retro and ct_retro.startswith("image/"):
        try:
            from PIL import Image
            from io import BytesIO
            img_f = Image.open(BytesIO(file_for_ocr))
            img_r = Image.open(BytesIO(contents_retro))
            # Normalizza alla stessa larghezza
            w = max(img_f.width, img_r.width)
            def _resize(i):
                if i.width == w: return i
                h = int(i.height * (w / i.width))
                return i.resize((w, h))
            img_f, img_r = _resize(img_f), _resize(img_r)
            combined = Image.new("RGB", (w, img_f.height + img_r.height + 10), "white")
            combined.paste(img_f.convert("RGB"), (0, 0))
            combined.paste(img_r.convert("RGB"), (0, img_f.height + 10))
            out = BytesIO()
            combined.save(out, format="JPEG", quality=85)
            file_for_ocr = out.getvalue()
            ct_for_ocr = "image/jpeg"
        except Exception as e:
            raise HTTPException(400, f"Errore combinazione fronte/retro: {e}")

    try:
        data = await ocr_ci.estrai_dati_ci(file_for_ocr, ct_for_ocr)
    except Exception as e:
        raise HTTPException(503, f"Errore OCR: {e}")

    documento_url = None
    if anagrafica_id:
        # salva file fronte (o PDF) come documento principale; retro come allegato
        ext = (file.filename or "doc.bin").rsplit(".", 1)[-1].lower() or "bin"
        path = f"{os.environ.get('APP_NAME', 'assicura')}/anagrafiche/{anagrafica_id}/{tipo}_{_uid()}.{ext}"
        try:
            result = obj_storage.put_object(path, contents, ct)
            documento_url = f"/api/storage/{result['path']}"
            doc_entry = {
                "url": documento_url, "storage_path": result["path"],
                "nome_file": file.filename, "mime": ct,
                "size_kb": round(len(contents) / 1024, 1),
                "data_caricamento": _now_iso(),
                "scadenza": data.get("data_scadenza"),
                "caricato_da": user.get("id"),
            }
            if contents_retro:
                # salva retro come allegato addizionale
                ext_r = (file_retro.filename or "retro.jpg").rsplit(".", 1)[-1].lower() or "jpg"
                path_r = f"{os.environ.get('APP_NAME', 'assicura')}/anagrafiche/{anagrafica_id}/{tipo}_retro_{_uid()}.{ext_r}"
                try:
                    res_r = obj_storage.put_object(path_r, contents_retro, ct_retro)
                    doc_entry["url_retro"] = f"/api/storage/{res_r['path']}"
                    doc_entry["nome_file_retro"] = file_retro.filename
                except Exception:
                    pass
            await db.anagrafiche.update_one(
                {"id": anagrafica_id},
                {"$set": {f"documenti.{tipo}": doc_entry, "updated_at": _now_iso()}},
            )
        except Exception:
            pass

    await log_attivita(user, "ocr", tipo, anagrafica_id,
                       f"OCR {tipo}{'(fronte+retro)' if contents_retro else ''}: {data.get('cognome')} {data.get('nome')}")
    if documento_url:
        data["_documento_salvato"] = documento_url
    return data


@api.post("/utility/ocr-polizza")
async def ocr_polizza_endpoint(
    file: UploadFile = File(...),
    salva_come_allegato: bool = Form(False),
    polizza_id: Optional[str] = Form(None),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """OCR di una polizza italiana. Estrae dati strutturati (contraente, veicolo, garanzie, premi).

    Se polizza_id + salva_come_allegato -> salva il file come allegato della polizza.
    """
    import ocr_polizza as ocr_pol
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 20 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    file_for_ocr = contents
    ct_for_ocr = ct
    if ct == "application/pdf":
        try:
            import pdfplumber
            from io import BytesIO
            with pdfplumber.open(BytesIO(contents)) as pdf:
                if not pdf.pages:
                    raise HTTPException(400, "PDF vuoto")
                # per polizze processa prime 2 pagine (frontespizio + dettagli)
                images = []
                for p in pdf.pages[:2]:
                    img = p.to_image(resolution=200).original
                    out = BytesIO()
                    img.save(out, format="JPEG", quality=85)
                    images.append(out.getvalue())
                # combina in unica immagine verticale se più pagine
                if len(images) == 1:
                    file_for_ocr = images[0]
                else:
                    from PIL import Image
                    imgs = [Image.open(BytesIO(x)) for x in images]
                    w = max(i.width for i in imgs)
                    h = sum(i.height for i in imgs)
                    combined = Image.new("RGB", (w, h), "white")
                    y = 0
                    for i in imgs:
                        combined.paste(i, (0, y))
                        y += i.height
                    out = BytesIO()
                    combined.save(out, format="JPEG", quality=80)
                    file_for_ocr = out.getvalue()
                ct_for_ocr = "image/jpeg"
        except Exception as e:
            raise HTTPException(400, f"Errore conversione PDF: {e}")
    elif not ct.startswith("image/"):
        raise HTTPException(400, "Formato non supportato (PDF/JPG/PNG)")
    try:
        data = await ocr_pol.estrai_dati_polizza(file_for_ocr, ct_for_ocr)
    except Exception as e:
        raise HTTPException(503, f"Errore OCR polizza: {e}")

    if salva_come_allegato and polizza_id:
        ext = (file.filename or "polizza.pdf").rsplit(".", 1)[-1].lower() or "pdf"
        path = f"{os.environ.get('APP_NAME', 'assicura')}/polizze/{polizza_id}/polizza_{_uid()}.{ext}"
        try:
            result = obj_storage.put_object(path, contents, ct)
            url = f"/api/storage/{result['path']}"
            alleg = Allegato(
                entita_tipo="polizza", entita_id=polizza_id,
                nome_file=file.filename or "polizza.pdf",
                url=url, storage_path=result["path"],
                mime=ct, size_kb=round(len(contents) / 1024, 1),
                caricato_da=user.get("id"),
            )
            await db.allegati.insert_one(alleg.model_dump())
        except Exception:
            pass

    await log_attivita(user, "ocr", "polizza", polizza_id,
                       f"OCR polizza: {data.get('numero_polizza')} - {data.get('compagnia')}")
    return data


@api.post("/utility/ocr-carta-identita")
async def ocr_carta_identita(
    file: UploadFile = File(...),
    anagrafica_id: Optional[str] = Form(None),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """OCR di una carta d'identità italiana. Restituisce dati anagrafici estratti.

    Se anagrafica_id è fornito, salva ANCHE il file come documento 'carta_identita'
    nella scheda cliente.
    """
    import ocr_ci
    contents = await file.read()
    if len(contents) > 8 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 8 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    file_for_ocr = contents
    ct_for_ocr = ct
    if ct == "application/pdf":
        try:
            import pdfplumber
            from io import BytesIO
            with pdfplumber.open(BytesIO(contents)) as pdf:
                if not pdf.pages:
                    raise HTTPException(400, "PDF vuoto")
                img = pdf.pages[0].to_image(resolution=200).original
                out = BytesIO()
                img.save(out, format="JPEG", quality=85)
                file_for_ocr = out.getvalue()
                ct_for_ocr = "image/jpeg"
        except Exception as e:
            raise HTTPException(400, f"Errore conversione PDF: {e}")
    elif not ct.startswith("image/"):
        raise HTTPException(400, "Formato non supportato (richiesto JPG/PNG/PDF)")
    try:
        data = await ocr_ci.estrai_dati_ci(file_for_ocr, ct_for_ocr)
    except Exception as e:
        raise HTTPException(503, f"Errore OCR: {e}")

    # Se collegato a un'anagrafica, salva il file originale come documento
    documento_url = None
    if anagrafica_id:
        ext = (file.filename or "ci.bin").rsplit(".", 1)[-1].lower() or "bin"
        path = f"{os.environ.get('APP_NAME', 'assicura')}/anagrafiche/{anagrafica_id}/carta_identita_{_uid()}.{ext}"
        try:
            result = obj_storage.put_object(path, contents, ct)
            documento_url = f"/api/storage/{result['path']}"
            doc_entry = {
                "url": documento_url, "storage_path": result["path"],
                "nome_file": file.filename, "mime": ct,
                "size_kb": round(len(contents) / 1024, 1),
                "data_caricamento": _now_iso(),
                "scadenza": data.get("data_scadenza"),
                "caricato_da": user.get("id"),
            }
            await db.anagrafiche.update_one(
                {"id": anagrafica_id},
                {"$set": {"documenti.carta_identita": doc_entry, "updated_at": _now_iso()}},
            )
        except Exception:
            pass  # OCR ha successo anche se l'upload fallisce

    await log_attivita(user, "ocr", "carta_identita", anagrafica_id,
                       f"OCR CI: {data.get('cognome')} {data.get('nome')}")
    if documento_url:
        data["_documento_salvato"] = documento_url
    return data






@api.post("/utility/ocr-visura-camerale")
async def ocr_visura_camerale(
    file: UploadFile = File(...),
    anagrafica_id: Optional[str] = Form(None),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """OCR di una visura camerale. Estrae dati ditta + amministratori.

    Se anagrafica_id è fornito, salva il file come documento 'visura_camerale'.
    Gli amministratori vengono restituiti per importazione anagrafiche legate.
    """
    import ocr_visura
    contents = await file.read()
    if len(contents) > 15 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 15 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    file_for_ocr = contents
    ct_for_ocr = ct
    if ct == "application/pdf":
        try:
            import pdfplumber
            from io import BytesIO
            with pdfplumber.open(BytesIO(contents)) as pdf:
                if not pdf.pages:
                    raise HTTPException(400, "PDF vuoto")
                # per visure camerali analizziamo solo la prima pagina (sintesi)
                img = pdf.pages[0].to_image(resolution=200).original
                out = BytesIO()
                img.save(out, format="JPEG", quality=85)
                file_for_ocr = out.getvalue()
                ct_for_ocr = "image/jpeg"
        except Exception as e:
            raise HTTPException(400, f"Errore conversione PDF: {e}")
    elif not ct.startswith("image/"):
        raise HTTPException(400, "Formato non supportato (PDF/JPG/PNG)")
    try:
        data = await ocr_visura.estrai_dati_visura(file_for_ocr, ct_for_ocr)
    except Exception as e:
        raise HTTPException(503, f"Errore OCR visura: {e}")

    # Salva il file come documento dell'azienda
    if anagrafica_id:
        ext = (file.filename or "visura.pdf").rsplit(".", 1)[-1].lower() or "pdf"
        path = f"{os.environ.get('APP_NAME', 'assicura')}/anagrafiche/{anagrafica_id}/visura_camerale_{_uid()}.{ext}"
        try:
            result = obj_storage.put_object(path, contents, ct)
            doc_entry = {
                "url": f"/api/storage/{result['path']}", "storage_path": result["path"],
                "nome_file": file.filename, "mime": ct,
                "size_kb": round(len(contents) / 1024, 1),
                "data_caricamento": _now_iso(),
                "caricato_da": user.get("id"),
            }
            await db.anagrafiche.update_one(
                {"id": anagrafica_id},
                {"$set": {"documenti.visura_camerale": doc_entry, "updated_at": _now_iso()}},
            )
        except Exception:
            pass

    await log_attivita(user, "ocr", "visura_camerale", anagrafica_id,
                       f"OCR visura: {data.get('ragione_sociale')} - {len(data.get('amministratori') or [])} amministratori")
    return data


# ----------------------------------------------------------
# Anagrafiche routes moved to routes/anagrafiche.py (iter23).
# Block: KPI custom + tags + stats + CRUD + network + relazioni +
# documenti + privacy GDPR + firma digitale + INPS auto + interviste.
# Registered below via: api.include_router(_anag_router.router)
# ----------------------------------------------------------



# ============================================================
# POLIZZE
# ============================================================
@api.get("/polizze")
async def list_polizze(
    q: Optional[str] = None,
    stato: Optional[str] = None,
    ramo: Optional[str] = None,
    prodotto: Optional[str] = None,
    contraente_id: Optional[str] = None,
    compagnia_id: Optional[str] = None,
    collaboratore_id: Optional[str] = None,
    categoria: Optional[str] = None,  # auto_priv|auto_az|altri_priv|altri_az|vita_inv|vita_prot
    catastrofale: Optional[bool] = None,
    check_up: Optional[bool] = None,
    inabilita_malattia: Optional[bool] = None,
    tutela_legale: Optional[bool] = None,
    infortuni_conducente: Optional[bool] = None,
    # filtri periodo (su scadenza)
    dal: Optional[str] = None,
    al: Optional[str] = None,
    in_scadenza_giorni: Optional[int] = None,
    scadenza_oltre_giorni: Optional[int] = None,
    scadute_oggi: Optional[bool] = None,
    scadute_da_min: Optional[int] = None,
    scadute_da_max: Optional[int] = None,
    limit: int = 50000,
    user=Depends(current_user),
):
    flt = await visibility_filter(user)
    if stato:
        flt["stato"] = stato
    if ramo:
        flt["ramo"] = ramo
    if prodotto:
        flt["prodotto"] = prodotto
    if contraente_id:
        flt["contraente_id"] = contraente_id
    if compagnia_id:
        flt["compagnia_id"] = compagnia_id
    if collaboratore_id:
        flt["collaboratore_id"] = collaboratore_id
    if q:
        qrx = {"$regex": q, "$options": "i"}
        ana_ids = [a["id"] async for a in db.anagrafiche.find(
            {"ragione_sociale": qrx}, {"_id": 0, "id": 1}
        )]
        or_cond = [{"numero_polizza": qrx}, {"targa": qrx}]
        if ana_ids:
            or_cond.append({"contraente_id": {"$in": ana_ids}})
        flt["$or"] = or_cond

    # filtri di scadenza (presets)
    from datetime import date, timedelta
    today = date.today()
    today_s = today.isoformat()
    scad_cond: dict = {}
    if in_scadenza_giorni is not None:
        limite = (today + timedelta(days=int(in_scadenza_giorni))).isoformat()
        scad_cond["$gte"] = today_s
        scad_cond["$lte"] = limite
    if scadenza_oltre_giorni is not None:
        oltre = (today + timedelta(days=int(scadenza_oltre_giorni))).isoformat()
        scad_cond["$gt"] = oltre
    if scadute_oggi:
        flt["scadenza"] = today_s
    if scadute_da_min is not None:
        scad_cond["$lte"] = (today - timedelta(days=int(scadute_da_min))).isoformat()
    if scadute_da_max is not None:
        scad_cond["$gte"] = (today - timedelta(days=int(scadute_da_max))).isoformat()
    if dal:
        scad_cond["$gte"] = dal
    if al:
        scad_cond["$lte"] = al
    if scad_cond and "scadenza" not in flt:
        flt["scadenza"] = scad_cond

    items = await db.polizze.find(flt, {"_id": 0}).sort("scadenza", 1).to_list(limit)
    # enrich con contraente, compagnia, collaboratore
    ana_ids = list({i["contraente_id"] for i in items if i.get("contraente_id")})
    comp_ids = list({i["compagnia_id"] for i in items if i.get("compagnia_id")})
    collab_ids = list({i.get("collaboratore_id") for i in items if i.get("collaboratore_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}},
                                                          {"_id": 0, "id": 1, "ragione_sociale": 1})}
    comps = {c["id"]: c async for c in db.compagnie.find({"id": {"$in": comp_ids}},
                                                         {"_id": 0, "id": 1, "ragione_sociale": 1, "codice": 1})}
    collabs = {u["id"]: u async for u in db.users.find({"id": {"$in": collab_ids}},
                                                       {"_id": 0, "id": 1, "name": 1, "avatar_url": 1})}
    for it in items:
        it["contraente_nome"] = anas.get(it.get("contraente_id"), {}).get("ragione_sociale")
        it["compagnia_nome"] = comps.get(it.get("compagnia_id"), {}).get("ragione_sociale")
        _c = collabs.get(it.get("collaboratore_id", ""), {})
        it["collaboratore_nome"] = _c.get("name")
        it["collaboratore_avatar_url"] = _c.get("avatar_url")
    # Filtro categoria business (post-processing — richiede tipo anagrafica)
    if catastrofale is not None:
        items = [p for p in items if (p.get("catastrofale") is True) == catastrofale]
    if check_up is not None:
        items = [p for p in items if (p.get("check_up") is True) == check_up]
    if inabilita_malattia is not None:
        items = [p for p in items if (p.get("inabilita_malattia") is True) == inabilita_malattia]
    if tutela_legale is not None:
        items = [p for p in items if (p.get("tutela_legale") is True) == tutela_legale]
    if infortuni_conducente is not None:
        items = [p for p in items if (p.get("infortuni_conducente") is True) == infortuni_conducente]
    if categoria:
        anag_tipo = {a["id"]: (a.get("tipo") or "persona_fisica")
                     async for a in db.anagrafiche.find(
                         {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "tipo": 1, "tags": 1})}
        # azienda override via tags
        async for a in db.anagrafiche.find(
                {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "tags": 1}):
            tags = [t.lower() for t in (a.get("tags") or [])]
            if "azienda" in tags or "condominio" in tags:
                anag_tipo[a["id"]] = "persona_giuridica"

        def _cat(p):
            ramo_u = (p.get("ramo") or "").upper()
            prod_u = (p.get("prodotto") or "").upper()
            tipo = anag_tipo.get(p.get("contraente_id"), "persona_fisica")
            is_priv = tipo == "persona_fisica"
            is_vita = "VITA" in ramo_u or "VITA" in prod_u
            is_auto = ("AUTO" in ramo_u and ("RC" in ramo_u or ramo_u.startswith("AUTO"))) \
                      or ramo_u == "RCA" or "RCAUTO" in ramo_u
            if is_vita:
                return "vita_inv" if ("INVEST" in prod_u or "INVEST" in ramo_u) else "vita_prot"
            if is_auto:
                return "auto_priv" if is_priv else "auto_az"
            return "altri_priv" if is_priv else "altri_az"

        items = [p for p in items if _cat(p) == categoria]
    return items


@api.get("/polizze/veicolo-by-targa")
async def veicolo_by_targa(
    targa: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Cerca l'ultima polizza con la stessa targa e restituisce i dati veicolo
    per auto-popolare il form 'Nuova polizza'. Restituisce {trovata: bool, ...campi}.
    """
    t = (targa or "").strip().upper().replace(" ", "")
    if not t or len(t) < 4:
        return {"trovata": False}
    # Trova polizze con targa che combacia (case-insensitive su forma normalizzata)
    polizze = await db.polizze.find(
        {"targa": {"$regex": f"^{t}$", "$options": "i"}},
        {"_id": 0},
    ).sort("created_at", -1).to_list(20)
    if not polizze:
        return {"trovata": False}
    last = polizze[0]
    return {
        "trovata": True,
        "n_polizze": len(polizze),
        "veicolo_marca": last.get("veicolo_marca") or last.get("marca"),
        "veicolo_modello": last.get("veicolo_modello") or last.get("modello"),
        "veicolo_tipo": last.get("veicolo_tipo") or last.get("tipo_veicolo"),
        "veicolo_alimentazione": last.get("veicolo_alimentazione") or last.get("alimentazione"),
        "veicolo_kw": last.get("veicolo_kw") or last.get("kw"),
        "veicolo_cv_fiscali": last.get("veicolo_cv_fiscali") or last.get("cv") or last.get("cv_fiscali"),
        "veicolo_cilindrata": last.get("veicolo_cilindrata") or last.get("cilindrata"),
        "veicolo_data_immatricolazione": last.get("veicolo_data_immatricolazione") or last.get("data_immatricolazione"),
        "veicolo_uso": last.get("veicolo_uso") or last.get("tipo_uso"),
        "veicolo_posti": last.get("veicolo_posti") or last.get("numero_posti"),
        "telaio": last.get("telaio"),
        "ultima_polizza_id": last.get("id"),
        "ultima_compagnia_id": last.get("compagnia_id"),
        "ultimo_contraente_id": last.get("contraente_id"),
    }


@api.get("/polizze/{pid}")
async def get_polizza(pid: str, user=Depends(current_user)):
    doc = await db.polizze.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Non trovata")
    if user["role"] == "cliente" and user.get("anagrafica_id") != doc.get("contraente_id"):
        raise HTTPException(403, "Permesso negato")
    # 🔧 FIX: se `prodotto` contiene un UUID di ProdottoLibreria (per errore di
    # importazione), risolvi con il nome vero + migra il dato in DB.
    prodotto_val = doc.get("prodotto")
    if prodotto_val and isinstance(prodotto_val, str) and len(prodotto_val) == 36 and prodotto_val.count("-") == 4:
        prod_lib = await db.prodotti.find_one(
            {"id": prodotto_val}, {"_id": 0, "nome": 1},
        )
        if prod_lib and prod_lib.get("nome"):
            doc["prodotto"] = prod_lib["nome"]
            # migrazione idempotente: aggiorna il record
            await db.polizze.update_one(
                {"id": pid},
                {"$set": {"prodotto": prod_lib["nome"], "updated_at": _now_iso()}},
            )
    doc["contraente"] = await db.anagrafiche.find_one({"id": doc["contraente_id"]}, {"_id": 0})
    doc["compagnia"] = await db.compagnie.find_one({"id": doc["compagnia_id"]}, {"_id": 0})
    # Arricchimento collaboratore (operatore)
    doc["collaboratore_nome"] = None
    if doc.get("collaboratore_id"):
        u = await db.users.find_one(
            {"id": doc["collaboratore_id"]}, {"_id": 0, "name": 1, "email": 1},
        )
        if u:
            doc["collaboratore_nome"] = u.get("name") or u.get("email")
    # Breakdown provvigione (totale REALE polizza -> collab + margine via schema)
    bk = await _provv_breakdown(
        float(doc.get("provvigioni") or 0),
        doc.get("collaboratore_id"),
        doc.get("compagnia_id"),
        doc.get("ramo"),
    )
    doc["provvigione_totale"] = bk["provvigione_totale"]
    doc["provvigione_collaboratore"] = bk["provvigione_collaboratore"]
    doc["provvigione_margine"] = bk["provvigione_margine"]
    doc["provvigione_pct_collab"] = bk["pct_collab"]
    doc["provvigione_schema_nome"] = bk["schema_nome"]
    # 📅 SCADENZA COPERTURA: se ci sono titoli incassati/coperti, la copertura effettiva
    # arriva fino alla scadenza dell'ULTIMO titolo pagato/coperto.
    ultimo_titolo = await db.titoli.find_one(
        {"polizza_id": pid, "stato": {"$in": ["incassato", "coperto", "abbuonato"]}},
        {"_id": 0, "scadenza": 1, "coperto_fino_a": 1, "data_incasso": 1, "stato": 1, "numero_titolo": 1},
        sort=[("scadenza", -1)],
    )
    if ultimo_titolo:
        doc["copertura_fino_a"] = ultimo_titolo.get("coperto_fino_a") or ultimo_titolo.get("scadenza")
        doc["ultimo_titolo_stato"] = ultimo_titolo.get("stato")
        doc["ultimo_titolo_scadenza"] = ultimo_titolo.get("scadenza")
        doc["ultimo_titolo_numero"] = ultimo_titolo.get("numero_titolo")
    else:
        doc["copertura_fino_a"] = None
        doc["ultimo_titolo_stato"] = None
    # Raccogli IDs della catena di polizze sostituite (storia)
    polizza_ids_catena = [pid]
    cursor_id = doc.get("sostituisce_polizza")
    visited = {pid}
    while cursor_id and cursor_id not in visited:
        visited.add(cursor_id)
        polizza_ids_catena.append(cursor_id)
        prev = await db.polizze.find_one({"id": cursor_id}, {"_id": 0, "sostituisce_polizza": 1, "numero_polizza": 1})
        if not prev:
            break
        cursor_id = prev.get("sostituisce_polizza")
    # Carica titoli di tutta la catena (polizza corrente + sostituite)
    doc["titoli"] = await db.titoli.find(
        {"polizza_id": {"$in": polizza_ids_catena}}, {"_id": 0},
    ).sort("effetto", -1).to_list(500)
    # Mappa numeri polizza per arricchire i titoli storici
    if len(polizza_ids_catena) > 1:
        pmap = {p["id"]: p async for p in db.polizze.find(
            {"id": {"$in": polizza_ids_catena}},
            {"_id": 0, "id": 1, "numero_polizza": 1, "stato": 1},
        )}
        for t in doc["titoli"]:
            if t.get("polizza_id") != pid:
                p_old = pmap.get(t["polizza_id"]) or {}
                t["polizza_origine_id"] = t["polizza_id"]
                t["polizza_origine_numero"] = p_old.get("numero_polizza")
                t["polizza_origine_stato"] = p_old.get("stato")
    # Arricchisci ogni titolo con breakdown provvigione (sul suo provvigioni reale)
    for t in doc["titoli"]:
        tbk = await _provv_breakdown(
            float(t.get("provvigioni") or 0),
            t.get("collaboratore_id") or doc.get("collaboratore_id"),
            doc.get("compagnia_id"),
            doc.get("ramo"),
        )
        t["provvigione_totale"] = tbk["provvigione_totale"]
        t["provvigione_collaboratore"] = tbk["provvigione_collaboratore"]
        t["provvigione_margine"] = tbk["provvigione_margine"]
        t["provvigione_pct_collab"] = tbk["pct_collab"]
    # Arricchimento count allegati per ogni titolo (visibilità in PolizzaDetail > Titoli)
    tids = [t["id"] for t in doc["titoli"]]
    if tids:
        ag_pipeline = [
            {"$match": {"entita_tipo": "titolo", "entita_id": {"$in": tids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]
        ag_counts = {r["_id"]: r["n"] async for r in db.allegati.aggregate(ag_pipeline)}
        for t in doc["titoli"]:
            t["allegati_count"] = ag_counts.get(t["id"], 0)
    doc["sinistri"] = await db.sinistri.find({"polizza_id": pid}, {"_id": 0}).sort("data_avvenimento", -1).to_list(100)
    return doc


@api.post("/polizze", status_code=201)
async def create_polizza(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    # Applica default termini_mora_giorni dal prodotto in libreria,
    # altrimenti dal ramo (Vita = 30gg, altri = 15gg).
    if body.get("termini_mora_giorni") in (None, 0, ""):
        mora = None
        if body.get("prodotto"):
            prod = await db.prodotti.find_one(
                {"nome": body["prodotto"]}, {"_id": 0, "termini_mora_giorni": 1},
            )
            if prod and prod.get("termini_mora_giorni"):
                mora = int(prod["termini_mora_giorni"])
        if mora is None:
            from db_models import default_mora_for_ramo
            mora = default_mora_for_ramo(body.get("ramo"))
        body["termini_mora_giorni"] = mora
    # Auto-set is_libro_matricola dal prodotto in libreria
    if body.get("prodotto") and "is_libro_matricola" not in body:
        prod_lm = await db.prodotti.find_one(
            {"nome": body["prodotto"]}, {"_id": 0, "is_libro_matricola": 1},
        )
        if prod_lm and prod_lm.get("is_libro_matricola"):
            body["is_libro_matricola"] = True
    # Auto-calcola provvigioni se non specificate manualmente e c'è schema applicabile
    if (not body.get("provvigioni")) and body.get("premio_lordo") and body.get("collaboratore_id"):
        calc = await _calcola_provvigione(
            float(body["premio_lordo"]),
            body.get("collaboratore_id"),
            body.get("compagnia_id"),
            body.get("ramo"),
        )
        if calc["provvigione_totale"]:
            body["provvigioni"] = calc["provvigione_totale"]
    obj = Polizza(**body)
    await db.polizze.insert_one(obj.model_dump())
    await log_attivita(user, "create", "polizza", obj.id, f"Polizza {obj.numero_polizza}")
    # Alert: polizza emessa
    from alert_dispatcher import safe_dispatch
    await safe_dispatch("polizza.emessa", {
        "entita_tipo": "polizza", "entita_id": obj.id,
        "anagrafica_id": obj.contraente_id,
        "polizza_id": obj.id,
        "collaboratore_id": obj.collaboratore_id,
        "numero_polizza": obj.numero_polizza,
        "ramo": obj.ramo,
        "data_effetto": obj.effetto,
        "scadenza": obj.scadenza,
        "premio_totale": obj.premio_totale,
        "link": f"/polizze/{obj.id}",
    })
    return obj.model_dump()


@api.put("/polizze/{pid}")
async def update_polizza(pid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body["updated_at"] = _now_iso()
    res = await db.polizze.update_one({"id": pid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Non trovata")
    await log_attivita(user, "update", "polizza", pid)
    return strip_mongo_id(await db.polizze.find_one({"id": pid}, {"_id": 0}))


@api.delete("/polizze/{pid}")
async def delete_polizza(pid: str, user=Depends(require_user("admin"))):
    await db.polizze.delete_one({"id": pid})
    await log_attivita(user, "delete", "polizza", pid)
    return {"ok": True}


# ============================================================
# TITOLI
# ============================================================
@api.get("/titoli")
async def list_titoli(
    polizza_id: Optional[str] = None,
    stato: Optional[str] = None,
    stato_not: Optional[str] = None,  # iter23: esclude stati (es. "incassato,stornato")
    compagnia_id: Optional[str] = None,
    collaboratore_id: Optional[str] = None,
    ramo: Optional[str] = None,
    prodotto: Optional[str] = None,
    mezzo_pagamento: Optional[str] = None,
    conto_cassa_id: Optional[str] = None,
    in_scadenza_giorni: Optional[int] = None,
    coperti_non_pagati: Optional[bool] = None,
    titolo_coperto: Optional[bool] = None,  # filtra i titoli con flag titolo_coperto=true (anticipi)
    # nuovi filtri scadenza dettagliata
    scadute_oggi: Optional[bool] = None,
    scadute_da_min: Optional[int] = None,    # es. 5 (scadute da almeno 5 giorni)
    scadute_da_max: Optional[int] = None,    # es. 14 (scadute al massimo da 14 giorni)
    scadenza_oltre_giorni: Optional[int] = None,  # in scadenza oltre N gg
    # filtro periodo (sulla scadenza)
    dal: Optional[str] = None,
    al: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50000,
    user=Depends(current_user),
):
    flt: dict = {}
    if polizza_id:
        flt["polizza_id"] = polizza_id
    if stato:
        flt["stato"] = stato
    if stato_not:
        # csv: "incassato,stornato"
        stati_excl = [s.strip() for s in stato_not.split(",") if s.strip()]
        if stati_excl:
            flt["stato"] = {"$nin": stati_excl}

    # filtri lato polizza (compagnia/collaboratore/ramo/prodotto/q-su-targa-numero)
    pol_filter: dict = {}
    if compagnia_id: pol_filter["compagnia_id"] = compagnia_id
    if collaboratore_id: pol_filter["collaboratore_id"] = collaboratore_id
    if ramo: pol_filter["ramo"] = ramo
    if prodotto: pol_filter["prodotto"] = prodotto
    if user["role"] == "cliente" and user.get("anagrafica_id"):
        pol_filter["contraente_id"] = user["anagrafica_id"]

    if q:
        # ricerca generica: numero polizza, targa, contraente
        qrx = {"$regex": q, "$options": "i"}
        pol_q = {"$or": [{"numero_polizza": qrx}, {"targa": qrx}]}
        # cerca anche tra anagrafiche
        ana_ids = [a["id"] async for a in db.anagrafiche.find(
            {"ragione_sociale": qrx}, {"_id": 0, "id": 1}
        )]
        if ana_ids:
            pol_q["$or"].append({"contraente_id": {"$in": ana_ids}})
        if pol_filter:
            pol_filter = {"$and": [pol_filter, pol_q]}
        else:
            pol_filter = pol_q

    if pol_filter:
        pol_ids = [p["id"] async for p in db.polizze.find(pol_filter, {"_id": 0, "id": 1})]
        flt["polizza_id"] = {"$in": pol_ids}

    if mezzo_pagamento:
        flt["mezzo_pagamento"] = mezzo_pagamento
    if conto_cassa_id:
        flt["conto_cassa_id"] = conto_cassa_id

    from datetime import date, timedelta
    today = date.today()
    today_s = today.isoformat()

    if in_scadenza_giorni is not None:
        limite = (today + timedelta(days=int(in_scadenza_giorni))).isoformat()
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        flt["scadenza"] = {"$gte": today_s, "$lte": limite}

    if scadenza_oltre_giorni is not None:
        oltre = (today + timedelta(days=int(scadenza_oltre_giorni))).isoformat()
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        flt["scadenza"] = {"$gt": oltre}

    if scadute_oggi:
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        flt["scadenza"] = today_s

    if scadute_da_min is not None or scadute_da_max is not None:
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        cond = {}
        if scadute_da_min is not None:
            # scadenza <= oggi - min
            cond["$lte"] = (today - timedelta(days=int(scadute_da_min))).isoformat()
        if scadute_da_max is not None:
            # scadenza >= oggi - max
            cond["$gte"] = (today - timedelta(days=int(scadute_da_max))).isoformat()
        flt["scadenza"] = cond if isinstance(flt.get("scadenza"), str) is False else {**(flt.get("scadenza") if isinstance(flt.get("scadenza"), dict) else {}), **cond}

    if coperti_non_pagati:
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        flt["effetto"] = {"$lte": today_s}
        flt["scadenza"] = {"$gte": today_s}

    # Filtra esplicitamente i titoli coperti (anticipi dall'agenzia)
    if titolo_coperto is True:
        flt["titolo_coperto"] = True
        flt["data_copertura"] = {"$ne": None}
    elif titolo_coperto is False:
        flt["$or"] = [{"titolo_coperto": {"$ne": True}}, {"titolo_coperto": None}]

    # filtro periodo (sulla scadenza, sovrascrive eventuali altri set sulla scadenza)
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        if isinstance(flt.get("scadenza"), dict):
            flt["scadenza"] = {**flt["scadenza"], **cond}
        else:
            flt["scadenza"] = cond

    items = await db.titoli.find(flt, {"_id": 0}).sort("scadenza", 1).to_list(limit)
    pol_ids = list({t["polizza_id"] for t in items if t.get("polizza_id")})
    pols = {p["id"]: p async for p in db.polizze.find(
        {"id": {"$in": pol_ids}}, {"_id": 0, "id": 1, "numero_polizza": 1, "contraente_id": 1,
                                   "ramo": 1, "prodotto": 1, "compagnia_id": 1,
                                   "collaboratore_id": 1, "targa": 1})}
    ana_ids = list({p.get("contraente_id") for p in pols.values() if p.get("contraente_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    com_ids = list({p.get("compagnia_id") for p in pols.values() if p.get("compagnia_id")})
    coms = {c["id"]: c async for c in db.compagnie.find(
        {"id": {"$in": com_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    collab_ids = list({p.get("collaboratore_id") for p in pols.values() if p.get("collaboratore_id")})
    collabs = {u["id"]: u async for u in db.users.find(
        {"id": {"$in": collab_ids}}, {"_id": 0, "id": 1, "name": 1})}
    # Lookup nomi prodotti (Polizza.prodotto è un FK id verso db.prodotti)
    prod_ids = list({p.get("prodotto") for p in pols.values() if p.get("prodotto")})
    prodotti_map = {pr["id"]: (pr.get("nome") or pr.get("ramo") or "")
                    async for pr in db.prodotti.find(
                        {"id": {"$in": prod_ids}}, {"_id": 0, "id": 1, "nome": 1, "ramo": 1})}
    for t in items:
        p = pols.get(t.get("polizza_id"), {})
        t["numero_polizza"] = p.get("numero_polizza")
        t["ramo"] = p.get("ramo")
        prod_raw = p.get("prodotto") or ""
        # Risolvi sempre il nome del prodotto, mai un UUID
        t["prodotto"] = prodotti_map.get(prod_raw) or (
            "" if (prod_raw and len(prod_raw) >= 32 and "-" in prod_raw and " " not in prod_raw)
            else prod_raw
        )
        t["targa"] = p.get("targa")
        t["contraente_id"] = p.get("contraente_id")
        t["contraente_nome"] = anas.get(p.get("contraente_id", ""), {}).get("ragione_sociale")
        t["compagnia_nome"] = coms.get(p.get("compagnia_id", ""), {}).get("ragione_sociale")
        t["collaboratore_id"] = p.get("collaboratore_id")
        t["collaboratore_nome"] = collabs.get(p.get("collaboratore_id", ""), {}).get("name")
        t["mezzo_pagamento_preferito"] = p.get("mezzo_pagamento_preferito")
        t["ultimo_mezzo_pagamento"] = p.get("ultimo_mezzo_pagamento")
        # Breakdown provvigione (totale REALE -> quota collab + margine via schema)
        tbk = await _provv_breakdown(
            float(t.get("provvigioni") or 0),
            t.get("collaboratore_id"),
            p.get("compagnia_id"),
            p.get("ramo"),
        )
        t["provvigione_totale"] = tbk["provvigione_totale"]
        t["provvigione_collaboratore"] = tbk["provvigione_collaboratore"]
        t["provvigione_margine"] = tbk["provvigione_margine"]
        t["provvigione_pct_collab"] = tbk["pct_collab"]
    # arricchimento count allegati
    tids = [t["id"] for t in items]
    if tids:
        pipeline = [
            {"$match": {"entita_tipo": "titolo", "entita_id": {"$in": tids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]
        counts = {r["_id"]: r["n"] async for r in db.allegati.aggregate(pipeline)}
        for t in items:
            t["allegati_count"] = counts.get(t["id"], 0)
    return items


@api.post("/titoli/bulk-azione-allegato")
async def bulk_azione_allegato(
    action: str = Query(...),  # "incassa" | "copertura"
    ids_json: str = Query(...),  # JSON list di titoli
    data_incasso: Optional[str] = Query(None),
    mezzo_pagamento: Optional[str] = Query("bonifico"),
    conto_cassa_id: Optional[str] = Query(None),
    coperto_fino_a: Optional[str] = Query(None),
    data_copertura: Optional[str] = Query(None),
    invia_cliente: bool = Query(False),
    invia_collaboratore: bool = Query(False),
    note_email: Optional[str] = Query(None),
    file: Optional[UploadFile] = File(None),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Esegue bulk incasso/copertura con allegato opzionale e invio email a cliente/collaboratore."""
    import json as _json
    try:
        ids = _json.loads(ids_json) if isinstance(ids_json, str) else ids_json
    except Exception:
        raise HTTPException(400, "ids_json non valido")
    if not ids:
        raise HTTPException(400, "Nessun titolo selezionato")

    # 1) carica file su object storage (se presente)
    allegato_meta = None
    if file:
        data = await file.read()
        if len(data) > 25 * 1024 * 1024:
            raise HTTPException(400, "File troppo grande (max 25 MB)")
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
        file_id = _uid()
        path = f"{os.environ.get('APP_NAME', 'assicura')}/titoli/{file_id}.{ext}"
        ct = file.content_type or obj_storage.mime_for(file.filename or "")
        try:
            result = obj_storage.put_object(path, data, ct)
        except Exception as e:
            raise HTTPException(503, f"Errore upload: {e}")
        allegato_meta = {
            "nome_file": file.filename, "storage_path": result["path"],
            "content_type": ct, "size": result.get("size", len(data)),
        }

    # 2) esegui l'azione sui titoli
    risultato = {}
    if action == "incassa":
        risultato = await bulk_incassa({
            "ids": ids, "data_incasso": data_incasso,
            "mezzo_pagamento": mezzo_pagamento, "conto_cassa_id": conto_cassa_id,
        }, user)
    elif action == "copertura":
        coperto = data_copertura or coperto_fino_a or _now_iso()[:10]
        risultato = await bulk_copertura({"ids": ids, "data_copertura": coperto}, user)
    else:
        raise HTTPException(400, "action non valida")

    # 3) raccogli contraenti e collaboratori coinvolti
    titoli = await db.titoli.find({"id": {"$in": ids}}, {"_id": 0}).to_list(2000)
    pol_ids = list({t["polizza_id"] for t in titoli})
    polizze = await db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0}).to_list(2000)
    ana_ids = list({p["contraente_id"] for p in polizze if p.get("contraente_id")})
    collab_ids = list({p["collaboratore_id"] for p in polizze if p.get("collaboratore_id")})
    anas = await db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0}).to_list(500)
    collabs = await db.users.find({"id": {"$in": collab_ids}}, {"_id": 0, "password_hash": 0}).to_list(100)

    azione_label = "Quietanza incasso titolo" if action == "incassa" else "Conferma copertura titolo"
    email_create = 0

    # 4) salva l'allegato come record collegato a OGNI titolo selezionato
    #    (iter23: prima era linkato solo alla prima anagrafica -> bug)
    allegato_id = None
    allegati_ids: list[str] = []
    if allegato_meta and ids:
        for tid in ids:
            al = Allegato(
                entita_tipo="titolo", entita_id=tid,
                nome_file=allegato_meta["nome_file"], storage_path=allegato_meta["storage_path"],
                content_type=allegato_meta["content_type"], size=allegato_meta["size"],
                descrizione=f"{azione_label}", autore_id=user["id"],
            )
            await db.allegati.insert_one(al.model_dump())
            allegati_ids.append(al.id)
        allegato_id = allegati_ids[0] if allegati_ids else None
        # link anche al primo cliente per visibilità nel diario/anagrafica
        if anas:
            al_ana = Allegato(
                entita_tipo="anagrafica", entita_id=anas[0]["id"],
                nome_file=allegato_meta["nome_file"], storage_path=allegato_meta["storage_path"],
                content_type=allegato_meta["content_type"], size=allegato_meta["size"],
                descrizione=f"{azione_label} — {len(ids)} titoli", autore_id=user["id"],
            )
            await db.allegati.insert_one(al_ana.model_dump())

    # 5) crea email in coda per ogni cliente
    if invia_cliente:
        for ana in anas:
            if not ana.get("email"):
                continue
            corpo = (note_email or
                     f"Gentile {ana['ragione_sociale']},\n\n"
                     f"in allegato {azione_label.lower()} relativa alle Sue polizze.\n\n"
                     f"Cordiali saluti.")
            e = EmailMessaggio(
                destinatario_anagrafica_id=ana["id"], destinatario_email=ana["email"],
                oggetto=azione_label, corpo=corpo,
                template="titolo_quietanza", stato="in_coda", autore_id=user["id"],
            )
            await db.email.insert_one(e.model_dump())
            email_create += 1
            # auto-diario sul cliente
            await log_diario_cliente(
                ana["id"], "documento",
                titolo=f"📎 {azione_label}",
                descrizione=f"{len(ids)} titoli interessati. File: {allegato_meta['nome_file'] if allegato_meta else '—'}",
                autore=user,
            )

    # 6) email ai collaboratori
    if invia_collaboratore:
        for c in collabs:
            if not c.get("email"):
                continue
            corpo = (f"Ciao {c.get('name', '')},\n\n"
                     f"{azione_label} per {len(ids)} titoli delle polizze da te gestite.\n"
                     f"Operazione effettuata da {user.get('name')}.\n\n"
                     + (note_email or ""))
            e = EmailMessaggio(
                destinatario_email=c["email"], oggetto=f"{azione_label} (notifica collaboratore)",
                corpo=corpo, template="titolo_notifica_collab", stato="in_coda",
                autore_id=user["id"],
            )
            await db.email.insert_one(e.model_dump())
            email_create += 1

    return {
        **risultato,
        "allegato_id": allegato_id,
        "allegato_nome": allegato_meta["nome_file"] if allegato_meta else None,
        "email_create": email_create,
        "clienti_notificati": len([a for a in anas if a.get("email")]) if invia_cliente else 0,
        "collaboratori_notificati": len([c for c in collabs if c.get("email")]) if invia_collaboratore else 0,
    }


# Bulk actions sui titoli
@api.get("/titoli/sospesi")
async def titoli_sospesi(
    collaboratore_id: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Restituisce i titoli COPERTI ma NON INCASSATI (sospesi/anticipati dall'agenzia).

    Per ogni titolo restituisce: contraente, collaboratore, data_copertura,
    scadenza polizza, importo lordo. Usato dalla pagina 'Sospesi'.
    """
    flt = {"titolo_coperto": True, "stato": {"$in": ["da_incassare", "insoluto"]}}
    items = await db.titoli.find(flt, {"_id": 0}).sort("data_copertura", 1).to_list(5000)
    if not items:
        return []
    pol_ids = list({t["polizza_id"] for t in items if t.get("polizza_id")})
    pols = {p["id"]: p async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0})}
    ana_ids = list({p.get("contraente_id") for p in pols.values() if p.get("contraente_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0})}
    collab_ids = list({(p.get("collaboratore_id") or t.get("collaboratore_id"))
                       for t, p in [(t, pols.get(t["polizza_id"], {})) for t in items]
                       if (p.get("collaboratore_id") or t.get("collaboratore_id"))})
    collab_map = {}
    if collab_ids:
        async for u in db.users.find({"id": {"$in": collab_ids}}, {"_id": 0, "id": 1, "name": 1}):
            collab_map[u["id"]] = u["name"]
    result = []
    for t in items:
        p = pols.get(t["polizza_id"], {})
        a = anas.get(p.get("contraente_id", ""), {})
        collab_id = p.get("collaboratore_id") or t.get("collaboratore_id")
        if collaboratore_id and collab_id != collaboratore_id:
            continue
        result.append({
            "id": t["id"],
            "polizza_id": t["polizza_id"],
            "numero_polizza": p.get("numero_polizza"),
            "ramo": p.get("ramo"),
            "targa": p.get("targa"),
            "contraente_id": a.get("id"),
            "contraente_nome": a.get("ragione_sociale"),
            "cellulare": a.get("cellulare") or a.get("telefono"),
            "collaboratore_id": collab_id,
            "collaboratore_nome": collab_map.get(collab_id) if collab_id else None,
            "data_copertura": t.get("data_copertura"),
            "scadenza_polizza": p.get("scadenza"),
            "scadenza_titolo": t.get("scadenza"),
            "importo_lordo": t.get("importo_lordo", 0.0),
            "provvigioni": t.get("provvigioni", 0.0),
            "giorni_anticipo": _giorni_da_oggi(t.get("data_copertura")),
            "mezzo_pagamento_preferito": p.get("mezzo_pagamento_preferito"),
            "ultimo_mezzo_pagamento": p.get("ultimo_mezzo_pagamento"),
        })
    # totali aggregati
    return result


def _giorni_da_oggi(data_iso: Optional[str]) -> Optional[int]:
    if not data_iso:
        return None
    try:
        from datetime import date as _date
        d = _date.fromisoformat(data_iso[:10])
        return (_date.today() - d).days
    except Exception:
        return None


async def _total_sospesi_as_of(data_giorno: str) -> float:
    """Somma importo_lordo dei titoli SOSPESI (anticipati dall'agenzia, non ancora incassati)
    alla data indicata. Coerente con l'endpoint /titoli/sospesi.

    Un titolo è 'sospeso' alla data X se:
      - titolo_coperto = True
      - data_copertura <= X
      - stato in [da_incassare, insoluto]  OPPURE  (stato=incassato AND data_incasso > X)
    """
    pipeline = [
        {"$match": {
            "titolo_coperto": True,
            "data_copertura": {"$lte": data_giorno},
            "$or": [
                {"stato": {"$in": ["da_incassare", "insoluto"]}},
                {"$and": [
                    {"stato": "incassato"},
                    {"data_incasso": {"$gt": data_giorno}},
                ]},
            ],
        }},
        {"$group": {"_id": None, "tot": {"$sum": "$importo_lordo"}}},
    ]
    agg = await db.titoli.aggregate(pipeline).to_list(1)
    return round(float(agg[0]["tot"]) if agg else 0.0, 2)


@api.post("/titoli/bulk-incassa")
async def bulk_incassa(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """body: {ids: [...], data_incasso, mezzo_pagamento, conto_cassa_id}"""
    ids = body.get("ids") or []
    if not ids:
        raise HTTPException(400, "Nessun titolo selezionato")
    data_incasso = body.get("data_incasso") or _now_iso()[:10]
    # Lock: non si può incassare in una giornata di Prima Nota chiusa.
    await assert_giornata_aperta(data_incasso, azione="incassare nel giorno (Prima Nota chiusa)")
    mezzo = body.get("mezzo_pagamento") or "bonifico"
    conto_id = body.get("conto_cassa_id")
    if not conto_id:
        conto_id = await _resolve_conto_cassa(mezzo)
    n_incassati = 0; tot = 0.0
    for tid in ids:
        titolo = await db.titoli.find_one({"id": tid}, {"_id": 0})
        if not titolo or titolo.get("stato") == "incassato":
            continue
        await db.titoli.update_one(
            {"id": tid},
            {"$set": {"stato": "incassato", "data_incasso": data_incasso,
                      "mezzo_pagamento": mezzo, "conto_cassa_id": conto_id,
                      "updated_at": _now_iso()}},
        )
        pol = await db.polizze.find_one({"id": titolo["polizza_id"]}, {"_id": 0})
        mov = MovimentoContabile(
            data_movimento=data_incasso, tipo="entrata", categoria="incasso_premio",
            importo=titolo.get("importo_lordo", 0.0),
            descrizione=f"Incasso polizza {pol['numero_polizza'] if pol else titolo['polizza_id']}",
            polizza_id=titolo["polizza_id"], titolo_id=tid,
            anagrafica_id=pol.get("contraente_id") if pol else None,
            compagnia_id=pol.get("compagnia_id") if pol else None,
            conto_cassa_id=conto_id, mezzo_pagamento=mezzo,
            provvigioni=titolo.get("provvigioni", 0.0),
        )
        await db.movimenti.insert_one(mov.model_dump())
        n_incassati += 1
        tot += titolo.get("importo_lordo", 0.0)
    await log_attivita(user, "bulk_incasso", "titolo", None, f"{n_incassati} titoli incassati per €{tot:.2f}")
    return {"incassati": n_incassati, "totale": round(tot, 2)}


@api.post("/titoli/bulk-copertura")
async def bulk_copertura(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Marca i titoli come 'coperti' (anticipo provvigioni) creando il movimento
    visibile nel brogliaccio / Prima Nota.
    """
    ids = body.get("ids") or []
    if not ids:
        raise HTTPException(400, "ids richiesti")
    data_copertura = body.get("data_copertura") or _now_iso()[:10]
    # Lock: la copertura è una scrittura su Prima Nota -> blocco se giornata chiusa.
    await assert_giornata_aperta(data_copertura, azione="coprire titoli nel giorno (Prima Nota chiusa)")
    note = body.get("note")
    conto_id = body.get("conto_cassa_id")
    # Conto cassa di default (primo attivo) se non specificato
    if not conto_id:
        default_conto = await db.conti_cassa.find_one(
            {"attivo": True}, {"_id": 0, "id": 1}, sort=[("ordine", 1)],
        )
        conto_id = default_conto["id"] if default_conto else None

    set_fields = {
        "titolo_coperto": True,
        "data_copertura": data_copertura,
        "updated_at": _now_iso(),
    }
    if note:
        set_fields["note_copertura"] = note

    # 1) Aggiorna i titoli
    res = await db.titoli.update_many(
        {"id": {"$in": ids}},
        {"$set": set_fields},
    )

    # 2) Crea i movimenti "anticipo" in Prima Nota per ciascun titolo (idempotente:
    #    skip se esiste già un movimento di copertura per quel titolo+data).
    n_movimenti = 0
    async for t in db.titoli.find({"id": {"$in": ids}}, {"_id": 0}):
        # idempotenza: se esiste già un movimento di copertura per questo titolo, salta
        existing = await db.movimenti.find_one(
            {"titolo_id": t["id"], "categoria": "anticipo",
             "data_movimento": data_copertura, "tipo": "uscita"},
            {"_id": 0, "id": 1},
        )
        if existing:
            continue
        pol = await db.polizze.find_one({"id": t.get("polizza_id")}, {"_id": 0}) or {}
        importo = float(t.get("importo_lordo") or 0.0)
        mov = MovimentoContabile(
            data_movimento=data_copertura,
            tipo="uscita", categoria="anticipo",
            importo=importo,
            descrizione=f"Copertura anticipata polizza {pol.get('numero_polizza') or t.get('polizza_id', '')}",
            polizza_id=t.get("polizza_id"), titolo_id=t["id"],
            anagrafica_id=pol.get("contraente_id"),
            compagnia_id=pol.get("compagnia_id"),
            conto_cassa_id=conto_id,
            note=note,
        )
        await db.movimenti.insert_one(mov.model_dump())
        n_movimenti += 1

    await log_attivita(user, "bulk_copertura", "titolo", None,
                       f"{res.modified_count} titoli coperti dall'agenzia il {data_copertura} "
                       f"({n_movimenti} movimenti Prima Nota creati)")
    return {
        "aggiornati": res.modified_count,
        "data_copertura": data_copertura,
        "movimenti_creati": n_movimenti,
    }


@api.post("/titoli/notifica-copertura")
async def notifica_copertura(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Invia email di notifica copertura titolo a operatori e/o contraenti.

    body: {id: titolo_id, a_operatori: bool, a_contraenti: bool}
    Se SMTP non configurato -> logga l'attività e ritorna ok=False con un avviso.
    """
    tid = body.get("id")
    if not tid:
        raise HTTPException(400, "id richiesto")
    a_op = bool(body.get("a_operatori"))
    a_cnt = bool(body.get("a_contraenti"))
    if not (a_op or a_cnt):
        return {"ok": True, "inviate": 0}

    titolo = await db.titoli.find_one({"id": tid}, {"_id": 0})
    if not titolo:
        raise HTTPException(404, "Titolo non trovato")
    pol = await db.polizze.find_one({"id": titolo.get("polizza_id")}, {"_id": 0}) or {}

    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    smtp_ok = bool(az.get("smtp_host") and az.get("smtp_user"))

    destinatari: list[str] = []
    if a_cnt and pol.get("contraente_id"):
        ana = await db.anagrafiche.find_one({"id": pol["contraente_id"]}, {"_id": 0}) or {}
        if ana.get("email"):
            destinatari.append(ana["email"])
    if a_op and titolo.get("collaboratore_id"):
        op = await db.collaboratori.find_one({"id": titolo["collaboratore_id"]}, {"_id": 0}) or {}
        if op.get("email"):
            destinatari.append(op["email"])

    await log_attivita(user, "notifica_copertura", "titolo", tid,
                       f"a_op={a_op} a_cnt={a_cnt} destinatari={len(destinatari)} smtp={smtp_ok}")

    if not smtp_ok:
        return {"ok": False, "errore": "SMTP non configurato", "destinatari": destinatari}
    if not destinatari:
        return {"ok": False, "errore": "Nessun destinatario con email valida"}

    # invio reale (best-effort)
    import smtplib
    from email.message import EmailMessage
    try:
        msg = EmailMessage()
        msg["From"] = az.get("smtp_from") or az.get("smtp_user")
        msg["To"] = ", ".join(destinatari)
        msg["Subject"] = f"Copertura titolo polizza {pol.get('numero_polizza','—')}"
        msg.set_content(
            f"Notifica di copertura titolo.\n\n"
            f"Polizza: {pol.get('numero_polizza','—')} ({pol.get('ramo','')})\n"
            f"Importo: €{titolo.get('importo_lordo',0):.2f}\n"
            f"Scadenza: {titolo.get('scadenza','—')}\n"
            f"Copertura: {titolo.get('data_copertura','—')}\n",
        )
        port = int(az.get("smtp_port") or 587)
        host = az["smtp_host"]
        if port == 465:
            srv = smtplib.SMTP_SSL(host, port, timeout=30)
        else:
            srv = smtplib.SMTP(host, port, timeout=30)
            if az.get("smtp_use_tls", True):
                srv.starttls()
        if az.get("smtp_user"):
            srv.login(az["smtp_user"], az.get("smtp_password") or "")
        srv.send_message(msg)
        srv.quit()
        return {"ok": True, "inviate": len(destinatari), "destinatari": destinatari}
    except Exception as e:
        return {"ok": False, "errore": str(e), "destinatari": destinatari}


@api.get("/export/titoli.csv")
async def export_titoli_csv(stato: Optional[str] = None,
                            user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    items = await list_titoli(stato=stato, limit=10000, user=user)
    import csv as _csv
    out = _io.StringIO()
    w = _csv.writer(out, delimiter=";")
    w.writerow(["Numero polizza", "Targa", "Contraente", "Compagnia", "Collaboratore",
                "Ramo", "Tipo", "Effetto", "Scadenza", "Stato",
                "Lordo", "Netto", "Imposte", "Provvigioni",
                "Mezzo pagamento", "Data incasso", "Coperto fino a"])
    for t in items:
        w.writerow([t.get("numero_polizza"), t.get("targa"), t.get("contraente_nome"),
                    t.get("compagnia_nome"), t.get("collaboratore_nome"),
                    t.get("ramo"), t.get("tipo"), t.get("effetto"), t.get("scadenza"),
                    t.get("stato"), t.get("importo_lordo"), t.get("importo_netto"),
                    t.get("imposte"), t.get("provvigioni"),
                    t.get("mezzo_pagamento"), t.get("data_incasso"), t.get("data_copertura")])
    csv_bytes = out.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        _io.BytesIO(csv_bytes), media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="titoli.csv"'},
    )


@api.get("/export/titoli.xlsx")
async def export_titoli_xlsx(stato: Optional[str] = None,
                             user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    items = await list_titoli(stato=stato, limit=10000, user=user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Titoli"
    headers = ["Numero polizza", "Targa", "Contraente", "Compagnia", "Collaboratore",
               "Ramo", "Tipo", "Effetto", "Scadenza", "Stato",
               "Lordo €", "Netto €", "Imposte €", "Provvigioni €",
               "Mezzo pag.", "Data incasso", "Coperto fino"]
    ws.append(headers)
    head_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    head_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal="center")
    for t in items:
        ws.append([t.get("numero_polizza"), t.get("targa"), t.get("contraente_nome"),
                   t.get("compagnia_nome"), t.get("collaboratore_nome"),
                   t.get("ramo"), t.get("tipo"), t.get("effetto"), t.get("scadenza"),
                   t.get("stato"), t.get("importo_lordo"), t.get("importo_netto"),
                   t.get("imposte"), t.get("provvigioni"),
                   t.get("mezzo_pagamento"), t.get("data_incasso"), t.get("coperto_fino_a")])
    # auto column widths
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 30)
    out = _io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="titoli.xlsx"'},
    )


def _polizze_export_filters(**kw) -> dict:
    """Pulisce i kwargs filtro dell'export polizze (rimuove None)."""
    return {k: v for k, v in kw.items() if v is not None and v != ""}


@api.get("/export/polizze.csv")
async def export_polizze_csv(
    q: Optional[str] = None, stato: Optional[str] = None,
    compagnia_id: Optional[str] = None, ramo: Optional[str] = None,
    prodotto: Optional[str] = None, collaboratore_id: Optional[str] = None,
    contraente_id: Optional[str] = None,
    dal: Optional[str] = None, al: Optional[str] = None,
    in_scadenza_giorni: Optional[int] = None,
    scadenza_oltre_giorni: Optional[int] = None,
    scadute_oggi: Optional[bool] = None,
    scadute_da_min: Optional[int] = None,
    scadute_da_max: Optional[int] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    items = await list_polizze(
        q=q, stato=stato, ramo=ramo, prodotto=prodotto,
        contraente_id=contraente_id, compagnia_id=compagnia_id,
        collaboratore_id=collaboratore_id, dal=dal, al=al,
        in_scadenza_giorni=in_scadenza_giorni,
        scadenza_oltre_giorni=scadenza_oltre_giorni,
        scadute_oggi=scadute_oggi,
        scadute_da_min=scadute_da_min, scadute_da_max=scadute_da_max,
        limit=10000, user=user,
    )
    import csv as _csv
    out = _io.StringIO()
    w = _csv.writer(out, delimiter=";")
    w.writerow(["Numero polizza", "Targa", "Contraente", "Compagnia", "Collaboratore",
                "Ramo", "Prodotto", "Stato", "Effetto", "Scadenza",
                "Premio lordo", "Premio netto", "Provvigioni", "Frazionamento"])
    for p in items:
        w.writerow([p.get("numero_polizza"), p.get("targa"), p.get("contraente_nome"),
                    p.get("compagnia_nome"), p.get("collaboratore_nome"),
                    p.get("ramo"), p.get("prodotto"), p.get("stato"),
                    p.get("effetto"), p.get("scadenza"),
                    p.get("premio_lordo"), p.get("premio_netto"),
                    p.get("provvigioni"), p.get("frazionamento")])
    csv_bytes = out.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        _io.BytesIO(csv_bytes), media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="polizze.csv"'},
    )


@api.get("/export/polizze.xlsx")
async def export_polizze_xlsx(
    q: Optional[str] = None, stato: Optional[str] = None,
    compagnia_id: Optional[str] = None, ramo: Optional[str] = None,
    prodotto: Optional[str] = None, collaboratore_id: Optional[str] = None,
    contraente_id: Optional[str] = None,
    dal: Optional[str] = None, al: Optional[str] = None,
    in_scadenza_giorni: Optional[int] = None,
    scadenza_oltre_giorni: Optional[int] = None,
    scadute_oggi: Optional[bool] = None,
    scadute_da_min: Optional[int] = None,
    scadute_da_max: Optional[int] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    items = await list_polizze(
        q=q, stato=stato, ramo=ramo, prodotto=prodotto,
        contraente_id=contraente_id, compagnia_id=compagnia_id,
        collaboratore_id=collaboratore_id, dal=dal, al=al,
        in_scadenza_giorni=in_scadenza_giorni,
        scadenza_oltre_giorni=scadenza_oltre_giorni,
        scadute_oggi=scadute_oggi,
        scadute_da_min=scadute_da_min, scadute_da_max=scadute_da_max,
        limit=10000, user=user,
    )
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Polizze"
    headers = ["Numero polizza", "Targa", "Contraente", "Compagnia", "Collaboratore",
               "Ramo", "Prodotto", "Stato", "Effetto", "Scadenza",
               "Premio lordo €", "Premio netto €", "Provvigioni €", "Frazionamento"]
    ws.append(headers)
    head_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    head_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal="center")
    for p in items:
        ws.append([p.get("numero_polizza"), p.get("targa"), p.get("contraente_nome"),
                   p.get("compagnia_nome"), p.get("collaboratore_nome"),
                   p.get("ramo"), p.get("prodotto"), p.get("stato"),
                   p.get("effetto"), p.get("scadenza"),
                   p.get("premio_lordo"), p.get("premio_netto"),
                   p.get("provvigioni"), p.get("frazionamento")])
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 30)
    out = _io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="polizze.xlsx"'},
    )


@api.delete("/titoli/{tid}")
async def delete_titolo(tid: str, user=Depends(require_user("admin", "collaboratore"))):
    # Blocco se prima nota del giorno di incasso (o competenza) è chiusa
    t = await db.titoli.find_one({"id": tid}, {"_id": 0, "data_incasso": 1, "data_competenza": 1})
    if t:
        await assert_giornata_aperta(t.get("data_incasso") or t.get("data_competenza"), azione="eliminare il titolo")
    res = await db.titoli.delete_one({"id": tid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Titolo non trovato")
    await db.movimenti.delete_many({"titolo_id": tid})
    await log_attivita(user, "delete", "titolo", tid)
    return {"ok": True}


@api.post("/titoli", status_code=201)
async def create_titolo(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    # Se scadenza_mora non è specificata, ricavala da: polizza.termini_mora_giorni
    # oppure dal default per ramo (Vita=30, altri=15).
    if not body.get("scadenza_mora") and body.get("scadenza") and body.get("polizza_id"):
        pol = await db.polizze.find_one(
            {"id": body["polizza_id"]}, {"_id": 0, "termini_mora_giorni": 1, "ramo": 1, "prodotto": 1},
        ) or {}
        giorni = pol.get("termini_mora_giorni")
        if not giorni and pol.get("prodotto"):
            prod = await db.prodotti.find_one(
                {"nome": pol["prodotto"]}, {"_id": 0, "termini_mora_giorni": 1},
            )
            if prod:
                giorni = prod.get("termini_mora_giorni")
        if not giorni:
            from db_models import default_mora_for_ramo
            giorni = default_mora_for_ramo(pol.get("ramo"))
        try:
            from datetime import date as _date, timedelta as _td
            d = _date.fromisoformat(body["scadenza"][:10]) + _td(days=int(giorni))
            body["scadenza_mora"] = d.isoformat()
        except Exception:
            pass
    # Auto-calc provvigioni se non specificate
    if (not body.get("provvigioni")) and body.get("importo_lordo") and body.get("polizza_id"):
        pol_pre = await db.polizze.find_one(
            {"id": body["polizza_id"]},
            {"_id": 0, "collaboratore_id": 1, "compagnia_id": 1, "ramo": 1},
        ) or {}
        collab_id = body.get("collaboratore_id") or pol_pre.get("collaboratore_id")
        calc = await _calcola_provvigione(
            float(body["importo_lordo"]),
            collab_id,
            pol_pre.get("compagnia_id"),
            pol_pre.get("ramo"),
        )
        if calc["provvigione_totale"]:
            body["provvigioni"] = calc["provvigione_totale"]
        # Eredita collaboratore_id dalla polizza se non specificato
        if not body.get("collaboratore_id") and pol_pre.get("collaboratore_id"):
            body["collaboratore_id"] = pol_pre["collaboratore_id"]
    obj = Titolo(**body)
    await db.titoli.insert_one(obj.model_dump())
    await log_attivita(user, "create", "titolo", obj.id)
    return obj.model_dump()


@api.put("/titoli/{tid}")
async def update_titolo(tid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    # Blocco se prima nota chiusa per la data di incasso/competenza del titolo
    cur = await db.titoli.find_one({"id": tid}, {"_id": 0, "data_incasso": 1, "data_competenza": 1})
    if cur:
        await assert_giornata_aperta(cur.get("data_incasso") or cur.get("data_competenza"), azione="modificare il titolo")
    # Se il body cambia la data_incasso, controlla anche la nuova data
    if body.get("data_incasso"):
        await assert_giornata_aperta(body["data_incasso"], azione="impostare data incasso nel giorno")
    body["updated_at"] = _now_iso()
    res = await db.titoli.update_one({"id": tid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Non trovato")
    await log_attivita(user, "update", "titolo", tid)
    return strip_mongo_id(await db.titoli.find_one({"id": tid}, {"_id": 0}))


@api.post("/titoli/{tid}/incassa")
async def incassa_titolo(tid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Marca un titolo come incassato, crea movimento contabile entrata e gestisce la
    differenza eventuale (importo_pagato < lordo) in uno di due modi:

      * ``tipo_chiusura="sconto"`` (default) -> crea un movimento uscita 'sconto_cliente'
        per il residuo (entra in Prima Nota spese).
      * ``tipo_chiusura="sospeso"`` -> genera un nuovo Titolo (tipo=regolazione) col residuo
        in stato 'da_incassare', così resta visibile nei Sospesi.

    body: {data_incasso?, mezzo_pagamento?, conto_cassa_id?, importo_pagato?,
            tipo_chiusura?, motivo_sconto?}
    Se importo_pagato è omesso -> si assume pagamento completo del lordo (no residuo).
    """
    titolo = await db.titoli.find_one({"id": tid}, {"_id": 0})
    if not titolo:
        raise HTTPException(404, "Titolo non trovato")
    data_incasso = body.get("data_incasso") or _now_iso()[:10]
    # Lock: non si può incassare in una giornata di Prima Nota chiusa.
    await assert_giornata_aperta(data_incasso, azione="incassare nel giorno (Prima Nota chiusa)")
    mezzo = body.get("mezzo_pagamento") or "bonifico"
    conto_id = body.get("conto_cassa_id")
    if not conto_id:
        conto_id = await _resolve_conto_cassa(mezzo)
    lordo = float(titolo.get("importo_lordo") or 0)
    importo_pagato_raw = body.get("importo_pagato")
    importo_pagato = float(importo_pagato_raw) if importo_pagato_raw is not None else lordo
    residuo = round(max(0.0, lordo - importo_pagato), 2)
    tipo_chiusura = (body.get("tipo_chiusura") or "sconto").lower()
    if tipo_chiusura not in ("sconto", "sospeso"):
        raise HTTPException(400, "tipo_chiusura deve essere 'sconto' o 'sospeso'")
    motivo_sconto = body.get("motivo_sconto")
    # Quando si lascia il residuo come sospeso non si applica sconto
    sconto_applicato = residuo if (residuo > 0 and tipo_chiusura == "sconto") else 0.0

    await db.titoli.update_one(
        {"id": tid},
        {"$set": {
            "stato": "incassato",
            "data_incasso": data_incasso,
            "mezzo_pagamento": mezzo,
            "conto_cassa_id": conto_id,
            "importo_pagato": importo_pagato,
            "sconto_applicato": sconto_applicato,
            "motivo_sconto": motivo_sconto if sconto_applicato > 0 else None,
            "updated_at": _now_iso(),
        }},
    )
    pol = await db.polizze.find_one({"id": titolo["polizza_id"]}, {"_id": 0})
    # Aggiorna ultimo mezzo pagamento sulla polizza; se manca il preferito, impostalo.
    if pol:
        pol_updates = {
            "ultimo_mezzo_pagamento": mezzo,
            "ultimo_mezzo_pagamento_data": data_incasso,
        }
        if not pol.get("mezzo_pagamento_preferito"):
            pol_updates["mezzo_pagamento_preferito"] = mezzo
        await db.polizze.update_one({"id": pol["id"]}, {"$set": pol_updates})

    # Movimento entrata = importo effettivamente pagato dal cliente
    if residuo > 0 and tipo_chiusura == "sospeso":
        nota_residuo = f" (era €{lordo:.2f}, residuo €{residuo:.2f} a sospeso)"
    elif residuo > 0:
        nota_residuo = f" (era €{lordo:.2f}, sconto €{residuo:.2f})"
    else:
        nota_residuo = ""
    # Movimento entrata = importo effettivamente pagato dal cliente (cash netto in cassa).
    # Lo sconto viene memorizzato come quota_sconto SULLA STESSA RIGA: nel brogliaccio
    # mostreremo Totale=lordo, Spese=quota_sconto, conto cassa=importo (pagato).
    if residuo > 0 and tipo_chiusura == "sospeso":
        nota_residuo = f" (era €{lordo:.2f}, residuo €{residuo:.2f} a sospeso)"
    elif residuo > 0:
        nota_residuo = f" (premio lordo €{lordo:.2f}, sconto €{residuo:.2f})"
    else:
        nota_residuo = ""
    mov_in = MovimentoContabile(
        data_movimento=data_incasso,
        tipo="entrata",
        categoria="incasso_premio",
        importo=importo_pagato,
        descrizione=f"Incasso titolo polizza {pol['numero_polizza'] if pol else titolo['polizza_id']}" + nota_residuo,
        polizza_id=titolo["polizza_id"],
        titolo_id=tid,
        anagrafica_id=pol.get("contraente_id") if pol else None,
        compagnia_id=pol.get("compagnia_id") if pol else None,
        conto_cassa_id=conto_id,
        mezzo_pagamento=mezzo,
        provvigioni=titolo.get("provvigioni", 0.0),
        quota_sconto=sconto_applicato,
        note=(f"Sconto applicato: €{sconto_applicato:.2f}"
              + (f" — {motivo_sconto}" if motivo_sconto else "")) if sconto_applicato > 0 else None,
    )
    await db.movimenti.insert_one(mov_in.model_dump())

    # aggiorna anagrafica con ultimo mezzo di pagamento usato
    if pol and pol.get("contraente_id"):
        await db.anagrafiche.update_one(
            {"id": pol["contraente_id"]},
            {"$set": {
                "ultimo_mezzo_pagamento": mezzo,
                "ultimo_mezzo_pagamento_data": data_incasso,
                "updated_at": _now_iso(),
            }},
        )

    # NOTA: per tipo_chiusura=sconto NON creiamo più un movimento uscita separato.
    # Lo sconto viaggia sulla stessa riga entrata via mov_in.quota_sconto.
    residuo_titolo_id = None
    if residuo > 0 and tipo_chiusura == "sospeso":
        # Crea nuovo titolo residuo a sospeso (anticipato dall'agenzia se l'originale lo era)
        base_num = titolo.get("numero_titolo") or (pol.get("numero_polizza") if pol else None)
        residuo_titolo = Titolo(
            polizza_id=titolo["polizza_id"],
            numero_titolo=(f"{base_num}-RES" if base_num else None),
            tipo="regolazione",
            effetto=data_incasso,
            scadenza=titolo.get("scadenza") or data_incasso,
            stato="da_incassare",
            importo_lordo=residuo,
            importo_netto=0.0,
            imposte=0.0,
            provvigioni=0.0,
            titolo_coperto=bool(titolo.get("titolo_coperto")),
            data_copertura=data_incasso if titolo.get("titolo_coperto") else None,
            collaboratore_id=titolo.get("collaboratore_id"),
            fonte="manuale",
        )
        await db.titoli.insert_one(residuo_titolo.model_dump())
        residuo_titolo_id = residuo_titolo.id
        # Mostra il residuo a sospeso anche nella colonna 'Sospesi' del Brogliaccio:
        # lo salviamo come quota_credito sul movimento entrata appena creato.
        await db.movimenti.update_one(
            {"id": mov_in.id},
            {"$set": {"quota_credito": residuo, "note": (
                f"Residuo €{residuo:.2f} lasciato a sospeso "
                f"(titolo {residuo_titolo.id[:8]})"
            )}},
        )

    # Audit nel diario cliente per tracciabilità storica
    if pol and pol.get("contraente_id"):
        desc = (f"Pagamento titolo polizza {pol.get('numero_polizza')} - "
                f"€{importo_pagato:.2f} via {mezzo} il {data_incasso}")
        if residuo > 0 and tipo_chiusura == "sconto":
            desc += f" (sconto applicato: €{residuo:.2f}"
            if motivo_sconto:
                desc += f" - {motivo_sconto}"
            desc += ")"
        elif residuo > 0 and tipo_chiusura == "sospeso":
            desc += f" (residuo €{residuo:.2f} a sospeso)"
        await log_diario_cliente(
            pol["contraente_id"], "documento",
            titolo=f"Incasso titolo - polizza {pol.get('numero_polizza')}",
            descrizione=desc, autore=user,
        )

    extra_label = ""
    if residuo > 0 and tipo_chiusura == "sconto":
        extra_label = f" sconto €{residuo:.2f}"
    elif residuo > 0 and tipo_chiusura == "sospeso":
        extra_label = f" residuo sospeso €{residuo:.2f}"
    await log_attivita(user, "incasso", "titolo", tid,
                       f"€{importo_pagato:.2f}{extra_label}".strip())
    # Auto-genera Lettera di Abbuono se è stato applicato uno sconto
    lettera_id = None
    if sconto_applicato > 0:
        try:
            lettera_id = await _crea_lettera_abbuono_auto(
                titolo=titolo, polizza=pol, importo_pagato=importo_pagato,
                sconto=sconto_applicato, motivo=motivo_sconto, data_incasso=data_incasso,
                user_id=user.get("id"),
            )
        except Exception as _e:
            logger.warning("Auto-gen lettera abbuono fallita: %s", _e)
    return {
        "ok": True,
        "importo_pagato": importo_pagato,
        "residuo": residuo,
        "tipo_chiusura": tipo_chiusura,
        "sconto_applicato": sconto_applicato,
        "movimento_entrata_id": mov_in.id,
        "movimento_sconto_id": None,  # legacy: sconto ora vive sulla stessa riga entrata
        "titolo_residuo_id": residuo_titolo_id,
        "lettera_abbuono_id": lettera_id,
    }


# ============================================================
# LETTERA DI ABBUONO (sconto su titolo) + firma digitale
# ============================================================
async def _crea_lettera_abbuono_auto(
    *, titolo: dict, polizza: Optional[dict], importo_pagato: float,
    sconto: float, motivo: Optional[str], data_incasso: str, user_id: Optional[str],
) -> str:
    """Crea (idempotente) la lettera di abbuono per un titolo con sconto.

    Restituisce sempre l'id della lettera (creata o esistente).
    """
    existing = await db.lettere_abbuono.find_one(
        {"titolo_id": titolo["id"]}, {"_id": 0, "id": 1},
    )
    if existing:
        return existing["id"]
    lordo = float(titolo.get("importo_lordo") or 0.0)
    rec = LetteraAbbuono(
        titolo_id=titolo["id"],
        polizza_id=titolo.get("polizza_id"),
        anagrafica_id=(polizza or {}).get("contraente_id"),
        compagnia_id=(polizza or {}).get("compagnia_id"),
        importo_lordo=lordo,
        importo_pagato=importo_pagato,
        importo_sconto=sconto,
        motivo_sconto=motivo,
        data_incasso=data_incasso,
        created_by=user_id,
    )
    await db.lettere_abbuono.insert_one(rec.model_dump())
    return rec.id


async def _user_firma_to_b64(firma_url: str) -> str:
    """Scarica la firma del collaboratore (PNG/JPG) dallo storage e la
    converte in data-URL base64 utilizzabile dalla lettera di abbuono
    (e da QUALSIASI altro documento che richiede la firma operatore).

    `firma_url` è nel formato `/api/storage/<path>` salvato in
    `users.firma_digitale_url`.
    """
    prefix = "/api/storage/"
    path = firma_url[len(prefix):] if firma_url.startswith(prefix) else firma_url
    try:
        data, _ct = obj_storage.get_object(path)
    except Exception as e:
        raise HTTPException(500, f"Impossibile leggere la firma: {e}")
    if not data:
        raise HTTPException(404, "Firma utente non trovata su storage")
    ext = (path.rsplit(".", 1)[-1] or "png").lower()
    mime = "image/png" if ext == "png" else (
        "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
    )
    return f"data:{mime};base64,{_b64lib.b64encode(data).decode('ascii')}"


async def _build_lettera_abbuono_pdf(letid: str) -> tuple[bytes, dict]:
    let = await db.lettere_abbuono.find_one({"id": letid}, {"_id": 0})
    if not let:
        raise HTTPException(404, "Lettera non trovata")
    azienda = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    titolo = await db.titoli.find_one({"id": let["titolo_id"]}, {"_id": 0}) or {}
    polizza = (await db.polizze.find_one({"id": let.get("polizza_id")}, {"_id": 0})
               if let.get("polizza_id") else None) or {}
    anagrafica = (await db.anagrafiche.find_one({"id": let.get("anagrafica_id")}, {"_id": 0})
                  if let.get("anagrafica_id") else None) or {}
    compagnia = (await db.compagnie.find_one({"id": let.get("compagnia_id")}, {"_id": 0})
                 if let.get("compagnia_id") else None) or {}
    operatore = (await db.users.find_one(
        {"id": let.get("firma_operatore_user_id") or let.get("created_by")},
        {"_id": 0, "password_hash": 0},
    ) if (let.get("firma_operatore_user_id") or let.get("created_by")) else None) or {}
    pdf = pdf_lettera_abbuono.generate_lettera_abbuono(
        azienda=azienda, lettera=let, titolo=titolo, polizza=polizza,
        anagrafica=anagrafica, compagnia=compagnia, operatore=operatore,
    )
    return pdf, let


@api.post("/titoli/{tid}/lettera-abbuono", status_code=201)
async def crea_lettera_abbuono_manuale(
    tid: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Endpoint MANUALE: genera una lettera di abbuono per un titolo che ha
    `sconto_applicato > 0`. Idempotente (restituisce quella esistente se presente).
    """
    titolo = await db.titoli.find_one({"id": tid}, {"_id": 0})
    if not titolo:
        raise HTTPException(404, "Titolo non trovato")
    sconto = float(titolo.get("sconto_applicato") or 0.0)
    if sconto <= 0:
        raise HTTPException(400, "Nessuno sconto applicato a questo titolo")
    pol = await db.polizze.find_one({"id": titolo.get("polizza_id")}, {"_id": 0})
    importo_pagato = float(titolo.get("importo_pagato") or 0.0)
    motivo = titolo.get("motivo_sconto")
    data_incasso = titolo.get("data_incasso") or _now_iso()[:10]
    lid = await _crea_lettera_abbuono_auto(
        titolo=titolo, polizza=pol, importo_pagato=importo_pagato,
        sconto=sconto, motivo=motivo, data_incasso=data_incasso,
        user_id=user.get("id"),
    )
    rec = await db.lettere_abbuono.find_one({"id": lid}, {"_id": 0})
    await log_attivita(user, "create", "lettera_abbuono", lid, f"Sconto €{sconto:.2f}")
    return rec


@api.get("/lettere-abbuono")
async def list_lettere_abbuono(
    titolo_id: Optional[str] = None,
    polizza_id: Optional[str] = None,
    anagrafica_id: Optional[str] = None,
    user=Depends(current_user),
):
    flt: dict = {}
    if titolo_id:
        flt["titolo_id"] = titolo_id
    if polizza_id:
        flt["polizza_id"] = polizza_id
    if anagrafica_id:
        flt["anagrafica_id"] = anagrafica_id
    items = await db.lettere_abbuono.find(flt, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return items


@api.get("/lettere-abbuono/{lid}")
async def get_lettera_abbuono(lid: str, user=Depends(current_user)):
    rec = await db.lettere_abbuono.find_one({"id": lid}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Lettera non trovata")
    return rec


@api.get("/lettere-abbuono/{lid}/pdf")
async def get_lettera_abbuono_pdf(lid: str, user=Depends(current_user)):
    pdf, let = await _build_lettera_abbuono_pdf(lid)
    fname = f"lettera_abbuono_{lid[:8]}.pdf"
    return StreamingResponse(
        _io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )


@api.post("/lettere-abbuono/{lid}/firma")
async def firma_lettera_abbuono(
    lid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """body: {
        tipo: "operatore" | "cliente",
        b64: "data:image/png;base64,..." (richiesto se !from_user_profile),
        nome?: str,
        from_user_profile?: bool  (solo per tipo=operatore: pesca firma_digitale_url
                                   dal profilo del collaboratore loggato)
    }
    """
    tipo = (body.get("tipo") or "").lower()
    if tipo not in ("operatore", "cliente"):
        raise HTTPException(400, "tipo deve essere 'operatore' o 'cliente'")
    b64 = body.get("b64") or ""
    nome = body.get("nome") or None

    # Pesca la firma dal profilo del collaboratore (firma_digitale_url su user)
    if tipo == "operatore" and body.get("from_user_profile"):
        u_doc = await db.users.find_one({"id": user.get("id")}, {"_id": 0})
        if not u_doc or not u_doc.get("firma_digitale_url"):
            raise HTTPException(
                400,
                "Nessuna firma digitale caricata sul tuo profilo. "
                "Vai in Librerie -> Utenti/Collaboratori -> Documenti e carica la firma.",
            )
        b64 = await _user_firma_to_b64(u_doc["firma_digitale_url"])
        nome = nome or u_doc.get("name") or u_doc.get("nome")

    if not b64 or not b64.startswith("data:image"):
        raise HTTPException(400, "Firma non valida (atteso PNG base64 data URL)")
    rec = await db.lettere_abbuono.find_one({"id": lid}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Lettera non trovata")
    now = _now_iso()
    upd: dict = {"updated_at": now}
    if tipo == "operatore":
        upd["firma_operatore_b64"] = b64
        upd["firma_operatore_user_id"] = user.get("id")
        upd["firma_operatore_nome"] = nome or user.get("name")
        upd["firma_operatore_at"] = now
    else:
        upd["firma_cliente_b64"] = b64
        upd["firma_cliente_nome"] = nome
        upd["firma_cliente_at"] = now
    await db.lettere_abbuono.update_one({"id": lid}, {"$set": upd})

    # Se entrambe le firme presenti, rigenera e salva il PDF firmato su storage
    updated = await db.lettere_abbuono.find_one({"id": lid}, {"_id": 0})
    if updated and updated.get("firma_operatore_b64") and updated.get("firma_cliente_b64"):
        try:
            pdf, _ = await _build_lettera_abbuono_pdf(lid)
            path = f"{os.environ.get('APP_NAME', 'assicura')}/lettere_abbuono/{lid}.pdf"
            result = obj_storage.put_object(path, pdf, "application/pdf")
            await db.lettere_abbuono.update_one(
                {"id": lid},
                {"$set": {"signed_pdf_storage_path": result["path"], "updated_at": _now_iso()}},
            )
            # crea anche un Allegato sul titolo per visibilità
            try:
                al = Allegato(
                    entita_tipo="titolo", entita_id=updated["titolo_id"],
                    nome_file=f"lettera_abbuono_{lid[:8]}.pdf",
                    storage_path=result["path"], content_type="application/pdf",
                    size=result.get("size", len(pdf)),
                    descrizione="Lettera di abbuono firmata",
                    autore_id=user.get("id"),
                )
                await db.allegati.insert_one(al.model_dump())
            except Exception:
                pass
        except Exception as e:
            logger.warning("Errore salvataggio PDF firmato: %s", e)

    await log_attivita(user, "firma", "lettera_abbuono", lid, f"Firma {tipo}")
    return await db.lettere_abbuono.find_one({"id": lid}, {"_id": 0})


@api.delete("/lettere-abbuono/{lid}")
async def delete_lettera_abbuono(lid: str, user=Depends(require_user("admin"))):
    res = await db.lettere_abbuono.delete_one({"id": lid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Lettera non trovata")
    return {"ok": True}


# ============================================================
# SINISTRI
# ============================================================
@api.get("/sinistri")
async def list_sinistri(
    stato: Optional[str] = None,
    polizza_id: Optional[str] = None,
    compagnia_id: Optional[str] = None,
    contraente_id: Optional[str] = None,
    collaboratore_id: Optional[str] = None,
    ramo: Optional[str] = None,
    tipologia: Optional[str] = None,
    q: Optional[str] = None,
    dal: Optional[str] = None,
    al: Optional[str] = None,
    limit: int = 50000,
    user=Depends(current_user),
):
    flt = await visibility_filter(user)
    if stato:
        flt["stato"] = stato
    if polizza_id:
        flt["polizza_id"] = polizza_id
    if compagnia_id:
        flt["compagnia_id"] = compagnia_id
    if contraente_id:
        flt["contraente_id"] = contraente_id
    if collaboratore_id:
        flt["collaboratore_id"] = collaboratore_id
    if ramo:
        flt["ramo"] = ramo
    if tipologia:
        flt["tipologia_sinistro"] = {"$regex": tipologia, "$options": "i"}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        flt["data_avvenimento"] = cond
    if q:
        qrx = {"$regex": q, "$options": "i"}
        flt["$or"] = [
            {"numero_sinistro": qrx},
            {"numero_interno": qrx},
            {"luogo": qrx},
            {"descrizione": qrx},
            {"tipologia_sinistro": qrx},
        ]
    items = await db.sinistri.find(flt, {"_id": 0}).sort("data_avvenimento", -1).to_list(limit)
    # enrichment ottimizzato
    pol_ids = list({s["polizza_id"] for s in items if s.get("polizza_id")})
    ana_ids = list({s["contraente_id"] for s in items if s.get("contraente_id")})
    comp_ids = list({s["compagnia_id"] for s in items if s.get("compagnia_id")})
    pols = {p["id"]: p async for p in db.polizze.find(
        {"id": {"$in": pol_ids}},
        {"_id": 0, "id": 1, "numero_polizza": 1, "targa": 1, "ramo": 1, "prodotto": 1})}
    anas = {a["id"]: a async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}},
        {"_id": 0, "id": 1, "ragione_sociale": 1, "cognome": 1, "nome": 1})}
    comps = {c["id"]: c async for c in db.compagnie.find(
        {"id": {"$in": comp_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    collab_ids = list({s.get("collaboratore_id") for s in items if s.get("collaboratore_id")})
    collabs = {u["id"]: u async for u in db.users.find(
        {"id": {"$in": collab_ids}}, {"_id": 0, "id": 1, "name": 1, "avatar_url": 1})}
    for s in items:
        p = pols.get(s.get("polizza_id"), {})
        s["numero_polizza"] = p.get("numero_polizza")
        s["targa"] = p.get("targa") or s.get("targa")
        a = anas.get(s.get("contraente_id"), {})
        s["contraente_nome"] = a.get("ragione_sociale") or \
            f"{a.get('cognome','')} {a.get('nome','')}".strip()
        s["compagnia_nome"] = comps.get(s.get("compagnia_id"), {}).get("ragione_sociale")
        _c = collabs.get(s.get("collaboratore_id", ""), {})
        s["collaboratore_nome"] = _c.get("name")
        s["collaboratore_avatar_url"] = _c.get("avatar_url")
        # primo danneggiato come campo riassuntivo
        if s.get("danneggiati"):
            d0 = s["danneggiati"][0]
            s["danneggiato_nome"] = d0.get("nome") or d0.get("ragione_sociale")
    return items


@api.post("/sinistri", status_code=201)
async def create_sinistro(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    obj = Sinistro(**body)
    await db.sinistri.insert_one(obj.model_dump())
    await log_attivita(user, "create", "sinistro", obj.id, f"Sinistro {obj.numero_sinistro}")
    # Alert: sinistro aperto
    pol = await db.polizze.find_one({"id": obj.polizza_id}, {"_id": 0, "numero_polizza": 1, "ramo": 1, "contraente_id": 1, "collaboratore_id": 1}) if obj.polizza_id else None
    from alert_dispatcher import safe_dispatch
    await safe_dispatch("sinistro.aperto", {
        "entita_tipo": "sinistro", "entita_id": obj.id,
        "anagrafica_id": (pol or {}).get("contraente_id"),
        "polizza_id": obj.polizza_id,
        "collaboratore_id": (pol or {}).get("collaboratore_id"),
        "numero_sinistro": obj.numero_sinistro,
        "numero_polizza": (pol or {}).get("numero_polizza"),
        "ramo": (pol or {}).get("ramo"),
        "link": f"/sinistri/{obj.id}",
    })
    return obj.model_dump()


@api.put("/sinistri/{sid}")
async def update_sinistro(sid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    prev = await db.sinistri.find_one({"id": sid}, {"_id": 0})
    if not prev:
        raise HTTPException(404, "Non trovato")
    body["updated_at"] = _now_iso()
    # Auto-transizione stato: se viene impostata data_liquidazione e lo stato
    # non è già chiuso/liquidato, marca il sinistro come "liquidato".
    if body.get("data_liquidazione"):
        curr_stato = body.get("stato") or prev.get("stato")
        if curr_stato not in ("liquidato", "chiuso_senza_seguito", "respinto"):
            body["stato"] = "liquidato"
    await db.sinistri.update_one({"id": sid}, {"$set": body})
    await log_attivita(user, "update", "sinistro", sid)
    new = await db.sinistri.find_one({"id": sid}, {"_id": 0})
    # Alert: stato cambiato a chiuso o pagato
    new_stato = (new or {}).get("stato")
    prev_stato = prev.get("stato")
    if new_stato != prev_stato and new_stato in ("chiuso", "liquidato", "pagato"):
        pol = await db.polizze.find_one(
            {"id": new.get("polizza_id")},
            {"_id": 0, "numero_polizza": 1, "ramo": 1, "contraente_id": 1, "collaboratore_id": 1},
        ) if new.get("polizza_id") else None
        from alert_dispatcher import safe_dispatch
        evento = "sinistro.pagato" if new_stato in ("liquidato", "pagato") else "sinistro.chiuso"
        await safe_dispatch(evento, {
            "entita_tipo": "sinistro", "entita_id": sid,
            "anagrafica_id": (pol or {}).get("contraente_id"),
            "polizza_id": new.get("polizza_id"),
            "collaboratore_id": (pol or {}).get("collaboratore_id"),
            "numero_sinistro": new.get("numero_sinistro"),
            "numero_polizza": (pol or {}).get("numero_polizza"),
            "importo_liquidato": new.get("importo_liquidato"),
            "ramo": (pol or {}).get("ramo"),
            "link": f"/sinistri/{sid}",
        })
    return strip_mongo_id(new)


@api.get("/sinistri/{sid}")
async def get_sinistro(sid: str, user=Depends(current_user)):
    """Restituisce il singolo sinistro arricchito con dati polizza/contraente/compagnia."""
    s = await db.sinistri.find_one({"id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Sinistro non trovato")
    # arricchimento contestuale
    if s.get("polizza_id"):
        pol = await db.polizze.find_one(
            {"id": s["polizza_id"]},
            {"_id": 0, "numero_polizza": 1, "ramo": 1, "prodotto": 1, "targa": 1,
             "compagnia_id": 1, "contraente_id": 1, "premio_lordo": 1, "decorrenza": 1,
             "scadenza": 1, "veicolo": 1, "marca": 1, "modello": 1, "telaio": 1},
        ) or {}
        s["polizza"] = pol
        s["numero_polizza"] = pol.get("numero_polizza")
    if s.get("contraente_id"):
        ana = await db.anagrafiche.find_one(
            {"id": s["contraente_id"]},
            {"_id": 0, "ragione_sociale": 1, "cognome": 1, "nome": 1, "codice_fiscale": 1,
             "partita_iva": 1, "indirizzo": 1, "cap": 1, "comune": 1, "provincia": 1,
             "telefono": 1, "email": 1, "tipo": 1},
        ) or {}
        s["contraente"] = ana
        s["contraente_nome"] = ana.get("ragione_sociale") or \
            f"{ana.get('cognome','')} {ana.get('nome','')}".strip()
    if s.get("compagnia_id"):
        c = await db.compagnie.find_one(
            {"id": s["compagnia_id"]}, {"_id": 0, "ragione_sociale": 1, "codice": 1},
        ) or {}
        s["compagnia"] = c
        s["compagnia_nome"] = c.get("ragione_sociale")
    if s.get("collaboratore_id"):
        u = await db.users.find_one(
            {"id": s["collaboratore_id"]}, {"_id": 0, "name": 1, "email": 1},
        ) or {}
        s["collaboratore_nome"] = u.get("name")
    # documenti collegati (dal modulo storage)
    docs = await db.storage_files.find(
        {"$or": [{"sinistro_id": sid}, {"entity_id": sid, "entity_type": "sinistro"}]},
        {"_id": 0},
    ).sort("uploaded_at", -1).to_list(500)
    s["documenti"] = docs
    return s


@api.put("/sinistri/{sid}/cid")
async def update_costatazione_amichevole(
    sid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Salva/aggiorna la Costatazione Amichevole (CID) di un sinistro RC Auto."""
    s = await db.sinistri.find_one({"id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Sinistro non trovato")
    cid = body or {}
    cid["updated_at"] = _now_iso()
    await db.sinistri.update_one(
        {"id": sid},
        {"$set": {"costatazione_amichevole": cid, "updated_at": _now_iso()}},
    )
    await log_attivita(user, "update_cid", "sinistro", sid)
    return await db.sinistri.find_one({"id": sid}, {"_id": 0})


@api.delete("/sinistri/{sid}")
async def delete_sinistro(sid: str, user=Depends(require_user("admin", "collaboratore"))):
    res = await db.sinistri.delete_one({"id": sid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Sinistro non trovato")
    await log_attivita(user, "delete", "sinistro", sid)
    return {"ok": True}


# ============================================================
# CONTABILITA
# ============================================================
@api.get("/contabilita/movimenti")
async def list_movimenti(
    dal: Optional[str] = None,
    al: Optional[str] = None,
    tipo: Optional[str] = None,
    anagrafica_id: Optional[str] = None,
    limit: int = 50000,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    flt: dict = {}
    if tipo:
        flt["tipo"] = tipo
    if anagrafica_id:
        flt["anagrafica_id"] = anagrafica_id
    if dal or al:
        cond = {}
        if dal:
            cond["$gte"] = dal
        if al:
            cond["$lte"] = al
        flt["data_movimento"] = cond
    items = await db.movimenti.find(flt, {"_id": 0}).sort("data_movimento", -1).to_list(limit)
    # arricchimento count allegati
    ids = [m["id"] for m in items]
    if ids:
        pipeline = [
            {"$match": {"entita_tipo": "movimento", "entita_id": {"$in": ids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]
        counts = {r["_id"]: r["n"] async for r in db.allegati.aggregate(pipeline)}
        for m in items:
            m["allegati_count"] = counts.get(m["id"], 0)
    return items


async def _risolvi_schema_provvigionale(
    collaboratore_id: Optional[str] = None,
    compagnia_id: Optional[str] = None,
    ramo: Optional[str] = None,
) -> Optional[dict]:
    """Risolve il SchemaProvvigionale più specifico applicabile.

    Ordine di preferenza (la più specifica vince):
      1. collaboratore + compagnia + ramo
      2. collaboratore + compagnia
      3. collaboratore + ramo
      4. collaboratore
      5. compagnia + ramo
      6. compagnia
      7. ramo
      8. default (tutto null)
    """
    candidates = [
        {"collaboratore_id": collaboratore_id, "compagnia_id": compagnia_id, "ramo": ramo},
        {"collaboratore_id": collaboratore_id, "compagnia_id": compagnia_id, "ramo": None},
        {"collaboratore_id": collaboratore_id, "compagnia_id": None, "ramo": ramo},
        {"collaboratore_id": collaboratore_id, "compagnia_id": None, "ramo": None},
        {"collaboratore_id": None, "compagnia_id": compagnia_id, "ramo": ramo},
        {"collaboratore_id": None, "compagnia_id": compagnia_id, "ramo": None},
        {"collaboratore_id": None, "compagnia_id": None, "ramo": ramo},
        {"collaboratore_id": None, "compagnia_id": None, "ramo": None},
    ]
    for q in candidates:
        if all(v is None for v in q.values()):
            continue  # salta combinazioni vuote
        q["attivo"] = True
        doc = await db.schema_provvigionale.find_one(q, {"_id": 0})
        if doc:
            return doc
    return None


async def _calcola_provvigione(
    premio_lordo: float,
    collaboratore_id: Optional[str],
    compagnia_id: Optional[str],
    ramo: Optional[str],
) -> dict:
    """Calcola provvigioni in base allo schema risolto.

    Ritorna {provvigione_totale, provvigione_collaboratore, schema_id, schema_nome}.
    Se non trova schema, ritorna 0.
    """
    if premio_lordo <= 0:
        return {"provvigione_totale": 0.0, "provvigione_collaboratore": 0.0,
                "schema_id": None, "schema_nome": None}
    schema = await _risolvi_schema_provvigionale(collaboratore_id, compagnia_id, ramo)
    if not schema:
        return {"provvigione_totale": 0.0, "provvigione_collaboratore": 0.0,
                "schema_id": None, "schema_nome": None}
    pct_su_premio = float(schema.get("percentuale_su_premio") or 0)
    pct_collab = float(schema.get("percentuale_collaboratore") or 0)
    provv_tot = round(premio_lordo * pct_su_premio / 100.0, 2)
    provv_collab = round(provv_tot * pct_collab / 100.0, 2)
    return {
        "provvigione_totale": provv_tot,
        "provvigione_collaboratore": provv_collab,
        "schema_id": schema.get("id"),
        "schema_nome": schema.get("nome"),
        "pct_su_premio": pct_su_premio,
        "pct_collab": pct_collab,
    }


async def _provv_breakdown(
    provv_totale_reale: float,
    collaboratore_id: Optional[str],
    compagnia_id: Optional[str],
    ramo: Optional[str],
) -> dict:
    """Suddivide una provvigione totale REALE (es. quella già impostata sulla polizza/
    titolo) in quota collaboratore e quota margine agenzia, usando la %
    `percentuale_collaboratore` dello schema applicabile.

    Ritorna:
      {
        provvigione_totale: float (= provv_totale_reale, arrotondato),
        provvigione_collaboratore: float,
        provvigione_margine: float,
        pct_collab: float,
        schema_id: Optional[str],
        schema_nome: Optional[str],
      }
    """
    tot = round(float(provv_totale_reale or 0), 2)
    if tot <= 0:
        return {
            "provvigione_totale": 0.0,
            "provvigione_collaboratore": 0.0,
            "provvigione_margine": 0.0,
            "pct_collab": 0.0,
            "schema_id": None,
            "schema_nome": None,
        }
    pct_collab = 0.0
    schema_id = None
    schema_nome = None
    if collaboratore_id:
        schema = await _risolvi_schema_provvigionale(collaboratore_id, compagnia_id, ramo)
        if schema:
            pct_collab = float(schema.get("percentuale_collaboratore") or 0)
            schema_id = schema.get("id")
            schema_nome = schema.get("nome")
    provv_collab = round(tot * pct_collab / 100.0, 2)
    provv_margine = round(tot - provv_collab, 2)
    return {
        "provvigione_totale": tot,
        "provvigione_collaboratore": provv_collab,
        "provvigione_margine": provv_margine,
        "pct_collab": pct_collab,
        "schema_id": schema_id,
        "schema_nome": schema_nome,
    }


@api.get("/provvigioni/calcola")
async def calcola_provvigione_endpoint(
    premio_lordo: float, collaboratore_id: Optional[str] = None,
    compagnia_id: Optional[str] = None, ramo: Optional[str] = None,
    user=Depends(current_user),
):
    """Calcola la provvigione attesa per un dato premio + collaboratore + compagnia + ramo."""
    return await _calcola_provvigione(premio_lordo, collaboratore_id, compagnia_id, ramo)


# ============================================================
# LIBRO MATRICOLA — Applicazioni veicoli su polizza RCA flotta
# ============================================================
@api.get("/polizze/{pid}/applicazioni")
async def list_applicazioni(
    pid: str,
    includi_storico: bool = False,
    q: Optional[str] = None,
    user=Depends(current_user),
):
    """Elenco applicazioni (veicoli) di un libro matricola.

    Default: solo applicazioni con stato 'attiva' o 'sospesa'.
    Se includi_storico=true ritorna anche annullate/sostituite.
    q: ricerca su targa/intestatario/marca/modello/numero/note.
    """
    flt: dict = {"polizza_id": pid}
    if not includi_storico:
        flt["stato"] = {"$in": ["attiva", "sospesa"]}
    if q:
        try:
            num = int(q)
            num_q = {"numero": num}
        except (TypeError, ValueError):
            num_q = None
        qrx = {"$regex": q, "$options": "i"}
        ors = [
            {"targa": qrx}, {"intestatario": qrx},
            {"marca": qrx}, {"modello": qrx}, {"note": qrx},
        ]
        if num_q:
            ors.append(num_q)
        flt["$or"] = ors
    items = await db.applicazioni.find(flt, {"_id": 0}).sort("numero", 1).to_list(5000)
    return items


@api.post("/polizze/{pid}/applicazioni/{aid}/sostituisci", status_code=201)
async def sostituisci_applicazione(
    pid: str, aid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Sostituisce un'applicazione: marca la corrente come 'sostituita'
    e crea una nuova applicazione collegata.

    body: campi della nuova applicazione (es. targa, marca, modello, valore_veicolo...)
          + motivo (opzionale)
    """
    old = await db.applicazioni.find_one({"id": aid, "polizza_id": pid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Applicazione non trovata")
    if old.get("stato") in ("annullata", "sostituita"):
        raise HTTPException(400, "Applicazione già non attiva")

    motivo = body.pop("motivo", None)
    today_iso = _now_iso()[:10]
    last = await db.applicazioni.find(
        {"polizza_id": pid}, {"_id": 0, "numero": 1},
    ).sort("numero", -1).limit(1).to_list(1)
    nuovo_numero = (last[0]["numero"] + 1) if last else 100

    # Eredita i campi tariffari dalla vecchia, sovrascritti dal body
    nuovo = {**{k: v for k, v in old.items() if k not in {"id", "created_at", "updated_at"}}, **body}
    nuovo["polizza_id"] = pid
    nuovo["numero"] = nuovo_numero
    nuovo["stato"] = "attiva"
    nuovo["data_inclusione"] = body.get("data_inclusione") or today_iso
    nuovo["data_esclusione"] = None
    nuovo["sostituisce_id"] = aid
    nuovo["sostituita_da_id"] = None
    nuovo["data_sostituzione"] = today_iso
    obj = ApplicazioneLibroMatricola(**nuovo)
    await db.applicazioni.insert_one(obj.model_dump())

    # marca old come sostituita
    await db.applicazioni.update_one(
        {"id": aid},
        {"$set": {
            "stato": "sostituita",
            "data_esclusione": today_iso,
            "sostituita_da_id": obj.id,
            "data_sostituzione": today_iso,
            "motivo_annullamento": motivo,
            "updated_at": _now_iso(),
        }},
    )
    await log_attivita(user, "sostituisci", "applicazione", obj.id,
                       f"Sostituita {old.get('targa')} -> {obj.targa}")
    return {"nuova": obj.model_dump(), "sostituita_id": aid}


@api.post("/polizze/{pid}/applicazioni/{aid}/annulla")
async def annulla_applicazione(
    pid: str, aid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Annulla un'applicazione del libro matricola (vendita/demolizione/…).

    body: { motivo: str, data_annullamento?: YYYY-MM-DD }
    """
    app = await db.applicazioni.find_one({"id": aid, "polizza_id": pid}, {"_id": 0})
    if not app:
        raise HTTPException(404, "Applicazione non trovata")
    if app.get("stato") in ("annullata", "sostituita"):
        raise HTTPException(400, "Applicazione già non attiva")
    motivo = (body.get("motivo") or "").strip()
    if not motivo:
        raise HTTPException(400, "Motivo obbligatorio")
    data_ann = body.get("data_annullamento") or _now_iso()[:10]
    await db.applicazioni.update_one(
        {"id": aid},
        {"$set": {
            "stato": "annullata",
            "data_esclusione": data_ann,
            "motivo_annullamento": motivo,
            "updated_at": _now_iso(),
        }},
    )
    await log_attivita(user, "annulla", "applicazione", aid,
                       f"Targa {app.get('targa')} — {motivo}")
    return {"ok": True, "stato": "annullata", "data_esclusione": data_ann}


@api.post("/polizze/{pid}/applicazioni", status_code=201)
async def create_applicazione(pid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    pol = await db.polizze.find_one({"id": pid}, {"_id": 0, "is_libro_matricola": 1})
    if not pol:
        raise HTTPException(404, "Polizza non trovata")
    body["polizza_id"] = pid
    # auto-progressivo se non fornito
    if not body.get("numero"):
        last = await db.applicazioni.find(
            {"polizza_id": pid}, {"_id": 0, "numero": 1},
        ).sort("numero", -1).limit(1).to_list(1)
        body["numero"] = (last[0]["numero"] + 1) if last else 100
    if not body.get("data_inclusione"):
        body["data_inclusione"] = _now_iso()[:10]
    obj = ApplicazioneLibroMatricola(**body)
    await db.applicazioni.insert_one(obj.model_dump())
    await log_attivita(user, "create", "applicazione", obj.id,
                       f"Veicolo {obj.targa} su polizza {pid[:6]}")
    return obj.model_dump()


@api.post("/polizze/{pid}/annulla")
async def annulla_polizza(
    pid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Annulla una polizza: setta stato=annullata, data e motivo."""
    pol = await db.polizze.find_one({"id": pid}, {"_id": 0})
    if not pol:
        raise HTTPException(404, "Polizza non trovata")
    if pol.get("stato") == "annullata":
        raise HTTPException(400, "Polizza già annullata")
    data_ann = body.get("data_annullamento") or _now_iso()[:10]
    motivo = body.get("motivo_annullamento") or "Annullamento"
    await db.polizze.update_one({"id": pid}, {"$set": {
        "stato": "annullata",
        "data_annullamento": data_ann,
        "motivo_annullamento": motivo,
        "updated_at": _now_iso(),
    }})
    await log_attivita(user, "annulla", "polizza", pid, f"Motivo: {motivo}")
    return {"ok": True, "stato": "annullata", "data_annullamento": data_ann}


@api.post("/polizze/{pid}/sospendi")
async def sospendi_polizza(
    pid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Mette in sospensione una polizza."""
    pol = await db.polizze.find_one({"id": pid}, {"_id": 0})
    if not pol:
        raise HTTPException(404, "Polizza non trovata")
    data_sosp = body.get("data_sospensione") or _now_iso()[:10]
    riatt = body.get("riattivazione_prevista")
    await db.polizze.update_one({"id": pid}, {"$set": {
        "stato": "sospesa",
        "data_sospensione": data_sosp,
        "riattivazione_prevista": riatt,
        "updated_at": _now_iso(),
    }})
    await log_attivita(user, "sospendi", "polizza", pid,
                       f"Riatt prev: {riatt or '—'}")
    return {"ok": True, "stato": "sospesa", "data_sospensione": data_sosp}


@api.post("/polizze/{pid}/riattiva")
async def riattiva_polizza(
    pid: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Riattiva una polizza sospesa."""
    res = await db.polizze.update_one(
        {"id": pid, "stato": "sospesa"},
        {"$set": {"stato": "attiva", "data_sospensione": None, "riattivazione_prevista": None,
                  "updated_at": _now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(400, "Polizza non sospesa o non trovata")
    await log_attivita(user, "riattiva", "polizza", pid)
    return {"ok": True, "stato": "attiva"}


@api.post("/polizze/{pid}/sostituisci")
async def sostituisci_polizza(
    pid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Sostituisce una polizza con una nuova polizza.

    body: {
        compagnia_id, ramo, numero_polizza, effetto,
        scadenza, prossima_quietanza?, coassicurazione?, premio_lordo?, ...
    }
    """
    old = await db.polizze.find_one({"id": pid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Polizza non trovata")
    if old.get("stato") in ("annullata", "sostituita"):
        raise HTTPException(400, "Polizza già non attiva")
    if not body.get("numero_polizza") or not body.get("effetto"):
        raise HTTPException(400, "numero_polizza ed effetto obbligatori")

    crea_titolo = bool(body.pop("crea_titolo", True))  # default: sempre crea titolo di sostituzione
    today = _now_iso()[:10]
    nuovo: dict = {
        # campi ereditati salvo override
        "contraente_id": old.get("contraente_id"),
        "assicurato_ids": old.get("assicurato_ids") or [],
        "collaboratore_id": old.get("collaboratore_id"),
        "targa": old.get("targa"),
        "frazionamento": old.get("frazionamento") or "annuale",
        "termini_mora_giorni": old.get("termini_mora_giorni") or 15,
        # override dal body (compagnia, ramo, prodotto, numero, date, premi, ecc.)
        **{k: v for k, v in body.items() if v not in (None, "")},
        # link sostituzione (sempre)
        "sostituisce_polizza": pid,
        "stato": "attiva",
    }
    # garantisci compagnia/ramo se non override
    nuovo.setdefault("compagnia_id", old.get("compagnia_id"))
    nuovo.setdefault("ramo", old.get("ramo"))
    obj = Polizza(**nuovo)
    await db.polizze.insert_one(obj.model_dump())
    # marca vecchia
    await db.polizze.update_one({"id": pid}, {"$set": {
        "stato": "sostituita",
        "sostituita_da_polizza_id": obj.id,
        "data_annullamento": today,
        "motivo_annullamento": body.get("motivo") or "Sostituzione contratto",
        "updated_at": _now_iso(),
    }})

    # Crea titolo iniziale (prima rata) se richiesto
    titolo_id = None
    if crea_titolo:
        prossima_quietanza = nuovo.get("prossima_quietanza")
        scad_titolo = prossima_quietanza if prossima_quietanza else _calcola_scadenza_titolo(
            nuovo.get("effetto"), nuovo.get("frazionamento") or "annuale",
        )
        titolo_obj = Titolo(
            polizza_id=obj.id,
            tipo="sostituzione",
            effetto=nuovo.get("effetto"),
            scadenza=scad_titolo or nuovo.get("scadenza") or nuovo.get("effetto"),
            stato="da_incassare",
            importo_lordo=float(nuovo.get("premio_lordo") or 0),
            importo_netto=float(nuovo.get("premio_netto") or 0),
            imposte=float(nuovo.get("premio_imposte") or 0),
            provvigioni=float(nuovo.get("provvigioni") or 0),
        )
        await db.titoli.insert_one(titolo_obj.model_dump())
        titolo_id = titolo_obj.id

    await log_attivita(user, "sostituisci", "polizza", obj.id,
                       f"{old.get('numero_polizza')} -> {obj.numero_polizza}"
                       + (f" + titolo {titolo_id[:8]}" if titolo_id else ""))
    return {
        "ok": True,
        "nuova_polizza_id": obj.id,
        "nuova": obj.model_dump(),
        "titolo_id": titolo_id,
    }


@api.put("/polizze/{pid}/applicazioni/{aid}")
async def update_applicazione(pid: str, aid: str, body: dict,
                              user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body.pop("id", None); body.pop("polizza_id", None); body.pop("created_at", None)
    body["updated_at"] = _now_iso()
    res = await db.applicazioni.update_one({"id": aid, "polizza_id": pid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Applicazione non trovata")
    doc = await db.applicazioni.find_one({"id": aid}, {"_id": 0})
    return doc


@api.delete("/polizze/{pid}/applicazioni/{aid}")
async def delete_applicazione(pid: str, aid: str,
                              user=Depends(require_user("admin", "collaboratore"))):
    res = await db.applicazioni.delete_one({"id": aid, "polizza_id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Applicazione non trovata")
    return {"ok": True}


@api.post("/contabilita/movimenti", status_code=201)
async def create_movimento(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    # Lock: blocco inserimento in giornata di Prima Nota chiusa.
    await assert_giornata_aperta(body.get("data_movimento"), azione="inserire un movimento nel giorno")
    obj = MovimentoContabile(**body)
    await db.movimenti.insert_one(obj.model_dump())
    await log_attivita(user, "create", "movimento", obj.id, f"€{obj.importo} {obj.descrizione}")
    return obj.model_dump()


@api.post("/contabilita/giroconto", status_code=201)
async def crea_giroconto(body: dict, user=Depends(require_user("admin", "collaboratore"))):
    """Trasferimento tra due conti cassa.

    body: {data_movimento, conto_da_id, conto_a_id, importo, descrizione?}
    Crea due movimenti gemelli:
      - uscita dal conto 'da'
      - entrata sul conto 'a'
    Entrambi categoria='giroconto', con `giroconto_pair_id` per accoppiarli.
    """
    da_id = body.get("conto_da_id")
    a_id = body.get("conto_a_id")
    importo = float(body.get("importo") or 0)
    data_mov = body.get("data_movimento") or _now_iso()[:10]
    # Lock: blocco inserimento giroconto in giornata di Prima Nota chiusa.
    await assert_giornata_aperta(data_mov, azione="creare un giroconto nel giorno")
    if not da_id or not a_id:
        raise HTTPException(400, "Indicare conto sorgente e destinazione")
    if da_id == a_id:
        raise HTTPException(400, "I conti devono essere diversi")
    if importo <= 0:
        raise HTTPException(400, "Importo non valido")

    conto_da = await db.conti_cassa.find_one({"id": da_id}, {"_id": 0, "nome": 1})
    conto_a = await db.conti_cassa.find_one({"id": a_id}, {"_id": 0, "nome": 1})
    if not conto_da or not conto_a:
        raise HTTPException(404, "Conto non trovato")

    nota = body.get("descrizione") or ""
    base_desc = (f"Giroconto: {conto_da['nome']} -> {conto_a['nome']}"
                 + (f" — {nota}" if nota else ""))
    pair_id = f"GR-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S%f')}"

    mov_out = MovimentoContabile(
        data_movimento=data_mov, tipo="uscita", categoria="giroconto",
        importo=importo, descrizione=base_desc, conto_cassa_id=da_id,
        mezzo_pagamento="giroconto",
        note=f"giroconto_pair_id={pair_id}; verso_conto_id={a_id}",
    )
    mov_in = MovimentoContabile(
        data_movimento=data_mov, tipo="entrata", categoria="giroconto",
        importo=importo, descrizione=base_desc, conto_cassa_id=a_id,
        mezzo_pagamento="giroconto",
        note=f"giroconto_pair_id={pair_id}; da_conto_id={da_id}",
    )
    # Salva il pair_id anche come campo top-level per supportare delete coppia
    out_doc = {**mov_out.model_dump(), "pair_id": pair_id}
    in_doc = {**mov_in.model_dump(), "pair_id": pair_id}
    await db.movimenti.insert_many([out_doc, in_doc])
    await log_attivita(user, "giroconto", "movimento", pair_id,
                       f"€{importo:.2f}: {conto_da['nome']} -> {conto_a['nome']}")
    return {
        "ok": True,
        "pair_id": pair_id,
        "movimento_uscita_id": mov_out.id,
        "movimento_entrata_id": mov_in.id,
        "descrizione_breve": f"{conto_da['nome']} -> {conto_a['nome']} {importo:.2f} €",
    }


@api.put("/contabilita/movimenti/{mid}")
async def update_movimento(mid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    cur = await db.movimenti.find_one({"id": mid}, {"_id": 0, "chiusura_id": 1})
    if cur and cur.get("chiusura_id"):
        raise HTTPException(400, "Movimento in giornata chiusa - riaprire la chiusura per modificare")
    body["updated_at"] = _now_iso()
    res = await db.movimenti.update_one({"id": mid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Movimento non trovato")
    await log_attivita(user, "update", "movimento", mid)
    return strip_mongo_id(await db.movimenti.find_one({"id": mid}, {"_id": 0}))


@api.delete("/contabilita/movimenti/{mid}")
async def delete_movimento(mid: str, user=Depends(require_user("admin", "collaboratore"))):
    cur = await db.movimenti.find_one({"id": mid}, {"_id": 0})
    if not cur:
        raise HTTPException(404, "Movimento non trovato")
    if cur.get("chiusura_id"):
        raise HTTPException(400, "Movimento in giornata chiusa - riaprire la chiusura per eliminare")
    # se è un giroconto, cancella anche la coppia
    pair_id = cur.get("pair_id") or cur.get("giroconto_pair_id")
    # Fallback: estrai pair_id dalla nota (per dati legacy)
    if not pair_id and cur.get("note"):
        import re as _re
        m_match = _re.search(r"giroconto_pair_id=([^;\s]+)", cur["note"])
        if m_match:
            pair_id = m_match.group(1)
    if pair_id:
        # cancella sia entries con pair_id top-level sia legacy con pair_id nella nota
        await db.movimenti.delete_many({
            "$or": [
                {"pair_id": pair_id},
                {"giroconto_pair_id": pair_id},
                {"note": {"$regex": f"giroconto_pair_id={pair_id}"}},
            ],
        })
        await log_attivita(user, "delete", "giroconto", pair_id)
        return {"ok": True, "deleted_pair": True}
    # se collegato a un rappel, blocca eliminazione
    if cur.get("rappel_id") or cur.get("is_rappel"):
        raise HTTPException(400, "Movimento collegato a un Rappel: usa 'Annulla incasso' dalla pagina Rappel")
    # se è un pagamento collaboratore, blocca eliminazione
    if cur.get("categoria") == "provvigioni" and cur.get("tipo") == "uscita":
        if cur.get("pagamento_provvigioni_id"):
            raise HTTPException(400, "Movimento collegato a Pagamento Provvigioni: elimina dal modulo Provvigioni")
    res = await db.movimenti.delete_one({"id": mid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Movimento non trovato")
    await log_attivita(user, "delete", "movimento", mid)
    return {"ok": True}


@api.get("/contabilita/estratto-conto/{anagrafica_id}")
async def estratto_conto(anagrafica_id: str, user=Depends(current_user)):
    if user["role"] == "cliente" and user.get("anagrafica_id") != anagrafica_id:
        raise HTTPException(403, "Permesso negato")
    movs = await db.movimenti.find({"anagrafica_id": anagrafica_id}, {"_id": 0}) \
        .sort("data_movimento", 1).to_list(1000)
    saldo = 0.0
    rows = []
    for m in movs:
        delta = m["importo"] if m["tipo"] == "entrata" else -m["importo"]
        saldo += delta
        rows.append({**m, "saldo_progressivo": round(saldo, 2)})
    ana = await db.anagrafiche.find_one({"id": anagrafica_id}, {"_id": 0})
    return {"anagrafica": ana, "movimenti": rows, "saldo_finale": round(saldo, 2)}


@api.get("/contabilita/prima-nota")
async def prima_nota(
    dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    flt: dict = {}
    if dal or al:
        cond = {}
        if dal:
            cond["$gte"] = dal
        if al:
            cond["$lte"] = al
        flt["data_movimento"] = cond
    items = await db.movimenti.find(flt, {"_id": 0}).sort("data_movimento", 1).to_list(2000)
    # arricchimento count allegati
    ids = [m["id"] for m in items]
    if ids:
        pipeline = [
            {"$match": {"entita_tipo": "movimento", "entita_id": {"$in": ids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]
        counts = {r["_id"]: r["n"] async for r in db.allegati.aggregate(pipeline)}
        for m in items:
            m["allegati_count"] = counts.get(m["id"], 0)
    totale_entrate = sum(m["importo"] for m in items if m["tipo"] == "entrata")
    totale_uscite = sum(m["importo"] for m in items if m["tipo"] == "uscita")
    return {
        "movimenti": items,
        "totale_entrate": round(totale_entrate, 2),
        "totale_uscite": round(totale_uscite, 2),
        "saldo": round(totale_entrate - totale_uscite, 2),
    }


# ============================================================
# BROGLIACCIO - Prima Nota giornaliera con colonne per conto cassa
# ============================================================
async def _compute_brogliaccio(data_giorno: str) -> dict:
    """Calcola il brogliaccio della Prima Nota con la semantica assicurativa corretta.

    Mappatura colonne (dal flusso descritto dall'utente):
      - TOTALE   = importo del movimento (per incassi premi: premio LORDO di polizza)
      - PROVV    = provvigioni totali agenzia maturate sul movimento
      - SALDO    = importo - provv  SE compagnia.trattiene_provvigioni
                 = importo intero  SE compagnia non trattiene (provvigioni ricevute a parte)
      - CREDITI  = sospesi / anticipi (categorie: anticipo, anticipo_cliente)
      - SPESE    = TUTTE le uscite (provvigioni collaboratori, stipendi, sconti polizze,
                   spese generali, rimesse compagnia). Sconto è un "di cui".
      - SCONTI   = subset di spese (categoria sconto_cliente)
      - RIMESSE  = pagamenti E/C compagnia (categoria pagamento_compagnia)
      - SALDO CASSA = sum(saldo) - sum(rimesse) (cumulativo periodo)

    Inoltre restituisce `saldi_compagnie`: per ogni compagnia che lavoriamo, il saldo
    cassa cumulativo (totale saldi su quella compagnia − totale rimesse a quella compagnia).
    """
    # 1) Conti cassa attivi ordinati
    conti = await db.conti_cassa.find(
        {"attivo": True}, {"_id": 0}
    ).sort([("ordine", 1), ("nome", 1)]).to_list(200)

    # 2) Movimenti del giorno + arricchimento polizza->compagnia
    movs = await db.movimenti.find(
        {"data_movimento": data_giorno}, {"_id": 0}
    ).sort("created_at", 1).to_list(2000)

    # raccogli polizza_ids per risolvere compagnie + numero polizza + contraente
    pol_ids = [m["polizza_id"] for m in movs if m.get("polizza_id")]
    polizze_map = {}
    if pol_ids:
        async for p in db.polizze.find(
            {"id": {"$in": pol_ids}},
            {"_id": 0, "id": 1, "compagnia_id": 1, "numero_polizza": 1, "contraente_id": 1},
        ):
            polizze_map[p["id"]] = p

    # contraenti da risolvere: da polizze + da movimento.anagrafica_id
    ana_ids = list(
        {p.get("contraente_id") for p in polizze_map.values() if p.get("contraente_id")}
        | {m.get("anagrafica_id") for m in movs if m.get("anagrafica_id")}
    )
    ana_map = {}
    if ana_ids:
        async for a in db.anagrafiche.find(
            {"id": {"$in": ana_ids}},
            {"_id": 0, "id": 1, "ragione_sociale": 1, "nome": 1, "cognome": 1},
        ):
            ana_map[a["id"]] = a

    # carica compagnie utilizzate (per trattiene_provvigioni)
    comp_ids = list({m.get("compagnia_id") for m in movs if m.get("compagnia_id")}
                    | {p.get("compagnia_id") for p in polizze_map.values() if p.get("compagnia_id")})
    comp_map = {}
    if comp_ids:
        async for c in db.compagnie.find(
            {"id": {"$in": comp_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1, "trattiene_provvigioni": 1}
        ):
            comp_map[c["id"]] = c

    def _compagnia_di(m: dict) -> Optional[dict]:
        if m.get("compagnia_id"):
            return comp_map.get(m["compagnia_id"])
        if m.get("polizza_id"):
            pid = m["polizza_id"]
            p = polizze_map.get(pid)
            if p and p.get("compagnia_id"):
                return comp_map.get(p["compagnia_id"])
        return None

    righe = []
    tot = {"totale": 0.0, "provv": 0.0, "saldo": 0.0,
           "crediti": 0.0, "spese": 0.0, "sconti": 0.0, "rimesse": 0.0}
    per_conto_tot: dict = {c["id"]: 0.0 for c in conti}

    CATS_ANTICIPI = ("anticipo",)  # sospesi/anticipi
    CATS_RIMESSE = ("pagamento_compagnia",)
    CATS_SCONTI = ("sconto_cliente",)

    for m in movs:
        cat = m["categoria"]
        is_entrata = (m["tipo"] == "entrata")
        importo = float(m.get("importo") or 0)
        signed = importo if is_entrata else -importo

        # default
        c_totale = 0.0; c_provv = 0.0; c_saldo = 0.0
        c_crediti = 0.0; c_spese = 0.0; c_sconti = 0.0; c_rimesse = 0.0

        compagnia_riga = _compagnia_di(m)
        polizza_riga = polizze_map.get(m.get("polizza_id"))
        # contraente: dalla polizza, altrimenti dal movimento
        contr_id = None
        if polizza_riga and polizza_riga.get("contraente_id"):
            contr_id = polizza_riga["contraente_id"]
        elif m.get("anagrafica_id"):
            contr_id = m["anagrafica_id"]
        contr = ana_map.get(contr_id) if contr_id else None
        contr_nome = (contr or {}).get("ragione_sociale") if contr else None

        if is_entrata and cat == "incasso_premio":
            # Premio lordo polizza
            provv_riga = float(m.get("provvigioni") or m.get("quota_provvigione") or 0)
            trattiene = (compagnia_riga or {}).get("trattiene_provvigioni", True)
            # Lo sconto eventualmente applicato è memorizzato sulla STESSA riga via
            # quota_sconto (l'agenzia ha incassato meno per uno sconto cliente).
            quota_sc = float(m.get("quota_sconto") or 0)
            lordo_riga = importo + quota_sc
            c_totale = lordo_riga                  # mostra il PREMIO LORDO
            c_provv = provv_riga
            c_spese = quota_sc                     # sconto in colonna Spese
            c_sconti = quota_sc
            if trattiene:
                # Caso A: incasso il premio, devo versare (premio - provv) alla compagnia
                # -> saldo = premio - provv (positivo = debito verso compagnia)
                c_saldo = importo - provv_riga
            else:
                # Caso B: pagamento Direzione/RID. Il cliente ha pagato direttamente
                # alla compagnia. La compagnia ha già il premio. Devo solo registrare la
                # provvigione che mi spetta -> saldo = -provv (negativo = compagnia mi deve)
                c_saldo = -provv_riga
            # Se l'incasso era parziale con residuo lasciato a sospeso, mostralo qui.
            quota_cred = float(m.get("quota_credito") or 0)
            if quota_cred > 0:
                c_crediti = quota_cred
        elif is_entrata and cat == "anticipo":
            # anticipo che entra in cassa: alimenta crediti positivamente
            c_totale = importo
            c_crediti = importo
        elif is_entrata and cat == "provvigioni":
            # Rappel (sovraprovvigione fittizia): NON in Entrate, va in Provvigioni.
            # Riduce il saldo verso la compagnia (avere fittizio).
            c_totale = 0.0
            c_provv = importo
            c_saldo = -importo
        elif is_entrata and cat == "giroconto":
            # Giroconto IN-side: NON in Entrate (è solo un trasferimento interno).
            # Il movimento di banca è già contabilizzato nella colonna conto specifico.
            c_totale = 0.0
        elif is_entrata:
            # Altre entrate generiche (rimborsi, voci manuali, sconti positivi):
            # NON contano in Entrate (la regola è: ENTRATE = solo polizze/titoli/appendici).
            # Mostriamo l'importo come informativo ma c_totale resta 0.
            c_totale = 0.0
        else:
            # USCITE
            if cat == "giroconto":
                # Giroconto OUT-side: NON va in TOTALE (è solo un trasferimento interno).
                # La banca uscita è già contabilizzata in per_conto.
                c_totale = 0.0
            elif cat in CATS_RIMESSE:
                # Versamento compagnia: NON va in TOTALE, solo in Rimesse + Banca uscita.
                # Riduce il saldo da versare alla compagnia (gestito da _compagnia_estratto_data).
                c_totale = 0.0
                c_rimesse = importo
            elif cat in CATS_ANTICIPI:
                # Restituzione anticipo: NON va in TOTALE, solo in colonna Sospesi (negativa).
                c_totale = 0.0
                c_crediti = -importo
            elif cat == "provvigioni":
                # Pagamento provvigioni a collaboratore: NON va in TOTALE.
                # Solo in Spese + Banca uscita.
                c_totale = 0.0
                c_spese = importo
            else:
                # Spese generali, sconti, stipendi, prelievi, altro: NON in Totale.
                # Solo in colonna Spese + Banca uscita.
                c_totale = 0.0
                c_spese = importo
                if cat in CATS_SCONTI:
                    c_sconti = importo

        # Per conto cassa
        per_conto = {}
        if m.get("conto_cassa_id"):
            per_conto[m["conto_cassa_id"]] = signed
            per_conto_tot[m["conto_cassa_id"]] = per_conto_tot.get(m["conto_cassa_id"], 0) + signed

        riga = {
            "id": m["id"],
            "descrizione": m.get("descrizione", "")[:120],
            "contraente": contr_nome,
            "numero_polizza": (polizza_riga or {}).get("numero_polizza"),
            "categoria": cat,
            "tipo": m["tipo"],
            "compagnia": (compagnia_riga or {}).get("ragione_sociale"),
            "compagnia_id": (compagnia_riga or {}).get("id"),
            "trattiene_provvigioni": (compagnia_riga or {}).get("trattiene_provvigioni"),
            "totale": c_totale,
            "provv": c_provv,
            "saldo": c_saldo,
            "crediti": c_crediti,
            "spese": c_spese,
            "sconti": c_sconti,
            "rimesse": c_rimesse,
            "per_conto": per_conto,
            "conto_cassa_id": m.get("conto_cassa_id"),
            "mezzo_pagamento": m.get("mezzo_pagamento"),
            "allegati_count": 0,
        }
        righe.append(riga)
        tot["totale"] += c_totale
        tot["provv"] += c_provv
        tot["saldo"] += c_saldo
        tot["crediti"] += c_crediti
        tot["spese"] += c_spese
        tot["sconti"] += c_sconti
        tot["rimesse"] += c_rimesse

    # arricchimento allegati count
    ids = [r["id"] for r in righe]
    if ids:
        pipeline = [
            {"$match": {"entita_tipo": "movimento", "entita_id": {"$in": ids}, "is_deleted": False}},
            {"$group": {"_id": "$entita_id", "n": {"$sum": 1}}},
        ]
        counts = {x["_id"]: x["n"] async for x in db.allegati.aggregate(pipeline)}
        for r in righe:
            r["allegati_count"] = counts.get(r["id"], 0)

    totali_giornata = {**{k: round(v, 2) for k, v in tot.items()},
                       "per_conto": {k: round(v, 2) for k, v in per_conto_tot.items()}}

    # 3) Riepilogo conti cassa: saldo precedente, giornata, totale periodo
    conti_riepilogo = []
    for c in conti:
        prev = await db.movimenti.aggregate([
            {"$match": {"conto_cassa_id": c["id"], "data_movimento": {"$lt": data_giorno}}},
            {"$group": {"_id": None,
                        "in": {"$sum": {"$cond": [{"$eq": ["$tipo", "entrata"]}, "$importo", 0]}},
                        "out": {"$sum": {"$cond": [{"$eq": ["$tipo", "uscita"]}, "$importo", 0]}}}},
        ]).to_list(1)
        prev_in = prev[0]["in"] if prev else 0
        prev_out = prev[0]["out"] if prev else 0
        imp_prec = float(c.get("saldo_iniziale", 0)) + prev_in - prev_out
        imp_giornata = per_conto_tot.get(c["id"], 0.0)
        conti_riepilogo.append({
            "id": c["id"],
            "nome": c["nome"],
            "imp_precedente": round(imp_prec, 2),
            "imp_giornata": round(imp_giornata, 2),
            "totale_periodo": round(imp_prec + imp_giornata, 2),
        })

    # 4) Saldi compagnie (cumulativi: dal'inizio a fine giornata)
    saldi_compagnie = await _compute_saldi_compagnie(data_giorno)

    # 5) KPI CUMULATIVI fino al giorno corrente compreso (totali progressivi che crescono ogni giorno)
    cum_movs = await db.movimenti.find(
        {"data_movimento": {"$lte": data_giorno}}, {"_id": 0}
    ).to_list(50000)
    # carica le compagnie per applicare "trattiene_provvigioni" correttamente
    all_comp_ids = list({m.get("compagnia_id") for m in cum_movs if m.get("compagnia_id")})
    pol_ids_cum = list({m.get("polizza_id") for m in cum_movs if m.get("polizza_id")})
    if pol_ids_cum:
        async for p in db.polizze.find(
            {"id": {"$in": pol_ids_cum}}, {"_id": 0, "id": 1, "compagnia_id": 1},
        ):
            if p.get("compagnia_id"):
                all_comp_ids.append(p["compagnia_id"])
                polizze_map[p["id"]] = polizze_map.get(p["id"], p)
    all_comp_ids = list(set(all_comp_ids))
    if all_comp_ids:
        async for c in db.compagnie.find(
            {"id": {"$in": all_comp_ids}}, {"_id": 0, "id": 1, "trattiene_provvigioni": 1, "ragione_sociale": 1},
        ):
            comp_map.setdefault(c["id"], c)

    cum_entrate = 0.0; cum_provv = 0.0; cum_sospesi = 0.0
    cum_rimesse = 0.0; cum_sconti = 0.0; cum_spese = 0.0; cum_saldo = 0.0
    for m in cum_movs:
        importo = float(m.get("importo") or 0)
        cat = m["categoria"]; is_e = (m["tipo"] == "entrata")
        # Regola contabile dell'utente:
        # ENTRATE = SOLO premi/titoli/appendici (incasso_premio)
        # Rappel (entrata cat=provvigioni) -> in PROVVIGIONI, non in Entrate
        # Tutte le altre entrate generiche NON contano in Entrate
        if is_e and cat == "incasso_premio":
            cum_entrate += importo
        if is_e and cat == "provvigioni":
            cum_provv += importo  # rappel
        if is_e and cat == "incasso_premio":
            provv_riga = float(m.get("provvigioni") or 0)
            quota_sc = float(m.get("quota_sconto") or 0)
            # determina compagnia (via movimento o polizza)
            ccid = m.get("compagnia_id")
            if not ccid and m.get("polizza_id"):
                ccid = polizze_map.get(m["polizza_id"], {}).get("compagnia_id")
            trattiene = (comp_map.get(ccid) or {}).get("trattiene_provvigioni", True)
            cum_provv += provv_riga
            cum_saldo += (importo - provv_riga) if trattiene else (-provv_riga)
            # Sconto cliente memorizzato sulla riga: alimenta sia Sconti che Spese
            if quota_sc > 0:
                cum_sconti += quota_sc
                cum_spese += quota_sc
        elif is_e and cat == "anticipo":
            cum_sospesi += importo
        elif (not is_e) and cat == "pagamento_compagnia":
            cum_rimesse += importo
        elif (not is_e) and cat == "anticipo":
            cum_sospesi -= importo
        elif not is_e:
            cum_spese += importo
            if cat == "sconto_cliente":
                cum_sconti += importo

    riepilogo_kpi = {
        "entrate": round(cum_entrate, 2),
        "provvigioni": round(cum_provv, 2),
        "crediti": round(cum_sospesi, 2),
        "rimesse": round(cum_rimesse, 2),
        "sconti": round(cum_sconti, 2),
        "spese": round(cum_spese, 2),
        "saldo_cassa": round(cum_saldo - cum_rimesse, 2),
    }

    # Override 'crediti' (KPI Sospesi) con la stessa fonte della pagina /titoli/sospesi:
    # somma importo_lordo dei titoli ancora sospesi alla data corrente.
    sospesi_titoli_tot = await _total_sospesi_as_of(data_giorno)
    riepilogo_kpi["crediti"] = sospesi_titoli_tot

    # 5b) Liquidità cumulative (fino al giorno corrente compreso)
    saldo_cassa_compagnie_tot = round(sum(s["saldo_cassa"] for s in saldi_compagnie), 2)
    conti_attivi = await db.conti_cassa.find({"attivo": True}, {"_id": 0}).to_list(200)
    sum_conti = 0.0
    for c in conti_attivi:
        agg = await db.movimenti.aggregate([
            {"$match": {"conto_cassa_id": c["id"], "data_movimento": {"$lte": data_giorno}}},
            {"$group": {"_id": None,
                        "in": {"$sum": {"$cond": [{"$eq": ["$tipo", "entrata"]}, "$importo", 0]}},
                        "out": {"$sum": {"$cond": [{"$eq": ["$tipo", "uscita"]}, "$importo", 0]}}}},
        ]).to_list(1)
        in_tot = agg[0]["in"] if agg else 0
        out_tot = agg[0]["out"] if agg else 0
        sum_conti += float(c.get("saldo_iniziale", 0)) + in_tot - out_tot

    sospesi_attivi = sospesi_titoli_tot

    liquidita_box = {
        "sum_conti": round(sum_conti, 2),
        "sospesi_attivi": round(sospesi_attivi, 2),
        "saldo_cassa_compagnie": saldo_cassa_compagnie_tot,
        "liquidita_disponibile": round(sum_conti - sospesi_attivi - saldo_cassa_compagnie_tot, 2),
        "liquidita_postera": round(sum_conti - saldo_cassa_compagnie_tot, 2),
    }

    # 6) Chiusura?
    chiusura = await db.chiusure_giorno.find_one({"data": data_giorno}, {"_id": 0})

    return {
        "data": data_giorno,
        "conti_cassa": [{"id": c["id"], "nome": c["nome"], "ordine": c.get("ordine", 0)} for c in conti],
        "righe": righe,
        "totali_giornata": totali_giornata,
        "conti_riepilogo": conti_riepilogo,
        "saldi_compagnie": saldi_compagnie,
        "riepilogo_kpi": riepilogo_kpi,
        "liquidita": liquidita_box,
        "chiusura": chiusura,
        "chiusa": bool(chiusura and not chiusura.get("riaperta_at")),
    }


async def _compute_saldi_compagnie(fino_a: str) -> list:
    """Saldo cassa cumulativo per ogni compagnia attiva (da inizio fino al giorno incluso).

    Per ogni compagnia: somma saldi (incassi premio) − somma rimesse pagate.
    Saldo > 0 = dobbiamo ancora versare alla compagnia.
    """
    out = []
    compagnie = await db.compagnie.find({"attiva": True}, {"_id": 0}).to_list(500)
    for c in compagnie:
        cid = c["id"]
        trattiene = c.get("trattiene_provvigioni", True)
        # incassi premio per polizze di questa compagnia (fino a fino_a)
        pol_ids = [p["id"] async for p in db.polizze.find(
            {"compagnia_id": cid}, {"_id": 0, "id": 1}
        )]
        incassi = 0.0; provv = 0.0
        if pol_ids:
            agg = await db.movimenti.aggregate([
                {"$match": {
                    "polizza_id": {"$in": pol_ids},
                    "tipo": "entrata",
                    "categoria": "incasso_premio",
                    "data_movimento": {"$lte": fino_a},
                }},
                {"$group": {"_id": None,
                            "tot": {"$sum": "$importo"},
                            "provv": {"$sum": "$provvigioni"}}},
            ]).to_list(1)
            if agg:
                incassi = agg[0]["tot"]
                provv = agg[0]["provv"] or 0.0
        # rimesse pagate a questa compagnia
        rim_agg = await db.movimenti.aggregate([
            {"$match": {
                "compagnia_id": cid,
                "tipo": "uscita",
                "categoria": "pagamento_compagnia",
                "data_movimento": {"$lte": fino_a},
            }},
            {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
        ]).to_list(1)
        rimesse = rim_agg[0]["tot"] if rim_agg else 0.0
        # rappel cumulativi per questa compagnia (riducono il saldo da versare)
        rap_agg = await db.rappel.aggregate([
            {"$match": {"compagnia_id": cid, "data": {"$lte": fino_a}}},
            {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
        ]).to_list(1)
        rappel = rap_agg[0]["tot"] if rap_agg else 0.0
        # saldo = (premi - provv) se trattiene else premi  ;  meno rimesse  ; meno rappel
        saldo_dovuto = (incassi - provv) if trattiene else incassi
        saldo_residuo = saldo_dovuto - rimesse - rappel
        if abs(saldo_residuo) < 0.01 and abs(incassi) < 0.01 and abs(rappel) < 0.01:
            continue  # skip compagnie senza movimenti
        out.append({
            "compagnia_id": cid,
            "compagnia": c["ragione_sociale"],
            "trattiene_provvigioni": trattiene,
            "incassi_lordi": round(incassi, 2),
            "provvigioni": round(provv, 2),
            "saldo_dovuto": round(saldo_dovuto, 2),
            "rimesse_pagate": round(rimesse, 2),
            "rappel": round(rappel, 2),
            "saldo_cassa": round(saldo_residuo, 2),
        })
    # ordina per saldo_cassa descendente (le più "da pagare" in cima)
    out.sort(key=lambda x: abs(x["saldo_cassa"]), reverse=True)
    return out


@api.get("/contabilita/dati-compagnie")
async def dati_compagnie(
    dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Modulo Dati Compagnie: per ogni compagnia incassi lordi/netti, provvigioni,
    rimesse pagate, saldo attuale (cumulativo).

    Usa la STESSA fonte dati di `/api/compagnie/saldi-cassa` (db.titoli con
    stato='incassato' e db.movimenti categoria='pagamento_compagnia') per
    garantire saldi identici fra le due viste.

    Se dal/al sono passati, gli importi sono filtrati per il periodo;
    il saldo attuale rimane sempre cumulativo (fino a oggi)."""
    compagnie = await db.compagnie.find({"attiva": True}, {"_id": 0}).sort("ragione_sociale", 1).to_list(500)
    rows = []

    for c in compagnie:
        cid = c["id"]
        trattiene = bool(c.get("trattiene_provvigioni", True))
        pol_ids = [p["id"] async for p in db.polizze.find(
            {"compagnia_id": cid}, {"_id": 0, "id": 1}
        )]

        # ---- importi del PERIODO (filtrati su data_incasso) ----
        incassi_lordi = 0.0; provv_periodo = 0.0
        if pol_ids:
            match_t: dict = {"polizza_id": {"$in": pol_ids}, "stato": "incassato"}
            if dal or al:
                cond_t: dict = {}
                if dal: cond_t["$gte"] = dal
                if al: cond_t["$lte"] = al
                match_t["data_incasso"] = cond_t
            agg = await db.titoli.aggregate([
                {"$match": match_t},
                {"$group": {"_id": None,
                            "lordo": {"$sum": "$importo_lordo"},
                            "provv": {"$sum": "$provvigioni"}}},
            ]).to_list(1)
            if agg:
                incassi_lordi = agg[0]["lordo"] or 0
                provv_periodo = agg[0]["provv"] or 0
        netto_dovuto_periodo = (incassi_lordi - provv_periodo) if trattiene else incassi_lordi

        # rimesse del periodo (data_movimento)
        rim_match: dict = {"compagnia_id": cid, "tipo": "uscita", "categoria": "pagamento_compagnia"}
        if dal or al:
            cond_r: dict = {}
            if dal: cond_r["$gte"] = dal
            if al: cond_r["$lte"] = al
            rim_match["data_movimento"] = cond_r
        rim_agg = await db.movimenti.aggregate([
            {"$match": rim_match}, {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
        ]).to_list(1)
        rimesse_periodo = rim_agg[0]["tot"] if rim_agg else 0

        # rappel del periodo
        rap_match: dict = {"compagnia_id": cid}
        if dal or al:
            cond_rap: dict = {}
            if dal: cond_rap["$gte"] = dal
            if al: cond_rap["$lte"] = al
            rap_match["data"] = cond_rap
        rap_agg = await db.rappel.aggregate([
            {"$match": rap_match}, {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
        ]).to_list(1)
        rappel_periodo = rap_agg[0]["tot"] if rap_agg else 0

        # ---- SALDO ATTUALE cumulativo (TUTTI i titoli incassati, senza filtro data)
        # Stessa logica di /api/compagnie/saldi-cassa per garantire numeri identici. ----
        saldo_attuale = 0.0
        if pol_ids:
            agg_cum = await db.titoli.aggregate([
                {"$match": {"polizza_id": {"$in": pol_ids}, "stato": "incassato"}},
                {"$group": {"_id": None,
                            "lordo": {"$sum": "$importo_lordo"},
                            "provv": {"$sum": "$provvigioni"}}},
            ]).to_list(1)
            if agg_cum:
                cum_lordo = agg_cum[0]["lordo"] or 0
                cum_provv = agg_cum[0]["provv"] or 0
                saldo_attuale = (cum_lordo - cum_provv) if trattiene else cum_lordo
        rim_cum_agg = await db.movimenti.aggregate([
            {"$match": {"compagnia_id": cid, "tipo": "uscita", "categoria": "pagamento_compagnia"}},
            {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
        ]).to_list(1)
        saldo_attuale -= (rim_cum_agg[0]["tot"] if rim_cum_agg else 0)
        # sottrai rappel cumulativi (tutti)
        rap_cum_agg = await db.rappel.aggregate([
            {"$match": {"compagnia_id": cid}},
            {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
        ]).to_list(1)
        saldo_attuale -= (rap_cum_agg[0]["tot"] if rap_cum_agg else 0)

        if abs(incassi_lordi) < 0.01 and abs(rimesse_periodo) < 0.01 and abs(rappel_periodo) < 0.01 and abs(saldo_attuale) < 0.01:
            continue  # skip compagnie senza nulla

        rows.append({
            "compagnia_id": cid,
            "compagnia": c["ragione_sociale"],
            "trattiene_provvigioni": trattiene,
            "incassi_lordi": round(incassi_lordi, 2),
            "incassi_netti": round(netto_dovuto_periodo, 2),
            "provvigioni": round(provv_periodo, 2),
            "rimesse_pagate": round(rimesse_periodo, 2),
            "rappel": round(rappel_periodo, 2),
            "saldo_attuale": round(saldo_attuale, 2),
        })
    rows.sort(key=lambda x: abs(x["saldo_attuale"]), reverse=True)
    totali = {
        "incassi_lordi": round(sum(r["incassi_lordi"] for r in rows), 2),
        "incassi_netti": round(sum(r["incassi_netti"] for r in rows), 2),
        "provvigioni": round(sum(r["provvigioni"] for r in rows), 2),
        "rimesse_pagate": round(sum(r["rimesse_pagate"] for r in rows), 2),
        "rappel": round(sum(r["rappel"] for r in rows), 2),
        "saldo_attuale": round(sum(r["saldo_attuale"] for r in rows), 2),
    }
    return {"periodo": {"dal": dal, "al": al}, "compagnie": rows, "totali": totali}


@api.get("/contabilita/dati-compagnie/stampa")
async def stampa_dati_compagnie(
    dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    payload = await dati_compagnie(dal=dal, al=al, user=user)
    headers = ["Compagnia", "Regime", "Incassi lordi €", "Incassi netti €",
               "Provvigioni €", "Rimesse pagate €", "Saldo attuale €"]
    rows = [[r["compagnia"],
             "Tratteniamo" if r["trattiene_provvigioni"] else "No trattenute",
             r["incassi_lordi"], r["incassi_netti"],
             r["provvigioni"], r["rimesse_pagate"], r["saldo_attuale"]]
            for r in payload["compagnie"]]
    t = payload["totali"]
    rows.append(["TOTALE", "", t["incassi_lordi"], t["incassi_netti"],
                 t["provvigioni"], t["rimesse_pagate"], t["saldo_attuale"]])
    pdf = pdf_report.stampa_elenco(
        "Dati Compagnie", f"Periodo: {dal or '-'} -> {al or '-'}", headers, rows,
        col_widths_mm=[55, 25, 25, 25, 25, 28, 28],
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "dati_compagnie.pdf")



async def statistiche_contabilita(
    dal: Optional[str] = None, al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
) -> dict:
    """Statistiche contabilità per Prima Nota.

    Periodo (default: dall'inizio fino ad oggi):
      - 7 KPI: entrate, provvigioni, crediti, rimesse, sconti, spese, saldo_cassa_compagnie
      - Saldo cassa per Compagnia (cumulativo fino ad oggi)
      - Saldo per Conto Cassa (saldo iniziale + entrate − uscite, cumulativo)
      - Liquidità Disponibile = sum(conti) − sospesi/anticipi attivi − saldo_cassa_compagnie
      - Liquidità Postera = sum(conti) − saldo_cassa_compagnie
    """
    # Filtri periodo (usati solo per KPI di periodo, non per cumulativi)
    flt_periodo = {"is_deleted": {"$ne": True}}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        flt_periodo["data_movimento"] = cond

    # KPI di periodo
    movs_per = await db.movimenti.find(flt_periodo, {"_id": 0}).to_list(20000)
    entrate_per = sum(m["importo"] for m in movs_per if m["tipo"] == "entrata")
    by_cat = {}
    for m in movs_per:
        if m["tipo"] == "uscita":
            by_cat[m["categoria"]] = by_cat.get(m["categoria"], 0) + m["importo"]
    crediti_per = sum(m["importo"] for m in movs_per
                      if m["categoria"] == "anticipo" and m["tipo"] == "entrata")

    today = _now_iso()[:10]
    # Saldi compagnie cumulativi (fino a oggi)
    saldi_comp = await _compute_saldi_compagnie(today)
    saldo_cassa_compagnie_tot = round(sum(s["saldo_cassa"] for s in saldi_comp), 2)

    # Saldo per conto cassa cumulativo (saldo iniziale + entrate − uscite, TUTTO il tempo)
    conti = await db.conti_cassa.find(
        {"attivo": True}, {"_id": 0}
    ).sort([("ordine", 1), ("nome", 1)]).to_list(200)
    saldi_conti = []
    sum_conti = 0.0
    for c in conti:
        agg = await db.movimenti.aggregate([
            {"$match": {"conto_cassa_id": c["id"]}},
            {"$group": {"_id": None,
                        "in": {"$sum": {"$cond": [{"$eq": ["$tipo", "entrata"]}, "$importo", 0]}},
                        "out": {"$sum": {"$cond": [{"$eq": ["$tipo", "uscita"]}, "$importo", 0]}}}},
        ]).to_list(1)
        in_tot = agg[0]["in"] if agg else 0
        out_tot = agg[0]["out"] if agg else 0
        saldo = float(c.get("saldo_iniziale", 0)) + in_tot - out_tot
        sum_conti += saldo
        saldi_conti.append({
            "id": c["id"], "nome": c["nome"],
            "saldo_iniziale": float(c.get("saldo_iniziale", 0)),
            "entrate": round(in_tot, 2),
            "uscite": round(out_tot, 2),
            "saldo_attuale": round(saldo, 2),
        })

    # Crediti attivi = sospesi/anticipi (entrate categoria=anticipo) non ancora chiusi
    crediti_agg = await db.movimenti.aggregate([
        {"$match": {"categoria": "anticipo", "tipo": "entrata"}},
        {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
    ]).to_list(1)
    crediti_storno = await db.movimenti.aggregate([
        {"$match": {"categoria": "anticipo", "tipo": "uscita"}},
        {"$group": {"_id": None, "tot": {"$sum": "$importo"}}},
    ]).to_list(1)
    crediti_attivi = (crediti_agg[0]["tot"] if crediti_agg else 0) - (crediti_storno[0]["tot"] if crediti_storno else 0)

    liquidita_disponibile = round(sum_conti - crediti_attivi - saldo_cassa_compagnie_tot, 2)
    liquidita_postera = round(sum_conti - saldo_cassa_compagnie_tot, 2)

    # KPI di periodo per le 7 cards (semantica brogliaccio)
    kpi = {
        "entrate": round(entrate_per, 2),
        "provvigioni": round(by_cat.get("provvigioni", 0), 2),
        "crediti": round(crediti_per, 2),
        "rimesse": round(by_cat.get("pagamento_compagnia", 0), 2),
        "sconti": round(by_cat.get("sconto_cliente", 0), 2),
        "spese": round(sum(v for k, v in by_cat.items() if k != "pagamento_compagnia"), 2),
        "saldo_cassa_compagnie": saldo_cassa_compagnie_tot,
    }

    return {
        "periodo": {"dal": dal, "al": al},
        "kpi": kpi,
        "saldi_conti": saldi_conti,
        "saldi_compagnie": saldi_comp,
        "sum_conti": round(sum_conti, 2),
        "crediti_attivi": round(crediti_attivi, 2),
        "liquidita_disponibile": liquidita_disponibile,
        "liquidita_postera": liquidita_postera,
    }


@api.get("/contabilita/brogliaccio")
async def brogliaccio(
    data: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    if not data:
        data = _now_iso()[:10]
    return await _compute_brogliaccio(data)


async def _generate_brogliaccio_pdf(payload: dict) -> bytes:
    """Genera il PDF brogliaccio da un payload di _compute_brogliaccio."""
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    logo_bytes = None
    if az.get("logo_storage_path"):
        try:
            logo_bytes, _ = obj_storage.get_object(az["logo_storage_path"])
        except Exception:
            logo_bytes = None
    chiusura_info = None
    if payload.get("chiusa"):
        ch = payload["chiusura"]
        chiusura_info = {"closed_at": ch.get("created_at", "")[:19].replace("T", " "),
                         "closed_by_name": ch.get("closed_by_name", "")}
    return pdf_brogliaccio.stampa_brogliaccio(
        data_giorno=payload["data"],
        azienda=az,
        conti_cassa=payload["conti_cassa"],
        righe=payload["righe"],
        totali_giornata=payload["totali_giornata"],
        conti_riepilogo=payload["conti_riepilogo"],
        riepilogo_kpi=payload["riepilogo_kpi"],
        saldi_compagnie=payload.get("saldi_compagnie"),
        chiusura_info=chiusura_info,
        logo_bytes=logo_bytes,
    )


@api.get("/contabilita/brogliaccio/stampa")
async def stampa_brogliaccio_pdf(
    data: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    if not data:
        data = _now_iso()[:10]
    payload = await _compute_brogliaccio(data)
    pdf_bytes = await _generate_brogliaccio_pdf(payload)
    return _pdf_response(pdf_bytes, f"brogliaccio_{data}.pdf")


@api.post("/contabilita/chiusura-giorno", status_code=201)
async def chiudi_giorno(
    body: dict, user=Depends(require_user("admin", "collaboratore")),
):
    """Chiude la prima nota del giorno: snapshot + PDF + segna movimenti come immutabili.

    body: { data: "YYYY-MM-DD", invia_commercialista?: bool }
    """
    data_giorno = body.get("data") or _now_iso()[:10]
    # già chiusa?
    existing = await db.chiusure_giorno.find_one(
        {"data": data_giorno, "riaperta_at": None}, {"_id": 0}
    )
    if existing:
        raise HTTPException(400, f"Giornata {data_giorno} già chiusa")

    payload = await _compute_brogliaccio(data_giorno)
    if not payload["righe"]:
        raise HTTPException(400, "Nessun movimento da chiudere per questo giorno")

    chiusura = ChiusuraGiorno(
        data=data_giorno,
        closed_by=user["id"],
        closed_by_name=user.get("name", user["email"]),
        riepilogo={
            "totali_giornata": payload["totali_giornata"],
            "riepilogo_kpi": payload["riepilogo_kpi"],
            "conti_riepilogo": payload["conti_riepilogo"],
            "n_movimenti": len(payload["righe"]),
        },
    )

    # genera PDF e salva in storage
    payload["chiusa"] = True
    payload["chiusura"] = chiusura.model_dump()
    pdf_bytes = await _generate_brogliaccio_pdf(payload)
    storage_path = (f"{os.environ.get('APP_NAME', 'assicura')}"
                    f"/brogliaccio/{data_giorno}/{chiusura.id}.pdf")
    try:
        obj_storage.put_object(storage_path, pdf_bytes, "application/pdf")
        chiusura.pdf_storage_path = storage_path
    except Exception as e:
        raise HTTPException(503, f"Errore salvataggio PDF: {e}")

    # 1) salva chiusura, 2) marca movimenti. Se 2 fallisce, almeno chiusura è ok.
    await db.chiusure_giorno.insert_one(chiusura.model_dump())
    await db.movimenti.update_many(
        {"data_movimento": data_giorno},
        {"$set": {"chiusura_id": chiusura.id, "updated_at": _now_iso()}},
    )
    await log_attivita(user, "chiusura_giorno", "contabilita", chiusura.id,
                       f"Chiusura brogliaccio {data_giorno} ({len(payload['righe'])} movimenti)")

    invia = body.get("invia_commercialista") or False
    invio_result = None
    if invia:
        invio_result = await _invia_brogliaccio_email(chiusura.id, pdf_bytes, data_giorno)

    return {**chiusura.model_dump(), "invio_commercialista": invio_result}


@api.post("/contabilita/chiusura-giorno/{chiusura_id}/invia")
async def invia_chiusura_giorno(
    chiusura_id: str, user=Depends(require_user("admin", "collaboratore")),
):
    ch = await db.chiusure_giorno.find_one({"id": chiusura_id}, {"_id": 0})
    if not ch:
        raise HTTPException(404, "Chiusura non trovata")
    if not ch.get("pdf_storage_path"):
        raise HTTPException(400, "PDF non disponibile")
    try:
        pdf_bytes, _ = obj_storage.get_object(ch["pdf_storage_path"])
    except Exception as e:
        raise HTTPException(503, f"Errore lettura PDF: {e}")
    res = await _invia_brogliaccio_email(chiusura_id, pdf_bytes, ch["data"])
    return res


async def _invia_brogliaccio_email(chiusura_id: str, pdf_bytes: bytes, data_giorno: str) -> dict:
    """Invia il brogliaccio chiuso al commercialista (se SMTP configurato)."""
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    to_addr = az.get("email_commercialista")
    if not to_addr:
        await db.chiusure_giorno.update_one(
            {"id": chiusura_id},
            {"$set": {"email_errore": "Email commercialista non configurata in Librerie/Azienda",
                      "updated_at": _now_iso()}},
        )
        return {"ok": False, "errore": "Email commercialista non configurata in Librerie/Azienda"}
    if not az.get("smtp_host") or not az.get("smtp_user"):
        await db.chiusure_giorno.update_one(
            {"id": chiusura_id},
            {"$set": {"email_errore": "SMTP non configurato (host/user mancanti)",
                      "updated_at": _now_iso()}},
        )
        return {"ok": False, "errore": "SMTP non configurato (Librerie/Azienda -> SMTP)"}

    import smtplib
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["From"] = az.get("smtp_from") or az["smtp_user"]
    msg["To"] = to_addr
    msg["Subject"] = f"Prima Nota chiusa - {data_giorno} - {az.get('ragione_sociale', '')}"
    nome_comm = az.get("nome_commercialista") or "Studio"
    msg.set_content(
        f"Gentile {nome_comm},\n\n"
        f"in allegato la prima nota chiusa del {data_giorno}.\n\n"
        f"Cordiali saluti,\n{az.get('ragione_sociale', '')}"
    )
    msg.add_attachment(
        pdf_bytes, maintype="application", subtype="pdf",
        filename=f"brogliaccio_{data_giorno}.pdf",
    )
    try:
        port = int(az.get("smtp_port") or 587)
        if port == 465:
            srv = smtplib.SMTP_SSL(az["smtp_host"], port, timeout=30)
        else:
            srv = smtplib.SMTP(az["smtp_host"], port, timeout=30)
            if az.get("smtp_use_tls", True):
                srv.starttls()
        srv.login(az["smtp_user"], az.get("smtp_password") or "")
        srv.send_message(msg)
        srv.quit()
        now = _now_iso()
        await db.chiusure_giorno.update_one(
            {"id": chiusura_id},
            {"$set": {"email_inviata_a": to_addr, "email_inviata_at": now,
                      "email_errore": None, "updated_at": now}},
        )
        return {"ok": True, "inviata_a": to_addr}
    except Exception as e:
        await db.chiusure_giorno.update_one(
            {"id": chiusura_id},
            {"$set": {"email_errore": str(e), "updated_at": _now_iso()}},
        )
        return {"ok": False, "errore": str(e)}


@api.post("/contabilita/chiusura-giorno/{chiusura_id}/riapri")
async def riapri_chiusura_giorno(
    chiusura_id: str, body: dict, user=Depends(require_user("admin")),
):
    """Riapre una giornata già chiusa (solo admin). Richiede motivo."""
    motivo = (body or {}).get("motivo")
    if not motivo:
        raise HTTPException(400, "Motivo riapertura obbligatorio")
    ch = await db.chiusure_giorno.find_one({"id": chiusura_id}, {"_id": 0})
    if not ch:
        raise HTTPException(404, "Chiusura non trovata")
    now = _now_iso()
    await db.chiusure_giorno.update_one(
        {"id": chiusura_id},
        {"$set": {"riaperta_at": now, "riaperta_by": user["id"],
                  "riaperta_motivo": motivo, "updated_at": now}},
    )
    await db.movimenti.update_many(
        {"chiusura_id": chiusura_id},
        {"$set": {"chiusura_id": None, "updated_at": now}},
    )
    await log_attivita(user, "riapri_chiusura", "contabilita", chiusura_id,
                       f"Riaperta chiusura {ch['data']} - motivo: {motivo}")
    return {"ok": True}


@api.get("/contabilita/giornata-stato/{data}")
async def giornata_stato(
    data: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Stato Prima Nota per una data specifica. Usato dai banner UI.

    Returns: { data, chiusa: bool, chiusura_id?, closed_by_name?, can_riapri? }
    """
    from shared import giornata_chiusa
    ch = await giornata_chiusa(data)
    if not ch:
        return {"data": data, "chiusa": False}
    full = await db.chiusure_giorno.find_one({"id": ch["id"]}, {"_id": 0})
    return {
        "data": data,
        "chiusa": True,
        "chiusura_id": full.get("id") if full else ch["id"],
        "closed_by_name": (full or {}).get("closed_by_name"),
        "closed_at": (full or {}).get("created_at"),
        "can_riapri": user.get("role") == "admin",
    }


@api.get("/contabilita/giornate-chiuse")
async def giornate_chiuse(
    dal: Optional[str] = None,
    al: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
) -> list[str]:
    """Lista date (YYYY-MM-DD) con Prima Nota chiusa (no riaperture)."""
    flt: dict = {"riaperta_at": None}
    if dal or al:
        flt["data"] = {}
        if dal:
            flt["data"]["$gte"] = dal
        if al:
            flt["data"]["$lte"] = al
    out = []
    async for ch in db.chiusure_giorno.find(flt, {"_id": 0, "data": 1}).sort("data", -1):
        out.append(ch["data"])
    return out


@api.get("/contabilita/chiusure-giorno")
async def list_chiusure_giorno(
    limit: int = 500,
    anno: Optional[int] = None,
    q: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Storico chiusure giornaliere (Prima Nota chiuse).

    Parametri:
      - anno: filtra per anno (es. 2026). Quando omesso ritorna tutto.
      - q: cerca per data o ID
    """
    flt: dict = {}
    if anno:
        flt["data"] = {"$gte": f"{anno}-01-01", "$lte": f"{anno}-12-31"}
    if q:
        flt["$or"] = [{"data": {"$regex": q, "$options": "i"}}, {"id": {"$regex": q, "$options": "i"}}]
    items = await db.chiusure_giorno.find(flt, {"_id": 0}).sort("data", -1).to_list(limit)
    # arricchisci con n_movimenti e progressivo
    for i, it in enumerate(items):
        it["progressivo"] = i + 1
        if not it.get("riepilogo"):
            it["riepilogo"] = {}
    return items


@api.delete("/contabilita/chiusura-giorno/{chiusura_id}")
async def delete_chiusura_giorno(
    chiusura_id: str, user=Depends(require_user("admin")),
):
    """Elimina definitivamente una chiusura giornaliera (admin only).
    NOTA: i movimenti sottostanti restano nel DB ma vengono rimarcati come riaperti.
    """
    ch = await db.chiusure_giorno.find_one({"id": chiusura_id}, {"_id": 0})
    if not ch:
        raise HTTPException(404, "Chiusura non trovata")
    # rimuovi PDF dallo storage (best-effort)
    try:
        if ch.get("pdf_storage_path"):
            obj_storage.delete_object(ch["pdf_storage_path"])
    except Exception:
        pass
    # riapri i movimenti
    await db.movimenti.update_many(
        {"chiusura_id": chiusura_id},
        {"$set": {"chiusura_id": None, "updated_at": _now_iso()}},
    )
    await db.chiusure_giorno.delete_one({"id": chiusura_id})
    await log_attivita(user, "delete", "chiusura_giorno", chiusura_id,
                       f"Eliminata chiusura del {ch['data']}")
    return {"ok": True}


@api.get("/contabilita/chiusura-giorno/{chiusura_id}/pdf")
async def download_chiusura_pdf(
    chiusura_id: str, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    ch = await db.chiusure_giorno.find_one({"id": chiusura_id}, {"_id": 0})
    if not ch or not ch.get("pdf_storage_path"):
        raise HTTPException(404, "PDF non trovato")
    try:
        data, _ = obj_storage.get_object(ch["pdf_storage_path"])
    except Exception as e:
        raise HTTPException(503, f"Errore: {e}")
    return StreamingResponse(
        _io.BytesIO(data), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="brogliaccio_{ch["data"]}.pdf"'},
    )


# ============================================================
# IMPORTAZIONE ANIA
# ============================================================
@api.post("/import/ania")
async def import_ania(file: UploadFile = File(...),
                      user=Depends(require_user("admin", "collaboratore"))):
    contents = await file.read()
    log = await ania_importer.importa_zip(db, contents, file.filename, user)
    await log_attivita(user, "import", "ania", log.id,
                       f"Import file {file.filename}", payload=log.record_types_processati)
    # Alert: import ANIA completato -> notifica collaboratori sinistri se ha creato sinistri
    if log.sinistri_creati > 0:
        from alert_dispatcher import safe_dispatch
        await safe_dispatch("sinistro.importato_ania", {
            "entita_tipo": "import", "entita_id": log.id,
            "nome_file": file.filename,
            "n_sinistri": log.sinistri_creati,
            "n_polizze": log.polizze_create + log.polizze_aggiornate,
            "link": "/importazioni",
        })
    return log.model_dump()


# Alias OMNIA (stesso endpoint, naming nuovo)
@api.post("/import/omnia")
async def import_omnia(file: UploadFile = File(...),
                       user=Depends(require_user("admin", "collaboratore"))):
    return await import_ania(file, user)


# Mapping per flussi (compagnia/ramo/collaboratore/prodotto)
@api.get("/import/mappings")
async def list_mappings(
    tipo: Optional[str] = None,
    flusso: str = "omnia",
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    flt: dict = {"flusso": flusso}
    if tipo:
        flt["tipo"] = tipo
    return await db.import_mappings.find(flt, {"_id": 0}).to_list(2000)


@api.get("/import/unmapped")
async def list_unmapped_entities(
    flusso: str = "omnia",
    include_mapped: bool = False,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Ritorna le entità del flusso ancora da mappare a un'entità DB.

    Se `include_mapped=true`, include anche quelle già associate (per permettere
    la modifica/eliminazione del mapping dal wizard).
    """
    by_tipo: dict[str, list[dict]] = {
        "compagnia": [], "ramo": [], "prodotto": [], "collaboratore": [], "garanzia": [],
    }
    flt: dict = {
        "flusso": flusso,
        "tipo": {"$in": ["compagnia", "ramo", "prodotto", "collaboratore", "garanzia"]},
    }
    if not include_mapped:
        flt["$or"] = [{"entita_id": None}, {"entita_id": ""}]
    async for m in db.import_mappings.find(flt, {"_id": 0}):
        tipo = m.get("tipo")
        if tipo in by_tipo:
            by_tipo[tipo].append({
                "id": m.get("id"),
                "valore_flusso": m.get("valore_flusso"),
                "label_flusso": m.get("label_flusso") or m.get("valore_flusso"),
                "occorrenze": m.get("occorrenze", 0),
                "entita_id": m.get("entita_id"),
                "label_programma": m.get("label_programma"),
            })
    # Candidati DB (librerie esistenti)
    candidates: dict[str, list[dict]] = {}
    candidates["compagnia"] = [
        {"id": c["id"], "label": c.get("ragione_sociale") or c.get("codice")}
        async for c in db.compagnie.find({}, {"_id": 0, "id": 1, "ragione_sociale": 1, "codice": 1}).sort("ragione_sociale", 1)
    ]
    candidates["ramo"] = [
        {"id": r["id"], "label": r.get("nome") or r.get("codice")}
        async for r in db.rami.find({"attivo": {"$ne": False}}, {"_id": 0, "id": 1, "nome": 1, "codice": 1}).sort("nome", 1)
    ]
    candidates["collaboratore"] = [
        {"id": u["id"], "label": u.get("name") or u.get("email")}
        async for u in db.users.find(
            {"role": {"$in": ["admin", "collaboratore", "dipendente"]}, "attivo": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "email": 1},
        ).sort("name", 1)
    ]
    candidates["prodotto"] = [
        {"id": p["id"], "label": p.get("nome")}
        async for p in db.prodotti.find({}, {"_id": 0, "id": 1, "nome": 1}).sort("nome", 1)
    ]
    # Garanzie: usa la libreria mapping_garanzie + lista distinct dalle polizze
    gar_labels: set[str] = set()
    async for g in db.mapping_garanzie.find({"nome_personalizzato": {"$ne": None}}, {"_id": 0, "nome_personalizzato": 1}):
        if g.get("nome_personalizzato"):
            gar_labels.add(g["nome_personalizzato"])
    candidates["garanzia"] = [{"id": n, "label": n} for n in sorted(gar_labels)]
    return {**by_tipo, "candidates": candidates}


@api.post("/import/mappings", status_code=201)
async def save_mapping(body: dict, user=Depends(require_user("admin", "collaboratore"))):
    """Crea o aggiorna una mappatura. Chiave univoca: (tipo, flusso, valore_flusso)."""
    tipo = body.get("tipo")
    flusso = body.get("flusso", "omnia")
    valore_flusso = body.get("valore_flusso")
    if not (tipo and valore_flusso):
        raise HTTPException(400, "tipo e valore_flusso richiesti")
    upd = {
        "tipo": tipo, "flusso": flusso, "valore_flusso": valore_flusso,
        "entita_id": body.get("entita_id"),
        "label_programma": body.get("label_programma"),
        "note": body.get("note"),
        "updated_at": _now_iso(),
    }
    await db.import_mappings.update_one(
        {"tipo": tipo, "flusso": flusso, "valore_flusso": valore_flusso},
        {"$set": upd, "$setOnInsert": {"created_at": _now_iso()}},
        upsert=True,
    )
    return upd


@api.delete("/import/mappings/{mid}")
async def delete_mapping(mid: str, user=Depends(require_user("admin"))):
    await db.import_mappings.delete_one({"id": mid})
    return {"ok": True}


@api.post("/import/mappings/apply")
async def apply_import_mappings(
    flusso: str = "omnia",
    user=Depends(require_user("admin", "collaboratore")),
):
    """Riapplica i mapping salvati ai record già importati.

    Esegue back-fill per:
      - compagnia -> aggiorna `polizze.compagnia_id` e `sinistri.compagnia_id` quando matcha
        un `valore_flusso` (es. codice compagnia exp) — usa un campo helper non disponibile,
        quindi il back-fill compagnia è limitato a polizze con compagnia_id vuoto e
        contraente con compagnia_exp == valore_flusso (best-effort).
      - collaboratore -> aggiorna `polizze.collaboratore_id` se `operatore_ania_codice` matcha
      - ramo -> aggiorna `polizze.ramo` su match esatto del ramo originale
      - garanzia -> rinomina garanzie nelle polizze (chiave su `codice_ania` o nome originale)
      - prodotto -> aggiorna `polizze.prodotto` su match esatto del valore originale
    """
    summary = {
        "polizze_compagnia": 0,
        "polizze_collaboratore": 0,
        "polizze_prodotto": 0,
        "polizze_garanzia": 0,
        "sinistri_compagnia": 0,
    }
    mappings = await db.import_mappings.find(
        {"flusso": flusso, "entita_id": {"$ne": None}},
        {"_id": 0},
    ).to_list(5000)

    gar_map: dict[str, str] = {}
    for m in mappings:
        tipo = m.get("tipo")
        valore = (m.get("valore_flusso") or "").strip()
        entita_id = m.get("entita_id")
        if not (tipo and valore and entita_id):
            continue

        if tipo == "collaboratore":
            r = await db.polizze.update_many(
                {"operatore_ania_codice": valore},
                {"$set": {"collaboratore_id": entita_id, "updated_at": _now_iso()}},
            )
            summary["polizze_collaboratore"] += r.modified_count

        elif tipo == "compagnia":
            # Back-fill: aggiorna le polizze con compagnia_codice_exp == valore_flusso
            r1 = await db.polizze.update_many(
                {"compagnia_codice_exp": valore},
                {"$set": {"compagnia_id": entita_id, "updated_at": _now_iso()}},
            )
            # Inoltre, aggiorna le polizze "spurie" che puntano a una compagnia
            # auto-creata con stesso codice (compatibilità con import legacy)
            spuria = await db.compagnie.find_one(
                {"codice": valore, "id": {"$ne": entita_id}},
                {"_id": 0, "id": 1},
            )
            if spuria:
                r2 = await db.polizze.update_many(
                    {"compagnia_id": spuria["id"]},
                    {"$set": {"compagnia_id": entita_id, "compagnia_codice_exp": valore, "updated_at": _now_iso()}},
                )
                # Aggiorna anche eventuali sinistri
                rs = await db.sinistri.update_many(
                    {"compagnia_id": spuria["id"]},
                    {"$set": {"compagnia_id": entita_id, "updated_at": _now_iso()}},
                )
                summary["polizze_compagnia"] += r2.modified_count
                summary["sinistri_compagnia"] += rs.modified_count
            summary["polizze_compagnia"] += r1.modified_count

        elif tipo == "prodotto":
            # 🔧 FIX: usa il nome del prodotto (non l'UUID) come valore in polizze.prodotto
            prod = await db.prodotti.find_one({"id": entita_id}, {"_id": 0, "nome": 1})
            nome_prod = (prod or {}).get("nome") or entita_id
            r = await db.polizze.update_many(
                {"$or": [{"prodotto": valore}, {"prodotto_originale": valore}]},
                {"$set": {"prodotto": nome_prod, "updated_at": _now_iso()}},
            )
            summary["polizze_prodotto"] += r.modified_count

        elif tipo == "ramo":
            # Back-fill: aggiorna polizze con ramo_originale == valore_flusso (codice importato es. "3D")
            # entita_id qui è il CODICE del RamoLibreria (es. "INFORTUNI"), non un UUID
            ramo_doc = await db.rami.find_one({"id": entita_id}, {"_id": 0, "codice": 1, "nome": 1})
            codice_ramo = (ramo_doc or {}).get("codice") or entita_id
            r = await db.polizze.update_many(
                {"$or": [{"ramo": valore}, {"ramo_originale": valore}]},
                {"$set": {"ramo": codice_ramo, "updated_at": _now_iso()}},
            )
            summary["polizze_ramo"] = summary.get("polizze_ramo", 0) + r.modified_count

        elif tipo == "garanzia":
            gar_map[valore.upper()] = entita_id

    if gar_map:
        async for p in db.polizze.find(
            {"garanzie": {"$exists": True, "$ne": []}},
            {"_id": 0, "id": 1, "garanzie": 1},
        ):
            changed = False
            for g in p.get("garanzie") or []:
                code = (g.get("codice_ania") or g.get("garanzia_originale") or "").strip().upper()
                if code in gar_map and g.get("garanzia") != gar_map[code]:
                    g["garanzia"] = gar_map[code]
                    changed = True
            if changed:
                await db.polizze.update_one(
                    {"id": p["id"]},
                    {"$set": {"garanzie": p["garanzie"], "updated_at": _now_iso()}},
                )
                summary["polizze_garanzia"] += 1

    await log_attivita(user, "apply", "import_mappings", None,
                       f"Applicate {len(mappings)} mappature", payload=summary)
    return summary


# Report dettagliato di un singolo import
@api.get("/import/log/{lid}")
async def get_import_log(lid: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    log = await db.import_logs.find_one({"id": lid}, {"_id": 0})
    if not log:
        raise HTTPException(404, "Log non trovato")
    return log


@api.get("/import/storico")
async def import_storico(limit: int = 50, user=Depends(require_user("admin", "collaboratore"))):
    items = await db.import_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return items


# ============================================================
# LIBRO MATRICOLA / STATO DI RISCHIO (Targhe)
# ============================================================
from libro_matricola import build_preview as _lm_preview, commit_import as _lm_commit  # type: ignore


@api.post("/import/libro-matricola/preview")
async def libro_matricola_preview(
    file: UploadFile = File(...),
    user=Depends(require_user("admin", "collaboratore")),
):
    """Upload XLSX/CSV libro matricola -> ritorna header, preview rows, suggested mapping."""
    content = await file.read()
    try:
        return _lm_preview(content, file.filename or "libro_matricola.xlsx")
    except Exception as e:
        raise HTTPException(400, f"Errore lettura file: {e}")


@api.post("/import/libro-matricola/commit")
async def libro_matricola_commit(
    file: UploadFile = File(...),
    mapping: str = Form(...),     # JSON string {header: field}
    polizza_id: Optional[str] = Form(None),
    user=Depends(require_user("admin", "collaboratore")),
):
    """Esegue l'import con il mapping confermato dall'utente."""
    import json as _json
    try:
        mapping_dict = _json.loads(mapping)
    except Exception:
        raise HTTPException(400, "mapping non valido (atteso JSON)")
    content = await file.read()
    try:
        stats = await _lm_commit(db, content, file.filename or "libro_matricola.xlsx",
                                 mapping_dict, user, polizza_id=polizza_id)
    except Exception as e:
        raise HTTPException(400, f"Errore import: {e}")
    return stats


# ---- Veicoli: lista + lookup per targa (per riuso tra polizze) ----
@api.get("/veicoli")
async def list_veicoli(
    targa: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 100,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    flt: dict = {}
    if targa:
        flt["targa"] = targa.upper().strip()
    if q:
        regex = {"$regex": q.upper().strip(), "$options": "i"}
        flt["$or"] = [
            {"targa": regex}, {"marca": regex}, {"modello": regex},
            {"proprietario": regex}, {"telaio": regex},
        ]
    items = await db.veicoli.find(flt, {"_id": 0}).sort("targa", 1).to_list(limit)
    return items


@api.get("/veicoli/by-targa/{targa}")
async def get_veicolo_by_targa(
    targa: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Lookup rapido per pre-compilare i dati veicolo su una nuova polizza."""
    v = await db.veicoli.find_one({"targa": targa.upper().strip()}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Veicolo non trovato")
    # Polizze già associate a questa targa
    pol = await db.polizze.find(
        {"$or": [{"veicoli_ids": v["id"]}, {"targa": v["targa"]}]},
        {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "stato": 1,
         "compagnia_id": 1, "effetto": 1, "scadenza_originale": 1},
    ).to_list(50)
    return {"veicolo": v, "polizze_collegate": pol}


@api.delete("/veicoli/{vid}")
async def delete_veicolo(vid: str, user=Depends(require_user("admin"))):
    await db.veicoli.delete_one({"id": vid})
    await db.polizze.update_many({}, {"$pull": {"veicoli_ids": vid}})
    return {"ok": True}


@api.get("/polizze/{pid}/libro-matricola/export")
async def export_libro_matricola(
    pid: str,
    includi_storico: bool = False,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Esporta tutte le applicazioni libro matricola di una polizza in formato Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io
    pol = await db.polizze.find_one({"id": pid}, {"_id": 0})
    if not pol:
        raise HTTPException(404, "Polizza non trovata")
    flt: dict = {"polizza_id": pid}
    if not includi_storico:
        flt["stato"] = {"$in": ["attiva", "sospesa"]}
    apps = await db.applicazioni.find(flt, {"_id": 0}).sort("numero", 1).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Matricola"
    headers = [
        "N°", "Targa", "Stato", "Inclusione", "Esclusione",
        "Marca", "Modello", "Tipo veicolo", "Alimentazione", "Uso",
        "Data immatr.", "CV fiscali", "KW", "Cilindrata", "Quintali", "Posti",
        "Classe BM", "BM provenienza", "Franchigia",
        "Valore veicolo", "Valore accessori",
        "Intestatario", "Provincia", "Massimali",
        "Leasing", "Scad. leasing", "Note",
    ]
    ws.append(headers)
    # Stile header
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E40AF")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for a in apps:
        ws.append([
            a.get("numero"), a.get("targa"), a.get("stato"),
            a.get("data_inclusione"), a.get("data_esclusione"),
            a.get("marca"), a.get("modello"), a.get("tipo_veicolo"),
            a.get("tipo_alimentazione"), a.get("tipo_uso"),
            a.get("data_immatricolazione"), a.get("cv_fiscali"), a.get("kw"),
            a.get("cilindrata"), a.get("quintali"), a.get("posti"),
            a.get("bm_assegnata"), a.get("bm_provenienza"), a.get("franchigia"),
            a.get("valore_veicolo"), a.get("valore_accessori"),
            a.get("intestatario"), a.get("provincia_intestatario"), a.get("massimali"),
            a.get("leasing"), a.get("scadenza_leasing"), a.get("note"),
        ])
    # Auto-width approssimativo
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(h) + 2)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_n = re.sub(r"[^A-Za-z0-9_-]", "_", str(pol.get("numero_polizza") or pid))
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="LibroMatricola_{safe_n}.xlsx"'},
    )


# ============================================================
# PENSIONI INPS
# ============================================================
class CalcoloRequest(BaseModel):
    tipo_pensione: Literal["invalidita", "inabilita", "superstite"]
    settimane_contributive: int = 0
    retribuzione_media_annua: float = 0.0
    eta: int = 0
    percentuale_invalidita: Optional[float] = None
    numero_familiari: int = 0
    anagrafica_id: Optional[str] = None


@api.post("/pensioni/calcola")
async def calcola_pensione(body: CalcoloRequest, user=Depends(current_user)):
    risultato = inps_calculator.calcola_pensione(
        tipo=body.tipo_pensione,
        settimane_contributive=body.settimane_contributive,
        retribuzione_media_annua=body.retribuzione_media_annua,
        eta=body.eta,
        percentuale_invalidita=body.percentuale_invalidita,
        numero_familiari=body.numero_familiari,
    )
    calc = CalcoloPensione(
        anagrafica_id=body.anagrafica_id,
        tipo_pensione=body.tipo_pensione,
        data_inizio_contribuzione="",
        settimane_contributive=body.settimane_contributive,
        retribuzione_media_annua=body.retribuzione_media_annua,
        eta_richiedente=body.eta,
        percentuale_invalidita=body.percentuale_invalidita,
        numero_familiari=body.numero_familiari,
        **risultato,
    )
    await db.calcoli_pensione.insert_one(calc.model_dump())
    await log_attivita(user, "calc_pensione", "pensione", calc.id, body.tipo_pensione)
    return calc.model_dump()


@api.post("/pensioni/parse-estratto")
async def parse_estratto(file: UploadFile = File(...), user=Depends(current_user)):
    raw = await file.read()
    fname = (file.filename or "").lower()
    text = ""

    if fname.endswith(".pdf"):
        try:
            import pdfplumber
            import io as _io
            with pdfplumber.open(_io.BytesIO(raw)) as pdf:
                pages_text = []
                for page in pdf.pages:
                    t = page.extract_text() or ""
                    pages_text.append(t)
                text = "\n".join(pages_text)
        except Exception as e:
            logger.error("PDF parse error: %s", e)
            raise HTTPException(400, f"Impossibile leggere il PDF: {e}")
        if not text.strip():
            return {
                "settimane_contributive": 0,
                "retribuzione_media_annua": 0.0,
                "anni_stimati": 0,
                "warning": "Il PDF non contiene testo estraibile (potrebbe essere scansionato). Inserisci manualmente i dati.",
            }
    else:
        try:
            text = raw.decode("utf-8", errors="ignore")
        except Exception:
            text = raw.decode("latin-1", errors="ignore")

    result = inps_calculator.parse_estratto_conto_inps(text)
    # se non ho trovato nulla nel PDF avviso
    if fname.endswith(".pdf") and not result.get("settimane_contributive") and not result.get("retribuzione_media_annua"):
        result["warning"] = (
            "PDF letto correttamente ma non sono riuscito a riconoscere settimane/retribuzione. "
            "Inseriscili manualmente o controlla che l'estratto contenga voci come "
            "\"Totale settimane\" e \"Retribuzione imponibile\"."
        )
    return result


@api.get("/pensioni/storico")
async def storico_pensioni(anagrafica_id: Optional[str] = None, user=Depends(current_user)):
    flt: dict = {}
    if anagrafica_id:
        flt["anagrafica_id"] = anagrafica_id
    items = await db.calcoli_pensione.find(flt, {"_id": 0}).sort("created_at", -1).to_list(200)
    return items


@api.get("/anagrafiche/{aid}/calcolo-pensione/preview")
async def calcolo_pensione_preview(aid: str, user=Depends(current_user)):
    """Restituisce dati pre-compilati per il calcolo pensione partendo dall'anagrafica.

    Calcola:
    - eta corrente
    - numero_familiari aventi diritto (coniuge se sposato + figli a carico)
    - settimane stimate (da anni contributivi se presenti)
    - flag requisiti_superstite_ok (se ha coniuge o figli a carico)
    """
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")

    from datetime import date
    eta = 0
    if ana.get("data_nascita"):
        try:
            d = date.fromisoformat(ana["data_nascita"])
            today = date.today()
            eta = today.year - d.year - (1 if (today.month, today.day) < (d.month, d.day) else 0)
        except Exception:
            pass

    # familiari aventi diritto per pensione superstite
    coniugato = (ana.get("stato_civile") or "").lower() in ("coniugato", "coniugata", "sposato", "sposata", "unito civilmente")
    figli_carico = int(ana.get("numero_figli_a_carico") or 0)
    n_familiari = (1 if coniugato else 0) + figli_carico
    requisiti_superstite_ok = n_familiari > 0

    # settimane stimate
    settimane = ana.get("settimane_contributive")
    if not settimane and ana.get("data_inizio_contribuzione"):
        try:
            di = date.fromisoformat(ana["data_inizio_contribuzione"])
            anni = (date.today() - di).days / 365.25
            settimane = int(anni * 52)
        except Exception:
            settimane = 0

    return {
        "anagrafica_id": aid,
        "nome": ana["ragione_sociale"],
        "eta": eta,
        "data_nascita": ana.get("data_nascita"),
        "stato_civile": ana.get("stato_civile"),
        "coniugato": coniugato,
        "tipo_lavoratore": ana.get("tipo_lavoratore"),
        "professione": ana.get("professione"),
        "reddito_annuo_lordo": ana.get("reddito_annuo_lordo") or 0.0,
        "numero_figli": ana.get("numero_figli") or 0,
        "numero_figli_a_carico": figli_carico,
        "numero_familiari": n_familiari,
        "requisiti_superstite_ok": requisiti_superstite_ok,
        "settimane_contributive": settimane or 0,
        "warnings": [
            "Nessun coniuge né figli a carico: la pensione superstite non spetta." if not requisiti_superstite_ok else None,
            "Reddito annuo non valorizzato: il calcolo del GAP non è significativo." if not ana.get("reddito_annuo_lordo") else None,
        ],
    }


@api.post("/anagrafiche/{aid}/calcolo-pensione/calcola")
async def calcolo_pensione_anagrafica(aid: str, body: dict, user=Depends(current_user)):
    """Esegue il calcolo per i 3 tipi di pensione + GAP di reddito.

    body può sovrascrivere i parametri ricavati dall'anagrafica.
    """
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    preview = await calcolo_pensione_preview(aid, user)

    settimane = int(body.get("settimane_contributive") or preview["settimane_contributive"] or 0)
    retribuzione = float(body.get("retribuzione_media_annua") or preview.get("reddito_annuo_lordo") or 0)
    eta = int(body.get("eta") or preview["eta"] or 0)
    invalidita = body.get("percentuale_invalidita")
    familiari = int(body.get("numero_familiari") if body.get("numero_familiari") is not None else preview["numero_familiari"])

    risultati = {}
    for tipo in ("invalidita", "inabilita", "superstite"):
        # superstite: se non ha familiari, restituisci 0 con avviso
        if tipo == "superstite" and familiari == 0:
            risultati[tipo] = {
                "pensione_lorda_mensile": 0,
                "pensione_lorda_annua": 0,
                "pensione_netta_stimata": 0,
                "metodologia": "Non spettante (assenza di coniuge/figli a carico)",
                "coefficiente_applicato": 0,
                "dettaglio": {"note": ["Requisiti soggettivi non soddisfatti."]},
            }
            continue
        risultati[tipo] = inps_calculator.calcola_pensione(
            tipo=tipo,
            settimane_contributive=settimane,
            retribuzione_media_annua=retribuzione,
            eta=eta,
            percentuale_invalidita=invalidita,
            numero_familiari=familiari,
        )

    # GAP di reddito: differenza tra reddito attuale e pensione lorda annua per ogni tipo
    gap = {}
    for tipo, r in risultati.items():
        diff = retribuzione - (r.get("pensione_lorda_annua") or 0)
        gap[tipo] = {
            "gap_annuo": round(diff, 2),
            "gap_mensile": round(diff / 12.0, 2),
            "copertura_percentuale": round(((r.get("pensione_lorda_annua") or 0) / retribuzione * 100), 1) if retribuzione else 0,
        }

    # salva storico (un calcolo per il principale - invalidità)
    main = risultati["invalidita"]
    calc = CalcoloPensione(
        anagrafica_id=aid,
        tipo_pensione="invalidita",
        data_inizio_contribuzione="",
        settimane_contributive=settimane,
        retribuzione_media_annua=retribuzione,
        eta_richiedente=eta,
        percentuale_invalidita=invalidita,
        numero_familiari=familiari,
        **main,
    )
    await db.calcoli_pensione.insert_one(calc.model_dump())
    await log_attivita(user, "calc_pensione", "pensione", aid)

    return {
        "anagrafica": preview,
        "parametri_usati": {
            "settimane": settimane, "retribuzione_annua": retribuzione,
            "eta": eta, "familiari": familiari, "invalidita": invalidita,
        },
        "risultati": risultati,
        "gap_reddito": gap,
    }



# ============================================================
# ANALISI CLIENTE COMPLETA (Diagnosi - SatorCRM-like)
# ============================================================
async def _ensure_analisi(aid: str) -> dict:
    """Carica o crea l'analisi cliente per un'anagrafica."""
    doc = await db.analisi_cliente.find_one({"anagrafica_id": aid}, {"_id": 0})
    if not doc:
        ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
        if not ana:
            raise HTTPException(404, "Anagrafica non trovata")
        # Pre-popola dai campi anagrafica
        ac = AnalisiCliente(
            anagrafica_id=aid,
            reddito_lordo_annuo=float(ana.get("reddito_annuo_lordo") or 0),
        )
        await db.analisi_cliente.insert_one(ac.model_dump())
        doc = ac.model_dump()
    return doc


@api.get("/anagrafiche/{aid}/analisi")
async def get_analisi_cliente(aid: str, user=Depends(current_user)):
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    return await _ensure_analisi(aid)


@api.put("/anagrafiche/{aid}/analisi")
async def update_analisi_cliente(
    aid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Aggiorna i campi dell'analisi cliente (qualunque sezione)."""
    await _ensure_analisi(aid)
    # Filtra solo i campi del modello
    allowed = set(AnalisiCliente.model_fields.keys()) - {"id", "anagrafica_id", "created_at", "updated_at"}
    update = {k: v for k, v in body.items() if k in allowed}
    update["updated_at"] = _now_iso()
    await db.analisi_cliente.update_one({"anagrafica_id": aid}, {"$set": update})
    await log_attivita(user, "update_analisi", "analisi_cliente", aid)
    doc = await db.analisi_cliente.find_one({"anagrafica_id": aid}, {"_id": 0})
    return doc


@api.post("/anagrafiche/{aid}/analisi/calcola-redditi")
async def calcola_redditi_analisi(aid: str, user=Depends(current_user)):
    """Calcola approfondimento redditi (contributi, IRPEF, netto) usando i dati
    salvati nell'analisi e nell'anagrafica."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ac = await _ensure_analisi(aid)
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0}) or {}

    # Determina coniuge a carico
    parente = ana.get("parente_di") or []
    coniuge_present = any((p.get("relazione") or "").lower() in ("coniuge", "sposo", "sposa") for p in parente)
    has_coniuge_a_carico = bool(coniuge_present)

    tipo_lav = ac.get("tipo_lavoratore") or ana.get("tipologia_lavoratore") or ana.get("tipo_lavoratore") or "altro"
    risultato = reddito_calc.calcola_redditi(
        reddito_lordo=float(ac.get("reddito_lordo_annuo") or 0),
        tipo_lavoratore=tipo_lav,
        altri_redditi=float(ac.get("altri_redditi_annuali") or 0) +
                      float(ac.get("dividendi_partecipazioni") or 0) +
                      float(ac.get("reddito_da_affitti") or 0),
        oneri_deducibili=float(ac.get("oneri_deducibili") or 0),
        oneri_fondo_pensione=float(ac.get("oneri_fondo_pensione") or 0),
        altre_detrazioni=float(ac.get("altre_detrazioni") or 0),
        ha_coniuge_a_carico=has_coniuge_a_carico,
        numero_figli_a_carico=int(ana.get("numero_figli_a_carico") or 0),
        regime_forfettario=bool(ac.get("regime_forfettario")),
    )
    return risultato


@api.post("/anagrafiche/{aid}/analisi/calcola-successione")
async def calcola_successione_analisi(aid: str, user=Depends(current_user)):
    """Calcola quote di successione (intestata e quote di legittima) basandosi
    sull'albero genealogico (parente_di) dell'anagrafica."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    ac = await _ensure_analisi(aid)

    parente = ana.get("parente_di") or []
    coniuge = any((p.get("relazione") or "").lower() in ("coniuge", "sposo", "sposa") for p in parente)
    n_figli = sum(1 for p in parente if (p.get("relazione") or "").lower() in ("figlio", "figlia"))
    n_genitori = sum(1 for p in parente if (p.get("relazione") or "").lower() in ("padre", "madre", "genitore"))
    n_fratelli = sum(1 for p in parente if (p.get("relazione") or "").lower() in ("fratello", "sorella"))

    # Calcola patrimonio totale
    patrimonio = _calcola_patrimonio_totale(ac)

    risultato = successione_calc.calcola_successione(
        coniuge=coniuge,
        numero_figli=n_figli,
        genitori_vivi=n_genitori,
        fratelli=n_fratelli,
        patrimonio=patrimonio,
    )
    risultato["componenti_familiari"] = {
        "coniuge": coniuge,
        "figli": n_figli,
        "genitori": n_genitori,
        "fratelli": n_fratelli,
    }
    return risultato


def _calcola_patrimonio_totale(ac: dict) -> dict | float:
    """Helper: somma valori patrimoniali (per uso interno)."""
    immobili_tot = sum(float(i.get("valore_commerciale") or 0) * float(i.get("percentuale_proprieta") or 100) / 100
                       for i in (ac.get("immobili") or []))
    veicoli_tot = sum(float(v.get("valore_commerciale") or 0) for v in (ac.get("veicoli") or []))
    beni_tot = sum(float(b.get("valore") or 0) for b in (ac.get("beni") or []))
    aziende_tot = sum(float(a.get("valore_ipotetico") or 0) * float(a.get("percentuale_partecipazione") or 100) / 100
                      for a in (ac.get("aziende") or []))
    return round(immobili_tot + veicoli_tot + beni_tot + aziende_tot, 2)


@api.get("/anagrafiche/{aid}/analisi/patrimonio")
async def get_patrimonio_riepilogo(aid: str, user=Depends(current_user)):
    """Riepilogo patrimoniale: liquidità, immobili, veicoli, beni, aziende, debiti, montante."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ac = await _ensure_analisi(aid)
    immobili_tot = sum(float(i.get("valore_commerciale") or 0) * float(i.get("percentuale_proprieta") or 100) / 100
                       for i in (ac.get("immobili") or []))
    veicoli_tot = sum(float(v.get("valore_commerciale") or 0) for v in (ac.get("veicoli") or []))
    beni_tot = sum(float(b.get("valore") or 0) for b in (ac.get("beni") or []))
    aziende_tot = sum(float(a.get("valore_ipotetico") or 0) * float(a.get("percentuale_partecipazione") or 100) / 100
                      for a in (ac.get("aziende") or []))
    liquidita = float(ac.get("liquidita") or 0)
    tfr = float(ac.get("tfr_maturato") or 0)
    debiti = float(ac.get("debiti") or 0)
    # Montante contributivo dallo storico redditi
    storico = ac.get("storico_redditi") or []
    montante = round(sum(float(r.get("reddito") or 0) * 0.33 for r in storico), 2)
    return {
        "patrimonio_liquido": round(liquidita + tfr, 2),
        "patrimonio_immobiliare": round(immobili_tot, 2),
        "patrimonio_aziendale": round(aziende_tot, 2),
        "patrimonio_veicoli": round(veicoli_tot, 2),
        "altri_beni": round(beni_tot, 2),
        "montante_contributivo": montante,
        "debiti": debiti,
        "patrimonio_netto": round(liquidita + tfr + immobili_tot + aziende_tot + veicoli_tot + beni_tot - debiti, 2),
        "patrimonio_totale": round(liquidita + tfr + immobili_tot + aziende_tot + veicoli_tot + beni_tot, 2),
    }


@api.post("/anagrafiche/{aid}/analisi/calcola-pensioni-future")
async def calcola_pensioni_future(aid: str, user=Depends(current_user)):
    """Calcola la proiezione delle pensioni di vecchiaia da 64 a 71 anni
    (anticipate e vecchiaia), oltre a invalidità/inabilità/superstiti
    sulla base dello storico redditi e dei periodi contributivi."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ac = await _ensure_analisi(aid)
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0}) or {}

    from datetime import date
    eta_attuale = 0
    if ana.get("data_nascita"):
        try:
            d = date.fromisoformat(ana["data_nascita"])
            today = date.today()
            eta_attuale = today.year - d.year - (1 if (today.month, today.day) < (d.month, d.day) else 0)
        except Exception:
            pass

    storico = ac.get("storico_redditi") or []
    periodi = ac.get("periodi_contributivi") or []

    # Calcolo anni e settimane dai periodi contributivi
    settimane = 0
    if periodi:
        from datetime import date as _date
        for p in periodi:
            try:
                ini = _date.fromisoformat(p["inizio_periodo"])
                fin = _date.fromisoformat(p["fine_periodo"]) if p.get("fine_periodo") else _date.today()
                gg = (fin - ini).days
                settimane += gg // 7
            except Exception:
                pass
    settimane = settimane or int(ana.get("settimane_contributive") or 0)
    anni_contrib = settimane / 52.0

    # Retribuzione media (ultimi 5 anni se disponibili)
    redditi = sorted([(r.get("anno") or 0, float(r.get("reddito") or 0)) for r in storico], reverse=True)
    if redditi:
        ultimi5 = redditi[:5]
        retribuzione_media = sum(r[1] for r in ultimi5) / len(ultimi5)
    else:
        retribuzione_media = float(ac.get("reddito_lordo_annuo") or 0)

    # Pensioni di OGGI (invalidità/inabilità/superstiti)
    parente = ana.get("parente_di") or []
    coniuge = any((p.get("relazione") or "").lower() in ("coniuge", "sposo", "sposa") for p in parente)
    figli_carico = int(ana.get("numero_figli_a_carico") or 0)
    n_familiari = (1 if coniuge else 0) + figli_carico

    pensioni_oggi = {}
    for tipo in ("invalidita", "inabilita", "superstite"):
        if tipo == "superstite" and n_familiari == 0:
            pensioni_oggi[tipo] = {
                "pensione_lorda_mensile": 0, "pensione_lorda_annua": 0,
                "pensione_netta_stimata": 0,
                "metodologia": "Non spettante (no familiari)", "coefficiente_applicato": 0, "dettaglio": {},
            }
            continue
        pensioni_oggi[tipo] = inps_calculator.calcola_pensione(
            tipo=tipo, settimane_contributive=settimane,
            retribuzione_media_annua=retribuzione_media, eta=eta_attuale,
            percentuale_invalidita=75, numero_familiari=n_familiari,
        )

    # Pensioni DOMANI: proiezione da 64 a 71
    pensioni_domani = []
    montante_attuale = round(retribuzione_media * 0.33 * anni_contrib, 2)
    for eta_pens in range(64, 72):
        anni_extra = eta_pens - eta_attuale
        if anni_extra <= 0:
            continue
        anni_tot = anni_contrib + anni_extra
        montante = retribuzione_media * 0.33 * anni_tot
        coeff = inps_calculator._coefficiente_trasformazione(eta_pens)
        importo_annuo = montante * coeff
        modalita = "Anticipata" if eta_pens < 67 else "Vecchiaia"
        pensioni_domani.append({
            "eta_pensionamento": eta_pens,
            "anno_pensionamento": (ana.get("data_nascita", "1990-01-01")[:4] if ana.get("data_nascita") else "?"),
            "modalita": modalita,
            "anni_contribuzione_totali": round(anni_tot, 1),
            "settimane_totali": int(anni_tot * 52),
            "importo_annuo": round(importo_annuo, 2),
            "importo_mensile": round(importo_annuo / 13.0, 2),
            "montante_contributivo": round(montante, 2),
            "coefficiente_trasformazione": coeff,
        })

    # Calcolo data anno pensionamento corretto
    if ana.get("data_nascita"):
        try:
            anno_nascita = int(ana["data_nascita"][:4])
            for p in pensioni_domani:
                p["anno_pensionamento"] = anno_nascita + p["eta_pensionamento"]
        except Exception:
            pass

    # Totale versato / rivalutato (basato su redditi storici + 24% aliquota commerciante)
    aliquota_storica = reddito_calc.ALIQUOTE_CONTRIBUTIVE.get(
        (ac.get("tipo_lavoratore") or "commerciante").lower(), 0.24
    )
    totale_versato = round(sum(r[1] for r in redditi) * aliquota_storica, 2)
    totale_rivalutato = round(totale_versato * 1.068, 2)  # rivalutazione media ~6.8%

    return {
        "eta_attuale": eta_attuale,
        "anni_contribuzione": round(anni_contrib, 1),
        "settimane_contributive": settimane,
        "retribuzione_media_annua": round(retribuzione_media, 2),
        "totale_versato": totale_versato,
        "totale_rivalutato": totale_rivalutato,
        "montante_contributivo_attuale": montante_attuale,
        "pensioni_oggi": pensioni_oggi,
        "pensioni_domani": pensioni_domani,
        "numero_familiari": n_familiari,
        "ha_coniuge": coniuge,
    }


@api.post("/anagrafiche/{aid}/analisi/calcola-scoperture")
async def calcola_scoperture_analisi(aid: str, user=Depends(current_user)):
    """Calcola scoperture pensionistiche (problema oggi/futuro) e capitale da assicurare."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ac = await _ensure_analisi(aid)
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0}) or {}
    from datetime import date

    # Recupera pensioni future
    pens_data = await calcola_pensioni_future(aid, user)
    pensioni_oggi = pens_data["pensioni_oggi"]

    # Pensione di vecchiaia (a 67 anni o prima disponibile)
    pensione_vecchiaia_annua = 0.0
    for p in pens_data["pensioni_domani"]:
        if p["eta_pensionamento"] == 67:
            pensione_vecchiaia_annua = p["importo_annuo"]
            break
    if not pensione_vecchiaia_annua and pens_data["pensioni_domani"]:
        pensione_vecchiaia_annua = pens_data["pensioni_domani"][0]["importo_annuo"]

    eta_attuale = pens_data["eta_attuale"]

    # Età coniuge
    eta_coniuge = None
    parente = ana.get("parente_di") or []
    has_coniuge = False
    has_figli = False
    eta_figlio_piu_piccolo = None
    for p in parente:
        rel = (p.get("relazione") or "").lower()
        if rel in ("coniuge", "sposo", "sposa"):
            has_coniuge = True
            # Recupera età coniuge dall'anagrafica collegata
            con = await db.anagrafiche.find_one({"id": p.get("anagrafica_id")}, {"_id": 0, "data_nascita": 1})
            if con and con.get("data_nascita"):
                try:
                    d = date.fromisoformat(con["data_nascita"])
                    today = date.today()
                    eta_coniuge = today.year - d.year - (1 if (today.month, today.day) < (d.month, d.day) else 0)
                except Exception:
                    pass
        elif rel in ("figlio", "figlia"):
            has_figli = True
            figlio = await db.anagrafiche.find_one({"id": p.get("anagrafica_id")}, {"_id": 0, "data_nascita": 1})
            if figlio and figlio.get("data_nascita"):
                try:
                    d = date.fromisoformat(figlio["data_nascita"])
                    today = date.today()
                    eta_f = today.year - d.year - (1 if (today.month, today.day) < (d.month, d.day) else 0)
                    if eta_figlio_piu_piccolo is None or eta_f < eta_figlio_piu_piccolo:
                        eta_figlio_piu_piccolo = eta_f
                except Exception:
                    pass

    return reddito_calc.calcola_scoperture_pensionistiche(
        reddito_lordo=float(ac.get("reddito_lordo_annuo") or 0),
        altri_redditi=float(ac.get("altri_redditi_annuali") or 0) + float(ac.get("reddito_da_affitti") or 0),
        dividendi=float(ac.get("dividendi_partecipazioni") or 0),
        pensione_invalidita_annua=pensioni_oggi.get("invalidita", {}).get("pensione_lorda_annua", 0),
        pensione_inabilita_annua=pensioni_oggi.get("inabilita", {}).get("pensione_lorda_annua", 0),
        pensione_superstite_annua=pensioni_oggi.get("superstite", {}).get("pensione_lorda_annua", 0),
        pensione_vecchiaia_annua=pensione_vecchiaia_annua,
        eta_attuale=eta_attuale,
        eta_pensionamento=67,
        eta_max_target=70,
        eta_coniuge=eta_coniuge,
        eta_figlio_piu_piccolo=eta_figlio_piu_piccolo,
        debiti=float(ac.get("debiti") or 0),
        has_coniuge=has_coniuge,
        has_figli=has_figli,
        has_convivente=False,
    )


@api.get("/anagrafiche/{aid}/analisi/pdf-diagnosi-reddito")
async def pdf_diagnosi_reddito(aid: str, user=Depends(current_user)):
    """Genera il PDF 'Diagnosi del Reddito' (17 pagine, stile HubSicura)."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    ac = await _ensure_analisi(aid)
    pens = await calcola_pensioni_future(aid, user)
    scop = await calcola_scoperture_analisi(aid, user)
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}

    pdf_bytes = pdf_diagnosi.genera_diagnosi_reddito(ana, ac, pens, scop, az, user)
    return StreamingResponse(
        _io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=diagnosi_reddito_{ana['ragione_sociale'].replace(' ', '_')}.pdf"},
    )


@api.get("/anagrafiche/{aid}/analisi/pdf-progetto-azzob")
async def pdf_progetto_azzob(aid: str, user=Depends(current_user)):
    """Genera il PDF 'Progetto Futuro Senza Sorprese - metodo AZZOB / ISO 31000'."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    ac = await _ensure_analisi(aid)
    pens = await calcola_pensioni_future(aid, user)
    scop = await calcola_scoperture_analisi(aid, user)
    patr = await get_patrimonio_riepilogo(aid, user)
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}

    pdf_bytes = pdf_diagnosi.genera_progetto_azzob(ana, ac, pens, scop, patr, az, user)
    return StreamingResponse(
        _io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=progetto_azzob_{ana['ragione_sociale'].replace(' ', '_')}.pdf"},
    )


@api.post("/anagrafiche/{aid}/analisi/upload-estratto-inps")
async def upload_estratto_inps(
    aid: str,
    file: UploadFile = File(...),
    anno_riferimento: Optional[int] = Form(None),
    sostituisci_storico: bool = Form(False),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Carica un estratto contributivo INPS (PDF), lo parsa, lo salva nello storage
    e popola/aggiunge i dati nell'analisi cliente.

    - sostituisci_storico=False (default): MERGE periodi/redditi nuovi che mancano
    - sostituisci_storico=True: sostituisce integralmente storico_redditi e
      periodi_contributivi con quelli appena letti
    """
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    contents = await file.read()
    if len(contents) > 15 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 15 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "") or "application/pdf"

    text = ""
    if ct == "application/pdf" or (file.filename or "").lower().endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(_io.BytesIO(contents)) as pdf:
                text = "\n".join((p.extract_text() or "") for p in pdf.pages)
        except Exception as e:
            raise HTTPException(400, f"Impossibile leggere il PDF: {e}")
    else:
        try:
            text = contents.decode("utf-8", errors="ignore")
        except Exception:
            text = contents.decode("latin-1", errors="ignore")

    parsed = inps_calculator.parse_estratto_conto_inps(text)

    ext = (file.filename or "estratto.pdf").rsplit(".", 1)[-1].lower() or "pdf"
    storage_path = f"{os.environ.get('APP_NAME', 'assicura')}/anagrafiche/{aid}/estratti_inps/{_uid()}.{ext}"
    file_url = None
    try:
        res = obj_storage.put_object(storage_path, contents, ct)
        file_url = f"/api/storage/{res['path']}"
        storage_path = res["path"]
    except Exception as e:
        logger.error("Errore salvataggio estratto INPS: %s", e)

    if not anno_riferimento and parsed.get("storico_redditi"):
        anno_riferimento = max((r["anno"] for r in parsed["storico_redditi"]), default=None)

    ac = await _ensure_analisi(aid)
    update = {"updated_at": _now_iso()}

    estratto_entry = {
        "id": _uid(),
        "url": file_url,
        "storage_path": storage_path,
        "nome_file": file.filename,
        "mime": ct,
        "size_kb": round(len(contents) / 1024, 1),
        "data_caricamento": _now_iso(),
        "anno_riferimento": anno_riferimento,
        "totale_settimane": parsed.get("settimane_contributive", 0),
        "totale_versato": parsed.get("totale_versato", 0),
        "totale_retribuzioni": parsed.get("totale_retribuzioni", 0),
        "montante_stimato": parsed.get("montante_stimato", 0),
        "anni_stimati": parsed.get("anni_stimati", 0),
        "caricato_da": user.get("id"),
        "caricato_da_nome": user.get("name") or user.get("email"),
    }
    estratti = list(ac.get("estratti_conto_inps") or []) + [estratto_entry]
    update["estratti_conto_inps"] = estratti

    if sostituisci_storico:
        update["storico_redditi"] = parsed.get("storico_redditi", [])
        update["periodi_contributivi"] = parsed.get("periodi_contributivi", [])
    else:
        existing_storico = {(r.get("anno"), r.get("cassa", "")): r for r in (ac.get("storico_redditi") or [])}
        for nuovo in parsed.get("storico_redditi") or []:
            key = (nuovo["anno"], nuovo.get("cassa", ""))
            existing_storico[key] = nuovo
        update["storico_redditi"] = list(existing_storico.values())

        existing_periodi = {(p.get("inizio_periodo"), p.get("fondo")): p
                            for p in (ac.get("periodi_contributivi") or [])}
        for nuovo in parsed.get("periodi_contributivi") or []:
            key = (nuovo["inizio_periodo"], nuovo["fondo"])
            existing_periodi[key] = nuovo
        update["periodi_contributivi"] = list(existing_periodi.values())

    if parsed.get("reddito_annuo_lordo"):
        # Reddito annualizzato (ultimo anno × 12/mesi se parziale)
        update["reddito_lordo_annuo"] = parsed["reddito_annuo_lordo"]
    elif parsed.get("storico_redditi"):
        ultimo = sorted(parsed["storico_redditi"], key=lambda x: x["anno"], reverse=True)[0]
        if ultimo.get("reddito"):
            update["reddito_lordo_annuo"] = ultimo["reddito"]

    await db.analisi_cliente.update_one({"anagrafica_id": aid}, {"$set": update})

    ana_upd = {
        "settimane_contributive": parsed.get("settimane_contributive", 0),
        "data_inizio_contribuzione": parsed.get("data_inizio_contribuzione"),
        "retribuzione_media_annua": parsed.get("retribuzione_media_annua"),
        "updated_at": _now_iso(),
    }
    # Reddito annuo lordo dell'anagrafica = ultimo anno annualizzato a 12 mesi
    if parsed.get("reddito_annuo_lordo"):
        ana_upd["reddito_annuo_lordo"] = parsed["reddito_annuo_lordo"]
    elif parsed.get("storico_redditi"):
        ultimo = sorted(parsed["storico_redditi"], key=lambda x: x["anno"], reverse=True)[0]
        if ultimo.get("reddito"):
            ana_upd["reddito_annuo_lordo"] = ultimo["reddito"]
    ana_upd = {k: v for k, v in ana_upd.items() if v}
    if ana_upd:
        await db.anagrafiche.update_one({"id": aid}, {"$set": ana_upd})

    await log_attivita(user, "upload_estratto_inps", "anagrafica", aid,
                       descrizione=f"Estratto INPS: {parsed.get('settimane_contributive', 0)} sett., "
                                   f"{len(parsed.get('storico_redditi', []))} anni di redditi.")

    return {
        "estratto": estratto_entry,
        "parsed": {
            "settimane_contributive": parsed.get("settimane_contributive", 0),
            "anni_stimati": parsed.get("anni_stimati", 0),
            "righe_contributive": parsed.get("righe_contributive", 0),
            "totale_retribuzioni": parsed.get("totale_retribuzioni", 0),
            "totale_versato": parsed.get("totale_versato", 0),
            "montante_stimato": parsed.get("montante_stimato", 0),
            "storico_redditi": parsed.get("storico_redditi", [])[:50],
            "periodi_contributivi_count": len(parsed.get("periodi_contributivi", [])),
        },
        "warning": parsed.get("warning"),
    }


@api.post("/anagrafiche/{aid}/analisi/piramide/auto-popola")
async def auto_popola_piramide(aid: str, user=Depends(current_user)):
    """Pre-popola la piramide delle soluzioni con coperture suggerite calcolate
    dai dati già presenti (scoperture pensionistiche, successione, patrimonio)."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ac = await _ensure_analisi(aid)
    try:
        scop = await calcola_scoperture_analisi(aid, user)
    except Exception:
        scop = {}
    try:
        patr = await get_patrimonio_riepilogo(aid, user)
    except Exception:
        patr = {}

    suggerimenti = []
    if scop.get("invalidita"):
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Invalidita",
            "titolo": "Capitale Invalidità Permanente",
            "capitale_assicurato": scop["invalidita"].get("capitale_da_assicurare", 0),
            "premio_annuo": 0,
            "durata_anni": scop.get("anni_a_70", 30),
            "compagnia": "",
            "note": f"Scopertura mensile: {scop['invalidita'].get('scopertura_mensile', 0)} €",
            "ordine": 1,
        })
    if scop.get("inabilita"):
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Invalidita",
            "titolo": "Inabilità totale (NA / LTC)",
            "capitale_assicurato": scop["inabilita"].get("capitale_da_assicurare", 0),
            "premio_annuo": 0,
            "durata_anni": scop.get("anni_a_70", 30),
            "compagnia": "",
            "note": "Rendita vitalizia per non autosufficienza",
            "ordine": 2,
        })
    if scop.get("superstiti"):
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Premorienza",
            "titolo": "TCM Temporanea Caso Morte",
            "capitale_assicurato": scop["superstiti"].get("capitale_da_assicurare", 0),
            "premio_annuo": 0,
            "durata_anni": 30,
            "compagnia": "",
            "note": "Capitale per famiglia + debiti residui",
            "ordine": 3,
        })
    if scop.get("vecchiaia"):
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Pensione",
            "titolo": "Fondo Pensione (integrazione)",
            "capitale_assicurato": 0,
            "premio_annuo": int(float(ac.get("capacita_risparmio_annuale") or 0) * 0.5),
            "durata_anni": max(1, 67 - (await db.anagrafiche.find_one({"id": aid}, {"_id": 0, "data_nascita": 1}) or {}).get("_eta", 30) if False else 30),
            "compagnia": "",
            "note": f"Scopertura mensile vecchiaia: {scop['vecchiaia'].get('scopertura_mensile', 0)} €",
            "ordine": 4,
        })
    # Responsabilità (basato su patrimonio)
    if patr.get("patrimonio_totale", 0) > 50000:
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Responsabilita",
            "titolo": "RC Capofamiglia / Professionale",
            "capitale_assicurato": 1000000,
            "premio_annuo": 0,
            "durata_anni": 1,
            "compagnia": "",
            "note": "Massimale standard 1 milione",
            "ordine": 5,
        })
    # Perdita beni (basato su patrimonio immobiliare)
    if patr.get("patrimonio_immobiliare", 0) > 0:
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Beni",
            "titolo": "Polizza Casa (incendio/furto)",
            "capitale_assicurato": patr.get("patrimonio_immobiliare", 0),
            "premio_annuo": 0,
            "durata_anni": 1,
            "compagnia": "",
            "note": "Sul valore commerciale immobili",
            "ordine": 6,
        })
    if patr.get("patrimonio_aziendale", 0) > 0:
        suggerimenti.append({
            "id": _uid(),
            "categoria": "Beni",
            "titolo": "Polizza Azienda (Property + RC)",
            "capitale_assicurato": patr.get("patrimonio_aziendale", 0),
            "premio_annuo": 0,
            "durata_anni": 1,
            "compagnia": "",
            "note": "Tutela attività commerciale/aziendale",
            "ordine": 7,
        })

    await db.analisi_cliente.update_one(
        {"anagrafica_id": aid},
        {"$set": {"piramide_soluzioni": suggerimenti, "updated_at": _now_iso()}},
    )
    return {"piramide_soluzioni": suggerimenti, "count": len(suggerimenti)}


@api.post("/anagrafiche/{aid}/analisi/trattativa/auto-popola")
async def auto_popola_trattativa(aid: str, user=Depends(current_user)):
    """Pre-popola la trattativa A/B (scenario "Non fai nulla" vs "Ti affidi a me")
    usando i dati di scoperture, redditi e patrimonio già presenti."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ac = await _ensure_analisi(aid)
    try:
        scop = await calcola_scoperture_analisi(aid, user)
    except Exception:
        scop = {}
    try:
        red = await calcola_redditi_analisi(aid, user)
    except Exception:
        red = {}
    try:
        pens = await calcola_pensioni_future(aid, user)
    except Exception:
        pens = {}

    # Scenario A: "Non fai nulla" (scoperture attuali)
    scen_a = {
        "invalidita": scop.get("invalidita", {}).get("pensione_annua", 0),
        "importo_pensione": scop.get("vecchiaia", {}).get("pensione_annua", 0),
        "premorienza": scop.get("superstiti", {}).get("pensione_annua", 0),
        "responsabilita": 0,
        "perdita_beni": 0,
        "prima_data_pensionabile": next(
            (p["anno_pensionamento"] for p in pens.get("pensioni_domani", []) if p.get("modalita") == "Vecchiaia"),
            None,
        ),
        "versamento_fondo": float(ac.get("oneri_fondo_pensione") or 0),
        "risparmio_annuo": float(ac.get("capacita_risparmio_annuale") or 0),
        "vantaggio_fiscale": 0,
        "reddito": red.get("reddito_netto", float(ac.get("reddito_lordo_annuo") or 0)),
    }
    # Scenario B: "Ti affidi a me" (coperture suggerite)
    scen_b = {
        "invalidita": scop.get("invalidita", {}).get("capitale_da_assicurare", 0),
        "importo_pensione": (scop.get("vecchiaia", {}).get("pensione_annua", 0)
                             + float(ac.get("capacita_risparmio_annuale") or 0) * 0.5 * 20),
        "premorienza": scop.get("superstiti", {}).get("capitale_da_assicurare", 0),
        "responsabilita": 1000000,
        "perdita_beni": sum(float(i.get("valore_commerciale") or 0) for i in (ac.get("immobili") or [])),
        "prima_data_pensionabile": next(
            (p["anno_pensionamento"] for p in pens.get("pensioni_domani", []) if p.get("modalita") == "Anticipata"),
            None,
        ),
        "versamento_fondo": float(ac.get("capacita_risparmio_annuale") or 0) * 0.5,
        "risparmio_annuo": float(ac.get("capacita_risparmio_annuale") or 0) * 0.5,
        "vantaggio_fiscale": float(ac.get("capacita_risparmio_annuale") or 0) * 0.5
                              * (red.get("aliquota_irpef_marginale_pct", 35) / 100.0),
        "reddito": red.get("reddito_netto", 0),
    }
    soglia = float(ac.get("danno_devastante_entrate_mensili") or 1000)
    soglie = {
        "trascurabile": round(soglia / 5, 2),
        "basso": round(soglia * 2 / 5, 2),
        "medio": round(soglia * 3 / 5, 2),
        "alto": round(soglia * 4 / 5, 2),
        "molto_alto": round(soglia, 2),
    }
    tratt = {
        "scenario_a": scen_a,
        "scenario_b": scen_b,
        "obiettivi": ac.get("cosa_renderebbe_felice") or "",
        "perdita_entrate": ac.get("cosa_non_vuoi_carriera") or "",
        "soglie_devastante": soglie,
    }
    await db.analisi_cliente.update_one(
        {"anagrafica_id": aid},
        {"$set": {"trattativa": tratt, "updated_at": _now_iso()}},
    )
    return tratt


@api.get("/anagrafiche/{aid}/analisi/pdf-sezione")
async def pdf_sezione_analisi(
    aid: str,
    sezione: str = "all",
    user=Depends(current_user),
):
    """Genera PDF di UNA sezione (finanza/patrimonio/contesto/redditi/pensione/
    scoperture/successione/trattativa/piramide) o di TUTTE (sezione=all).
    Logo aziendale in alto a sinistra, header personalizzato."""
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    ac = await _ensure_analisi(aid)
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}

    import pdf_sezioni
    if sezione not in pdf_sezioni.SEZIONI and sezione != "all":
        raise HTTPException(400, f"Sezione non valida. Usa: {list(pdf_sezioni.SEZIONI.keys())} o 'all'")

    # Calcola dati on-demand se la sezione li richiede
    dati = {}
    sezioni_da_calcolare = [sezione] if sezione != "all" else list(pdf_sezioni.SEZIONI.keys())
    if any(pdf_sezioni.SEZIONI.get(s, (None, None, False))[2] for s in sezioni_da_calcolare):
        try:
            dati["redditi"] = await calcola_redditi_analisi(aid, user)
        except Exception:
            dati["redditi"] = {}
        try:
            dati["pensioni"] = await calcola_pensioni_future(aid, user)
        except Exception:
            dati["pensioni"] = {}
        try:
            dati["scoperture"] = await calcola_scoperture_analisi(aid, user)
        except Exception:
            dati["scoperture"] = {}
        try:
            dati["successione"] = await calcola_successione_analisi(aid, user)
        except Exception:
            dati["successione"] = {}

    pdf_bytes = pdf_sezioni.genera_pdf_sezione(sezione, ana, ac, az, dati)
    fname = f"analisi_{sezione}_{ana.get('codice_fiscale') or aid}.pdf"
    return StreamingResponse(
        _io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={fname}"},
    )


@api.delete("/anagrafiche/{aid}/analisi/estratto-inps/{estratto_id}")
async def delete_estratto_inps(
    aid: str, estratto_id: str,
    pulisci_storico: bool = False,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Rimuove un estratto INPS dall'archivio (file + entry).
    Se pulisci_storico=true, azzera anche storico_redditi e periodi_contributivi
    (utile se erano stati popolati solo da questo estratto)."""
    ac = await _ensure_analisi(aid)
    estratti = ac.get("estratti_conto_inps") or []
    target = next((e for e in estratti if e.get("id") == estratto_id), None)
    if not target:
        raise HTTPException(404, "Estratto non trovato")
    try:
        if target.get("storage_path"):
            obj_storage.delete_object(target["storage_path"])
    except Exception:
        pass
    new_estratti = [e for e in estratti if e.get("id") != estratto_id]
    update = {"estratti_conto_inps": new_estratti, "updated_at": _now_iso()}
    if pulisci_storico:
        update["storico_redditi"] = []
        update["periodi_contributivi"] = []
        # se restano altri estratti, ripopola con quelli più recenti
    await db.analisi_cliente.update_one(
        {"anagrafica_id": aid}, {"$set": update},
    )
    await log_attivita(user, "delete_estratto_inps", "anagrafica", aid,
                       descrizione=f"Estratto INPS {estratto_id} rimosso (pulisci={pulisci_storico})")
    return {"ok": True, "remaining": len(new_estratti)}


@api.post("/anagrafiche/{aid}/analisi/debug-estratto-inps")
async def debug_estratto_inps(
    aid: str,
    file: UploadFile = File(...),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Endpoint diagnostico: estrae il testo grezzo del PDF e mostra cosa il parser
    riconosce, SENZA salvare niente. Utile per capire perché un PDF non viene parsato."""
    contents = await file.read()
    ct = file.content_type or "application/pdf"
    text = ""
    if ct == "application/pdf" or (file.filename or "").lower().endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(_io.BytesIO(contents)) as pdf:
                text = "\n".join((p.extract_text() or "") for p in pdf.pages)
        except Exception as e:
            return {"errore": f"Impossibile leggere il PDF: {e}", "testo_estratto": ""}
    else:
        text = contents.decode("utf-8", errors="ignore")

    parsed = inps_calculator.parse_estratto_conto_inps(text)
    return {
        "testo_estratto_lunghezza": len(text),
        "testo_anteprima": text[:3000],
        "righe_contributive": parsed.get("righe_contributive", 0),
        "settimane_contributive": parsed.get("settimane_contributive", 0),
        "periodi_count": len(parsed.get("periodi_contributivi") or []),
        "storico_count": len(parsed.get("storico_redditi") or []),
        "periodi_sample": (parsed.get("periodi_contributivi") or [])[:5],
        "storico_sample": (parsed.get("storico_redditi") or [])[:10],
        "totale_versato": parsed.get("totale_versato", 0),
        "totale_retribuzioni": parsed.get("totale_retribuzioni", 0),
    }



# ============================================================
# PIPELINE (kanban-like data)
# ============================================================
PIPELINE_STATI_VALIDI = {
    "polizze": {"in_emissione", "attiva", "sospesa", "scaduta", "annullata"},
    "sinistri": {"aperto", "in_istruttoria", "liquidato", "chiuso_senza_seguito", "respinto"},
    "titoli": {"da_incassare", "insoluto", "incassato", "stornato"},
}


@api.post("/pipeline/{entita}/{eid}/move")
async def pipeline_move(
    entita: str, eid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Sposta una card di pipeline in una nuova colonna (cambia stato).

    body: {nuovo_stato: 'attiva' | ...}
    Supporta sia pipeline built-in (polizze/sinistri/titoli) che pipeline custom.
    """
    nuovo = body.get("nuovo_stato") or body.get("colonna_key")
    if entita in PIPELINE_STATI_VALIDI:
        if nuovo not in PIPELINE_STATI_VALIDI[entita]:
            raise HTTPException(400, f"Stato non valido per {entita}: {nuovo}")
        coll = {"polizze": db.polizze, "sinistri": db.sinistri, "titoli": db.titoli}[entita]
        res = await coll.update_one(
            {"id": eid},
            {"$set": {"stato": nuovo, "updated_at": _now_iso()}},
        )
        if res.matched_count == 0:
            raise HTTPException(404, "Elemento non trovato")
        await log_attivita(user, "move_pipeline", entita, eid, f"Stato -> {nuovo}")
        return {"ok": True, "id": eid, "stato": nuovo}
    # Pipeline custom: entita = pipeline_id, eid = card_id
    return await move_card(entita, eid, {"colonna_key": nuovo}, user)


@api.get("/pipeline/{entita}")
async def pipeline_data(entita: str, user=Depends(current_user)):
    """Ritorna dati per visualizzazione pipeline/kanban.

    entita: 'polizze' | 'sinistri' | 'titoli' | 'clienti' | 'email' | <id pipeline custom>
    """
    flt = await visibility_filter(user) if entita in ("polizze", "sinistri") else {}
    if entita == "polizze":
        stages = [
            ("in_emissione", "In emissione"), ("attiva", "Attive"),
            ("sospesa", "Sospese"), ("scaduta", "Scadute"), ("annullata", "Annullate"),
        ]
        items = await db.polizze.find(flt, {"_id": 0}).to_list(5000)
        ana_ids = list({i.get("contraente_id") for i in items})
        anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
        com_ids = list({i.get("compagnia_id") for i in items})
        coms = {c["id"]: c async for c in db.compagnie.find({"id": {"$in": com_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
        cols = []
        for stato, label in stages:
            cards = [i for i in items if i.get("stato") == stato]
            cols.append({
                "key": stato, "label": label, "count": len(cards),
                "cards": [{
                    "id": p["id"], "title": p["numero_polizza"],
                    "subtitle": anas.get(p.get("contraente_id", ""), {}).get("ragione_sociale", ""),
                    "footer": coms.get(p.get("compagnia_id", ""), {}).get("ragione_sociale", ""),
                    "extra": f"{p.get('ramo','')} · €{p.get('premio_lordo', 0):.2f}",
                    "date": p.get("scadenza"), "link": f"/polizze/{p['id']}",
                } for p in cards[:50]],
            })
        return {"entita": "polizze", "colonne": cols}

    if entita == "sinistri":
        stages = [
            ("aperto", "Aperti"), ("in_istruttoria", "In istruttoria"),
            ("liquidato", "Liquidati"), ("chiuso_senza_seguito", "Chiusi"), ("respinto", "Respinti"),
        ]
        items = await db.sinistri.find(flt, {"_id": 0}).to_list(5000)
        pol_ids = list({i.get("polizza_id") for i in items})
        pols = {p["id"]: p async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0, "id": 1, "numero_polizza": 1, "contraente_id": 1})}
        ana_ids = list({p.get("contraente_id") for p in pols.values()})
        anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
        cols = []
        for stato, label in stages:
            cards = [i for i in items if i.get("stato") == stato]
            cols.append({
                "key": stato, "label": label, "count": len(cards),
                "cards": [{
                    "id": s["id"], "title": s["numero_sinistro"],
                    "subtitle": anas.get(pols.get(s.get("polizza_id",""), {}).get("contraente_id", ""), {}).get("ragione_sociale", ""),
                    "footer": pols.get(s.get("polizza_id", ""), {}).get("numero_polizza", ""),
                    "extra": f"Riserva: €{s.get('riserva', 0):.2f}",
                    "date": s.get("data_avvenimento"), "link": f"/polizze/{s.get('polizza_id')}",
                } for s in cards[:50]],
            })
        return {"entita": "sinistri", "colonne": cols}

    if entita == "titoli":
        stages = [
            ("da_incassare", "Da incassare"), ("insoluto", "Insoluti"),
            ("incassato", "Incassati"), ("stornato", "Stornati"),
        ]
        items = await db.titoli.find({}, {"_id": 0}).to_list(5000)
        pol_ids = list({i.get("polizza_id") for i in items})
        pols = {p["id"]: p async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0, "id": 1, "numero_polizza": 1, "contraente_id": 1})}
        ana_ids = list({p.get("contraente_id") for p in pols.values()})
        anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
        cols = []
        for stato, label in stages:
            cards = [i for i in items if i.get("stato") == stato]
            cols.append({
                "key": stato, "label": label, "count": len(cards),
                "cards": [{
                    "id": t["id"],
                    "title": pols.get(t.get("polizza_id",""), {}).get("numero_polizza", "—"),
                    "subtitle": anas.get(pols.get(t.get("polizza_id",""), {}).get("contraente_id",""), {}).get("ragione_sociale", ""),
                    "footer": f"€{t.get('importo_lordo', 0):.2f}",
                    "extra": t.get("tipo", ""),
                    "date": t.get("scadenza"),
                    "link": f"/polizze/{t.get('polizza_id')}",
                } for t in cards[:50]],
            })
        return {"entita": "titoli", "colonne": cols}

    if entita == "clienti":
        # raggruppamento per quantità polizze
        stages = [
            ("prospect", "Prospect (0 polizze)"),
            ("nuovo", "Nuovo (1 polizza)"),
            ("attivo", "Attivo (2-3)"),
            ("top", "Top cliente (4+)"),
        ]
        anas = await db.anagrafiche.find({}, {"_id": 0}).to_list(5000)
        from collections import defaultdict
        counts = defaultdict(int)
        async for p in db.polizze.find({"stato": "attiva"}, {"contraente_id": 1, "_id": 0}):
            counts[p.get("contraente_id", "")] += 1
        cols_data = {k: [] for k, _ in stages}
        for a in anas:
            n = counts.get(a["id"], 0)
            if n == 0: cols_data["prospect"].append(a)
            elif n == 1: cols_data["nuovo"].append(a)
            elif n in (2, 3): cols_data["attivo"].append(a)
            else: cols_data["top"].append(a)
        cols = [{
            "key": k, "label": label, "count": len(cols_data[k]),
            "cards": [{
                "id": a["id"], "title": a["ragione_sociale"],
                "subtitle": f"{a.get('comune', '')} ({a.get('provincia', '')})" if a.get("comune") else "",
                "footer": a.get("email") or a.get("cellulare") or "",
                "extra": f"{counts.get(a['id'], 0)} polizze attive",
                "link": f"/anagrafiche/{a['id']}",
            } for a in cols_data[k][:50]],
        } for k, label in stages]
        return {"entita": "clienti", "colonne": cols}

    # ===== PIPELINE CUSTOM (entita = id pipeline) =====
    pipeline = await db.pipelines_custom.find_one({"id": entita}, {"_id": 0})
    if pipeline:
        # Visibilità: admin vede tutto, gli altri vedono solo le proprie
        if user["role"] != "admin" and pipeline.get("operatore_id") not in (None, user.get("id")):
            raise HTTPException(403, "Pipeline non accessibile")
        colonne_def = sorted(pipeline.get("colonne", []), key=lambda c: c.get("ordine", 0))
        cards = await db.pipeline_cards.find(
            {"pipeline_id": entita, "archiviata": {"$ne": True}},
            {"_id": 0},
        ).sort([("colonna_key", 1), ("ordine", 1), ("created_at", -1)]).to_list(5000)
        # Arricchimento anagrafica / operatore
        ana_ids = list({c.get("anagrafica_id") for c in cards if c.get("anagrafica_id")})
        anas = {}
        if ana_ids:
            async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1}):
                anas[a["id"]] = a["ragione_sociale"]
        op_ids = list({c.get("operatore_id") for c in cards if c.get("operatore_id")})
        ops = {}
        if op_ids:
            async for u in db.users.find({"id": {"$in": op_ids}}, {"_id": 0, "id": 1, "name": 1}):
                ops[u["id"]] = u["name"]
        cols_out = []
        for cd in colonne_def:
            col_cards = [c for c in cards if c.get("colonna_key") == cd["key"]]
            cols_out.append({
                "key": cd["key"], "label": cd["label"], "colore": cd.get("colore"),
                "count": len(col_cards),
                "cards": [{
                    "id": c["id"], "title": c["titolo"],
                    "subtitle": anas.get(c.get("anagrafica_id", "")) or (c.get("descrizione") or "")[:50],
                    "footer": ops.get(c.get("operatore_id", "")) or "",
                    "extra": (f"€{c.get('valore_stimato', 0):.0f}" if c.get('valore_stimato') else "")
                             + (" · " + c.get("priorita", "") if c.get("priorita") else ""),
                    "date": c.get("scadenza"),
                    "link": f"/anagrafiche/{c['anagrafica_id']}" if c.get("anagrafica_id") else None,
                    "tags": c.get("tags", []),
                    "priorita": c.get("priorita"),
                    "card_id": c["id"],
                } for c in col_cards[:50]],
            })
        return {
            "entita": entita,
            "pipeline": {"id": pipeline["id"], "nome": pipeline["nome"], "tipo": pipeline.get("tipo"),
                         "icona": pipeline.get("icona"), "colore": pipeline.get("colore"),
                         "descrizione": pipeline.get("descrizione")},
            "colonne": cols_out,
            "editabile_struttura": True,  # frontend sa che può modificare colonne / cards
        }

    raise HTTPException(400, "Entità pipeline non supportata")


# ============================================================
# PIPELINE CUSTOM — CRUD pipelines + colonne + cards
# ============================================================
@api.get("/pipelines")
async def list_pipelines_custom(user=Depends(current_user)):
    """Elenco pipeline custom (built-in escluse). Filtra per visibilità."""
    flt = {}
    if user["role"] != "admin":
        flt = {"$or": [{"operatore_id": user.get("id")}, {"operatore_id": None}]}
    items = await db.pipelines_custom.find(flt, {"_id": 0}).sort("nome", 1).to_list(200)
    # arricchimento conteggio cards
    for p in items:
        n = await db.pipeline_cards.count_documents({"pipeline_id": p["id"], "archiviata": {"$ne": True}})
        p["cards_count"] = n
    return items


@api.post("/pipelines", status_code=201)
async def crea_pipeline(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    if not body.get("nome"):
        raise HTTPException(400, "Nome richiesto")
    # se non passa colonne, usa il template del tipo
    if not body.get("colonne"):
        body["colonne"] = _template_colonne(body.get("tipo", "generico"))
    body.setdefault("operatore_id", user.get("id"))
    obj = PipelineCustom(**body)
    await db.pipelines_custom.insert_one(obj.model_dump())
    await log_attivita(user, "create", "pipeline", obj.id, obj.nome)
    return obj.model_dump()


def _template_colonne(tipo: str) -> list:
    templates = {
        "marketing": [
            {"key": "lead", "label": "Lead", "colore": "#94A3B8", "ordine": 1},
            {"key": "contattato", "label": "Contattato", "colore": "#0EA5E9", "ordine": 2},
            {"key": "interessato", "label": "Interessato", "colore": "#F59E0B", "ordine": 3},
            {"key": "trattativa", "label": "Trattativa", "colore": "#7C3AED", "ordine": 4},
            {"key": "vinto", "label": "Vinto", "colore": "#10B981", "ordine": 5},
            {"key": "perso", "label": "Perso", "colore": "#EF4444", "ordine": 6},
        ],
        "vendita": [
            {"key": "qualificazione", "label": "Qualificazione", "colore": "#94A3B8", "ordine": 1},
            {"key": "proposta", "label": "Proposta", "colore": "#0EA5E9", "ordine": 2},
            {"key": "negoziazione", "label": "Negoziazione", "colore": "#F59E0B", "ordine": 3},
            {"key": "chiuso_vinto", "label": "Chiuso vinto", "colore": "#10B981", "ordine": 4},
            {"key": "chiuso_perso", "label": "Chiuso perso", "colore": "#EF4444", "ordine": 5},
        ],
        "onboarding": [
            {"key": "da_iniziare", "label": "Da iniziare", "colore": "#94A3B8", "ordine": 1},
            {"key": "documenti", "label": "Raccolta documenti", "colore": "#0EA5E9", "ordine": 2},
            {"key": "verifica", "label": "Verifica", "colore": "#F59E0B", "ordine": 3},
            {"key": "completato", "label": "Completato", "colore": "#10B981", "ordine": 4},
        ],
        "supporto": [
            {"key": "aperto", "label": "Aperto", "colore": "#0EA5E9", "ordine": 1},
            {"key": "in_lavorazione", "label": "In lavorazione", "colore": "#F59E0B", "ordine": 2},
            {"key": "in_attesa_cliente", "label": "In attesa cliente", "colore": "#94A3B8", "ordine": 3},
            {"key": "risolto", "label": "Risolto", "colore": "#10B981", "ordine": 4},
        ],
    }
    return templates.get(tipo, [
        {"key": "todo", "label": "Da fare", "colore": "#94A3B8", "ordine": 1},
        {"key": "in_corso", "label": "In corso", "colore": "#0EA5E9", "ordine": 2},
        {"key": "fatto", "label": "Fatto", "colore": "#10B981", "ordine": 3},
    ])


@api.get("/pipelines/{pid}")
async def get_pipeline(pid: str, user=Depends(current_user)):
    p = await db.pipelines_custom.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Pipeline non trovata")
    return p


@api.put("/pipelines/{pid}")
async def update_pipeline(pid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body.pop("id", None)
    body["updated_at"] = _now_iso()
    res = await db.pipelines_custom.update_one({"id": pid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Pipeline non trovata")
    return await db.pipelines_custom.find_one({"id": pid}, {"_id": 0})


@api.delete("/pipelines/{pid}")
async def delete_pipeline(pid: str, user=Depends(require_user("admin"))):
    await db.pipeline_cards.delete_many({"pipeline_id": pid})
    res = await db.pipelines_custom.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Pipeline non trovata")
    return {"ok": True}


# ----- Colonne (manipolazione singola colonna nella pipeline) -----
@api.post("/pipelines/{pid}/colonne")
async def aggiungi_colonna(pid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    if not body.get("label"):
        raise HTTPException(400, "Label colonna richiesto")
    p = await db.pipelines_custom.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Pipeline non trovata")
    key = body.get("key") or _slugify(body["label"])
    if any(c["key"] == key for c in p.get("colonne", [])):
        raise HTTPException(400, f"Esiste già una colonna con chiave '{key}'")
    nuova = PipelineColonna(
        key=key, label=body["label"], colore=body.get("colore"),
        ordine=body.get("ordine", len(p.get("colonne", [])) + 1),
        descrizione=body.get("descrizione"),
    ).model_dump()
    await db.pipelines_custom.update_one({"id": pid}, {"$push": {"colonne": nuova}, "$set": {"updated_at": _now_iso()}})
    return nuova


@api.put("/pipelines/{pid}/colonne/{col_key}")
async def modifica_colonna(pid: str, col_key: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    # rinomina, sposta, cambia colore
    set_ops = {}
    for k, v in body.items():
        if k in ("label", "colore", "descrizione", "ordine"):
            set_ops[f"colonne.$.{k}"] = v
    if not set_ops:
        raise HTTPException(400, "Nessun campo da modificare")
    set_ops["updated_at"] = _now_iso()
    res = await db.pipelines_custom.update_one(
        {"id": pid, "colonne.key": col_key},
        {"$set": set_ops},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Colonna non trovata")
    return {"ok": True}


@api.delete("/pipelines/{pid}/colonne/{col_key}")
async def elimina_colonna(pid: str, col_key: str,
                           sposta_in: Optional[str] = None,
                           user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Elimina una colonna. Se sposta_in è specificato, le card vengono spostate; altrimenti vengono archiviate."""
    p = await db.pipelines_custom.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Pipeline non trovata")
    rimanenti = [c for c in p.get("colonne", []) if c["key"] != col_key]
    if len(rimanenti) == len(p.get("colonne", [])):
        raise HTTPException(404, "Colonna non trovata")
    await db.pipelines_custom.update_one({"id": pid},
        {"$set": {"colonne": rimanenti, "updated_at": _now_iso()}})
    if sposta_in:
        if not any(c["key"] == sposta_in for c in rimanenti):
            raise HTTPException(400, f"Colonna destinazione '{sposta_in}' non esiste")
        await db.pipeline_cards.update_many(
            {"pipeline_id": pid, "colonna_key": col_key},
            {"$set": {"colonna_key": sposta_in, "updated_at": _now_iso()}},
        )
    else:
        await db.pipeline_cards.update_many(
            {"pipeline_id": pid, "colonna_key": col_key},
            {"$set": {"archiviata": True, "updated_at": _now_iso()}},
        )
    return {"ok": True}


# ----- Cards CRUD -----
@api.post("/pipelines/{pid}/cards", status_code=201)
async def crea_card(pid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    p = await db.pipelines_custom.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Pipeline non trovata")
    cols = p.get("colonne", [])
    if not cols:
        raise HTTPException(400, "La pipeline non ha colonne")
    col_key = body.get("colonna_key") or cols[0]["key"]
    if not any(c["key"] == col_key for c in cols):
        raise HTTPException(400, f"Colonna '{col_key}' non esiste")
    body["pipeline_id"] = pid
    body["colonna_key"] = col_key
    body.setdefault("operatore_id", user.get("id"))
    obj = PipelineCard(**body)
    await db.pipeline_cards.insert_one(obj.model_dump())
    await log_attivita(user, "create", "pipeline_card", obj.id, obj.titolo)
    return obj.model_dump()


@api.put("/pipelines/{pid}/cards/{card_id}")
async def update_card(pid: str, card_id: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body.pop("id", None)
    body.pop("pipeline_id", None)
    body["updated_at"] = _now_iso()
    res = await db.pipeline_cards.update_one(
        {"id": card_id, "pipeline_id": pid}, {"$set": body},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Card non trovata")
    return await db.pipeline_cards.find_one({"id": card_id}, {"_id": 0})


@api.delete("/pipelines/{pid}/cards/{card_id}")
async def delete_card(pid: str, card_id: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    res = await db.pipeline_cards.delete_one({"id": card_id, "pipeline_id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Card non trovata")
    return {"ok": True}


@api.post("/pipelines/{pid}/cards/{card_id}/move")
async def move_card(pid: str, card_id: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Sposta una card in una nuova colonna della stessa pipeline."""
    nuova_col = body.get("nuovo_stato") or body.get("colonna_key")
    if not nuova_col:
        raise HTTPException(400, "colonna_key richiesto")
    p = await db.pipelines_custom.find_one({"id": pid}, {"_id": 0, "colonne": 1})
    if not p:
        raise HTTPException(404, "Pipeline non trovata")
    if not any(c["key"] == nuova_col for c in p.get("colonne", [])):
        raise HTTPException(400, f"Colonna '{nuova_col}' non esiste in questa pipeline")
    res = await db.pipeline_cards.update_one(
        {"id": card_id, "pipeline_id": pid},
        {"$set": {"colonna_key": nuova_col, "updated_at": _now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Card non trovata")
    await log_attivita(user, "move_card", "pipeline_card", card_id, f"-> {nuova_col}")
    return {"ok": True}


def _slugify(s: str) -> str:
    import re as _re
    s = _re.sub(r"[^a-zA-Z0-9]+", "_", s.lower()).strip("_")
    return s[:40] or "col"


# ============================================================
# EMAIL PIPELINE
# ============================================================
@api.get("/email")
async def list_email(stato: Optional[str] = None, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    flt: dict = {}
    if stato:
        flt["stato"] = stato
    items = await db.email.find(flt, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.post("/email", status_code=201)
async def create_email(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body["autore_id"] = user["id"]
    obj = EmailMessaggio(**body)
    await db.email.insert_one(obj.model_dump())
    await log_attivita(user, "create", "email", obj.id, f"Email a {obj.destinatario_email}")
    return obj.model_dump()


@api.post("/email/{eid}/invia")
async def invia_email(eid: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Invio email (MOCK - non invia realmente, marca come inviata)."""
    email_doc = await db.email.find_one({"id": eid}, {"_id": 0})
    if not email_doc:
        raise HTTPException(404, "Non trovata")
    await db.email.update_one(
        {"id": eid},
        {"$set": {"stato": "inviata", "data_invio": _now_iso(), "updated_at": _now_iso()}},
    )
    await log_attivita(user, "invio", "email", eid, "Email inviata (mock)")
    # auto-traccia nel diario del cliente
    if email_doc.get("destinatario_anagrafica_id"):
        await log_diario_cliente(
            email_doc["destinatario_anagrafica_id"], "email",
            titolo=f"📧 Email inviata: {email_doc.get('oggetto', '')}",
            descrizione=(email_doc.get("corpo") or "")[:500],
            autore=user,
        )
    return {"ok": True, "note": "Invio simulato. Configurare SMTP per invio reale."}


@api.post("/email/invia-singola")
async def email_invia_singola(
    body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Invia un'email immediata via SMTP configurato in AziendaConfig.

    body: {to: str, subject: str, body_text?: str, body_html?: str}
    """
    to = (body.get("to") or "").strip()
    subject = (body.get("subject") or "").strip()
    body_text = body.get("body_text") or body.get("body") or ""
    body_html = body.get("body_html") or ""
    if not to or not subject:
        raise HTTPException(400, "to e subject obbligatori")
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    if not az.get("smtp_host"):
        raise HTTPException(503, "SMTP non configurato in Librerie/Azienda")
    from avvisi_scadenze import _invia_email
    try:
        _invia_email(az, to, subject, body_text or _strip_tags(body_html), body_html or f"<pre>{body_text}</pre>")
    except Exception as e:
        raise HTTPException(500, f"Errore invio: {e}")
    await log_attivita(user, "email_send", "email", None, f"to={to} subject={subject[:60]}")
    # Storico invio email singola
    from datetime import datetime as _dt2, timezone as _tz2
    await db.storico_avvisi.insert_one({
        "id": str(uuid.uuid4()),
        "tipo": body.get("tipo_avviso") or "email_singola",
        "canale": "email",
        "contraente_id": body.get("contraente_id"),
        "contraente_nome": body.get("contraente_nome"),
        "destinatario": to,
        "soggetto": subject,
        "titoli_ids": body.get("titoli_ids") or [],
        "polizza_id": body.get("polizza_id"),
        "stato": "inviato",
        "sent_at": _dt2.now(_tz2.utc).isoformat(),
        "utente_id": user.get("id"),
        "utente_nome": user.get("name") or user.get("email"),
    })
    return {"ok": True}


def _strip_tags(html: str) -> str:
    import re as _re
    return _re.sub(r"<[^>]+>", "", html or "")


@api.post("/avvisi/pdf-bulk")
async def avvisi_pdf_bulk(
    body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Genera un PDF con un avviso per ciascun contraente (aggrega per contraente).

    body: {titoli_ids: [str], corpo_lettera?: str, soggetto?: str}
    Risposta: file PDF (application/pdf).
    """
    titoli_ids = body.get("titoli_ids") or []
    if not titoli_ids:
        raise HTTPException(400, "titoli_ids obbligatorio")
    soggetto = body.get("soggetto") or "Promemoria pagamento polizza/e in scadenza"
    corpo = body.get("corpo_lettera") or _CORPO_LETTERA_DEFAULT

    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    titoli = await db.titoli.find({"id": {"$in": titoli_ids}}, {"_id": 0}).to_list(5000)
    if not titoli:
        raise HTTPException(404, "Nessun titolo trovato")
    pol_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
    pols = {p["id"]: p async for p in db.polizze.find(
        {"id": {"$in": pol_ids}},
        {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "prodotto": 1,
         "targa": 1, "contraente_id": 1},
    )}
    ana_ids = list({p.get("contraente_id") for p in pols.values() if p.get("contraente_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}},
        {"_id": 0, "id": 1, "ragione_sociale": 1, "email": 1, "indirizzo": 1,
         "cap": 1, "comune": 1, "provincia": 1, "cellulare": 1},
    )}

    from collections import defaultdict
    bucket = defaultdict(list)
    for t in titoli:
        p = pols.get(t.get("polizza_id"), {})
        cid = p.get("contraente_id")
        if cid:
            bucket[cid].append({
                **t,
                "numero_polizza": p.get("numero_polizza"),
                "ramo": p.get("ramo"),
                "prodotto": p.get("prodotto"),
                "targa": p.get("targa"),
            })

    gruppi = []
    for cid, ts in bucket.items():
        a = anas.get(cid, {})
        indirizzo = ", ".join(filter(None, [
            a.get("indirizzo"),
            " ".join(filter(None, [a.get("cap"), a.get("comune"), f"({a['provincia']})" if a.get("provincia") else None])),
        ]))
        gruppi.append({
            "contraente_id": cid,
            "contraente_nome": a.get("ragione_sociale") or "—",
            "contraente_indirizzo": indirizzo or None,
            "contraente_email": a.get("email"),
            "titoli": ts,
        })

    from pdf_avvisi import genera_pdf_avvisi
    from datetime import date as _date_pdf
    pdf_bytes = genera_pdf_avvisi(
        gruppi=gruppi, azienda=az, corpo_lettera=corpo, soggetto=soggetto,
    )
    await log_attivita(user, "avvisi_pdf", "report", None,
                       f"contraenti={len(gruppi)} titoli={len(titoli_ids)}")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="avvisi_scadenza_{_date_pdf.today().isoformat()}.pdf"',
        },
    )


@api.post("/avvisi/invia-bulk-titoli")
async def invia_bulk_avvisi_titoli(
    body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Invia email avviso titoli aggregando per contraente.

    body: {
        titoli_ids: [str],
        soggetto?: str (template),
        corpo_lettera?: str (testo libero modificabile inserito sopra la tabella),
    }
    Per ogni contraente con titoli coinvolti, viene inviata UNA sola email che
    elenca tutti i suoi titoli e somma gli importi.
    """
    titoli_ids = body.get("titoli_ids") or []
    if not titoli_ids:
        raise HTTPException(400, "titoli_ids obbligatorio (lista)")
    soggetto_tpl = body.get("soggetto") or "Promemoria pagamento polizza/e in scadenza"
    corpo_lettera = body.get("corpo_lettera") or _CORPO_LETTERA_DEFAULT

    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    if not az.get("smtp_host"):
        raise HTTPException(503, "SMTP non configurato in Librerie/Azienda")

    titoli = await db.titoli.find({"id": {"$in": titoli_ids}}, {"_id": 0}).to_list(5000)
    if not titoli:
        raise HTTPException(404, "Nessun titolo trovato")

    pol_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
    pols = {p["id"]: p async for p in db.polizze.find(
        {"id": {"$in": pol_ids}},
        {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "targa": 1,
         "prodotto": 1, "contraente_id": 1, "compagnia_id": 1},
    )}
    ana_ids = list({p.get("contraente_id") for p in pols.values() if p.get("contraente_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}},
        {"_id": 0, "id": 1, "ragione_sociale": 1, "email": 1},
    )}

    # raggruppa titoli per contraente
    from collections import defaultdict
    bucket: dict = defaultdict(list)
    for t in titoli:
        p = pols.get(t.get("polizza_id"), {})
        contraente_id = p.get("contraente_id")
        if not contraente_id:
            continue
        bucket[contraente_id].append({**t, "_pol": p})

    from avvisi_scadenze import _invia_email
    inviate = 0
    skipped: list[dict] = []
    for contraente_id, ts in bucket.items():
        a = anas.get(contraente_id, {})
        to_addr = (a.get("email") or "").strip()
        if not to_addr:
            skipped.append({"contraente_id": contraente_id, "motivo": "email mancante"})
            continue
        text, html = _render_avviso_titoli_email(
            ragione_sociale=a.get("ragione_sociale") or "Cliente",
            corpo_lettera=corpo_lettera,
            titoli=ts,
            azienda=az,
        )
        try:
            _invia_email(az, to_addr, soggetto_tpl, text, html)
            inviate += 1
        except Exception as e:
            skipped.append({"contraente_id": contraente_id, "motivo": str(e)})
    await log_attivita(user, "avvisi_bulk", "email", None,
                       f"contraenti={len(bucket)} inviate={inviate} skipped={len(skipped)}")
    # Storico invii: salva 1 record per contraente coinvolto (anche skipped)
    from datetime import datetime as _dt, timezone as _tz
    now_iso = _dt.now(_tz.utc).isoformat()
    log_docs: list[dict] = []
    for contraente_id, ts in bucket.items():
        a = anas.get(contraente_id, {})
        to_addr = (a.get("email") or "").strip()
        sk = next((s for s in skipped if s.get("contraente_id") == contraente_id), None)
        log_docs.append({
            "id": str(uuid.uuid4()),
            "tipo": "email_avviso_titoli",
            "canale": "email",
            "contraente_id": contraente_id,
            "contraente_nome": a.get("ragione_sociale"),
            "destinatario": to_addr or None,
            "soggetto": soggetto_tpl,
            "titoli_ids": [t.get("id") for t in ts],
            "n_titoli": len(ts),
            "totale_importo": round(sum(float(t.get("importo_lordo") or 0) for t in ts), 2),
            "stato": "errore" if sk else "inviato",
            "errore": sk.get("motivo") if sk else None,
            "sent_at": now_iso,
            "utente_id": user.get("id"),
            "utente_nome": user.get("name") or user.get("email"),
        })
    if log_docs:
        await db.storico_avvisi.insert_many(log_docs)
    return {"ok": True, "contraenti_totali": len(bucket), "inviate": inviate, "skipped": skipped}


def _render_avviso_titoli_email(*, ragione_sociale: str, corpo_lettera: str,
                                 titoli: list[dict], azienda: dict) -> tuple[str, str]:
    """Genera (text, html) per email avviso titoli aggregati per un contraente."""
    from html import escape as _esc
    rows_html: list[str] = []
    rows_txt: list[str] = []
    totale = 0.0
    for t in titoli:
        p = t.get("_pol") or {}
        rischio = (p.get("prodotto") or p.get("ramo") or "").upper()
        targa = p.get("targa") or ""
        scad = t.get("scadenza") or ""
        try:
            from datetime import date as _date
            scad_fmt = _date.fromisoformat(scad[:10]).strftime("%d-%m-%Y") if scad else ""
        except Exception:
            scad_fmt = scad
        imp = float(t.get("importo_lordo") or 0)
        totale += imp
        rows_html.append(
            f"<tr>"
            f"<td style='padding:6px 8px;border-bottom:1px dotted #999'>{_esc(str(p.get('numero_polizza') or ''))}</td>"
            f"<td style='padding:6px 8px;border-bottom:1px dotted #999'>{_esc(rischio)}</td>"
            f"<td style='padding:6px 8px;border-bottom:1px dotted #999'>{_esc(targa)}</td>"
            f"<td style='padding:6px 8px;border-bottom:1px dotted #999'>{_esc(scad_fmt)}</td>"
            f"<td style='padding:6px 8px;border-bottom:1px dotted #999;text-align:right'>{imp:,.2f}</td>"
            f"</tr>"
        )
        rows_txt.append(
            f"- {p.get('numero_polizza') or '—'} | {rischio} | {targa} | {scad_fmt} | {imp:,.2f} €"
        )
    iban = azienda.get("iban") or ""
    ragione_az = azienda.get("ragione_sociale") or ""
    intestaz_pagamenti = ""
    if ragione_az or iban:
        intestaz_pagamenti = (
            "<p>Il pagamento del/i premio/i potrà essere effettuato presso la nostra sede o tramite bonifico:</p>"
            f"<p><strong>{_esc(ragione_az)}</strong>"
            + (f"<br/><strong>IBAN: {_esc(iban)}</strong>" if iban else "")
            + "</p>"
        )
    body_lettera_html = "<p>" + _esc(corpo_lettera).replace("\n\n", "</p><p>").replace("\n", "<br/>") + "</p>"
    html = f"""<!DOCTYPE html>
<html><body style="font-family:Georgia,serif;color:#111;line-height:1.45;max-width:680px;margin:0 auto;padding:20px">
  {body_lettera_html}
  {intestaz_pagamenti}
  <table style="width:100%;border-collapse:collapse;margin-top:18px;font-size:13px">
    <thead>
      <tr style="border-bottom:2px solid #333">
        <th style="text-align:left;padding:8px">Num.Contratto</th>
        <th style="text-align:left;padding:8px">Rischio</th>
        <th style="text-align:left;padding:8px">Targa</th>
        <th style="text-align:left;padding:8px">Rata del</th>
        <th style="text-align:right;padding:8px">Imp.Totale</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows_html)}
    </tbody>
    <tfoot>
      <tr><td colspan="5" style="padding-top:14px"></td></tr>
      <tr style="border-top:2px solid #333">
        <td colspan="4" style="text-align:right;padding:8px;font-size:14px;color:#777">Totale Complessivo</td>
        <td style="text-align:right;padding:8px;font-size:14px;font-weight:bold">{totale:,.2f}</td>
      </tr>
    </tfoot>
  </table>
</body></html>"""
    text = (
        corpo_lettera
        + "\n\n"
        + (f"{ragione_az}\n" if ragione_az else "")
        + (f"IBAN: {iban}\n" if iban else "")
        + "\nNum.Contratto | Rischio | Targa | Rata del | Imp.Totale\n"
        + "\n".join(rows_txt)
        + f"\n\nTotale Complessivo: {totale:,.2f} €"
    )
    return text, html


@api.post("/email/avvisi-scadenze")
async def genera_avvisi_scadenze(giorni: int = 30, user=Depends(require_user("admin", "collaboratore"))):
    """Genera email di avviso per polizze in scadenza nei prossimi N giorni."""
    from datetime import date, timedelta
    oggi = date.today()
    limite = (oggi + timedelta(days=giorni)).isoformat()
    polizze = await db.polizze.find(
        {"stato": "attiva", "scadenza": {"$gte": oggi.isoformat(), "$lte": limite}},
        {"_id": 0},
    ).to_list(1000)
    creati = 0
    for p in polizze:
        ana = await db.anagrafiche.find_one({"id": p["contraente_id"]}, {"_id": 0})
        if not ana or not ana.get("email"):
            continue
        email = EmailMessaggio(
            destinatario_anagrafica_id=ana["id"],
            destinatario_email=ana["email"],
            oggetto=f"Promemoria scadenza polizza {p['numero_polizza']}",
            corpo=(f"Gentile {ana['ragione_sociale']},\n\n"
                   f"la informiamo che la sua polizza n. {p['numero_polizza']} "
                   f"({p['ramo']}) scadrà il {p['scadenza']}.\n\n"
                   f"La invitiamo a contattarci per il rinnovo.\n\n"
                   f"Cordiali saluti."),
            template="scadenza_polizza",
            stato="in_coda",
            polizza_id=p["id"],
            autore_id=user["id"],
        )
        await db.email.insert_one(email.model_dump())
        creati += 1
    await log_attivita(user, "genera_avvisi", "email", None,
                       f"Generati {creati} avvisi di scadenza")
    return {"avvisi_creati": creati}


# ============================================================
# AVVISI SCADENZE — CRON ADMIN (08:00 daily)
# ============================================================
@api.get("/avvisi-scadenze/preview")
async def preview_avvisi_scadenze(
    giorni: Optional[int] = None,
    dal: Optional[str] = None,
    al: Optional[str] = None,
    collaboratore_id: Optional[str] = None,
    mezzo_pagamento: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Anteprima: ritorna le scadenze (polizze + titoli) nei prossimi N giorni.

    Se ``giorni`` non specificato, usa il valore da AziendaConfig (default 15).
    Filtri opzionali:
      - dal / al: range scadenza (YYYY-MM-DD) — sovrascrive `giorni`
      - collaboratore_id: filtra polizze e titoli per collaboratore
      - mezzo_pagamento: filtra titoli per mezzo di pagamento
    """
    if giorni is None:
        az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
        giorni = int(az.get("notifica_scadenze_giorni") or 15)
    giorni = max(1, min(giorni, 365))
    data = await avvisi_scadenze.cerca_scadenze(db, giorni)

    pol = data["polizze"]
    tit = data["titoli"]
    # Applicazione filtri post-query (semplice: la dataset è già limitato)
    if dal:
        tit = [t for t in tit if (t.get("scadenza") or "") >= dal]
        pol = [p for p in pol if (p.get("scadenza") or "") >= dal]
    if al:
        tit = [t for t in tit if (t.get("scadenza") or "") <= al]
        pol = [p for p in pol if (p.get("scadenza") or "") <= al]
    if collaboratore_id and collaboratore_id != "all":
        tit = [t for t in tit if t.get("collaboratore_id") == collaboratore_id]
        pol = [p for p in pol if p.get("collaboratore_id") == collaboratore_id]
    if mezzo_pagamento and mezzo_pagamento != "all":
        tit = [t for t in tit if (t.get("mezzo_pagamento") or "").lower() == mezzo_pagamento.lower()]

    return {
        "giorni": giorni,
        "polizze": pol,
        "titoli": tit,
        "n_polizze": len(pol),
        "n_titoli": len(tit),
    }


@api.post("/avvisi-scadenze/esegui")
async def esegui_avvisi_scadenze(user=Depends(require_user("admin"))):
    """Esegue manualmente il job di invio email avvisi scadenze (bypassa cron 08:00)."""
    res = await avvisi_scadenze.esegui_job_scadenze(db, manuale=True)
    await log_attivita(user, "esegui_avvisi_scadenze", "notifiche", None,
                       f"polizze={res.get('n_polizze',0)} titoli={res.get('n_titoli',0)} "
                       f"inviata={res.get('email_inviata') or res.get('ok')}")
    return res


@api.get("/avvisi-scadenze/log")
async def list_log_avvisi_scadenze(
    limit: int = 50, user=Depends(require_user("admin", "collaboratore")),
):
    """Storico esecuzioni del job avvisi scadenze."""
    items = await db.notifiche_scadenze.find({}, {"_id": 0}).sort("eseguito_at", -1).to_list(limit)
    return items



# ============================================================
# ATTIVITA LOG
# ============================================================
@api.get("/attivita")
async def list_attivita(
    entita: Optional[str] = None,
    utente_id: Optional[str] = None,
    limit: int = 200,
    user=Depends(require_user("admin", "collaboratore")),
):
    flt: dict = {}
    if entita:
        flt["entita"] = entita
    if utente_id:
        flt["utente_id"] = utente_id
    items = await db.attivita_log.find(flt, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return items


# ============================================================
# STATISTICHE / DASHBOARD
# ============================================================
@api.get("/stats/dashboard")
async def stats_dashboard(user=Depends(current_user)):
    flt = await visibility_filter(user)
    today = _now_iso()[:10]
    from datetime import date, timedelta
    sessanta_gg = (date.today() + timedelta(days=60)).isoformat()

    n_anagrafiche = await db.anagrafiche.count_documents({} if user["role"] != "cliente" else {"id": user.get("anagrafica_id", "_none_")})
    n_polizze = await db.polizze.count_documents(flt)
    polizze_attive = await db.polizze.count_documents({**flt, "stato": "attiva"})
    in_scadenza = await db.polizze.count_documents({
        **flt, "stato": "attiva",
        "scadenza": {"$gte": today, "$lte": sessanta_gg},
    })
    sinistri_aperti = await db.sinistri.count_documents({**flt, "stato": {"$in": ["aperto", "in_istruttoria"]}})

    # Premi incassati (anno corrente) - filtra per cliente se necessario
    anno = today[:4]
    pol_match = {}
    if user["role"] == "cliente" and user.get("anagrafica_id"):
        cli_pol_ids = [p["id"] async for p in db.polizze.find(
            {"contraente_id": user["anagrafica_id"]}, {"_id": 0, "id": 1})]
        pol_match = {"polizza_id": {"$in": cli_pol_ids}}

    pipeline = [
        {"$match": {**pol_match, "stato": "incassato", "data_incasso": {"$regex": f"^{anno}"}}},
        {"$group": {"_id": None, "totale": {"$sum": "$importo_lordo"}}},
    ]
    cur = db.titoli.aggregate(pipeline)
    res = await cur.to_list(1)
    premi_anno = res[0]["totale"] if res else 0.0

    # Distribuzione polizze per ramo
    pipeline_ramo = [
        {"$match": flt},
        {"$group": {"_id": "$ramo", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 8},
    ]
    rami = await db.polizze.aggregate(pipeline_ramo).to_list(20)

    # Incassi ultimi 6 mesi
    from datetime import date
    sei_mesi = []
    for i in range(5, -1, -1):
        d = date.today().replace(day=1)
        # naive: scendi di i mesi
        m = d.month - i
        y = d.year
        while m <= 0:
            m += 12
            y -= 1
        mese = f"{y}-{m:02d}"
        agg = await db.titoli.aggregate([
            {"$match": {**pol_match, "stato": "incassato", "data_incasso": {"$regex": f"^{mese}"}}},
            {"$group": {"_id": None, "totale": {"$sum": "$importo_lordo"}}},
        ]).to_list(1)
        sei_mesi.append({"mese": mese, "totale": round(agg[0]["totale"], 2) if agg else 0.0})

    return {
        "anagrafiche": n_anagrafiche,
        "polizze_totali": n_polizze,
        "polizze_attive": polizze_attive,
        "polizze_in_scadenza": in_scadenza,
        "sinistri_aperti": sinistri_aperti,
        "premi_anno_corrente": round(premi_anno, 2),
        "polizze_per_ramo": [{"ramo": r["_id"] or "N/D", "count": r["count"]} for r in rami],
        "incassi_mensili": sei_mesi,
    }


@api.get("/stats/dashboard-admin")
async def stats_dashboard_admin(user=Depends(require_user("admin"))):
    """KPI estesi solo admin: provvigioni, nuova produzione per ramo, sinistri liquidati,
    clienti per categoria, scadenze 5/10/15 gg."""
    from datetime import date, timedelta
    today = date.today()
    today_iso = today.isoformat()
    inizio_anno = f"{today.year}-01-01"

    # Provvigioni totali annue (somma provvigioni titoli incassati anno corrente)
    pipe_provv = [
        {"$match": {"stato": "incassato", "data_incasso": {"$gte": inizio_anno}}},
        {"$group": {"_id": None,
                    "provv_totali": {"$sum": {"$ifNull": ["$provvigione_lorda", 0]}},
                    "premi_totali": {"$sum": "$importo_lordo"},
                    "n_titoli": {"$sum": 1}}},
    ]
    pres = await db.titoli.aggregate(pipe_provv).to_list(1)
    provv_anno = pres[0]["provv_totali"] if pres else 0
    premi_anno = pres[0]["premi_totali"] if pres else 0

    # Nuova produzione per ramo (polizze emesse anno corrente con premio)
    pipe_prod = [
        {"$match": {"data_decorrenza": {"$gte": inizio_anno}}},
        {"$group": {"_id": "$ramo", "count": {"$sum": 1}, "premio": {"$sum": {"$ifNull": ["$premio_lordo", 0]}}}},
        {"$sort": {"premio": -1}},
    ]
    prod_per_ramo = await db.polizze.aggregate(pipe_prod).to_list(50)

    # Nuova produzione giornaliera (oggi)
    pipe_oggi = [
        {"$match": {"data_decorrenza": today_iso}},
        {"$group": {"_id": "$ramo", "count": {"$sum": 1}, "premio": {"$sum": {"$ifNull": ["$premio_lordo", 0]}}}},
    ]
    prod_oggi = await db.polizze.aggregate(pipe_oggi).to_list(20)

    # Sinistri (numero + totale per stato)
    pipe_sin = [
        {"$group": {"_id": "$stato", "n": {"$sum": 1}, "tot": {"$sum": {"$ifNull": ["$importo_risarcito", 0]}}}},
    ]
    sin = await db.sinistri.aggregate(pipe_sin).to_list(20)
    sinistri_by_stato = {s["_id"] or "aperto": {"n": s["n"], "totale": round(s["tot"], 2)} for s in sin}

    # Clienti attivi suddivisi per categoria
    cat_counts = {"Condominio": 0, "Azienda": 0, "Parrocchia": 0, "Privato": 0}
    async for a in db.anagrafiche.find({}, {"_id": 0, "ragione_sociale": 1, "tipo": 1, "tags": 1, "categoria": 1}):
        rs = (a.get("ragione_sociale") or "").upper()
        tags = a.get("tags") or []
        cat = (a.get("categoria") or "").lower()
        if "CONDOMINIO" in rs or "condominio" in tags or cat == "condominio":
            cat_counts["Condominio"] += 1
        elif "PARROCCHIA" in rs or "parrocchia" in tags or cat == "parrocchia":
            cat_counts["Parrocchia"] += 1
        elif a.get("tipo") == "persona_giuridica" or "azienda" in tags or cat == "azienda":
            cat_counts["Azienda"] += 1
        else:
            cat_counts["Privato"] += 1

    # Scadenze a 5, 10, 15 giorni
    scadenze = {}
    for g in (5, 10, 15, 30, 60):
        fine = (today + timedelta(days=g)).isoformat()
        cnt = await db.polizze.count_documents({
            "stato": "attiva", "scadenza": {"$gte": today_iso, "$lte": fine},
        })
        scadenze[f"{g}gg"] = cnt

    return {
        "provvigioni_anno": round(provv_anno, 2),
        "premi_anno": round(premi_anno, 2),
        "n_titoli_anno": pres[0]["n_titoli"] if pres else 0,
        "produzione_per_ramo": [{"ramo": p["_id"] or "N/D", "n": p["count"], "premio": round(p["premio"], 2)} for p in prod_per_ramo],
        "produzione_oggi": [{"ramo": p["_id"] or "N/D", "n": p["count"], "premio": round(p["premio"], 2)} for p in prod_oggi],
        "sinistri_per_stato": sinistri_by_stato,
        "clienti_per_categoria": cat_counts,
        "scadenze": scadenze,
    }


# ============================================================
# LIBRERIE: Banche, Conti cassa, Prodotti, Rami
# ============================================================
# ====== LIBRERIE (extracted to routes/librerie.py) ======
from routes import librerie as _librerie_router  # noqa: E402
api.include_router(_librerie_router.router)
_seed_mezzi_pagamento = _librerie_router._seed_mezzi_pagamento  # backward-compat
_seed_tipi_pagamento = _librerie_router._seed_tipi_pagamento
_seed_conti_deposito_estesi = _librerie_router._seed_conti_deposito_estesi



# ============================================================
# UTENTI — DOCUMENTI E CORSI (firma digitale, CI, casellario, ecc.)
# ============================================================
ALLOWED_USER_DOCS = {
    "firma_digitale": "firma_digitale_url",
    "carta_identita": "carta_identita_url",
    "casellario": "casellario_url",
    "carichi_pendenti": "carichi_pendenti_url",
    "documento_iban": "documento_iban_url",
}


@api.post("/auth/users/{uid}/avatar")
async def upload_avatar_utente(
    uid: str,
    file: UploadFile = File(...),
    user=Depends(current_user),
):
    """Carica l'avatar (immagine profilo) di un utente.
    Permesso: admin oppure l'utente stesso.
    """
    if user["role"] != "admin" and user.get("id") != uid:
        raise HTTPException(403, "Permesso negato")
    target = await db.users.find_one({"id": uid}, {"_id": 0, "id": 1})
    if not target:
        raise HTTPException(404, "Utente non trovato")
    data = await file.read()
    if len(data) > 4 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 4 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    if not (ct or "").startswith("image/"):
        raise HTTPException(400, "Formato non supportato (richiesto JPG/PNG/WEBP)")
    ext = (file.filename or "avatar.jpg").rsplit(".", 1)[-1].lower() or "jpg"
    path = f"{os.environ.get('APP_NAME', 'assicura')}/users/{uid}/avatar_{_uid()}.{ext}"
    try:
        result = obj_storage.put_object(path, data, ct)
    except Exception as e:
        raise HTTPException(503, f"Errore upload: {e}")
    url = f"/api/storage/{result['path']}"
    await db.users.update_one({"id": uid}, {"$set": {"avatar_url": url, "updated_at": _now_iso()}})
    await log_attivita(user, "upload", "user_avatar", uid, "Avatar caricato")
    return {"avatar_url": url}


@api.delete("/auth/users/{uid}/avatar")
async def delete_avatar_utente(uid: str, user=Depends(current_user)):
    if user["role"] != "admin" and user.get("id") != uid:
        raise HTTPException(403, "Permesso negato")
    await db.users.update_one({"id": uid}, {"$set": {"avatar_url": None, "updated_at": _now_iso()}})
    return {"ok": True}



@api.post("/auth/users/{uid}/documenti/{doc_tipo}")
async def upload_documento_utente(
    uid: str, doc_tipo: str,
    file: UploadFile = File(...),
    user=Depends(require_user("admin")),
):
    """Carica un documento (firma_digitale | carta_identita | casellario | carichi_pendenti | documento_iban)."""
    if doc_tipo not in ALLOWED_USER_DOCS:
        raise HTTPException(400, f"Tipo documento non valido: {doc_tipo}")
    target = await db.users.find_one({"id": uid}, {"_id": 0, "id": 1})
    if not target:
        raise HTTPException(404, "Utente non trovato")
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 10 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    ext = (file.filename or "doc.bin").rsplit(".", 1)[-1].lower() or "bin"
    path = f"{os.environ.get('APP_NAME', 'assicura')}/users/{uid}/{doc_tipo}_{_uid()}.{ext}"
    try:
        result = obj_storage.put_object(path, data, ct)
    except Exception as e:
        raise HTTPException(503, f"Errore upload: {e}")
    url = f"/api/storage/{result['path']}"
    field = ALLOWED_USER_DOCS[doc_tipo]
    await db.users.update_one({"id": uid}, {"$set": {field: url, "updated_at": _now_iso()}})
    await log_attivita(user, "upload", "user_doc", uid, f"Caricato {doc_tipo} per {uid}")
    return {field: url}


@api.delete("/auth/users/{uid}/documenti/{doc_tipo}")
async def delete_documento_utente(uid: str, doc_tipo: str,
                                   user=Depends(require_user("admin"))):
    if doc_tipo not in ALLOWED_USER_DOCS:
        raise HTTPException(400, "Tipo documento non valido")
    field = ALLOWED_USER_DOCS[doc_tipo]
    await db.users.update_one({"id": uid}, {"$set": {field: None, "updated_at": _now_iso()}})
    return {"ok": True}


@api.post("/auth/users/{uid}/corsi")
async def aggiungi_corso_utente(
    uid: str, body: dict,
    user=Depends(require_user("admin")),
):
    """Aggiunge un corso/attestato al collaboratore. body: {titolo, ente, data_scadenza, url_attestato}"""
    target = await db.users.find_one({"id": uid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Utente non trovato")
    corso = {
        "id": _uid(),
        "titolo": body.get("titolo") or "Corso",
        "ente": body.get("ente"),
        "data": body.get("data"),
        "data_scadenza": body.get("data_scadenza"),
        "url_attestato": body.get("url_attestato"),
        "note": body.get("note"),
    }
    await db.users.update_one({"id": uid}, {"$push": {"corsi": corso}, "$set": {"updated_at": _now_iso()}})
    return corso


@api.post("/auth/users/{uid}/corsi/upload")
async def upload_attestato_corso(
    uid: str,
    file: UploadFile = File(...),
    titolo: str = "",
    ente: str = "",
    data_scadenza: str = "",
    user=Depends(require_user("admin")),
):
    target = await db.users.find_one({"id": uid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Utente non trovato")
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 10 MB)")
    ct = file.content_type or obj_storage.mime_for(file.filename or "")
    ext = (file.filename or "doc.pdf").rsplit(".", 1)[-1].lower() or "pdf"
    path = f"{os.environ.get('APP_NAME', 'assicura')}/users/{uid}/corso_{_uid()}.{ext}"
    try:
        result = obj_storage.put_object(path, data, ct)
    except Exception as e:
        raise HTTPException(503, f"Errore upload: {e}")
    corso = {
        "id": _uid(),
        "titolo": titolo or (file.filename or "Corso"),
        "ente": ente or None,
        "data_scadenza": data_scadenza or None,
        "url_attestato": f"/api/storage/{result['path']}",
        "nome_file": file.filename,
    }
    await db.users.update_one({"id": uid}, {"$push": {"corsi": corso}, "$set": {"updated_at": _now_iso()}})
    return corso


@api.delete("/auth/users/{uid}/corsi/{corso_id}")
async def elimina_corso_utente(uid: str, corso_id: str,
                                user=Depends(require_user("admin"))):
    await db.users.update_one({"id": uid}, {"$pull": {"corsi": {"id": corso_id}}})
    return {"ok": True}


# ============================================================
# POSTA — Inbox (popolata dal poller IMAP, in arrivo Step 2)
# ============================================================
@api.get("/email/inbox/stats")
async def email_inbox_stats(user=Depends(current_user)):
    """KPI per la pagina Posta dell'utente loggato."""
    uid = user["id"]
    is_admin = user.get("role") == "admin"
    # personale: smistato_a contiene l'utente
    pers_total = await db.email_inbox.count_documents({"smistato_a": uid})
    pers_unread = await db.email_inbox.count_documents({
        "smistato_a": uid, "letta_da": {"$nin": [uid]},
    })
    # condivisa: visibile a tutti (no smistamento)
    cond_total = await db.email_inbox.count_documents({"categoria": "condivisa"})
    cond_unread = await db.email_inbox.count_documents({
        "categoria": "condivisa", "letta_da": {"$nin": [uid]},
    })
    # totale globale (admin only)
    globale = await db.email_inbox.count_documents({}) if is_admin else None
    return {
        "personale": {"totale": pers_total, "non_lette": pers_unread},
        "condivisa": {"totale": cond_total, "non_lette": cond_unread},
        "globale": globale,
    }


@api.get("/email/inbox")
async def email_inbox_list(
    categoria: Optional[str] = None,  # "personale" | "condivisa"
    q: Optional[str] = None,
    limit: int = 500,
    user=Depends(current_user),
):
    uid = user["id"]
    flt: dict = {}
    if categoria == "personale":
        flt["smistato_a"] = uid
    elif categoria == "condivisa":
        flt["categoria"] = "condivisa"
    else:
        # default: tutte le email visibili all'utente
        flt["$or"] = [{"smistato_a": uid}, {"categoria": "condivisa"}]
    if q:
        flt["$and"] = [{
            "$or": [
                {"subject": {"$regex": q, "$options": "i"}},
                {"from_address": {"$regex": q, "$options": "i"}},
                {"from_name": {"$regex": q, "$options": "i"}},
            ],
        }]
    items = await db.email_inbox.find(flt, {"_id": 0, "body_html": 0}).sort("date", -1).limit(limit).to_list(limit)
    # marca "non letta" per l'utente loggato
    for it in items:
        it["non_letta"] = uid not in (it.get("letta_da") or [])
    return items


@api.get("/email/inbox/{eid}")
async def email_inbox_detail(eid: str, user=Depends(current_user)):
    rec = await db.email_inbox.find_one({"id": eid}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Email non trovata")
    # autorizzazione: admin OR (smistato_a OR condivisa)
    uid = user["id"]
    if user.get("role") != "admin":
        if uid not in (rec.get("smistato_a") or []) and rec.get("categoria") != "condivisa":
            raise HTTPException(403, "Email non accessibile")
    rec["non_letta"] = uid not in (rec.get("letta_da") or [])
    return rec


@api.post("/email/inbox/{eid}/leggi")
async def email_inbox_mark_read(eid: str, user=Depends(current_user)):
    rec = await db.email_inbox.find_one({"id": eid}, {"_id": 0, "letta_da": 1})
    if not rec:
        raise HTTPException(404, "Email non trovata")
    await db.email_inbox.update_one(
        {"id": eid},
        {"$addToSet": {"letta_da": user["id"]}},
    )
    return {"ok": True}


@api.delete("/email/inbox/{eid}")
async def delete_email_inbox(eid: str, user=Depends(current_user)):
    """Elimina un'email dall'inbox locale (non agisce su Gmail/IMAP).

    Cancella anche gli allegati associati nello storage.
    """
    em = await db.email_inbox.find_one({"id": eid}, {"_id": 0})
    if not em:
        raise HTTPException(404, "Email non trovata")
    # ACL: solo admin o destinatario personale
    if user["role"] != "admin":
        if user["id"] not in (em.get("smistato_a") or []):
            raise HTTPException(403, "Permesso negato")
    # rimuovi allegati dallo storage (best-effort)
    for att in (em.get("attachments") or []):
        sp = att.get("storage_path")
        if sp:
            try:
                obj_storage.delete_object(sp)
            except Exception:
                pass
    await db.email_inbox.delete_one({"id": eid})
    return {"ok": True}


# ============================================================
# IMAP POLLER — controlli scheduler + smistamento manuale
# ============================================================
@api.get("/email/poller/status")
async def imap_poller_status(user=Depends(require_user("admin"))):
    az = await db.azienda_config.find_one(
        {}, {"_id": 0, "imap_poller_enabled": 1, "imap_poller_minutes": 1,
             "imap_poller_last_run": 1, "imap_poller_last_uid": 1,
             "imap_host": 1, "imap_user": 1},
    ) or {}
    return {
        "running": imap_poller.is_running(),
        "enabled": bool(az.get("imap_poller_enabled")),
        "minutes": int(az.get("imap_poller_minutes") or 5),
        "last_run": az.get("imap_poller_last_run"),
        "last_uid": az.get("imap_poller_last_uid"),
        "imap_host": az.get("imap_host"),
        "imap_user": az.get("imap_user"),
    }


@api.post("/email/poller/start")
async def imap_poller_start(
    body: Optional[dict] = None, user=Depends(require_user("admin")),
):
    body = body or {}
    minutes = int(body.get("minutes") or 5)
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    if not (az.get("imap_host") and az.get("imap_user") and az.get("imap_password")):
        raise HTTPException(400, "IMAP non configurato (host/user/password)")
    imap_poller.start_scheduler(db, minutes=minutes)
    if az.get("id"):
        await db.azienda_config.update_one(
            {"id": az["id"]},
            {"$set": {"imap_poller_enabled": True, "imap_poller_minutes": minutes}},
        )
    return {"ok": True, "running": True, "minutes": minutes}


@api.post("/email/poller/stop")
async def imap_poller_stop(user=Depends(require_user("admin"))):
    imap_poller.stop_scheduler()
    az = await db.azienda_config.find_one({}, {"_id": 0, "id": 1}) or {}
    if az.get("id"):
        await db.azienda_config.update_one(
            {"id": az["id"]}, {"$set": {"imap_poller_enabled": False}},
        )
    return {"ok": True, "running": False}


@api.post("/email/poller/run-now")
async def imap_poller_run_now(user=Depends(require_user("admin"))):
    """Esegue un singolo ciclo di polling on-demand (utile per test)."""
    res = await imap_poller.poll_once(db)
    return res


@api.post("/email/poller/backfill-diario")
async def imap_backfill_diario(user=Depends(require_user("admin"))):
    """Crea le voci di diario mancanti per le email già ricevute.

    Itera su tutte le ``email_inbox`` con ``anagrafica_id`` valorizzato e
    inserisce una voce di tipo ``email_in`` in ``db.diario`` se non esiste già.
    Idempotente.
    """
    from shared import log_diario_cliente
    cursor = db.email_inbox.find(
        {"anagrafica_id": {"$ne": None}},
        {"_id": 0, "id": 1, "anagrafica_id": 1, "from_address": 1,
         "from_name": 1, "subject": 1, "body_text": 1, "date": 1},
    )
    creati = 0
    saltati = 0
    async for em in cursor:
        titolo = f"Email ricevuta da {em.get('from_name') or em.get('from_address')}: {em.get('subject') or '(no subject)'}"[:200]
        # idempotenza: salta se già presente una voce con stesso titolo+anag
        already = await db.diario.find_one(
            {"anagrafica_id": em["anagrafica_id"], "titolo": titolo}, {"_id": 0, "id": 1},
        )
        if already:
            saltati += 1
            continue
        await log_diario_cliente(
            em["anagrafica_id"], "email", titolo,
            (em.get("body_text") or "")[:2000], autore=None,
        )
        creati += 1
    return {"ok": True, "creati": creati, "saltati": saltati}


# ============================================================
# AVVISI — Generazione PDF (usa template "pdf_avviso" da Gestioni Modelli)
# ============================================================
@api.post("/avvisi/pdf")
async def avvisi_genera_pdf(body: dict, user=Depends(current_user)):
    """Genera il PDF avviso di scadenza per un contraente.

    Body:
      - ``contraente_id`` (str)
      - ``titoli_ids`` (list[str]) — IDs dei titoli da includere
      - ``template_id`` (str, opzionale) — usa il modello default se assente
    """
    contraente_id = body.get("contraente_id")
    titoli_ids = body.get("titoli_ids") or []
    template_id = body.get("template_id")
    if not contraente_id or not titoli_ids:
        raise HTTPException(400, "contraente_id e titoli_ids obbligatori")

    contraente = await db.anagrafiche.find_one({"id": contraente_id}, {"_id": 0})
    if not contraente:
        raise HTTPException(404, "Contraente non trovato")

    azienda = await db.azienda_config.find_one({}, {"_id": 0}) or {}

    # Carica titoli + arricchisci con polizza
    titoli = await db.titoli.find(
        {"id": {"$in": titoli_ids}}, {"_id": 0},
    ).to_list(500)
    polizze_ids = list({t.get("polizza_id") for t in titoli if t.get("polizza_id")})
    polizze = await db.polizze.find(
        {"id": {"$in": polizze_ids}}, {"_id": 0},
    ).to_list(500)
    pmap = {p["id"]: p for p in polizze}

    # Lookup nomi prodotti (campo Polizza.prodotto è un FK id verso db.prodotti)
    prodotti_ids = list({p.get("prodotto") for p in polizze if p.get("prodotto")})
    prodotti_map: dict = {}
    if prodotti_ids:
        async for pr in db.prodotti.find(
            {"id": {"$in": prodotti_ids}}, {"_id": 0, "id": 1, "nome": 1, "ramo": 1},
        ):
            prodotti_map[pr["id"]] = pr.get("nome") or pr.get("ramo") or ""

    righe = []
    for t in titoli:
        p = pmap.get(t.get("polizza_id"), {})
        # Risolvi nome prodotto: prima da prodotti_map, poi fallback su ramo
        prodotto_raw = p.get("prodotto") or ""
        rischio = prodotti_map.get(prodotto_raw) or prodotto_raw or p.get("ramo") or ""
        # Se ancora sembra un UUID, fallback finale su ramo
        if rischio and len(rischio) >= 32 and "-" in rischio and " " not in rischio:
            rischio = p.get("ramo") or ""
        # Rata del = data scadenza del titolo (campo .scadenza in Titolo)
        rata_del = (
            t.get("scadenza")
            or t.get("data_scadenza")
            or t.get("data_decorrenza")
        )
        righe.append({
            "numero_contratto": p.get("numero_polizza") or "",
            "rischio": rischio[:40],
            "targa": p.get("targa") or "",
            "rata_del": rata_del,
            "importo": float(t.get("importo_lordo") or 0),
        })

    # Carica template
    template: Optional[dict] = None
    if template_id:
        template = await db.template_modelli.find_one({"id": template_id}, {"_id": 0})
    if not template:
        from routes.modelli import get_default_template
        template = await get_default_template("pdf_avviso")

    dest = {
        "ragione_sociale": contraente.get("ragione_sociale") or "",
        "indirizzo": contraente.get("indirizzo") or "",
        "cap": contraente.get("cap") or "",
        "comune": contraente.get("comune") or "",
        "provincia": contraente.get("provincia") or "",
        "codice_fiscale": contraente.get("codice_fiscale") or "",
    }

    pdf_bytes = pdf_avviso.generate_avviso_pdf(
        azienda=azienda, destinatario=dest, righe=righe, template=template,
    )
    nome_file = f"avviso_{(contraente.get('ragione_sociale') or 'cliente').replace(' ', '_')[:40]}.pdf"
    return StreamingResponse(
        _io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{nome_file}"'},
    )


# ============================================================
# WHATSAPP DISPATCH — wa.me (click-to-chat) OR Evolution API (server invio)
# ============================================================
@api.post("/comunicazioni/whatsapp/invia")
async def whatsapp_invia(body: dict, user=Depends(current_user)):
    """Invia messaggio WhatsApp.

    Body: ``{numero, messaggio, anagrafica_id?, polizza_id?, provider?, instance_name?}``.

    Provider supportati:
    - ``wame`` (default): ritorna link ``https://wa.me/…`` da aprire nel browser
    - ``evolution``: invio server-side tramite Evolution API (multi-agenzia)
    """
    numero = (body.get("numero") or "").strip()
    messaggio = (body.get("messaggio") or "").strip()
    if not numero or not messaggio:
        raise HTTPException(400, "numero e messaggio obbligatori")
    az = await db.azienda_config.find_one({}, {"_id": 0}) or {}
    provider = (body.get("provider") or az.get("whatsapp_provider") or "wame").lower()

    cleaned = re.sub(r"[^\d+]", "", numero)
    if cleaned.startswith("00"):
        cleaned = "+" + cleaned[2:]
    if not cleaned.startswith("+"):
        cleaned = "+39" + cleaned  # default Italia

    if provider == "wame":
        from urllib.parse import quote
        wa_number = cleaned.lstrip("+")
        url = f"https://wa.me/{wa_number}?text={quote(messaggio)}"
        await db.storico_avvisi.insert_one({
            "id": _uid_local(),
            "canale": "whatsapp",
            "provider": "wame",
            "destinatario": cleaned,
            "messaggio": messaggio[:2000],
            "anagrafica_id": body.get("anagrafica_id"),
            "polizza_id": body.get("polizza_id"),
            "utente_id": user["id"],
            "created_at": _now_iso(),
        })
        return {"ok": True, "provider": "wame", "url": url}

    if provider == "evolution":
        # Sceglie l'istanza: esplicita nel body oppure la prima "open"
        instance_name = body.get("instance_name")
        if instance_name:
            inst = await db.whatsapp_instances.find_one(
                {"instance_name": instance_name}, {"_id": 0, "token": 1},
            )
        else:
            inst = (await db.whatsapp_instances.find_one(
                {"state": {"$in": ["open", "connected"]}}, {"_id": 0, "instance_name": 1, "token": 1},
            ) or await db.whatsapp_instances.find_one({}, {"_id": 0, "instance_name": 1, "token": 1}))
            if inst:
                instance_name = inst.get("instance_name")
        if not inst or not instance_name:
            raise HTTPException(400, "Nessuna istanza WhatsApp configurata (vai in WhatsApp Agenzie)")

        url_base = (os.environ.get("WHATSAPP_API_URL") or "").rstrip("/")
        api_key = inst.get("token") or os.environ.get("WHATSAPP_API_KEY") or ""
        if not url_base or not api_key:
            raise HTTPException(400, "Evolution API non configurata")

        import httpx
        to_clean = cleaned.lstrip("+")
        try:
            async with httpx.AsyncClient(timeout=20) as http:
                resp = await http.post(
                    f"{url_base}/message/sendText/{instance_name}",
                    headers={"apikey": api_key, "Content-Type": "application/json"},
                    json={"number": to_clean, "text": messaggio[:2000]},
                )
            if resp.status_code >= 400:
                raise HTTPException(503, f"Errore Evolution API ({resp.status_code}): {resp.text[:300]}")
            provider_id = ((resp.json() or {}).get("key") or {}).get("id")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(503, f"Errore Evolution API: {e}")
        await db.storico_avvisi.insert_one({
            "id": _uid_local(),
            "canale": "whatsapp",
            "provider": "evolution",
            "instance_name": instance_name,
            "provider_id": provider_id,
            "destinatario": cleaned,
            "messaggio": messaggio[:2000],
            "anagrafica_id": body.get("anagrafica_id"),
            "polizza_id": body.get("polizza_id"),
            "utente_id": user["id"],
            "created_at": _now_iso(),
        })
        # traccia anche in inbox
        try:
            await db.whatsapp_messages.insert_one({
                "id": _uid_local(),
                "instance_name": instance_name,
                "direction": "out",
                "number": to_clean,
                "text": messaggio[:2000],
                "created_at": _now_iso(),
                "anagrafica_id": body.get("anagrafica_id"),
                "polizza_id": body.get("polizza_id"),
            })
        except Exception:
            pass
        return {"ok": True, "provider": "evolution", "instance_name": instance_name, "provider_id": provider_id}

    raise HTTPException(400, f"Provider non supportato: {provider} (usa 'wame' o 'evolution')")


def _uid_local() -> str:
    return str(uuid.uuid4())


# ============================================================
# DIARIO COLLABORATORE (note + aggregato comunicazioni inviate + chat)
# ============================================================
@api.get("/diario")
async def list_diario(
    user_id: Optional[str] = None,
    tipo: Optional[str] = None,       # filtra: note|email|sms|whatsapp|chat
    q: Optional[str] = None,
    limit: int = 500,
    user=Depends(current_user),
):
    """Restituisce il feed cronologico (desc) del diario personale dell'utente.

    Aggrega 3 sorgenti:
      • `db.diario_note`         -> tipo='nota'
      • `db.storico_avvisi`      -> tipo='email'|'sms'|'whatsapp' (utente_id=user)
      • `db.chat`                -> tipo='chat' (mittente_id=user)

    Solo admin può specificare `user_id` di un altro utente, altrimenti viene
    forzato l'id dell'utente loggato.
    """
    uid = user_id if (user.get("role") == "admin" and user_id) else user["id"]
    items: list[dict] = []

    if not tipo or tipo == "nota":
        flt: dict = {"user_id": uid}
        if q:
            flt["$or"] = [
                {"titolo": {"$regex": q, "$options": "i"}},
                {"contenuto": {"$regex": q, "$options": "i"}},
            ]
        async for n in db.diario_note.find(flt, {"_id": 0}).sort("created_at", -1):
            items.append({
                "id": n["id"], "tipo": "nota",
                "at": n.get("created_at"),
                "titolo": n.get("titolo"),
                "contenuto": n.get("contenuto"),
                "tags": n.get("tags") or [],
                "anagrafica_id": n.get("anagrafica_id"),
                "polizza_id": n.get("polizza_id"),
            })

    if not tipo or tipo in ("email", "sms", "whatsapp"):
        flt_sa: dict = {"utente_id": uid}
        if tipo in ("email", "sms", "whatsapp"):
            flt_sa["canale"] = tipo
        if q:
            flt_sa["$or"] = [
                {"soggetto": {"$regex": q, "$options": "i"}},
                {"destinatario": {"$regex": q, "$options": "i"}},
                {"contraente_nome": {"$regex": q, "$options": "i"}},
            ]
        async for r in db.storico_avvisi.find(flt_sa, {"_id": 0}).sort("sent_at", -1).limit(limit):
            items.append({
                "id": r["id"], "tipo": r.get("canale") or "email",
                "at": r.get("sent_at"),
                "titolo": r.get("soggetto") or f"Invio {r.get('canale')}",
                "contenuto": (
                    f"A: {r.get('destinatario')}"
                    + (f" — {r.get('contraente_nome')}" if r.get('contraente_nome') else "")
                ),
                "stato": r.get("stato"),
                "anagrafica_id": r.get("contraente_id"),
                "polizza_id": r.get("polizza_id"),
            })

    if not tipo or tipo == "chat":
        flt_ch: dict = {"mittente_id": uid}
        if q:
            flt_ch["testo"] = {"$regex": q, "$options": "i"}
        async for c in db.chat.find(flt_ch, {"_id": 0}).sort("at", -1).limit(limit):
            items.append({
                "id": c["id"], "tipo": "chat",
                "at": c.get("at"),
                "titolo": f"Chat -> {c.get('destinatario_nome') or c.get('destinatario_id') or '?'}",
                "contenuto": c.get("testo"),
            })

    # ordina cronologicamente desc (string ISO confronto lessicografico OK)
    items.sort(key=lambda x: x.get("at") or "", reverse=True)
    return items[:limit]


@api.post("/diario", status_code=201)
async def crea_nota_diario(body: dict, user=Depends(current_user)):
    titolo = (body.get("titolo") or "").strip()
    if not titolo:
        raise HTTPException(400, "Titolo obbligatorio")
    nota = DiarioNota(
        user_id=user["id"],
        titolo=titolo,
        contenuto=body.get("contenuto"),
        anagrafica_id=body.get("anagrafica_id"),
        polizza_id=body.get("polizza_id"),
        tags=body.get("tags") or [],
    )
    await db.diario_note.insert_one(nota.model_dump())
    return nota.model_dump()


@api.delete("/diario/{nid}")
async def elimina_nota_diario(nid: str, user=Depends(current_user)):
    flt: dict = {"id": nid}
    if user.get("role") != "admin":
        flt["user_id"] = user["id"]  # un utente può eliminare solo le proprie
    res = await db.diario_note.delete_one(flt)
    if res.deleted_count == 0:
        raise HTTPException(404, "Nota non trovata")
    return {"ok": True}


# ============================================================
# DIARIO CLIENTE
# ============================================================
@api.get("/anagrafiche/{aid}/riepilogo")
async def riepilogo_cliente(aid: str, user=Depends(current_user)):
    """Riepilogo economico: premi lordi pagati + provvigioni incassate.

    Se l'anagrafica ha relazioni con persone giuridiche (aziende), aggrega anche
    quelle ma le riporta separatamente. Risultato:
    { privato: {premi, provvigioni, polizze_count, titoli_count},
      azienda: {premi, provvigioni, ...},
      totale: {...} }
    """
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")

    # raccogli id privati e aziende collegate
    privati = [aid] if ana.get("tipo") != "persona_giuridica" else []
    aziende = [aid] if ana.get("tipo") == "persona_giuridica" else []
    for rel in (ana.get("parente_di") or []):
        rid = rel.get("anagrafica_id")
        if not rid: continue
        rel_doc = await db.anagrafiche.find_one({"id": rid}, {"_id": 0, "tipo": 1})
        if rel_doc:
            if rel_doc.get("tipo") == "persona_giuridica":
                aziende.append(rid)
            else:
                privati.append(rid)

    async def aggrega(ids: list) -> dict:
        if not ids:
            return {"premi_lordi": 0.0, "provvigioni": 0.0, "polizze_count": 0,
                    "titoli_incassati": 0, "titoli_aperti": 0}
        pol_ids = [p["id"] async for p in db.polizze.find(
            {"contraente_id": {"$in": ids}}, {"_id": 0, "id": 1})]
        if not pol_ids:
            return {"premi_lordi": 0.0, "provvigioni": 0.0,
                    "polizze_count": 0, "titoli_incassati": 0, "titoli_aperti": 0}
        # somma titoli incassati
        agg = await db.titoli.aggregate([
            {"$match": {"polizza_id": {"$in": pol_ids}, "stato": "incassato"}},
            {"$group": {"_id": None,
                        "premi": {"$sum": "$importo_lordo"},
                        "provv": {"$sum": "$provvigioni"},
                        "n": {"$sum": 1}}},
        ]).to_list(1)
        n_aperti = await db.titoli.count_documents({"polizza_id": {"$in": pol_ids},
                                                     "stato": {"$in": ["da_incassare", "insoluto"]}})
        return {
            "premi_lordi": round((agg[0]["premi"] if agg else 0.0), 2),
            "provvigioni": round((agg[0]["provv"] if agg else 0.0), 2),
            "polizze_count": len(pol_ids),
            "titoli_incassati": agg[0]["n"] if agg else 0,
            "titoli_aperti": n_aperti,
        }

    priv = await aggrega(privati)
    az = await aggrega(aziende)
    tot = {
        "premi_lordi": round(priv["premi_lordi"] + az["premi_lordi"], 2),
        "provvigioni": round(priv["provvigioni"] + az["provvigioni"], 2),
        "polizze_count": priv["polizze_count"] + az["polizze_count"],
        "titoli_incassati": priv["titoli_incassati"] + az["titoli_incassati"],
        "titoli_aperti": priv["titoli_aperti"] + az["titoli_aperti"],
    }
    return {"anagrafica": ana, "privato": priv, "azienda": az, "totale": tot,
            "ids_privati": privati, "ids_aziende": aziende}


@api.post("/anagrafiche/tags/auto-genera")
async def auto_genera_tags(user=Depends(require_user("admin", "collaboratore"))):
    """Genera tag automatici per tutte le anagrafiche.

    Tag generati:
      - genitore_con_figli_minori (se ha relazioni 'figlio' con anagrafiche con data_nascita < 18 anni fa)
      - figli_minori (alias, su richiesta utente)
      - gen_anni_XX
      - cliente_attivo / prospect / top_cliente
      - dipendente / autonomo / professionista / imprenditore / pensionato (da tipologia_lavoratore)
    """
    from datetime import date
    today = date.today()
    AUTO_TAGS_GESTITI = {
        "genitore_con_figli_minori", "figli_minori",
        "cliente_attivo", "prospect", "top_cliente",
        "anziano_over_65", "giovane_under_30",
        "dipendente", "autonomo", "professionista", "imprenditore",
        "pensionato", "disoccupato", "studente", "casalinga",
    }
    aggiornate = 0
    cursor = db.anagrafiche.find({}, {"_id": 0})
    async for ana in cursor:
        existing_manual = [t for t in (ana.get("tags") or [])
                           if not (t.startswith("gen_") or t in AUTO_TAGS_GESTITI)]
        auto = []
        # generazione
        if ana.get("data_nascita"):
            try:
                d = date.fromisoformat(ana["data_nascita"])
                eta = today.year - d.year - (1 if (today.month, today.day) < (d.month, d.day) else 0)
                decade = (d.year // 10) * 10
                auto.append(f"gen_anni_{str(decade)[-2:]}")
                if eta >= 65: auto.append("anziano_over_65")
                if eta < 30: auto.append("giovane_under_30")
            except Exception:
                pass
        # genitore con figli minori (genera entrambi i tag: completo + breve)
        for rel in (ana.get("parente_di") or []):
            if rel.get("relazione") != "figlio": continue
            child = await db.anagrafiche.find_one({"id": rel.get("anagrafica_id")}, {"_id": 0, "data_nascita": 1})
            if not child or not child.get("data_nascita"): continue
            try:
                cd = date.fromisoformat(child["data_nascita"])
                eta_f = today.year - cd.year
                if eta_f < 18:
                    auto.append("genitore_con_figli_minori")
                    auto.append("figli_minori")
                    break
            except Exception:
                continue
        # tipologia lavoratore
        tl = ana.get("tipologia_lavoratore")
        if tl:
            auto.append(tl)
        # stato cliente
        n_attive = await db.polizze.count_documents({"contraente_id": ana["id"], "stato": "attiva"})
        if n_attive == 0:
            auto.append("prospect")
        elif n_attive >= 4:
            auto.append("top_cliente"); auto.append("cliente_attivo")
        else:
            auto.append("cliente_attivo")
        nuovi_tags = sorted(set(existing_manual + auto))
        await db.anagrafiche.update_one({"id": ana["id"]}, {"$set": {"tags": nuovi_tags, "updated_at": _now_iso()}})
        aggiornate += 1
    await log_attivita(user, "auto_tag", "anagrafica", None, f"Tag auto generati su {aggiornate} anagrafiche")
    return {"aggiornate": aggiornate}


@api.get("/anagrafiche/tags/elenco")
async def elenco_tags(user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Lista tag distinti con conteggio."""
    pipeline = [
        {"$unwind": "$tags"},
        {"$group": {"_id": "$tags", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    res = await db.anagrafiche.aggregate(pipeline).to_list(500)
    return [{"tag": r["_id"], "count": r["count"]} for r in res]


@api.post("/newsletter/invia")
async def newsletter_invia(body: dict, user=Depends(require_user("admin", "collaboratore"))):
    """Crea email in coda per tutti i clienti che matchano i tag (any-of).

    body: { tags: [...], oggetto, corpo }
    """
    tags = body.get("tags") or []
    oggetto = body.get("oggetto", "").strip()
    corpo = body.get("corpo", "").strip()
    if not oggetto or not corpo:
        raise HTTPException(400, "Oggetto e corpo richiesti")
    flt = {"tags": {"$in": tags}} if tags else {}
    flt["email"] = {"$nin": ["", None]}
    targets = await db.anagrafiche.find(flt, {"_id": 0, "id": 1, "email": 1, "ragione_sociale": 1}).to_list(10000)
    creati = 0
    for t in targets:
        if not t.get("email"): continue
        e = EmailMessaggio(
            destinatario_anagrafica_id=t["id"], destinatario_email=t["email"],
            oggetto=oggetto, corpo=corpo.replace("{{nome}}", t.get("ragione_sociale", "")),
            template="newsletter", stato="in_coda", autore_id=user["id"],
        )
        await db.email.insert_one(e.model_dump())
        creati += 1
    await log_attivita(user, "newsletter", "email", None,
                       f"Newsletter su tags {tags}: {creati} email create")
    return {"email_create": creati, "tags": tags}


@api.get("/anagrafiche/{aid}/diario")
async def list_diario(aid: str, user=Depends(current_user)):
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    items = await db.diario.find({"anagrafica_id": aid}, {"_id": 0}).sort("data_evento", -1).to_list(500)
    return items


@api.post("/anagrafiche/{aid}/diario", status_code=201)
async def create_diario(aid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body["anagrafica_id"] = aid
    body["autore_id"] = user["id"]
    body["autore_nome"] = user.get("name")
    obj = DiarioVoce(**body)
    await db.diario.insert_one(obj.model_dump())
    await log_attivita(user, "create", "diario", obj.id, obj.titolo)
    return obj.model_dump()


@api.put("/diario/{did}")
async def update_diario(did: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body["updated_at"] = _now_iso()
    res = await db.diario.update_one({"id": did}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Voce non trovata")
    return strip_mongo_id(await db.diario.find_one({"id": did}, {"_id": 0}))


@api.delete("/diario/{did}")
async def delete_diario(did: str, user=Depends(require_user("admin", "collaboratore"))):
    await db.diario.delete_one({"id": did})
    return {"ok": True}


# ============================================================
# ALLEGATI (object storage)
# ============================================================
@api.get("/allegati")
async def list_allegati(
    entita_tipo: str, entita_id: str,
    applicazione_matricola_id: Optional[str] = None,
    user=Depends(current_user),
):
    flt: dict = {"entita_tipo": entita_tipo, "entita_id": entita_id, "is_deleted": False}
    if applicazione_matricola_id:
        flt["applicazione_matricola_id"] = applicazione_matricola_id
    items = await db.allegati.find(
        flt,
        {"_id": 0},
    ).sort("created_at", -1).to_list(200)
    return items


@api.post("/allegati", status_code=201)
async def upload_allegato(
    entita_tipo: str = Query(...),
    entita_id: str = Query(...),
    descrizione: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    visibile_cliente: bool = Query(False),
    applicazione_matricola_id: Optional[str] = Query(None),
    file: UploadFile = File(...),
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    if entita_tipo not in ("anagrafica", "polizza", "sinistro", "compagnia", "corso", "movimento", "titolo"):
        raise HTTPException(400, "Tipo entità non valido")
    data = await file.read()
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(400, "File troppo grande (max 25 MB)")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
    file_id = _uid()
    path = f"{os.environ.get('APP_NAME', 'assicura')}/{entita_tipo}/{entita_id}/{file_id}.{ext}"
    content_type = file.content_type or obj_storage.mime_for(file.filename or "")
    try:
        result = obj_storage.put_object(path, data, content_type)
    except Exception as e:
        raise HTTPException(503, f"Errore upload: {e}")
    obj = Allegato(
        entita_tipo=entita_tipo, entita_id=entita_id,
        nome_file=file.filename, storage_path=result["path"],
        content_type=content_type, size=result.get("size", len(data)),
        descrizione=descrizione, autore_id=user["id"],
        categoria=categoria, visibile_cliente=visibile_cliente,
        applicazione_matricola_id=applicazione_matricola_id,
    )
    await db.allegati.insert_one(obj.model_dump())
    await log_attivita(user, "upload", "allegato", obj.id, f"{file.filename} su {entita_tipo}:{entita_id}")
    # auto-traccia nel diario se l'allegato è su anagrafica/polizza/sinistro di un cliente
    anag_id_for_diario = None
    if entita_tipo == "anagrafica":
        anag_id_for_diario = entita_id
    elif entita_tipo in ("polizza", "sinistro"):
        coll = db.polizze if entita_tipo == "polizza" else db.sinistri
        target = await coll.find_one({"id": entita_id}, {"_id": 0, "contraente_id": 1})
        if target:
            anag_id_for_diario = target.get("contraente_id")
    if anag_id_for_diario:
        await log_diario_cliente(
            anag_id_for_diario, "documento_inviato" if False else "altro",
            titolo=f"📎 Documento caricato: {file.filename}",
            descrizione=f"Allegato a {entita_tipo} (size: {result.get('size', len(data))} bytes)" + (f" - {descrizione}" if descrizione else ""),
            autore=user,
        )
    return obj.model_dump()


@api.get("/allegati/{aid}/download")
async def download_allegato(aid: str, user=Depends(current_user)):
    rec = await db.allegati.find_one({"id": aid, "is_deleted": False}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Allegato non trovato")
    # ACL su cliente: vede solo allegati relativi alla sua anagrafica/polizze/sinistri
    if user["role"] == "cliente":
        ok = False
        if rec["entita_tipo"] == "anagrafica" and rec["entita_id"] == user.get("anagrafica_id"):
            ok = True
        elif rec["entita_tipo"] in ("polizza", "sinistro"):
            target = await db.polizze.find_one({"id": rec["entita_id"]}, {"_id": 0, "contraente_id": 1}) \
                if rec["entita_tipo"] == "polizza" else \
                await db.sinistri.find_one({"id": rec["entita_id"]}, {"_id": 0, "contraente_id": 1})
            ok = target and target.get("contraente_id") == user.get("anagrafica_id")
        if not ok:
            raise HTTPException(403, "Permesso negato")
    try:
        data, ctype = obj_storage.get_object(rec["storage_path"])
    except Exception as e:
        raise HTTPException(503, f"Errore download: {e}")
    return StreamingResponse(
        _io.BytesIO(data),
        media_type=rec.get("content_type") or ctype,
        headers={"Content-Disposition": f'inline; filename="{rec["nome_file"]}"'},
    )


@api.delete("/allegati/{aid}")
async def delete_allegato(aid: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    res = await db.allegati.update_one({"id": aid}, {"$set": {"is_deleted": True, "updated_at": _now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(404, "Allegato non trovato")
    await log_attivita(user, "delete", "allegato", aid)
    return {"ok": True}


@api.patch("/allegati/{aid}/visibilita")
async def toggle_allegato_visibilita(
    aid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Toggle visibile_cliente di un allegato (per spostarlo tra 'visibile' e 'interno')."""
    new_val = bool(body.get("visibile_cliente", False))
    res = await db.allegati.update_one(
        {"id": aid}, {"$set": {"visibile_cliente": new_val, "updated_at": _now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Allegato non trovato")
    return {"ok": True, "visibile_cliente": new_val}


@api.patch("/allegati/{aid}")
async def patch_allegato(
    aid: str, body: dict,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Aggiorna campi descrittivi (categoria, descrizione, visibile_cliente, applicazione_matricola_id)."""
    allowed = {"categoria", "descrizione", "visibile_cliente", "applicazione_matricola_id"}
    upd = {k: v for k, v in body.items() if k in allowed}
    if not upd:
        raise HTTPException(400, "Nessun campo aggiornabile fornito")
    upd["updated_at"] = _now_iso()
    res = await db.allegati.update_one({"id": aid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Allegato non trovato")
    return await db.allegati.find_one({"id": aid}, {"_id": 0})


# ============================================================
# STORAGE GENERICO (logo azienda, documenti utente, attestati)
# Path canonico: /api/storage/{full_path}
# ============================================================
@api.get("/storage/{full_path:path}")
async def serve_storage(full_path: str, user=Depends(current_user)):
    """Serve file dallo storage. ACL:
    - file in /users/{uid}/...   -> solo admin o proprietario
    - file in /azienda/...       -> qualunque utente autenticato (logo per stampe)
    - file in /titoli/... ecc.   -> qualunque utente autenticato (gli allegati hanno ACL specifica)
    """
    # ACL sui documenti utente
    parts = full_path.split("/")
    if "users" in parts:
        try:
            idx = parts.index("users")
            owner_uid = parts[idx + 1]
        except (ValueError, IndexError):
            owner_uid = None
        if user["role"] != "admin" and user["id"] != owner_uid:
            raise HTTPException(403, "Permesso negato")
    try:
        data, ctype = obj_storage.get_object(full_path)
    except Exception as e:
        raise HTTPException(404, f"File non trovato: {e}")
    filename = full_path.rsplit("/", 1)[-1]
    return StreamingResponse(
        _io.BytesIO(data),
        media_type=ctype or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


# ============================================================
# CHAT INTERNA
# ============================================================
@api.get("/chat/utenti")
async def chat_utenti(user=Depends(current_user)):
    """Restituisce gli utenti con cui posso chattare."""
    if user["role"] == "cliente":
        flt = {"role": {"$in": ["admin", "collaboratore", "dipendente"]}}
    else:
        flt = {"id": {"$ne": user["id"]}}
    users = await db.users.find(flt, {"_id": 0, "password_hash": 0}).sort("name", 1).to_list(500)
    # arricchisci con count non letti e ultimo messaggio
    for u in users:
        unread = await db.chat.count_documents(
            {"destinatario_id": user["id"], "mittente_id": u["id"], "letto": False}
        )
        u["unread"] = unread
        last = await db.chat.find_one(
            {"$or": [
                {"mittente_id": user["id"], "destinatario_id": u["id"]},
                {"mittente_id": u["id"], "destinatario_id": user["id"]},
            ]},
            {"_id": 0, "testo": 1, "created_at": 1},
            sort=[("created_at", -1)],
        )
        u["ultimo_messaggio"] = last
    return users


@api.get("/chat/messaggi")
async def chat_messaggi(con: str, limit: int = 200, user=Depends(current_user)):
    """Messaggi della conversazione con un altro utente."""
    items = await db.chat.find(
        {"$or": [
            {"mittente_id": user["id"], "destinatario_id": con},
            {"mittente_id": con, "destinatario_id": user["id"]},
        ]},
        {"_id": 0},
    ).sort("created_at", 1).to_list(limit)
    # marca come letti i messaggi a me
    await db.chat.update_many(
        {"mittente_id": con, "destinatario_id": user["id"], "letto": False},
        {"$set": {"letto": True, "letto_at": _now_iso()}},
    )
    return items


@api.post("/chat/messaggi", status_code=201)
async def chat_invia(
    destinatario_id: str = Query(None),
    testo: str = Query(""),
    file: Optional[UploadFile] = File(None),
    body: Optional[dict] = None,
    user=Depends(current_user),
):
    # supporta sia JSON (body) che multipart (con file)
    if body is None:
        try:
            from fastapi import Request as _Req
            pass
        except Exception:
            pass
    # multipart: prendi destinatario_id e testo dai query/form param
    dest_id = destinatario_id or (body or {}).get("destinatario_id")
    txt = testo or (body or {}).get("testo", "")
    if not dest_id:
        raise HTTPException(400, "destinatario_id richiesto")
    if not txt.strip() and not file:
        raise HTTPException(400, "Testo o allegato richiesto")
    dest = await db.users.find_one({"id": dest_id}, {"_id": 0})
    if not dest:
        raise HTTPException(404, "Destinatario non trovato")
    if user["role"] == "cliente" and dest.get("role") == "cliente":
        raise HTTPException(403, "I clienti possono scrivere solo allo staff")

    allegato_id = allegato_nome = allegato_ct = None
    if file:
        data = await file.read()
        if len(data) > 25 * 1024 * 1024:
            raise HTTPException(400, "File troppo grande (max 25 MB)")
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
        file_id = _uid()
        path = f"{os.environ.get('APP_NAME', 'assicura')}/chat/{user['id']}/{file_id}.{ext}"
        ct = file.content_type or obj_storage.mime_for(file.filename or "")
        try:
            result = obj_storage.put_object(path, data, ct)
        except Exception as e:
            raise HTTPException(503, f"Errore upload: {e}")
        al = Allegato(
            entita_tipo="anagrafica", entita_id=dest_id,  # uso anagrafica come fallback
            nome_file=file.filename, storage_path=result["path"],
            content_type=ct, size=result.get("size", len(data)),
            descrizione=f"Allegato chat tra {user.get('name')} e {dest.get('name')}",
            autore_id=user["id"],
        )
        await db.allegati.insert_one(al.model_dump())
        allegato_id = al.id
        allegato_nome = file.filename
        allegato_ct = ct

    msg = MessaggioChat(
        mittente_id=user["id"], mittente_nome=user.get("name", ""),
        destinatario_id=dest_id, destinatario_nome=dest.get("name", ""),
        testo=txt, allegato_id=allegato_id, allegato_nome=allegato_nome,
        allegato_content_type=allegato_ct,
    )
    await db.chat.insert_one(msg.model_dump())
    # auto-traccia nel diario se l'interlocutore è un cliente
    cliente_user = None
    if dest.get("role") == "cliente" and dest.get("anagrafica_id"):
        cliente_user = dest
    elif user.get("role") == "cliente" and user.get("anagrafica_id"):
        cliente_user = user
    if cliente_user:
        direction = "->" if user.get("role") != "cliente" else "←"
        titolo = f"💬 Chat {direction} {cliente_user.get('name')}"
        desc = txt[:300] + (" [+ allegato]" if allegato_nome else "")
        await log_diario_cliente(cliente_user["anagrafica_id"], "chat", titolo, desc, autore=user)
    return msg.model_dump()


@api.get("/notifiche/sommario")
async def notifiche_sommario(user=Depends(current_user)):
    """Aggrega notifiche per il badge in sidebar."""
    chat_unread = await db.chat.count_documents({"destinatario_id": user["id"], "letto": False})
    res = {"chat": chat_unread, "totale": chat_unread}
    if user["role"] in ("admin", "collaboratore", "dipendente"):
        from datetime import date, timedelta
        oggi = date.today().isoformat()
        in_scad = (date.today() + timedelta(days=15)).isoformat()
        polizze_scad = await db.polizze.count_documents({
            "stato": "attiva", "scadenza": {"$gte": oggi, "$lte": in_scad},
        })
        titoli_scaduti = await db.titoli.count_documents({
            "stato": {"$in": ["da_incassare", "insoluto"]}, "scadenza": {"$lt": oggi},
        })
        email_coda = await db.email.count_documents({"stato": "in_coda"})
        sinistri_aperti = await db.sinistri.count_documents({"stato": "aperto"})
        res["polizze_scadenza"] = polizze_scad
        res["titoli_scaduti"] = titoli_scaduti
        res["email_coda"] = email_coda
        res["sinistri_aperti"] = sinistri_aperti
        res["totale"] = chat_unread + polizze_scad + titoli_scaduti + sinistri_aperti
    return res


@api.get("/search")
async def search_globale(q: str, limit: int = 8, user=Depends(current_user)):
    """Ricerca rapida cross-entità per la barra in alto.
    Cerca su: anagrafiche (nome, CF, P.IVA, email, telefono, cellulare, comune, indirizzo),
    polizze (numero, ramo, prodotto, targa, contraente), sinistri (numero, luogo),
    titoli (numero), compagnie (ragione sociale)."""
    if not q or len(q.strip()) < 2:
        return {"anagrafiche": [], "polizze": [], "sinistri": [], "titoli": [], "compagnie": []}
    qclean = q.strip()
    qrx = {"$regex": qclean, "$options": "i"}
    is_client = user["role"] == "cliente"
    # === ANAGRAFICHE: cerca su molti campi ===
    ana_filter = {"id": user.get("anagrafica_id")} if is_client else {"$or": [
        {"ragione_sociale": qrx}, {"nome": qrx}, {"cognome": qrx},
        {"codice_fiscale": qrx}, {"partita_iva": qrx},
        {"email": qrx}, {"telefono": qrx}, {"cellulare": qrx},
        {"comune": qrx}, {"indirizzo": qrx}, {"professione": qrx},
        {"tags": qrx},
    ]}
    anas = await db.anagrafiche.find(
        ana_filter, {"_id": 0, "id": 1, "ragione_sociale": 1, "codice_fiscale": 1,
                     "comune": 1, "email": 1, "cellulare": 1, "telefono": 1},
    ).limit(limit).to_list(limit)

    # === POLIZZE: numero, targa, ramo, prodotto, oggetto assicurato, note ===
    pol_filter = {"$or": [
        {"numero_polizza": qrx}, {"targa": qrx},
        {"ramo": qrx}, {"prodotto": qrx},
        {"oggetto_assicurato": qrx},
        {"veicolo_marca": qrx}, {"veicolo_modello": qrx},
    ]}
    if is_client:
        pol_filter = {"$and": [pol_filter, {"contraente_id": user.get("anagrafica_id")}]}
    polizze = await db.polizze.find(
        pol_filter, {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "prodotto": 1,
                     "stato": 1, "targa": 1, "contraente_id": 1},
    ).limit(limit).to_list(limit)
    # arricchisci con contraente
    ana_ids = [p.get("contraente_id") for p in polizze if p.get("contraente_id")]
    ana_map = {a["id"]: a["ragione_sociale"] async for a in db.anagrafiche.find(
        {"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    for p in polizze:
        p["contraente_nome"] = ana_map.get(p.get("contraente_id", ""))

    # Aggiungi polizze trovate tramite anagrafica matchata (per nome contraente)
    if anas and not is_client:
        anag_ids_to_search = [a["id"] for a in anas]
        extra_pol = await db.polizze.find(
            {"contraente_id": {"$in": anag_ids_to_search}},
            {"_id": 0, "id": 1, "numero_polizza": 1, "ramo": 1, "prodotto": 1, "stato": 1, "targa": 1, "contraente_id": 1},
        ).limit(limit).to_list(limit)
        existing_pol_ids = {p["id"] for p in polizze}
        for ep in extra_pol:
            if ep["id"] not in existing_pol_ids:
                ep["contraente_nome"] = ana_map.get(ep.get("contraente_id")) or next(
                    (a["ragione_sociale"] for a in anas if a["id"] == ep.get("contraente_id")), None)
                polizze.append(ep)
        polizze = polizze[:limit * 2]

    # === SINISTRI ===
    sin_filter = {"$or": [{"numero_sinistro": qrx}, {"luogo": qrx}, {"descrizione": qrx}]}
    if is_client:
        sin_filter = {"$and": [sin_filter, {"contraente_id": user.get("anagrafica_id")}]}
    sinistri = await db.sinistri.find(
        sin_filter, {"_id": 0, "id": 1, "numero_sinistro": 1, "polizza_id": 1, "stato": 1, "data_avvenimento": 1},
    ).limit(limit).to_list(limit)

    # === TITOLI ===
    titoli = []
    if not is_client:
        titoli = await db.titoli.find(
            {"$or": [{"numero_titolo": qrx}, {"numero_polizza": qrx}, {"contraente_nome": qrx}]},
            {"_id": 0, "id": 1, "numero_titolo": 1, "numero_polizza": 1, "polizza_id": 1,
             "contraente_nome": 1, "importo_lordo": 1, "stato": 1, "data_scadenza": 1},
        ).limit(limit).to_list(limit)

    # === COMPAGNIE ===
    compagnie = []
    if not is_client:
        compagnie = await db.compagnie.find(
            {"$or": [{"ragione_sociale": qrx}, {"codice": qrx}]},
            {"_id": 0, "id": 1, "ragione_sociale": 1, "codice": 1},
        ).limit(limit).to_list(limit)

    return {
        "anagrafiche": anas, "polizze": polizze, "sinistri": sinistri,
        "titoli": titoli, "compagnie": compagnie,
    }


@api.get("/chat/unread")
async def chat_unread(user=Depends(current_user)):
    n = await db.chat.count_documents({"destinatario_id": user["id"], "letto": False})
    return {"unread": n}


# ============================================================
# CORSI
# ============================================================
@api.get("/corsi")
async def list_corsi(user=Depends(current_user)):
    flt: dict = {"pubblicato": True}
    if user["role"] != "admin":
        flt["$or"] = [
            {"visibile_ruoli": user["role"]},
            {"visibile_utenti": user["id"]},
        ]
    items = await db.corsi.find(flt, {"_id": 0}).sort("created_at", -1).to_list(500)
    # arricchisci con progresso utente
    ids = [c["id"] for c in items]
    progressi = {p["corso_id"]: p async for p in db.progressi_corso.find(
        {"corso_id": {"$in": ids}, "utente_id": user["id"]}, {"_id": 0})}
    for c in items:
        c["progresso"] = progressi.get(c["id"])
    return items


@api.get("/corsi/{cid}")
async def get_corso(cid: str, user=Depends(current_user)):
    doc = await db.corsi.find_one({"id": cid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Corso non trovato")
    # check visibility
    if user["role"] != "admin":
        ok = user["role"] in (doc.get("visibile_ruoli") or []) or user["id"] in (doc.get("visibile_utenti") or [])
        if not ok:
            raise HTTPException(403, "Corso non visibile")
    doc["progresso"] = await db.progressi_corso.find_one(
        {"corso_id": cid, "utente_id": user["id"]}, {"_id": 0}
    )
    return doc


@api.post("/corsi", status_code=201)
async def create_corso(body: dict, user=Depends(require_user("admin", "collaboratore"))):
    body["autore_id"] = user["id"]
    obj = Corso(**body)
    await db.corsi.insert_one(obj.model_dump())
    await log_attivita(user, "create", "corso", obj.id, obj.titolo)
    return obj.model_dump()


@api.put("/corsi/{cid}")
async def update_corso(cid: str, body: dict, user=Depends(require_user("admin", "collaboratore"))):
    body["updated_at"] = _now_iso()
    res = await db.corsi.update_one({"id": cid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Corso non trovato")
    return strip_mongo_id(await db.corsi.find_one({"id": cid}, {"_id": 0}))


@api.delete("/corsi/{cid}")
async def delete_corso(cid: str, user=Depends(require_user("admin"))):
    await db.corsi.delete_one({"id": cid})
    await db.progressi_corso.delete_many({"corso_id": cid})
    return {"ok": True}


@api.post("/corsi/{cid}/progresso")
async def upsert_progresso(cid: str, body: dict, user=Depends(current_user)):
    """body: {secondi_visti, durata_totale_sec, ultima_posizione_sec}"""
    pos = int(body.get("ultima_posizione_sec", 0))
    dur = int(body.get("durata_totale_sec", 0))
    seen = int(body.get("secondi_visti", pos))
    perc = (seen / dur * 100.0) if dur > 0 else 0.0
    completato = perc >= 95.0
    existing = await db.progressi_corso.find_one({"corso_id": cid, "utente_id": user["id"]}, {"_id": 0})
    if existing:
        new_seen = max(existing.get("secondi_visti", 0), seen)
        new_dur = max(existing.get("durata_totale_sec", 0), dur)
        new_perc = (new_seen / new_dur * 100.0) if new_dur > 0 else 0.0
        await db.progressi_corso.update_one(
            {"id": existing["id"]},
            {"$set": {
                "secondi_visti": new_seen, "durata_totale_sec": new_dur,
                "percentuale": round(new_perc, 1), "completato": new_perc >= 95.0,
                "ultima_posizione_sec": pos,
                "ultima_visualizzazione": _now_iso(), "updated_at": _now_iso(),
            }},
        )
        return await db.progressi_corso.find_one({"id": existing["id"]}, {"_id": 0})
    obj = ProgressoCorso(
        corso_id=cid, utente_id=user["id"], secondi_visti=seen,
        durata_totale_sec=dur, percentuale=round(perc, 1),
        completato=completato, ultima_posizione_sec=pos,
    )
    await db.progressi_corso.insert_one(obj.model_dump())
    return obj.model_dump()


@api.get("/corsi/{cid}/progressi")
async def list_progressi(cid: str, user=Depends(require_user("admin", "collaboratore"))):
    """Stato di avanzamento di tutti gli utenti su un corso."""
    items = await db.progressi_corso.find({"corso_id": cid}, {"_id": 0}).to_list(2000)
    uids = list({p["utente_id"] for p in items})
    users = {u["id"]: u async for u in db.users.find(
        {"id": {"$in": uids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1})}
    for p in items:
        u = users.get(p["utente_id"], {})
        p["utente_nome"] = u.get("name")
        p["utente_email"] = u.get("email")
        p["utente_ruolo"] = u.get("role")
    return items


# ============================================================
# GEOLOCALIZZAZIONE
# ============================================================
@api.get("/geo/suggest")
async def geo_suggest(
    q: str,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """Autocomplete indirizzo via Nominatim (italiano). Min 3 char."""
    return await geocoder_svc.cerca_suggerimenti(q or "", paese="it", limit=6)


@api.get("/geo/anagrafiche")
async def geo_anagrafiche(user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Restituisce solo anagrafiche con lat/lng valide."""
    items = await db.anagrafiche.find(
        {"lat": {"$ne": None}, "lng": {"$ne": None}},
        {"_id": 0, "id": 1, "ragione_sociale": 1, "comune": 1, "provincia": 1,
         "indirizzo": 1, "lat": 1, "lng": 1, "telefono": 1, "cellulare": 1, "email": 1,
         "tipo_persona": 1, "tags": 1, "categoria": 1},
    ).to_list(5000)
    # Marca clienti vs prospect: ha almeno una polizza attiva?
    ids = [a["id"] for a in items]
    if ids:
        with_pol = set()
        async for p in db.polizze.find(
            {"contraente_id": {"$in": ids}, "stato": {"$in": ["attiva", "in_emissione"]}},
            {"_id": 0, "contraente_id": 1},
        ):
            with_pol.add(p["contraente_id"])
        for a in items:
            a["is_cliente"] = a["id"] in with_pol
    return items


@api.post("/geo/anagrafiche/{aid}/geocode")
async def geocode_anagrafica(aid: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Geocoding gratuito con Nominatim (OpenStreetMap)."""
    import requests as _rq
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    parts = [ana.get("indirizzo"), ana.get("cap"), ana.get("comune"), ana.get("provincia"), ana.get("nazione") or "Italia"]
    q = ", ".join([p for p in parts if p])
    if not q:
        raise HTTPException(400, "Indirizzo insufficiente per geocoding")
    try:
        r = _rq.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": q, "format": "json", "limit": 1, "countrycodes": "it"},
            headers={"User-Agent": "AssicuraApp/1.0"}, timeout=15,
        )
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        raise HTTPException(503, f"Geocoding fallito: {e}")
    if not data:
        return {"found": False, "query": q}
    lat = float(data[0]["lat"]); lng = float(data[0]["lon"])
    await db.anagrafiche.update_one(
        {"id": aid},
        {"$set": {"lat": lat, "lng": lng, "indirizzo_geocoded": data[0].get("display_name"),
                  "updated_at": _now_iso()}},
    )
    return {"found": True, "lat": lat, "lng": lng, "address": data[0].get("display_name")}


# ============================================================
# STAMPE PDF
# ============================================================
def _pdf_response(pdf_bytes: bytes, filename: str):
    return StreamingResponse(
        _io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@api.get("/stampa/anagrafiche")
async def stampa_anagrafiche(q: Optional[str] = None, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    flt = {}
    if q:
        flt["$or"] = [
            {"ragione_sociale": {"$regex": q, "$options": "i"}},
            {"codice_fiscale": {"$regex": q, "$options": "i"}},
        ]
    items = await db.anagrafiche.find(flt, {"_id": 0}).sort("ragione_sociale", 1).to_list(5000)
    headers = ["Ragione sociale", "CF / P.IVA", "Tipo", "Comune", "Email", "Telefono", "Cellulare"]
    rows = [[a["ragione_sociale"], a.get("codice_fiscale") or a.get("partita_iva") or "",
             "PG" if a.get("tipo") == "persona_giuridica" else "PF",
             f"{a.get('comune','')} ({a.get('provincia','')})" if a.get("comune") else "",
             a.get("email") or "", a.get("telefono") or "", a.get("cellulare") or ""] for a in items]
    pdf = pdf_report.stampa_elenco(
        "Elenco Anagrafiche", f"{len(items)} clienti", headers, rows,
        col_widths_mm=[60, 35, 12, 45, 45, 35, 35],
        filtri_attivi={"Ricerca": q} if q else None,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "anagrafiche.pdf")


@api.get("/stampa/polizze")
async def stampa_polizze(
    q: Optional[str] = None,
    stato: Optional[str] = None, compagnia_id: Optional[str] = None,
    ramo: Optional[str] = None, prodotto: Optional[str] = None,
    collaboratore_id: Optional[str] = None, contraente_id: Optional[str] = None,
    dal: Optional[str] = None, al: Optional[str] = None,
    in_scadenza_giorni: Optional[int] = None,
    scadenza_oltre_giorni: Optional[int] = None,
    scadute_oggi: Optional[bool] = None,
    scadute_da_min: Optional[int] = None,
    scadute_da_max: Optional[int] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    items = await list_polizze(
        q=q, stato=stato, ramo=ramo, prodotto=prodotto,
        contraente_id=contraente_id, compagnia_id=compagnia_id,
        collaboratore_id=collaboratore_id, dal=dal, al=al,
        in_scadenza_giorni=in_scadenza_giorni,
        scadenza_oltre_giorni=scadenza_oltre_giorni,
        scadute_oggi=scadute_oggi,
        scadute_da_min=scadute_da_min, scadute_da_max=scadute_da_max,
        limit=10000, user=user,
    )
    headers = ["N. polizza", "Contraente", "Compagnia", "Collaboratore", "Ramo", "Stato", "Effetto", "Scadenza", "Premio €", "Provv. €"]
    rows = [[p.get("numero_polizza", ""),
             p.get("contraente_nome", ""),
             p.get("compagnia_nome", ""),
             p.get("collaboratore_nome", ""),
             p.get("ramo", ""), p.get("stato", ""),
             p.get("effetto", ""), p.get("scadenza", ""),
             p.get("premio_lordo", 0), p.get("provvigioni", 0)] for p in items]
    pdf = pdf_report.stampa_elenco(
        "Elenco Polizze", f"{len(items)} polizze", headers, rows,
        col_widths_mm=[28, 45, 40, 30, 22, 22, 22, 22, 22, 22],
        filtri_attivi={"Stato": stato, "Ramo": ramo, "Compagnia": compagnia_id,
                       "Collaboratore": collaboratore_id, "Dal": dal, "Al": al},
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "polizze.pdf")


@api.get("/stampa/titoli")
async def stampa_titoli(
    stato: Optional[str] = None, in_scadenza_giorni: Optional[int] = None,
    coperti_non_pagati: Optional[bool] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    # riusa la logica list_titoli replicando i filtri base
    flt: dict = {}
    if stato: flt["stato"] = stato
    from datetime import date, timedelta
    today = date.today().isoformat()
    if in_scadenza_giorni is not None:
        limite = (date.today() + timedelta(days=int(in_scadenza_giorni))).isoformat()
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        flt["scadenza"] = {"$gte": today, "$lte": limite}
    if coperti_non_pagati:
        flt["stato"] = {"$in": ["da_incassare", "insoluto"]}
        flt["effetto"] = {"$lte": today}
        flt["scadenza"] = {"$gte": today}
    items = await db.titoli.find(flt, {"_id": 0}).sort("scadenza", 1).to_list(5000)
    pol_ids = list({t["polizza_id"] for t in items})
    pols = {p["id"]: p async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0})}
    ana_ids = list({p.get("contraente_id") for p in pols.values() if p.get("contraente_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    headers = ["Polizza", "Contraente", "Ramo", "Tipo", "Effetto", "Scadenza", "Stato", "Lordo €", "Provv. €"]
    rows = []
    for t in items:
        p = pols.get(t["polizza_id"], {})
        rows.append([
            p.get("numero_polizza", ""),
            anas.get(p.get("contraente_id",""), {}).get("ragione_sociale", ""),
            p.get("ramo", ""), t.get("tipo", ""),
            t.get("effetto", ""), t.get("scadenza", ""), t.get("stato", ""),
            t.get("importo_lordo", 0), t.get("provvigioni", 0),
        ])
    pdf = pdf_report.stampa_elenco(
        "Elenco Titoli",
        f"{len(items)} titoli" + (" - in scadenza" if in_scadenza_giorni else "") + (" - coperti non pagati" if coperti_non_pagati else ""),
        headers, rows,
        col_widths_mm=[32, 55, 25, 22, 22, 22, 25, 25, 25],
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "titoli.pdf")


@api.get("/stampa/titoli/sospesi")
async def stampa_titoli_sospesi(
    collaboratore_id: Optional[str] = None,
    user=Depends(require_user("admin", "collaboratore", "dipendente")),
):
    """PDF dell'elenco Sospesi Anticipi con data di stampa odierna in testata."""
    from datetime import date as _date
    flt = {"titolo_coperto": True, "stato": {"$in": ["da_incassare", "insoluto"]}}
    items = await db.titoli.find(flt, {"_id": 0}).sort("data_copertura", 1).to_list(5000)
    pol_ids = list({t["polizza_id"] for t in items if t.get("polizza_id")})
    pols = {p["id"]: p async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0})}
    ana_ids = list({p.get("contraente_id") for p in pols.values() if p.get("contraente_id")})
    anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0})}
    collab_map = {}
    cids = list({(p.get("collaboratore_id") or t.get("collaboratore_id"))
                 for t, p in [(t, pols.get(t["polizza_id"], {})) for t in items]
                 if (p.get("collaboratore_id") or t.get("collaboratore_id"))})
    if cids:
        async for u in db.users.find({"id": {"$in": cids}}, {"_id": 0, "id": 1, "name": 1}):
            collab_map[u["id"]] = u["name"]

    rows = []
    totale = 0.0
    for t in items:
        p = pols.get(t["polizza_id"], {})
        a = anas.get(p.get("contraente_id", ""), {})
        cid = p.get("collaboratore_id") or t.get("collaboratore_id")
        if collaboratore_id and cid != collaboratore_id:
            continue
        gg = _giorni_da_oggi(t.get("data_copertura"))
        imp = float(t.get("importo_lordo") or 0)
        totale += imp
        rows.append([
            a.get("ragione_sociale", ""),
            collab_map.get(cid, "—") if cid else "—",
            p.get("numero_polizza", ""),
            p.get("ramo", "") + (f" {p.get('targa')}" if p.get("targa") else ""),
            (t.get("data_copertura") or "")[:10],
            f"{gg} gg" if gg is not None else "—",
            (p.get("scadenza") or "")[:10],
            imp,
        ])
    rows.append(["", "", "", "", "", "", "TOTALE", totale])

    headers = ["Cliente", "Collaboratore", "Polizza", "Ramo / Targa", "Coperto il", "Anticipo", "Scad. polizza", "Importo €"]
    sottotitolo = f"{len(rows) - 1} titoli sospesi — Data stampa: {_date.today().strftime('%d/%m/%Y')}"
    pdf = pdf_report.stampa_elenco(
        "Sospesi Anticipi", sottotitolo, headers, rows,
        col_widths_mm=[45, 35, 28, 30, 22, 18, 22, 25],
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "sospesi_anticipi.pdf")


@api.get("/stampa/sinistri")
async def stampa_sinistri(stato: Optional[str] = None,
                          user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    flt = await visibility_filter(user)
    if stato: flt["stato"] = stato
    items = await db.sinistri.find(flt, {"_id": 0}).sort("data_avvenimento", -1).to_list(5000)
    pol_ids = list({s["polizza_id"] for s in items if s.get("polizza_id")})
    pols = {p["id"]: p async for p in db.polizze.find({"id": {"$in": pol_ids}}, {"_id": 0, "id": 1, "numero_polizza": 1, "contraente_id": 1})}
    ana_ids = list({p.get("contraente_id") for p in pols.values()})
    anas = {a["id"]: a async for a in db.anagrafiche.find({"id": {"$in": ana_ids}}, {"_id": 0, "id": 1, "ragione_sociale": 1})}
    headers = ["N. sinistro", "Polizza", "Contraente", "Avvenimento", "Denuncia", "Stato", "Riserva €", "Liquidato €"]
    rows = []
    for s in items:
        p = pols.get(s.get("polizza_id", ""), {})
        rows.append([s["numero_sinistro"], p.get("numero_polizza", ""),
                     anas.get(p.get("contraente_id", ""), {}).get("ragione_sociale", ""),
                     s.get("data_avvenimento", ""), s.get("data_denuncia", ""),
                     s.get("stato", ""), s.get("riserva", 0), s.get("liquidazione", 0)])
    pdf = pdf_report.stampa_elenco(
        "Elenco Sinistri", f"{len(items)} sinistri", headers, rows,
        col_widths_mm=[30, 32, 55, 25, 25, 25, 28, 28],
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "sinistri.pdf")


@api.get("/stampa/sinistro/{sid}")
async def stampa_singolo_sinistro(sid: str, user=Depends(current_user)):
    """PDF scheda singolo sinistro: dati generali + soggetti + anagrafiche + note + liquidazione."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from io import BytesIO
    s = await db.sinistri.find_one({"id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Sinistro non trovato")
    # arricchimento
    pol = await db.polizze.find_one({"id": s.get("polizza_id")}, {"_id": 0}) or {}
    ana = await db.anagrafiche.find_one({"id": s.get("contraente_id")}, {"_id": 0}) or {}
    comp = await db.compagnie.find_one({"id": s.get("compagnia_id")}, {"_id": 0}) or {}
    coll = await db.users.find_one({"id": s.get("collaboratore_id")}, {"_id": 0}) if s.get("collaboratore_id") else None

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    y = H - 18 * mm
    intest = await _intestazione_pdf()
    if intest.get("logo_path"):
        try:
            c.drawImage(intest["logo_path"], 15 * mm, y - 12 * mm, height=20 * mm, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    c.setFont("Helvetica-Bold", 16)
    c.drawString(60 * mm, y, "SCHEDA SINISTRO")
    c.setFont("Helvetica", 9)
    c.drawString(60 * mm, y - 6 * mm, intest.get("intestazione", "") or "")
    y -= 28 * mm

    def line(label, value, dy=5.5):
        nonlocal y
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(15 * mm, y, label + ":")
        c.setFont("Helvetica", 9)
        c.drawString(55 * mm, y, str(value or "—")[:120])
        y -= dy * mm

    c.setFillColor(colors.HexColor("#1e40af"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(15 * mm, y, f"Sinistro N. {s.get('numero_sinistro','—')}   ·   Stato: {s.get('stato','—').upper()}")
    c.setFillColor(colors.black)
    y -= 7 * mm

    line("Numero interno", s.get("numero_interno"))
    line("Anno / Tipologia", f"{s.get('anno') or ''}  ·  {s.get('tipologia_sinistro') or '—'}")
    line("Polizza", f"{pol.get('numero_polizza','—')} · {pol.get('ramo','')} · {pol.get('prodotto','') or ''}")
    line("Contraente", ana.get("ragione_sociale") or f"{ana.get('cognome','')} {ana.get('nome','')}".strip())
    line("Compagnia", comp.get("ragione_sociale"))
    if coll:
        line("Collaboratore", coll.get("name"))
    line("Data avvenimento", s.get("data_avvenimento"))
    line("Data denuncia", s.get("data_denuncia"))
    line("Luogo", s.get("luogo"))
    line("Targa", pol.get("targa") or s.get("targa"))
    line("Riserva €", f"{float(s.get('riserva') or 0):,.2f}")
    line("Liquidazione €", f"{float(s.get('liquidazione') or 0):,.2f}")

    if s.get("descrizione"):
        y -= 2 * mm
        c.setFont("Helvetica-Bold", 9); c.drawString(15 * mm, y, "Descrizione:")
        y -= 5 * mm
        c.setFont("Helvetica", 8.5)
        for chunk in [s["descrizione"][i:i + 110] for i in range(0, len(s["descrizione"]), 110)][:8]:
            c.drawString(15 * mm, y, chunk); y -= 4 * mm

    # Soggetti coinvolti
    if s.get("soggetti_coinvolti"):
        y -= 3 * mm
        c.setFont("Helvetica-Bold", 10); c.drawString(15 * mm, y, "SOGGETTI COINVOLTI"); y -= 5 * mm
        c.setFont("Helvetica", 8)
        for sog in s["soggetti_coinvolti"][:10]:
            c.drawString(15 * mm, y,
                f"· {sog.get('soggetto','—')} ({sog.get('ruolo','—')}) · Polizza {sog.get('numero_polizza','—')} · "
                f"Ris. € {sog.get('riserva',0):.2f}  Pag. € {sog.get('pagato',0):.2f}")
            y -= 4.5 * mm

    # Anagrafiche associate
    if s.get("anagrafiche_associate"):
        y -= 3 * mm
        c.setFont("Helvetica-Bold", 10); c.drawString(15 * mm, y, "ANAGRAFICHE ASSOCIATE"); y -= 5 * mm
        c.setFont("Helvetica", 8)
        for ag in s["anagrafiche_associate"][:8]:
            c.drawString(15 * mm, y,
                f"· {ag.get('nome','—')} — {ag.get('tipo','—')} — {ag.get('telefono','') or ''} {ag.get('email','') or ''}")
            y -= 4.5 * mm

    # Note
    if s.get("note"):
        y -= 3 * mm
        c.setFont("Helvetica-Bold", 10); c.drawString(15 * mm, y, "NOTE"); y -= 5 * mm
        c.setFont("Helvetica", 8)
        for n in s["note"][:8]:
            c.drawString(15 * mm, y, f"[{n.get('data','—')}] {n.get('operatore','')}: {(n.get('descrizione') or '')[:110]}")
            y -= 4.5 * mm

    # Liquidazione dettaglio
    ld = s.get("liquidazione_dettaglio") or {}
    if ld:
        y -= 3 * mm
        c.setFont("Helvetica-Bold", 10); c.drawString(15 * mm, y, "LIQUIDAZIONE"); y -= 5 * mm
        c.setFont("Helvetica", 8.5)
        for k in ("tipo_definizione", "data_definizione", "franchigia", "scoperto",
                  "importo_denunciato", "riserva_corrente", "data_riserva", "data_prescrizione"):
            if ld.get(k) not in (None, "", 0):
                c.drawString(15 * mm, y, f"{k.replace('_', ' ').capitalize()}: {ld[k]}")
                y -= 4.5 * mm

    c.setFont("Helvetica-Oblique", 7); c.setFillColor(colors.grey)
    c.drawString(15 * mm, 15 * mm, f"Stampato il {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.showPage(); c.save()
    return _pdf_response(buf.getvalue(), f"sinistro_{s.get('numero_sinistro','x')}.pdf")


@api.get("/stampa/sinistro/{sid}/cid")
async def stampa_cid(sid: str, user=Depends(current_user)):
    """PDF della Costatazione Amichevole compilata (RC Auto)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from io import BytesIO
    s = await db.sinistri.find_one({"id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Sinistro non trovato")
    cid = s.get("costatazione_amichevole") or {}
    pol = await db.polizze.find_one({"id": s.get("polizza_id")}, {"_id": 0}) or {}
    ana = await db.anagrafiche.find_one({"id": s.get("contraente_id")}, {"_id": 0}) or {}

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    # Header
    c.setFillColor(colors.HexColor("#0c4a6e"))
    c.rect(0, H - 18 * mm, W, 18 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(15 * mm, H - 11 * mm, "Costatazione Amichevole di incidente — Denuncia di sinistro")
    c.setFont("Helvetica", 7.5)
    c.drawString(15 * mm, H - 15 * mm, "(art. 143 D.Lgs. n. 209 del 2005 - Codice delle assicurazioni private)")
    c.setFillColor(colors.black)

    y = H - 25 * mm

    def fld(x, label, val, w=70):
        c.setFont("Helvetica-Bold", 7); c.drawString(x * mm, y, label)
        c.setFont("Helvetica", 8); c.drawString((x + 25) * mm, y, str(val or "—")[:w])

    # Sezione 1-5: data, luogo, feriti, danni
    c.setFont("Helvetica-Bold", 9)
    c.drawString(15 * mm, y, "1. Data incidente:")
    c.setFont("Helvetica", 9)
    c.drawString(55 * mm, y, str(cid.get("data_incidente") or s.get("data_avvenimento") or "—"))
    c.setFont("Helvetica-Bold", 9); c.drawString(100 * mm, y, "Ora:")
    c.setFont("Helvetica", 9); c.drawString(115 * mm, y, str(cid.get("ora", "—")))
    c.setFont("Helvetica-Bold", 9); c.drawString(140 * mm, y, "3. Feriti:")
    c.setFont("Helvetica", 9); c.drawString(160 * mm, y, "SI" if cid.get("feriti") else "NO")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 9); c.drawString(15 * mm, y, "2. Luogo:")
    c.setFont("Helvetica", 9); c.drawString(35 * mm, y, str(cid.get("luogo") or s.get("luogo") or "—")[:90])
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 9); c.drawString(15 * mm, y, "4. Danni materiali ad altri veicoli:")
    c.setFont("Helvetica", 9); c.drawString(80 * mm, y, "SI" if cid.get("danni_altri_veicoli") else "NO")
    c.setFont("Helvetica-Bold", 9); c.drawString(110 * mm, y, "Oggetti diversi:")
    c.setFont("Helvetica", 9); c.drawString(140 * mm, y, "SI" if cid.get("danni_oggetti_diversi") else "NO")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 9); c.drawString(15 * mm, y, "5. Testimoni:")
    c.setFont("Helvetica", 9); c.drawString(40 * mm, y, str(cid.get("testimoni") or "—")[:90])
    y -= 8 * mm

    # Veicoli A e B affiancati
    veicolo_a = cid.get("veicolo_a") or {}
    veicolo_b = cid.get("veicolo_b") or {}
    # Precompila A da polizza/anagrafica se mancano
    if not veicolo_a.get("contraente_cognome"):
        veicolo_a["contraente_cognome"] = ana.get("cognome") or ana.get("ragione_sociale", "")
        veicolo_a["contraente_nome"] = ana.get("nome", "")
        veicolo_a["codice_fiscale"] = ana.get("codice_fiscale") or ana.get("partita_iva", "")
        veicolo_a["indirizzo"] = ana.get("indirizzo", "")
        veicolo_a["targa"] = pol.get("targa", "")
        veicolo_a["compagnia"] = (await db.compagnie.find_one({"id": s.get("compagnia_id")}, {"_id": 0, "ragione_sociale": 1}) or {}).get("ragione_sociale", "")
        veicolo_a["numero_polizza"] = pol.get("numero_polizza", "")

    # Colonne A (sinistra) e B (destra)
    col_a_x, col_b_x = 15, 110
    c.setFillColor(colors.HexColor("#bae6fd"))
    c.rect((col_a_x - 1) * mm, y - 2 * mm, 90 * mm, 6 * mm, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#fde68a"))
    c.rect((col_b_x - 1) * mm, y - 2 * mm, 90 * mm, 6 * mm, fill=1, stroke=0)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(col_a_x * mm, y, "VEICOLO A")
    c.drawString(col_b_x * mm, y, "VEICOLO B")
    y -= 8 * mm

    def block(x, v):
        nonlocal y
        sy = y
        c.setFont("Helvetica-Bold", 7.5)
        rows = [
            ("Cognome", v.get("contraente_cognome", "")),
            ("Nome", v.get("contraente_nome", "")),
            ("CF / P.IVA", v.get("codice_fiscale", "")),
            ("Indirizzo", v.get("indirizzo", "")),
            ("Telefono / Email", v.get("contatto", "")),
            ("Marca/Tipo", v.get("marca_tipo", "")),
            ("Targa", v.get("targa", "")),
            ("Compagnia", v.get("compagnia", "")),
            ("N. polizza", v.get("numero_polizza", "")),
            ("Conducente", v.get("conducente_nome", "")),
            ("Patente N.", v.get("patente", "")),
            ("Punto urto", v.get("punto_urto", "")),
            ("Danni visibili", v.get("danni_visibili", "")),
            ("Osservazioni", v.get("osservazioni", "")),
        ]
        for label, val in rows:
            c.setFont("Helvetica-Bold", 7); c.drawString(x * mm, sy, label + ":")
            c.setFont("Helvetica", 7.5); c.drawString((x + 22) * mm, sy, str(val or "—")[:50])
            sy -= 4.2 * mm
        return sy
    yA = block(col_a_x, veicolo_a)
    yB = block(col_b_x, veicolo_b)
    y = min(yA, yB) - 3 * mm

    # Circostanze (17) — checkbox compatte
    circ_labels = [
        "in sosta / fermato", "ripartiva dopo sosta / aprendo portiera",
        "stava parcheggiando", "usciva da parcheggio/luogo privato",
        "entrava in parcheggio/luogo privato", "si immetteva in piazza/rotatoria",
        "circolava in piazza/rotatoria", "tamponava nella stessa fila",
        "procedeva stessa direzione su fila diversa", "cambiava fila",
        "sorpassava", "girava a destra", "girava a sinistra",
        "retrocedeva", "invadeva sede stradale opposta",
        "proveniva da destra", "non osservava precedenza/semaforo rosso",
    ]
    circ_a = cid.get("circostanze_a") or []
    circ_b = cid.get("circostanze_b") or []
    c.setFont("Helvetica-Bold", 9)
    c.drawString(15 * mm, y, "12. Circostanze dell'incidente (crocette utili)")
    y -= 5 * mm
    c.setFont("Helvetica", 7.5)
    for i, lab in enumerate(circ_labels):
        n = i + 1
        ck_a = "X" if n in circ_a else " "
        ck_b = "X" if n in circ_b else " "
        c.drawString(15 * mm, y, f"[{ck_a}] {n}.  {lab}  [{ck_b}]")
        y -= 4 * mm
        if y < 25 * mm:
            c.showPage(); y = H - 20 * mm

    # Footer / firme
    if y < 35 * mm:
        c.showPage(); y = H - 25 * mm
    y -= 4 * mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(15 * mm, y, "15. Firma del conducente A: ___________________________")
    c.drawString(110 * mm, y, "Firma del conducente B: ___________________________")
    y -= 8 * mm
    c.setFont("Helvetica-Oblique", 7); c.setFillColor(colors.grey)
    c.drawString(15 * mm, 12 * mm, f"Documento generato il {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sinistro N. {s.get('numero_sinistro','')}")

    c.showPage(); c.save()
    return _pdf_response(buf.getvalue(), f"cid_{s.get('numero_sinistro','x')}.pdf")


@api.get("/stampa/prima-nota")
async def stampa_prima_nota(dal: Optional[str] = None, al: Optional[str] = None,
                            user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    flt: dict = {}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al
        flt["data_movimento"] = cond
    items = await db.movimenti.find(flt, {"_id": 0}).sort("data_movimento", 1).to_list(5000)
    headers = ["Data", "Tipo", "Categoria", "Descrizione", "Mezzo", "Entrata €", "Uscita €"]
    rows = []
    for m in items:
        rows.append([m.get("data_movimento", ""), m.get("tipo", ""), m.get("categoria", ""),
                     (m.get("descrizione") or "")[:60], m.get("mezzo_pagamento") or "",
                     m["importo"] if m["tipo"] == "entrata" else "",
                     m["importo"] if m["tipo"] == "uscita" else ""])
    sub = f"Periodo: {dal or '—'} -> {al or '—'}"
    pdf = pdf_report.stampa_elenco(
        "Prima Nota", sub, headers, rows,
        col_widths_mm=[25, 18, 35, 90, 35, 28, 28],
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, "prima_nota.pdf")


@api.get("/stampa/brogliaccio")
async def stampa_brogliaccio(data: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    """Brogliaccio del giorno (stile Schiantarelli)."""
    conti = await db.conti_cassa.find({"attivo": True}, {"_id": 0}).sort("ordine", 1).to_list(50)
    pdf = await brog.genera_brogliaccio_pdf(db, data, conti)
    return _pdf_response(pdf, f"brogliaccio_{data}.pdf")


@api.get("/stampa/estratto-conto/{aid}")
async def stampa_estratto_conto(aid: str, user=Depends(current_user)):
    if user["role"] == "cliente" and user.get("anagrafica_id") != aid:
        raise HTTPException(403, "Permesso negato")
    ana = await db.anagrafiche.find_one({"id": aid}, {"_id": 0})
    if not ana:
        raise HTTPException(404, "Anagrafica non trovata")
    movs = await db.movimenti.find({"anagrafica_id": aid}, {"_id": 0}).sort("data_movimento", 1).to_list(5000)
    saldo = 0.0
    headers = ["Data", "Descrizione", "Dare €", "Avere €", "Saldo €"]
    rows = []
    for m in movs:
        delta = m["importo"] if m["tipo"] == "entrata" else -m["importo"]
        saldo += delta
        rows.append([m["data_movimento"], (m.get("descrizione") or "")[:80],
                     m["importo"] if m["tipo"] == "entrata" else "",
                     m["importo"] if m["tipo"] == "uscita" else "",
                     round(saldo, 2)])
    rows.append(["", "SALDO FINALE", "", "", round(saldo, 2)])
    pdf = pdf_report.stampa_elenco(
        f"Estratto conto - {ana['ragione_sociale']}",
        f"CF/P.IVA: {ana.get('codice_fiscale') or ana.get('partita_iva') or '—'}",
        headers, rows,
        col_widths_mm=[25, 130, 28, 28, 32], landscape_mode=False,
        **(await _intestazione_pdf()),
    )
    return _pdf_response(pdf, f"estratto_{ana['ragione_sociale']}.pdf")


# ============================================================
# HEALTH
# ============================================================
@api.get("/")
async def root():
    return {"app": "Programma Assicurativo", "status": "ok"}


# ============================================================
# CALENDARIO — Eventi agenzia / per operatore
# ============================================================
@api.get("/calendario")
async def list_eventi(
    dal: Optional[str] = None,
    al: Optional[str] = None,
    operatore_id: Optional[str] = None,
    tipo: Optional[str] = None,
    includi_scadenze: bool = True,
    user=Depends(current_user),
):
    """Restituisce eventi calendario. Se includi_scadenze=True aggiunge auto-eventi
    sintetici per scadenze polizze e titoli."""
    flt: dict = {}
    if dal or al:
        cond = {}
        if dal: cond["$gte"] = dal
        if al: cond["$lte"] = al + "T23:59:59"
        flt["inizio"] = cond
    if operatore_id:
        flt["$or"] = [{"operatore_id": operatore_id},
                      {"partecipanti_user_ids": operatore_id}]
    if tipo:
        flt["tipo"] = tipo
    if user["role"] == "cliente":
        flt["anagrafica_id"] = user.get("anagrafica_id")
    items = await db.calendario.find(flt, {"_id": 0}).sort("inizio", 1).to_list(2000)

    # arricchimento nome operatore
    op_ids = {e.get("operatore_id") for e in items if e.get("operatore_id")}
    if op_ids:
        op_map = {u["id"]: u["name"] async for u in
                  db.users.find({"id": {"$in": list(op_ids)}}, {"_id": 0, "id": 1, "name": 1})}
        for e in items:
            if e.get("operatore_id"):
                e["operatore_nome"] = op_map.get(e["operatore_id"])

    # scadenze auto
    if includi_scadenze and user["role"] != "cliente":
        scad_flt: dict = {"stato": "attiva"}
        if dal or al:
            cond = {}
            if dal: cond["$gte"] = dal
            if al: cond["$lte"] = al
            scad_flt["scadenza"] = cond
        async for p in db.polizze.find(scad_flt, {"_id": 0, "id": 1, "numero_polizza": 1, "scadenza": 1, "collaboratore_id": 1, "contraente_id": 1, "ramo": 1}):
            if operatore_id and p.get("collaboratore_id") != operatore_id:
                continue
            if not p.get("scadenza"):
                continue
            items.append({
                "id": f"_pol_{p['id']}", "titolo": f"Scadenza polizza {p.get('numero_polizza')}",
                "inizio": p.get("scadenza") + "T09:00:00",
                "tutto_il_giorno": True,
                "tipo": "scadenza_polizza",
                "polizza_id": p.get("id"),
                "anagrafica_id": p.get("contraente_id"),
                "operatore_id": p.get("collaboratore_id"),
                "colore": "#dc2626",
                "descrizione": f"Ramo {p.get('ramo')}",
                "stato": "confermato",
                "_auto": True,
            })
    items.sort(key=lambda e: e.get("inizio", ""))
    return items


@api.post("/calendario", status_code=201)
async def crea_evento(body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body.setdefault("operatore_id", user.get("id"))
    obj = EventoCalendario(**body)
    await db.calendario.insert_one(obj.model_dump())
    await log_attivita(user, "create", "evento", obj.id, obj.titolo)
    return obj.model_dump()


@api.put("/calendario/{eid}")
async def aggiorna_evento(eid: str, body: dict, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    body.pop("id", None)
    body["updated_at"] = _now_iso()
    res = await db.calendario.update_one({"id": eid}, {"$set": body})
    if res.matched_count == 0:
        raise HTTPException(404, "Evento non trovato")
    return await db.calendario.find_one({"id": eid}, {"_id": 0})


@api.delete("/calendario/{eid}")
async def elimina_evento(eid: str, user=Depends(require_user("admin", "collaboratore", "dipendente"))):
    res = await db.calendario.delete_one({"id": eid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Evento non trovato")
    return {"ok": True}


# ============================================================
# MIGRAZIONE — Unifica sconto su singola riga del Brogliaccio
# ============================================================
async def _migrate_unify_sconti() -> dict:
    """One-shot: prende ogni uscita sconto_cliente legata a un titolo, trova la
    corrispondente entrata incasso_premio sullo stesso titolo+data e:
      1) somma l'importo dello sconto in quota_sconto dell'entrata
      2) cancella la riga uscita (non serve più: il dato vive sulla stessa riga)
    Idempotente. Salva un flag in db.migrazioni quando completata.
    """
    flag = await db.migrazioni.find_one({"id": "unify_sconti_on_row_v1"})
    if flag:
        return {"ok": True, "already_done": True}

    sconti = await db.movimenti.find(
        {"tipo": "uscita", "categoria": "sconto_cliente", "titolo_id": {"$ne": None}},
        {"_id": 0},
    ).to_list(100000)
    merged = 0
    for s in sconti:
        entrata = await db.movimenti.find_one({
            "tipo": "entrata",
            "categoria": "incasso_premio",
            "titolo_id": s["titolo_id"],
            "data_movimento": s["data_movimento"],
        }, {"_id": 0})
        if not entrata:
            continue
        new_qs = float(entrata.get("quota_sconto") or 0) + float(s.get("importo") or 0)
        await db.movimenti.update_one(
            {"id": entrata["id"]},
            {"$set": {"quota_sconto": round(new_qs, 2)}},
        )
        await db.movimenti.delete_one({"id": s["id"]})
        merged += 1
    await db.migrazioni.insert_one({
        "id": "unify_sconti_on_row_v1",
        "eseguita_at": _now_iso(),
        "merged": merged,
        "totale_sconti": len(sconti),
    })
    return {"ok": True, "merged": merged, "totale_sconti": len(sconti)}


@api.post("/contabilita/migra-sconti")
async def migra_sconti(user=Depends(require_user("admin"))):
    """Endpoint admin per re-eseguire la migrazione (dopo aver pulito il flag)."""
    return await _migrate_unify_sconti()


@api.post("/admin/migra-prodotti-uuid")
async def migra_polizze_prodotti_uuid(user=Depends(require_user("admin"))):
    """Migrazione idempotente: risolve i campi Polizza.prodotto che contengono
    un UUID di ProdottoLibreria (bug import ANIA) sostituendoli col nome corretto.
    """
    import re
    r = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$")
    # mappa UUID → nome dal catalogo prodotti
    id_to_name = {}
    async for p in db.prodotti.find({}, {"_id": 0, "id": 1, "nome": 1}):
        if p.get("id") and p.get("nome"):
            id_to_name[p["id"]] = p["nome"]
    migrate_count = 0
    unresolved = 0
    async for pol in db.polizze.find(
        {"prodotto": {"$exists": True, "$ne": None}},
        {"_id": 0, "id": 1, "prodotto": 1},
    ):
        v = pol.get("prodotto")
        if not (v and isinstance(v, str) and r.match(v)):
            continue
        nome = id_to_name.get(v)
        if nome:
            await db.polizze.update_one(
                {"id": pol["id"]},
                {"$set": {"prodotto": nome, "updated_at": _now_iso()}},
            )
            migrate_count += 1
        else:
            unresolved += 1
    return {"migrate": migrate_count, "unresolved": unresolved}


@api.post("/admin/ricalcola-titoli-netto")
async def ricalcola_titoli_netto(user=Depends(require_user("admin"))):
    """Ricalcola `importo_netto` dei titoli come `importo_lordo - imposte`
    quando la somma non torna. Fix dei titoli importati con imposte incomplete
    (solo tasse assicurative invece di tasse+imposte+ssn).
    """
    fixed = 0
    checked = 0
    async for t in db.titoli.find(
        {}, {"_id": 0, "id": 1, "importo_lordo": 1, "importo_netto": 1, "imposte": 1},
    ):
        checked += 1
        lordo = float(t.get("importo_lordo") or 0)
        netto = float(t.get("importo_netto") or 0)
        imposte = float(t.get("imposte") or 0)
        if lordo <= 0:
            continue
        # Se la somma netto + imposte non torna al lordo (tolleranza 1 cent),
        # ricalcola netto = lordo - imposte
        if abs((netto + imposte) - lordo) > 0.01:
            new_netto = round(lordo - imposte, 2)
            await db.titoli.update_one(
                {"id": t["id"]},
                {"$set": {"importo_netto": new_netto, "updated_at": _now_iso()}},
            )
            fixed += 1
    return {"checked": checked, "fixed": fixed}


# ----- Mount -----


# ====================== DASHBOARD LINKS (link utili rapidi) ======================












# Modular routers (extracted from server.py)
from routes import dashboard as _dash_router  # noqa: E402
from routes import ocr as _ocr_router  # noqa: E402
from routes import anagrafiche as _anag_router  # noqa: E402
from routes import alert as _alert_router  # noqa: E402
from routes import modelli as _modelli_router  # noqa: E402
from routes import kpi as _kpi_router  # noqa: E402
from routes import permessi as _perm_router  # noqa: E402
from routes import insights as _insights_router  # noqa: E402
from routes import cervello as _cervello_router  # noqa: E402
from routes import marketing_pro as _mktp_router  # noqa: E402
from routes import commerciale as _comm_router  # noqa: E402
from routes import agenzie as _age_router  # noqa: E402
from routes import setup_scambio as _setup_router  # noqa: E402
from routes import documenti_inbox as _docinbox_router  # noqa: E402
from routes import extras_p1p2 as _extras_router  # noqa: E402
from routes import whatsapp_evo as _wa_evo_router  # noqa: E402
from routes import ai_chat as _ai_chat_router  # noqa: E402
from routes import tenants as _tenants_router  # noqa: E402
from routes import super_admin as _super_admin_router  # noqa: E402
from routes import marketplace as _marketplace_router  # noqa: E402
from routes import tickets as _tickets_router  # noqa: E402
from routes import super_admin_logs as _sa_logs_router  # noqa: E402
api.include_router(_dash_router.router)
api.include_router(_ocr_router.router)
api.include_router(_anag_router.router)
api.include_router(_alert_router.router)
api.include_router(_modelli_router.router)
api.include_router(_kpi_router.router)
api.include_router(_perm_router.router)
api.include_router(_insights_router.router)
api.include_router(_cervello_router.router)
api.include_router(_mktp_router.router)
api.include_router(_comm_router.router)
api.include_router(_age_router.router)
api.include_router(_setup_router.router)
api.include_router(_docinbox_router.router)
api.include_router(_extras_router.router)
api.include_router(_wa_evo_router.router)
api.include_router(_ai_chat_router.router)
api.include_router(_tenants_router.router)
api.include_router(_super_admin_router.router)
api.include_router(_marketplace_router.router)
api.include_router(_tickets_router.router)
api.include_router(_sa_logs_router.router)

app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        o.strip() for o in os.environ.get("CORS_ORIGINS", "*").split(",") if o.strip()
    ],
    allow_origin_regex=os.environ.get("CORS_ORIGIN_REGEX") or None,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# Multi-tenant context middleware
# ------------------------------------------------------------
# Estrae l'utente dal JWT (cookie o Bearer) e lo mette in `_current_user_ctx`.
# Il wrapper `TenantAwareCollection` legge il ContextVar e inietta il filtro
# `agenzia_tenant_id` automaticamente in tutte le query DB dei router.
#
# In caso di request anonima (login, webhook WhatsApp, health-check) → nessun
# filtro applicato → il wrapper fa passthrough puro.
# ============================================================
from auth import get_token_from_request, decode_token as _decode_tok  # noqa: E402
from database import raw_db, _current_user_ctx  # noqa: E402


@app.middleware("http")
async def tenant_context_middleware(request, call_next):
    user = None
    token = get_token_from_request(request)
    if token:
        try:
            payload = _decode_tok(token)
            if payload.get("type") == "access":
                user = await raw_db.users.find_one(
                    {"id": payload["sub"]}, {"password_hash": 0, "_id": 0},
                )
        except Exception:
            user = None
    ctx_tok = _current_user_ctx.set(user)
    try:
        response = await call_next(request)
    finally:
        try:
            _current_user_ctx.reset(ctx_tok)
        except (ValueError, LookupError):
            pass
    return response


# ----- Startup -----
@app.on_event("startup")
async def startup():
    # indexes
    await db.users.create_index("email", unique=True)
    await db.anagrafiche.create_index("ragione_sociale")
    await db.anagrafiche.create_index("codice_fiscale")
    await db.anagrafiche.create_index("id_anagrafica_exp")
    await db.polizze.create_index("numero_polizza")
    await db.polizze.create_index("contraente_id")
    await db.polizze.create_index("id_polizza_exp")
    await db.titoli.create_index("polizza_id")
    await db.titoli.create_index("id_titolo_exp")
    await db.sinistri.create_index("polizza_id")
    await db.sinistri.create_index("id_sinistro_exp")
    await db.attivita_log.create_index("created_at")

    # admin seed
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@assicura.it").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin123!")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        admin = UserPublic(email=admin_email, name="Amministratore", role="admin",
                           is_super_admin=True).model_dump()
        admin["password_hash"] = hash_password(admin_password)
        await db.users.insert_one(admin)
        logger.info("Admin seeded: %s", admin_email)
    else:
        updates: dict = {}
        if not verify_password(admin_password, existing.get("password_hash", "")):
            updates["password_hash"] = hash_password(admin_password)
        if not existing.get("is_super_admin"):
            updates["is_super_admin"] = True
        if updates:
            await db.users.update_one({"email": admin_email}, {"$set": updates})

    # Multi-tenant seed + migrazione dati legacy
    from tenant import seed_tenants, migrate_existing_data_to_principale
    await seed_tenants()
    await migrate_existing_data_to_principale()
    await db.tenants.create_index("id", unique=True)
    # Marketplace default modules
    from routes.marketplace import seed_default_moduli
    await seed_default_moduli()

    # Dedicated super admin user (idempotent)
    super_email = "superadmin@assicura.it"
    super_pwd = "Superadmin123!"
    existing_su = await db.users.find_one({"email": super_email})
    if not existing_su:
        su_doc = UserPublic(email=super_email, name="Super Admin", role="admin",
                            is_super_admin=True).model_dump()
        su_doc["password_hash"] = hash_password(super_pwd)
        su_doc["attivo"] = True
        await db.users.insert_one(su_doc)
        logger.info("Super admin seeded: %s", super_email)
    else:
        updates = {}
        if not verify_password(super_pwd, existing_su.get("password_hash", "")):
            updates["password_hash"] = hash_password(super_pwd)
        if not existing_su.get("is_super_admin"):
            updates["is_super_admin"] = True
        if updates:
            await db.users.update_one({"email": super_email}, {"$set": updates})

    # demo users (dipendente + cliente collegato a anagrafica demo)
    from seed_demo import seed_demo
    await seed_demo(db)

    # Seed alert preset rules (idempotent)
    from alert_presets import seed_alert_presets
    await seed_alert_presets()

    # Seed mezzi pagamento (libreria unica) — idempotente
    await _seed_mezzi_pagamento()

    # seed librerie (banche, conti cassa, rami) - solo se vuote
    if await db.conti_cassa.count_documents({}) == 0:
        for i, c in enumerate([
            {"nome": "Cassa Contanti", "tipo": "cassa", "ordine": 1},
            {"nome": "Assegni", "tipo": "cassa", "ordine": 2},
            {"nome": "BPER Sondrio", "tipo": "banca", "ordine": 3},
            {"nome": "Intesa Sanpaolo", "tipo": "banca", "ordine": 4},
            {"nome": "Credit Agricole", "tipo": "banca", "ordine": 5},
            {"nome": "Pagamento Direzione", "tipo": "rid", "ordine": 6},
        ]):
            await db.conti_cassa.insert_one(ContoCassa(**c).model_dump())

    # Seed conti deposito estesi (BPER VILLA, AGOS, ecc.) — idempotente
    await _seed_conti_deposito_estesi()
    # Seed tipi pagamento (combinazione modalità × conto) — solo se vuoto
    await _seed_tipi_pagamento()

    # migrazione idempotente conti cassa: rinomina legacy + disattiva PayPal
    await db.conti_cassa.update_one(
        {"nome": "RID Direzione"}, {"$set": {"nome": "Pagamento Direzione"}},
    )
    await db.conti_cassa.update_one(
        {"nome": "PayPal / Online"}, {"$set": {"attivo": False}},
    )

    if await db.rami.count_documents({}) == 0:
        for r in [
            {"codice": "RCA", "nome": "RC Auto"},
            {"codice": "INCENDIO", "nome": "Incendio"},
            {"codice": "FURTO", "nome": "Furto"},
            {"codice": "VITA", "nome": "Vita"},
            {"codice": "MALATTIA", "nome": "Malattia"},
            {"codice": "INFORTUNI", "nome": "Infortuni"},
            {"codice": "RC_PROF", "nome": "RC Professionale"},
            {"codice": "RC_GEN", "nome": "RC Generale"},
            {"codice": "MULTIRISCHIO", "nome": "Multirischio Casa/Azienda"},
            {"codice": "TUTELA_LEGALE", "nome": "Tutela Legale"},
            {"codice": "ASSISTENZA", "nome": "Assistenza"},
        ]:
            await db.rami.insert_one(RamoLibreria(**r).model_dump())

    if await db.banche.count_documents({}) == 0:
        for b in [
            {"nome": "BPER Banca", "codice_abi": "05387"},
            {"nome": "Intesa Sanpaolo", "codice_abi": "03069"},
            {"nome": "Credit Agricole Italia", "codice_abi": "06230"},
            {"nome": "Unicredit", "codice_abi": "02008"},
        ]:
            await db.banche.insert_one(Banca(**b).model_dump())

    # init object storage (non bloccante)
    try:
        obj_storage.init_storage()
    except Exception as e:
        logger.warning("Object storage non disponibile: %s", e)

    # Scheduler: avvisi scadenze giornalieri (08:00 locale)
    try:
        avvisi_scadenze.start_scheduler(db, hour=8, minute=0)
    except Exception as e:
        logger.warning("Scheduler avvisi scadenze non avviato: %s", e)

    # Seed modelli di default (idempotente)
    try:
        from routes.modelli import seed_default_models
        await seed_default_models()
    except Exception as e:
        logger.warning("Seed modelli fallito (non bloccante): %s", e)

    # Seed profili permessi di default (Admin/Collaboratore/Sola lettura)
    try:
        from routes.permessi import seed_default_profili
        await seed_default_profili()
    except Exception as e:
        logging.warning(f"seed permessi profili skipped: {e}")
    try:
        from routes.librerie import seed_default_tipologie_sinistri
        await seed_default_tipologie_sinistri()
    except Exception as e:
        logger.warning("Seed tipologie sinistri fallito: %s", e)

    # IMAP Poller — avviato automaticamente solo se abilitato in config
    try:
        az_cfg = await db.azienda_config.find_one(
            {}, {"_id": 0, "imap_poller_enabled": 1, "imap_poller_minutes": 1,
                 "imap_host": 1, "imap_user": 1, "imap_password": 1},
        ) or {}
        if (az_cfg.get("imap_poller_enabled")
                and az_cfg.get("imap_host")
                and az_cfg.get("imap_user")
                and az_cfg.get("imap_password")):
            imap_poller.start_scheduler(
                db, minutes=int(az_cfg.get("imap_poller_minutes") or 5),
            )
            logger.info("IMAP poller auto-avviato all'avvio del backend")
    except Exception as e:
        logger.warning("IMAP poller non avviato: %s", e)

    # One-shot: unifica vecchi sconti su singola riga del brogliaccio
    try:
        res = await _migrate_unify_sconti()
        if res.get("merged"):
            logger.info("Migrazione unify_sconti: %s sconti uniti su riga incasso", res["merged"])
    except Exception as e:
        logger.warning("Migrazione unify_sconti fallita (non bloccante): %s", e)

    # One-shot: applica termini_mora_giorni di default ai prodotti esistenti
    try:
        from db_models import default_mora_for_ramo
        async for p in db.prodotti.find({"termini_mora_giorni": {"$exists": False}}, {"_id": 0, "id": 1, "ramo": 1}):
            await db.prodotti.update_one(
                {"id": p["id"]},
                {"$set": {"termini_mora_giorni": default_mora_for_ramo(p.get("ramo"))}},
            )
    except Exception as e:
        logger.warning("Migrazione termini_mora_giorni fallita (non bloccante): %s", e)


@app.on_event("shutdown")
async def shutdown():
    try:
        avvisi_scadenze.stop_scheduler()
    except Exception:
        pass
    try:
        imap_poller.stop_scheduler()
    except Exception:
        pass
    client.close()
